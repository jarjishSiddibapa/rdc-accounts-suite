"""
Persistence layer for the Trial Balance tool's 4 mapping tables — now backed
by the suite's shared MySQL database (see models.py for the 4 ORM models,
which already existed from a prior port attempt and are reused as-is here)
instead of a single location-mapping.xlsx workbook read/written with a file
lock.

Cascade
───────
Location Code → Location Name → Region → Accounts Incharge
Account Code  → Head Office Assigned Person   (independent lookup)

Original xlsx sheet layout (kept only for the internal startup seed parser —
see the desktop app's own mapping_store.py docstring)
────────────────────────────────────────────────────────────────────────────
1. "Location Code Map"   : Location Code   | Location Name
2. "Region Incharge Map" : Region          | Accounts Incharge
3. "Location Region Map" : Location Name   | Region
4. "Account HO Map"      : Account Code    | Head Office Assigned Person

Internal in-memory formats (unchanged by the DB migration — processor.py's
business logic consumes these exact shapes and needed zero changes)
────────────────────────────────────────────────────────────────────────────
loc_name_map        : { "location_code": "Location Name" }
region_incharge_map : { "Region": "Accounts Incharge" }
name_region_map     : { "Location Name": "Region" }
account_ho_map      : { "account_code": "Head Office Assigned Person" }

load_all()'s return order — (loc_name_map, region_incharge_map,
name_region_map, account_ho_map) — matches processor.process_report's
parameter order exactly, which is what the original desktop app's own
mapping_store.load_all() also returned.
"""

from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from app.config import SEED_DIR
from app.soft_delete import delete_keyed_row, seed_missing_keyed_rows, sync_keyed_rows, upsert_keyed_row

from .models import AccountHoMap, LocationCodeMap, LocationRegionMap, RegionInchargeMap

SHEET_LOC_NAME    = "Location Code Map"
SHEET_INCHARGE    = "Region Incharge Map"
SHEET_NAME_REGION = "Location Region Map"
SHEET_ACCOUNT_HO  = "Account HO Map"

# Bundled, one-way deployment seed — copied from the desktop app's own
# bundled location-mapping-seed.xlsx. The centralized database is the sole
# runtime source of truth; there is no user-facing bulk import/export (see
# router module docstring for why that feature was intentionally not
# re-added).
SEED_MAPPING_PATH = SEED_DIR / "trial-balance-location-mapping-seed.xlsx"


# ── helpers ────────────────────────────────────────────────────────────────

def _s(v): return str(v).strip() if v is not None else ""


# ── DB -> in-memory ──────────────────────────────────────────────────────────

def _maybe_seed(db: Session) -> None:
    """First-run only: if the Location Code Map table is completely empty (a
    brand-new MySQL database), load the existing real mapping data from the
    seed workbook (backend/seed_data/trial-balance-location-mapping-seed.xlsx)
    so the suite launches with zero data loss from what the desktop app
    already had bundled."""
    if db.query(LocationCodeMap.id).first() is not None:
        return
    if SEED_MAPPING_PATH.exists():
        seed_missing_from_excel(db, SEED_MAPPING_PATH)


def load_all(db: Session) -> tuple:
    """Return (loc_name_map, region_incharge_map, name_region_map,
    account_ho_map) read from the shared MySQL tables. Seeds from the seed
    workbook on first run if the database is empty."""
    _maybe_seed(db)

    loc_name_map: dict = {}
    for row in db.query(LocationCodeMap).filter(LocationCodeMap.is_deleted == False).all():  # noqa: E712
        loc_name_map[_s(row.location_code)] = _s(row.location_name)

    region_incharge_map: dict = {}
    for row in db.query(RegionInchargeMap).filter(RegionInchargeMap.is_deleted == False).all():  # noqa: E712
        region_incharge_map[_s(row.region)] = _s(row.accounts_incharge)

    name_region_map: dict = {}
    for row in db.query(LocationRegionMap).filter(LocationRegionMap.is_deleted == False).all():  # noqa: E712
        name_region_map[_s(row.location_name)] = _s(row.region)

    account_ho_map: dict = {}
    for row in db.query(AccountHoMap).filter(AccountHoMap.is_deleted == False).all():  # noqa: E712
        account_ho_map[_s(row.account_code)] = _s(row.ho_person)

    return loc_name_map, region_incharge_map, name_region_map, account_ho_map


# ── in-memory -> DB (always rewrites all 4 tables, mirroring the original
#    xlsx save_all's always-rewrite-everything semantics) ───────────────────

def save_all(db: Session, loc_name_map: dict, region_incharge_map: dict,
             name_region_map: dict, account_ho_map: dict) -> None:
    """Sync all 4 mapping tables in MySQL to match these in-memory dicts —
    soft-delete-aware (see app/soft_delete.py): a row missing from the
    incoming dict gets is_deleted=True, never actually removed; a row whose
    key reappears (even if previously soft-deleted) is revived and updated
    in place."""
    sync_keyed_rows(db, LocationCodeMap, ("location_code",), {
        code: {"location_name": name} for code, name in loc_name_map.items()
    })
    sync_keyed_rows(db, RegionInchargeMap, ("region",), {
        region: {"accounts_incharge": incharge} for region, incharge in region_incharge_map.items()
    })
    sync_keyed_rows(db, LocationRegionMap, ("location_name",), {
        name: {"region": region} for name, region in name_region_map.items()
    })
    sync_keyed_rows(db, AccountHoMap, ("account_code",), {
        code: {"ho_person": person} for code, person in account_ho_map.items()
    })
    db.commit()


# ── Single-row CRUD (safe under concurrent edits — see app/soft_delete.py's
#    upsert_keyed_row/delete_keyed_row) ──────────────────────────────────────
#
# save_all() above always rewrites all 4 tables from one combined in-memory
# snapshot - fixing/adding/editing/deleting a single row anywhere used to go
# through load_all() -> mutate one entry -> save_all(), which races: two
# such requests close together (completely normal usage) can each read a
# snapshot that doesn't yet include the other's write, then each write their
# own stale snapshot back over all 4 tables - whichever commits last
# silently soft-deletes every row the other one had just added. The
# functions below touch only the ONE row being changed and are what every
# single-row endpoint should call instead.

def upsert_location_code(db: Session, code: str, location_name: str) -> None:
    upsert_keyed_row(db, LocationCodeMap, ("location_code",), code.strip(), {
        "location_name": location_name.strip(),
    })
    db.commit()


def delete_location_code(db: Session, code: str) -> bool:
    result = delete_keyed_row(db, LocationCodeMap, ("location_code",), code.strip())
    db.commit()
    return result


def upsert_location_region(db: Session, location_name: str, region: str) -> None:
    upsert_keyed_row(db, LocationRegionMap, ("location_name",), location_name.strip(), {
        "region": region.strip(),
    })
    db.commit()


def delete_location_region(db: Session, location_name: str) -> bool:
    result = delete_keyed_row(db, LocationRegionMap, ("location_name",), location_name.strip())
    db.commit()
    return result


def upsert_region_incharge(db: Session, region: str, accounts_incharge: str) -> None:
    upsert_keyed_row(db, RegionInchargeMap, ("region",), region.strip(), {
        "accounts_incharge": accounts_incharge.strip(),
    })
    db.commit()


def delete_region_incharge(db: Session, region: str) -> bool:
    result = delete_keyed_row(db, RegionInchargeMap, ("region",), region.strip())
    db.commit()
    return result


def upsert_account_ho(db: Session, account_code: str, ho_person: str) -> None:
    upsert_keyed_row(db, AccountHoMap, ("account_code",), account_code.strip(), {
        "ho_person": ho_person.strip(),
    })
    db.commit()


def delete_account_ho(db: Session, account_code: str) -> bool:
    result = delete_keyed_row(db, AccountHoMap, ("account_code",), account_code.strip())
    db.commit()
    return result


# ── Internal legacy workbook parser ─────────────────────────────────────────
# The workbook is a bundled, one-way deployment seed. The centralized database
# is the sole runtime source of truth; there is no user-facing import/export.

def _load_two_col_sheet(wb, sheet_name) -> dict:
    d: dict = {}
    if sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            key = _s(row[0])
            val = _s(row[1]) if len(row) > 1 else ""
            d[key] = val
    return d


def _parse_workbook(path) -> tuple:
    """Parse a 4-sheet location-mapping-seed.xlsx workbook into the same 4
    in-memory structures load_all() returns. Tolerates a missing file or
    missing sheets gracefully (same tolerance the original file-based
    mapping_store had)."""
    path = Path(path)
    if not path.exists():
        return {}, {}, {}, {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    loc_name_map        = _load_two_col_sheet(wb, SHEET_LOC_NAME)
    region_incharge_map = _load_two_col_sheet(wb, SHEET_INCHARGE)
    name_region_map     = _load_two_col_sheet(wb, SHEET_NAME_REGION)
    account_ho_map      = _load_two_col_sheet(wb, SHEET_ACCOUNT_HO)

    wb.close()
    return loc_name_map, region_incharge_map, name_region_map, account_ho_map


def seed_missing_from_excel(db: Session, path=None) -> dict:
    """Add legacy workbook rows that have never existed in the database.

    It never changes an existing row and never revives an archived row, so
    it is safe to run on every application startup and remains idempotent.
    """
    source = Path(path or SEED_MAPPING_PATH)
    loc_name_map, region_incharge_map, name_region_map, account_ho_map = _parse_workbook(source)

    inserted = {
        "location_code_map": seed_missing_keyed_rows(
            db, LocationCodeMap, ("location_code",), {
                code: {"location_name": name} for code, name in loc_name_map.items()
            },
        ),
        "region_incharge_map": seed_missing_keyed_rows(
            db, RegionInchargeMap, ("region",), {
                region: {"accounts_incharge": incharge}
                for region, incharge in region_incharge_map.items()
            },
        ),
        "location_region_map": seed_missing_keyed_rows(
            db, LocationRegionMap, ("location_name",), {
                name: {"region": region} for name, region in name_region_map.items()
            },
        ),
        "account_ho_map": seed_missing_keyed_rows(
            db, AccountHoMap, ("account_code",), {
                code: {"ho_person": person} for code, person in account_ho_map.items()
            },
        ),
    }
    db.commit()
    return inserted
