"""Give SUBTOTAL/SUM formula cells a cached value, without ever touching the
live formula itself.

openpyxl deliberately never calculates or caches formula results - a cell
written as ``=SUBTOTAL(9,C2:C13)`` is saved with only the formula text, no
``<v>`` (value) element. Excel is *supposed* to recalculate every formula
itself the moment the file is opened (openpyxl already sets the workbook's
``fullCalcOnLoad`` flag by default, which asks for exactly that) - but a
file downloaded from an email attachment normally opens in Excel's
"Protected View" first, and Protected View is well known to suppress
automatic recalculation until the user clicks "Enable Editing", and even
then behavior is inconsistent across Excel versions/builds. The practical
result: the report opens showing blank Grand Total / footer cells until
the user manually forces a recalculation (Ctrl+Alt+F9), which looks like a
broken report.

The fix here computes the same values these formulas already show
correctly in the email preview (see app.services.mailer_shared.sheet_to_html,
which resolves the identical formula set for HTML display) and writes them
as a real cached ``<v>`` into the saved .xlsx's XML, directly after
openpyxl finishes writing the file. This is a pure, additive post-processing
step over openpyxl's own output - no external application (Excel/LibreOffice
COM or otherwise) is involved, and the formula text is never modified, so
anyone who edits a referenced cell still gets a live, correctly
recalculating SUBTOTAL/SUM exactly as before.

Only the same small, controlled set of aggregate forms the email renderer
already understands are resolved: ``SUBTOTAL(9, range)``,
``SUBTOTAL(109, range)`` (hidden-row-excluding), and ``SUM(...)`` over one or
more comma-separated ranges/cells (e.g. ``SUM(D5:E5,G5:H5)`` for a row total
that skips non-adjacent columns). Any other formula is left exactly as
openpyxl wrote it (no cached value forced), so this can never mask a
genuinely wrong/unsupported formula behind a fabricated number.
"""

import re
import shutil
import tempfile
import zipfile
from pathlib import Path

from openpyxl.utils.cell import range_boundaries

_CELL_OR_RANGE = r"\$?[A-Z]{1,3}\$?\d+(?:\s*:\s*\$?[A-Z]{1,3}\$?\d+)?"
_SUBTOTAL_RE = re.compile(
    r"^\s*=\s*SUBTOTAL\s*\(\s*(9|109)\s*,\s*"
    r"(\$?[A-Z]{1,3}\$?\d+\s*:\s*\$?[A-Z]{1,3}\$?\d+)\s*\)\s*$",
    re.IGNORECASE,
)
_SUM_RE = re.compile(
    r"^\s*=\s*SUM\s*\(\s*"
    r"(" + _CELL_OR_RANGE + r"(?:\s*,\s*" + _CELL_OR_RANGE + r")*)"
    r"\s*\)\s*$",
    re.IGNORECASE,
)


def _iter_sum_parts(ws, parts_text: str):
    """Yield every cell in a SUM(...) formula's comma-separated ranges/cells."""
    for part in re.split(r"\s*,\s*", parts_text.strip()):
        part = part.replace(" ", "")
        ref = part if ":" in part else f"{part}:{part}"
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            yield from row


def _resolve_sheet_formulas(ws) -> dict[str, float]:
    """Return {cell_coordinate: numeric_value} for every SUBTOTAL/SUM cell
    on `ws` that this module knows how to evaluate. Mirrors
    mailer_shared.sheet_to_html's _display_value exactly, so the number a
    user sees in the email preview and the number cached into the actual
    .xlsx are always the same value, computed the same way."""
    resolved: dict[str, object] = {}
    resolving: set[str] = set()

    def _value(cell):
        coordinate = cell.coordinate
        if coordinate in resolved:
            return resolved[coordinate]

        raw = cell.value
        if not (isinstance(raw, str) and raw.startswith("=")):
            resolved[coordinate] = raw
            return raw
        if coordinate in resolving:
            return None

        subtotal_match = _SUBTOTAL_RE.match(raw)
        sum_match = _SUM_RE.match(raw)
        match = subtotal_match or sum_match
        if not match:
            resolved[coordinate] = None
            return None

        function_num = int(subtotal_match.group(1)) if subtotal_match else None

        resolving.add(coordinate)
        try:
            total = 0
            if subtotal_match:
                range_ref = subtotal_match.group(2)
                min_col, min_row, max_col, max_row = range_boundaries(range_ref.replace(" ", ""))
                source_cells = (
                    source_cell
                    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
                    for source_cell in row
                )
            else:
                source_cells = _iter_sum_parts(ws, sum_match.group(1))
            for source_cell in source_cells:
                source_raw = source_cell.value
                if subtotal_match and isinstance(source_raw, str) and _SUBTOTAL_RE.match(source_raw):
                    continue
                if function_num == 109 and ws.row_dimensions[source_cell.row].hidden:
                    continue
                value = _value(source_cell)
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    total += value
        finally:
            resolving.discard(coordinate)

        resolved[coordinate] = total
        return total

    values: dict[str, float] = {}
    for row in ws.iter_rows():
        for cell in row:
            raw = cell.value
            if isinstance(raw, str) and raw.startswith("="):
                result = _value(cell)
                if isinstance(result, (int, float)) and not isinstance(result, bool):
                    values[cell.coordinate] = result
    return values


def cache_formula_values(wb) -> dict[str, dict[str, float]]:
    """Resolve every supported formula cell across every sheet in `wb`.
    Call this on the in-memory workbook BEFORE or AFTER wb.save(path) (the
    object is unchanged either way, this never mutates `wb`) and pass the
    result to inject_cached_values(path, ...) once the file is on disk."""
    return {sheet: _resolve_sheet_formulas(wb[sheet]) for sheet in wb.sheetnames}


_TAG_RE = re.compile(r"<(sheet|Relationship)\b([^>]*)/?>", re.IGNORECASE)
_ATTR_RE = re.compile(r'([\w:]+)="([^"]*)"')


def _attrs(tag_body: str) -> dict[str, str]:
    return {name: value for name, value in _ATTR_RE.findall(tag_body)}


def _sheet_name_to_xml_path(workbook_xml: str, rels_xml: str) -> dict[str, str]:
    # Attribute order in these elements isn't guaranteed (e.g. openpyxl emits
    # <sheet ... name="Main" ... r:id="rId1"/> but a Relationship's Target can
    # come before or after its Id) - parse attributes by name, not position.
    rel_target: dict[str, str] = {}
    for tag, body in _TAG_RE.findall(rels_xml):
        if tag.lower() != "relationship":
            continue
        attrs = _attrs(body)
        rid, target = attrs.get("Id"), attrs.get("Target")
        if rid and target:
            rel_target[rid] = target

    mapping: dict[str, str] = {}
    for tag, body in _TAG_RE.findall(workbook_xml):
        if tag.lower() != "sheet":
            continue
        attrs = _attrs(body)
        name, rid = attrs.get("name"), attrs.get("r:id")
        target = rel_target.get(rid) if rid else None
        if name and target:
            # Target may be package-root-relative ("/xl/worksheets/sheet1.xml")
            # or relative to xl/ ("worksheets/sheet1.xml") - normalize both to
            # the zip-internal path used by ZipFile.namelist().
            mapping[name] = target.lstrip("/") if target.startswith("/") else f"xl/{target}"
    return mapping


def _inject_into_sheet_xml(xml_text: str, cell_values: dict[str, float]) -> str:
    for coordinate, value in cell_values.items():
        # Match <c r="C14" ...>...<f ...>...</f>[<v>...</v>]</c> for this exact
        # cell only, and (re)write its <v> right after the formula - leaving
        # the <f> element, and every other cell in the sheet, untouched.
        pattern = re.compile(
            r'(<c r="' + re.escape(coordinate) + r'"[^>]*>.*?<f[^>]*>.*?</f>)(?:<v>.*?</v>)?(</c>)',
            re.DOTALL,
        )
        xml_text, count = pattern.subn(
            lambda m: f"{m.group(1)}<v>{value}</v>{m.group(2)}", xml_text, count=1
        )
        # count == 0 means this coordinate wasn't a formula cell in the saved
        # XML after all (shouldn't happen since we resolved it from the same
        # in-memory sheet) - silently skip rather than risk corrupting the file.
    return xml_text


def inject_cached_values(path: str, values_by_sheet: dict[str, dict[str, float]]) -> None:
    """Patch the already-saved .xlsx at `path` in place, adding a cached
    <v> value to each formula cell resolved by cache_formula_values. Every
    other byte of the file - styles, merges, every other sheet, the
    formulas themselves - is preserved exactly as openpyxl wrote it."""
    values_by_sheet = {sheet: cells for sheet, cells in values_by_sheet.items() if cells}
    if not values_by_sheet:
        return

    with zipfile.ZipFile(path, "r") as archive:
        names = archive.namelist()
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
        rels_xml = archive.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        sheet_paths = _sheet_name_to_xml_path(workbook_xml, rels_xml)

        replacements: dict[str, str] = {}
        for sheet_name, cell_values in values_by_sheet.items():
            sheet_path = sheet_paths.get(sheet_name)
            if not sheet_path or sheet_path not in names:
                continue
            original = archive.read(sheet_path).decode("utf-8")
            replacements[sheet_path] = _inject_into_sheet_xml(original, cell_values)

        if not replacements:
            return

        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=str(Path(path).parent))
        try:
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as out:
                for item in archive.infolist():
                    data = replacements.get(item.filename)
                    if data is not None:
                        out.writestr(item, data.encode("utf-8"))
                    else:
                        out.writestr(item, archive.read(item.filename))
        finally:
            import os
            os.close(fd)

    shutil.move(tmp_path, path)
