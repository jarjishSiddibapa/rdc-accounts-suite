"""Durable DMS invoice-booking scans, workbook generation, and scheduled mail."""

from __future__ import annotations

import html
import json
import logging
import re
import time
import shutil
import uuid
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from string import Formatter
from typing import Any
from urllib.parse import quote, urljoin

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app import audit_middleware, security
from app.config import SCRATCH_DIR
from app.database import SessionLocal
from app.models import (
    InvoiceBookingTrackerCheck,
    InvoiceBookingTrackerMapping,
    InvoiceBookingTrackerNotification,
    InvoiceBookingTrackerSettings,
)
from app.regional import now_ist
from app.services.mailer_shared import send_mail

logger = logging.getLogger(__name__)

APP_KEY = "invoice-booking-tracker"
DEFAULT_LOGIN_URL = "https://dms.rdc.in/"
DEFAULT_SUBJECT = "Ultrafine Pending Invoice Booking Tracker as on {date}"
DEFAULT_BODY = """Dear Team,

Please find below the Ultrafine pending invoice booking tracker as on {date}.

{tracker_table}

Total pending invoices: {total_pending}

Thanks,
Ultrafine Team"""
TEMPLATE_FIELDS = frozenset({"date", "total_pending", "location_count", "tracker_table"})
CHECK_MAX_ATTEMPTS = 3
ACCOUNT_IN_USE_ERROR_PREFIX = "DMS_ACCOUNT_IN_USE:"
ACCOUNT_IN_USE_PUBLIC_MESSAGE = (
    "The DMS account is already logged in. Please update the tracker after the current DMS user signs out."
)

# These rows are transcribed from the supplied, proven manual tracker. Seeds
# add missing natural keys only: an administrator's edits and archived rows
# always win over bundled defaults.
DEFAULT_MAPPINGS = (
    ("CAPEX", "Khushi", "Accounts payment ultrafine CAPEX invoices", "ACCOUNTS_PAYMENT_ULTRAFINE_CAPEX_INVOICES"),
    ("ANDHRA/Nellore", "Jaysukh", "Accounts payment ultrafine invoices ANDHRA/Nellore", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_ANDHRA_NELLORE"),
    ("BANGALORE", "Hitanshi", "Accounts payment ultrafine invoices BANGALORE", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_BANGALORE"),
    ("FlyAsh", "Hitanshi", "Accounts payment ultrafine invoices FlyAsh", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_FLYASH"),
    ("GOA", "Rakesh", "Accounts payment ultrafine invoices GOA", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_GOA"),
    ("HO", "Hitanshi", "Accounts payment ultrafine invoices HO", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_HO"),
    ("NAGPUR", "Vishal", "Accounts payment ultrafine invoices NAGPUR", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_NAGPUR"),
    ("ODISHA", "Hitanshi", "Accounts payment ultrafine invoices ODISHA", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_ODISHA"),
    ("RAIPUR", "Hitanshi", "Accounts payment ultrafine invoices RAIPUR", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_RAIPUR"),
    ("SURAT", "Jaysukh", "Accounts payment ultrafine invoices SURAT", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_SURAT"),
    ("TAMIL NADU", "Hitanshi", "Accounts payment ultrafine invoices TAMILNADU", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_TAMILNADU"),
    ("TELANGANA", "Hitanshi", "Accounts payment ultrafine invoices TELANGANA", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_TELANGANA"),
    ("VIZAG/VISAKHAPATNAM", "Jaysukh", "Accounts payment ultrafine invoices VIZAG/VISAKHAPATNAM", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_VIZAG_VISAKHAPATNAM"),
    ("WADA", "Vishal", "Accounts payment ultrafine invoices WADA", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_WADA"),
    ("West Bengal (Aggregate, Microsilica & Howrah)", "Ashutosh", "Accounts payment ultrafine invoices West Bengal (Aggregate, Microsilica & Howrah)", "ACCOUNTS_PAYMENT_ULTRAFINE_INVOICES_WEST_BENGAL"),
)


def _utcnow() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.strip().casefold()).strip("-")


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", value.casefold())).strip()


def parse_recipients(raw: str | None) -> list[str]:
    if not raw:
        return []
    try:
        values = json.loads(raw)
    except (TypeError, ValueError):
        return []
    return [str(value).strip() for value in values if str(value).strip()] if isinstance(values, list) else []


def validate_template(template: str, *, allow_table: bool = True) -> None:
    try:
        fields = {
            name.split(".", 1)[0].split("[", 1)[0]
            for _, name, _, _ in Formatter().parse(template)
            if name
        }
    except ValueError as exc:
        raise ValueError("Template contains unmatched braces") from exc
    allowed = TEMPLATE_FIELDS if allow_table else TEMPLATE_FIELDS - {"tracker_table"}
    unsupported = fields - allowed
    if unsupported:
        raise ValueError(f"Unsupported template placeholder: {sorted(unsupported)[0]}")


def seed_settings(db: Session) -> None:
    if db.query(InvoiceBookingTrackerSettings).filter_by(id=1).first() is None:
        db.add(InvoiceBookingTrackerSettings(
            id=1,
            enabled=False,
            login_url=DEFAULT_LOGIN_URL,
            login_timeout_seconds=90,
            scheduled_email_enabled=True,
            scheduled_email_time="08:00",
            mail_to="[]",
            mail_cc="[]",
            subject_template=DEFAULT_SUBJECT,
            body_template=DEFAULT_BODY,
            version=1,
            is_deleted=False,
        ))
    existing = {
        row.location_key
        for row in db.query(InvoiceBookingTrackerMapping).all()
    }
    for order, (location, owner, queue_label, queue_key) in enumerate(DEFAULT_MAPPINGS, 1):
        key = normalize_key(location)
        if key not in existing:
            db.add(InvoiceBookingTrackerMapping(
                location_key=key,
                location=location,
                responsible_person=owner,
                queue_label=queue_label,
                queue_key=queue_key,
                sort_order=order,
                is_active=True,
                is_deleted=False,
            ))
    db.commit()


def get_or_create_settings(db: Session) -> InvoiceBookingTrackerSettings:
    row = db.query(InvoiceBookingTrackerSettings).filter_by(id=1).first()
    if row is None:
        seed_settings(db)
        row = db.query(InvoiceBookingTrackerSettings).filter_by(id=1).one()
    elif row.is_deleted:
        row.is_deleted = False
        db.commit()
    return row


def _business_date(value: date | None = None) -> str:
    current = value or now_ist().date()
    suffix = "th" if 10 < current.day % 100 < 14 else {1: "st", 2: "nd", 3: "rd"}.get(current.day % 10, "th")
    return f"{current.day}{suffix} {current.strftime('%B %Y')}"


def _table_html(rows: list[dict[str, Any]]) -> str:
    body = "".join(
        "<tr>"
        f"<td style='padding:8px 10px;border:1px solid #cbd5e1'>{html.escape(str(row['location']))}</td>"
        f"<td style='padding:8px 10px;border:1px solid #cbd5e1'>{html.escape(str(row['responsible_person']))}</td>"
        f"<td style='padding:8px 10px;border:1px solid #cbd5e1;text-align:right'>{row['pending'] or '-'}</td>"
        "</tr>"
        for row in rows
    )
    total = sum(int(row["pending"]) for row in rows)
    return (
        "<table style='border-collapse:collapse;font-family:Arial,sans-serif;font-size:13px;color:#172033'>"
        "<thead><tr style='background:#203864;color:#fff'>"
        "<th style='padding:9px 10px;border:1px solid #203864;text-align:left'>Location</th>"
        "<th style='padding:9px 10px;border:1px solid #203864;text-align:left'>Responsible Person</th>"
        "<th style='padding:9px 10px;border:1px solid #203864;text-align:right'>Pending</th>"
        f"</tr></thead><tbody>{body}"
        "<tr style='font-weight:700;background:#e9eff7'><td colspan='2' style='padding:9px 10px;border:1px solid #cbd5e1'>Grand Total</td>"
        f"<td style='padding:9px 10px;border:1px solid #cbd5e1;text-align:right'><strong>{total}</strong></td></tr>"
        "</tbody></table>"
    )


def render_templates(subject_template: str, body_template: str, rows: list[dict[str, Any]], on_date: date | None = None) -> tuple[str, str]:
    values = {
        "date": _business_date(on_date),
        "total_pending": str(sum(int(row["pending"]) for row in rows)),
        "location_count": str(len(rows)),
        "tracker_table": _table_html(rows),
    }
    subject = subject_template.format_map({**values, "tracker_table": ""})
    marker = "__TRACKER_TABLE__"
    escaped_body = html.escape(body_template.replace("{tracker_table}", marker)).replace("\n", "<br>\n")
    for key in ("date", "total_pending", "location_count"):
        escaped_body = escaped_body.replace("{" + key + "}", html.escape(values[key]))
    return subject, escaped_body.replace(marker, values["tracker_table"])


def create_workbook(rows: list[dict[str, Any]], output_path: Path, on_date: date | None = None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice Booking Tracker"
    current = on_date or now_ist().date()
    sheet.merge_cells("A1:C1")
    sheet["A1"] = f"UF PENDING INVOICE BOOKING TRACKER AS ON {current.strftime('%d-%m-%Y')}"
    sheet["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="203864")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 25
    headers = ("Locations", "Responsible Person", current.strftime("%d-%m-%Y"))
    thin = Side(style="thin", color="A6A6A6")
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(2, column, value)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_number, result in enumerate(rows, 3):
        values = (result["location"], result["responsible_person"], int(result["pending"]))
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_number, column, value)
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="right" if column == 3 else "left", vertical="center")
        sheet.cell(row_number, 3).number_format = '0;-0;-'
    total_row = len(rows) + 3
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    sheet.cell(total_row, 1, "Grand Total")
    sheet.cell(total_row, 3, f"=SUM(C3:C{total_row - 1})")
    for column in range(1, 4):
        cell = sheet.cell(total_row, column)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E2F3")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.cell(total_row, 3).number_format = '0;-0;-'
    sheet.column_dimensions["A"].width = 48
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 16
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:C{total_row - 1}"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def count_unbooked(statuses: list[str]) -> int:
    return sum(1 for status in statuses if status.strip().casefold() != "booked")


def _find_login_fields(page):
    password = page.locator("input[type=password]:visible").first
    if password.count() == 0:
        return None, None
    username = page.locator('input[type=email]:visible, input[name*="user" i]:visible, input[type=text]:visible').first
    return (username if username.count() else None), password


def _login(page, username: str, password: str, timeout_seconds: int) -> None:
    username_field = password_field = None
    detection_deadline = time.monotonic() + min(12, timeout_seconds)
    while time.monotonic() < detection_deadline:
        if "/console" in page.url.casefold():
            return
        username_field, password_field = _find_login_fields(page)
        if password_field is not None:
            break
        page.wait_for_timeout(250)
    if password_field is None:
        raise RuntimeError("The DMS login form did not become available")
    if username_field is None:
        raise RuntimeError("The DMS username field could not be detected")
    username_field.fill(username)
    password_field.fill(password)
    buttons = page.locator("button:visible, input[type=submit]:visible")
    clicked = False
    for index in range(buttons.count()):
        candidate = buttons.nth(index)
        label = ((candidate.inner_text() or candidate.get_attribute("value") or "").strip().casefold())
        if any(word in label for word in ("login", "log in", "sign in")):
            candidate.click()
            clicked = True
            break
    if not clicked:
        password_field.press("Enter")
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        page.wait_for_timeout(500)
        notices = page.locator(".alert:visible, .error:visible, .toast-message:visible").all_inner_texts()
        if any("same user name is already logged" in notice.casefold() for notice in notices):
            raise RuntimeError(
                f"{ACCOUNT_IN_USE_ERROR_PREFIX} The DMS portal reports that this username already has an active session."
            )
        if "/console" in page.url.casefold():
            return
        _, current_password = _find_login_fields(page)
        if current_password is None and "/login" not in page.url.casefold():
            return
    raise RuntimeError("DMS login timed out; verify the saved credentials or session")


def _logout(page) -> bool:
    """Release DMS's single-login server session before closing Chromium.

    The live portal keeps an account marked logged-in after a browser process
    merely closes. Prefer its visible logout control; the URL fallback covers
    layouts that hide the control in a collapsed menu. Failure remains best-
    effort because it must never replace the scan's real result/error.
    """
    try:
        controls = page.get_by_text(re.compile(r"^\s*(?:log\s*out|logout|sign\s*out)\s*$", re.IGNORECASE))
        for index in range(controls.count()):
            control = controls.nth(index)
            if control.is_visible():
                control.click(timeout=5000)
                page.wait_for_timeout(500)
                return True
    except Exception:
        pass
    try:
        page.goto(urljoin(page.url, "/logout"), wait_until="domcontentloaded", timeout=10_000)
        return "/console" not in page.url.casefold()
    except Exception:
        return False


def _discover_queue_url(page, login_url: str, label: str, queue_key: str | None) -> str:
    console_url = urljoin(login_url, "/console")
    page.goto(console_url, wait_until="domcontentloaded", timeout=60_000)
    page.wait_for_timeout(800)
    target = normalize_text(label)
    links = page.locator("a[href]:visible")
    best: str | None = None
    for index in range(links.count()):
        link = links.nth(index)
        text = normalize_text(link.inner_text() or "")
        href = link.get_attribute("href") or ""
        if "workflowcases" not in href.casefold():
            continue
        if text == target:
            return urljoin(page.url, href)
        if target in text or text in target:
            best = best or urljoin(page.url, href)
    if best:
        return best
    if queue_key:
        return urljoin(login_url, f"/console/workflowcases?Q={quote(queue_key)}")
    raise RuntimeError(f"The DMS work queue '{label}' could not be found")


def _status_table(page):
    tables = page.locator("table:visible")
    for index in range(tables.count()):
        table = tables.nth(index)
        headers = [normalize_text(value) for value in table.locator("thead th").all_inner_texts()]
        if "status" in headers:
            return table, headers.index("status")
    raise RuntimeError("The DMS results table does not contain a visible Status column")


def _scan_queue(page, queue_url: str) -> tuple[int, int, int]:
    page.goto(queue_url, wait_until="domcontentloaded", timeout=60_000)
    page.wait_for_timeout(800)
    statuses: list[str] = []
    pages = 0
    seen: set[str] = set()
    expected_total: int | None = None
    while True:
        table, status_index = _status_table(page)
        rows = table.locator("tbody tr:visible")
        page_rows: list[list[str]] = []
        for index in range(rows.count()):
            cells = rows.nth(index).locator("td").all_inner_texts()
            if cells and status_index < len(cells) and not (len(cells) == 1 and "no data" in cells[0].casefold()):
                page_rows.append(cells)
        fingerprint = json.dumps(page_rows, ensure_ascii=False)
        if fingerprint in seen:
            raise RuntimeError("DMS pagination repeated a page before reaching the final page")
        seen.add(fingerprint)
        pages += 1
        statuses.extend(cells[status_index] for cells in page_rows)

        info = page.locator(".dataTables_info:visible, [id$=_info]:visible")
        if info.count():
            match = re.search(r"of\s+([\d,]+)\s+entries", info.first.inner_text(), re.IGNORECASE)
            if match:
                expected_total = int(match.group(1).replace(",", ""))

        next_buttons = page.locator("a.paginate_button.next:visible, button:has-text('Next'):visible, a:has-text('Next'):visible")
        next_button = next_buttons.first if next_buttons.count() else None
        disabled = next_button is None
        if next_button is not None:
            classes = (next_button.get_attribute("class") or "").casefold()
            aria_disabled = (next_button.get_attribute("aria-disabled") or "").casefold()
            disabled = "disabled" in classes or aria_disabled == "true"
        if disabled:
            break
        previous = fingerprint
        next_button.click()
        deadline = time.monotonic() + 20
        while time.monotonic() < deadline:
            page.wait_for_timeout(250)
            try:
                candidate, _ = _status_table(page)
                candidate_rows = [candidate.locator("tbody tr:visible").nth(i).locator("td").all_inner_texts() for i in range(candidate.locator("tbody tr:visible").count())]
                if json.dumps(candidate_rows, ensure_ascii=False) != previous:
                    break
            except Exception:
                pass
        else:
            raise RuntimeError("DMS did not load the next results page")
        if pages >= 10_000:
            raise RuntimeError("DMS pagination exceeded the safety limit")
    if expected_total is not None and len(statuses) != expected_total:
        raise RuntimeError(f"DMS pagination was incomplete: scanned {len(statuses)} of {expected_total} records")
    return count_unbooked(statuses), len(statuses), pages


def fetch_tracker(snapshot: dict[str, Any], mappings: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError("Playwright is not installed on the server") from exc
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        context = None
        try:
            context = browser.new_context(viewport={"width": 1440, "height": 1000}, storage_state=snapshot.get("saved_session"))
            page = context.new_page()
            page.goto(snapshot["login_url"], wait_until="domcontentloaded", timeout=60_000)
            _login(page, snapshot["username"], snapshot["password"], snapshot["login_timeout_seconds"])
            results = []
            for mapping in mappings:
                queue_url = _discover_queue_url(page, snapshot["login_url"], mapping["queue_label"], mapping.get("queue_key"))
                pending, records, pages = _scan_queue(page, queue_url)
                results.append({**mapping, "pending": pending, "records_scanned": records, "pages_scanned": pages})
            # A saved login is useful only to enter this run. Explicit logout
            # is required so the 08:00 automation never blocks the employee
            # who uses the same DMS ID after arriving at the office.
            return results, context.storage_state()
        finally:
            if context is not None:
                for current_page in reversed(context.pages):
                    if _logout(current_page):
                        break
            browser.close()


def fetch_tracker_with_retries(snapshot: dict[str, Any], mappings: list[dict[str, Any]], progress_cb=None):
    last_error: Exception | None = None
    for attempt in range(1, CHECK_MAX_ATTEMPTS + 1):
        if progress_cb:
            progress_cb(0.05 + ((attempt - 1) * 0.12), f"Checking every DMS page (attempt {attempt} of {CHECK_MAX_ATTEMPTS})")
        current = dict(snapshot)
        if attempt > 1:
            current["saved_session"] = None
        try:
            rows, session = fetch_tracker(current, mappings)
            return rows, session, attempt
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            logger.warning("Invoice tracker portal attempt %s/%s failed: %s", attempt, CHECK_MAX_ATTEMPTS, str(exc)[:500])
    raise RuntimeError(f"DMS tracker check failed after {CHECK_MAX_ATTEMPTS} attempts. Last error: {last_error}") from last_error


def _deliver_notification(notification_id: int) -> dict[str, Any]:
    db = SessionLocal()
    output_path: Path | None = None
    output_dir: Path | None = None
    try:
        notification = db.query(InvoiceBookingTrackerNotification).filter_by(id=notification_id).with_for_update().first()
        if notification is None or notification.is_deleted or notification.status not in {"pending", "failed"}:
            return {"id": notification_id, "status": "skipped"}
        settings = get_or_create_settings(db)
        notification.status = "sending"
        notification.attempted_at = _utcnow()
        notification.error_message = None
        db.commit()
        rows = json.loads(notification.result_json)
        output_dir = SCRATCH_DIR / f"invoice-booking-tracker-{uuid.uuid4().hex}"
        output_path = output_dir / notification.attachment_filename
        create_workbook(rows, output_path)
        if not settings.sender_email or not settings.sender_app_password_encrypted:
            raise RuntimeError("The dedicated tracker sender email and app password are not configured")
        send_mail(
            from_email=settings.sender_email,
            app_password=security.decrypt(settings.sender_app_password_encrypted),
            to_addresses=parse_recipients(notification.to_recipients),
            cc_addresses=parse_recipients(notification.cc_recipients),
            subject=notification.subject,
            html_body=notification.body,
            attachments=[str(output_path)],
        )
        notification = db.query(InvoiceBookingTrackerNotification).filter_by(id=notification_id).one()
        notification.status = "sent"
        notification.sent_at = _utcnow()
        settings = get_or_create_settings(db)
        settings.last_scheduled_sent_date = now_ist().date()
        db.commit()
        audit_middleware.log_event("invoice_booking_tracker.notification.sent", details={"notification_id": notification_id})
        return {"id": notification_id, "status": "sent"}
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        notification = db.query(InvoiceBookingTrackerNotification).filter_by(id=notification_id).first()
        if notification:
            notification.status = "failed"
            notification.error_message = str(exc)[:2000]
            db.commit()
        audit_middleware.log_event("invoice_booking_tracker.notification.failed", details={"notification_id": notification_id, "error": type(exc).__name__})
        return {"id": notification_id, "status": "failed", "message": str(exc)}
    finally:
        if output_dir:
            shutil.rmtree(output_dir, ignore_errors=True)
        db.close()


def _release_lock(token: str) -> None:
    db = SessionLocal()
    try:
        row = get_or_create_settings(db)
        if row.check_lock_token == token:
            row.check_lock_token = None
            row.check_lock_expires_at = None
            db.commit()
    finally:
        db.close()


def run_check_job(trigger: str = "scheduled", progress_cb=None) -> dict[str, Any]:
    started = time.monotonic()
    token = str(uuid.uuid4())
    db = SessionLocal()
    try:
        settings = db.query(InvoiceBookingTrackerSettings).filter_by(id=1).with_for_update().first()
        if settings is None:
            seed_settings(db)
            settings = db.query(InvoiceBookingTrackerSettings).filter_by(id=1).with_for_update().one()
        now = _utcnow()
        if settings.check_lock_token and settings.check_lock_expires_at and settings.check_lock_expires_at > now:
            db.add(InvoiceBookingTrackerCheck(trigger=trigger, status="skipped", error_message="Another tracker check is already running.", checked_at=now, duration_seconds=round(time.monotonic() - started, 3)))
            db.commit()
            return {"skipped": True, "message": "Another tracker check is already running."}
        username = (settings.username or "").strip()
        password = security.decrypt(settings.password_encrypted or "")
        if not username or not password:
            from app.jobs import JobUserError
            raise JobUserError("Configure the DMS username and password before checking the tracker")
        try:
            saved_session = json.loads(security.decrypt(settings.session_state_encrypted or "")) if settings.session_state_encrypted else None
        except (TypeError, ValueError):
            saved_session = None
        mappings = [{"id": row.id, "location": row.location, "responsible_person": row.responsible_person, "queue_label": row.queue_label, "queue_key": row.queue_key} for row in db.query(InvoiceBookingTrackerMapping).filter(InvoiceBookingTrackerMapping.is_deleted.is_(False), InvoiceBookingTrackerMapping.is_active.is_(True)).order_by(InvoiceBookingTrackerMapping.sort_order, InvoiceBookingTrackerMapping.id).all()]
        if not mappings:
            from app.jobs import JobUserError
            raise JobUserError("Configure at least one active tracker mapping before checking DMS")
        snapshot = {"login_url": settings.login_url, "username": username, "password": password, "saved_session": saved_session, "login_timeout_seconds": settings.login_timeout_seconds}
        settings.check_lock_token = token
        settings.check_lock_expires_at = now + timedelta(seconds=max(1800, CHECK_MAX_ATTEMPTS * (settings.login_timeout_seconds + (len(mappings) * 90))))
        db.commit()
    finally:
        db.close()

    try:
        rows, session, attempts = fetch_tracker_with_retries(snapshot, mappings, progress_cb)
        total_pending = sum(row["pending"] for row in rows)
        total_records = sum(row["records_scanned"] for row in rows)
        total_pages = sum(row["pages_scanned"] for row in rows)
        db = SessionLocal()
        try:
            settings = db.query(InvoiceBookingTrackerSettings).filter_by(id=1).with_for_update().one()
            check = InvoiceBookingTrackerCheck(trigger=trigger, status="success", total_pending=total_pending, total_records_scanned=total_records, total_pages_scanned=total_pages, result_json=json.dumps(rows, ensure_ascii=False), checked_at=_utcnow(), duration_seconds=round(time.monotonic() - started, 3))
            db.add(check)
            db.flush()
            # fetch_tracker explicitly logs out to release DMS's single-login
            # session, so its captured pre-logout state must not be reused.
            settings.session_state_encrypted = None
            settings.last_total_pending = total_pending
            settings.last_checked_at = check.checked_at
            settings.last_check_status = "success"
            settings.last_error = None
            settings.updated_at = _utcnow()
            notification_id = None
            if trigger == "scheduled" and settings.scheduled_email_enabled:
                today = now_ist().date()
                key = f"scheduled:{today.isoformat()}"
                notification = db.query(InvoiceBookingTrackerNotification).filter_by(notification_key=key).first()
                subject, body = render_templates(settings.subject_template, settings.body_template, rows, today)
                filename = f"Ultrafine Pending Invoice Booking Tracker as on {today.strftime('%d.%m.%Y')}.xlsx"
                if notification is None:
                    notification = InvoiceBookingTrackerNotification(notification_key=key, check_id=check.id, subject=subject, body=body, to_recipients=settings.mail_to or "[]", cc_recipients=settings.mail_cc or "[]", result_json=json.dumps(rows, ensure_ascii=False), attachment_filename=filename, status="pending")
                    db.add(notification)
                    db.flush()
                elif notification.status == "failed":
                    notification.check_id = check.id
                    notification.subject = subject
                    notification.body = body
                    notification.to_recipients = settings.mail_to or "[]"
                    notification.cc_recipients = settings.mail_cc or "[]"
                    notification.result_json = json.dumps(rows, ensure_ascii=False)
                    notification.attachment_filename = filename
                    notification.status = "pending"
                    notification.error_message = None
                notification_id = notification.id
            db.commit()
            check_id = check.id
        finally:
            db.close()
        delivery = _deliver_notification(notification_id) if notification_id else None
        if progress_cb:
            progress_cb(1.0, "Every configured DMS page has been checked")
        audit_middleware.log_event("invoice_booking_tracker.checked", details={"trigger": trigger, "check_id": check_id, "attempts": attempts, "total_pending": total_pending})
        return {"check_id": check_id, "total_pending": total_pending, "total_records_scanned": total_records, "total_pages_scanned": total_pages, "rows": rows, "attempts": attempts, "notification": delivery}
    except Exception as exc:
        logger.exception("Invoice Booking Tracker check failed")
        db = SessionLocal()
        try:
            settings = get_or_create_settings(db)
            check = InvoiceBookingTrackerCheck(trigger=trigger, status="error", error_message=str(exc)[:2000], checked_at=_utcnow(), duration_seconds=round(time.monotonic() - started, 3))
            db.add(check)
            settings.last_checked_at = check.checked_at
            settings.last_check_status = "error"
            settings.last_error = str(exc)[:2000]
            # A failed browser run may have invalidated or logged out an
            # imported state; never keep retrying a known-stale session.
            settings.session_state_encrypted = None
            settings.updated_at = _utcnow()
            db.commit()
        finally:
            db.close()
        audit_middleware.log_event("invoice_booking_tracker.failed", details={"trigger": trigger, "error": type(exc).__name__})
        from app.jobs import JobUserError
        raise JobUserError(str(exc)) from exc
    finally:
        _release_lock(token)


def enqueue_due_check() -> str | None:
    db = SessionLocal()
    try:
        settings = db.query(InvoiceBookingTrackerSettings).filter(InvoiceBookingTrackerSettings.id == 1, InvoiceBookingTrackerSettings.is_deleted.is_(False)).with_for_update().first()
        if settings is None or not settings.enabled or not settings.scheduled_email_enabled or not settings.username or not settings.password_encrypted:
            return None
        current = now_ist()
        if current.strftime("%H:%M") < settings.scheduled_email_time or settings.last_scheduled_attempt_date == current.date():
            return None
        settings.last_scheduled_attempt_date = current.date()
        db.commit()
    finally:
        db.close()
    from app.jobs import submit_job
    return submit_job(run_check_job, "scheduled", owner_id=0, cancel_on_disconnect=False)
