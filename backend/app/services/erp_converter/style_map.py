"""Maps the flat CSS declarations produced by css_parser (plus structural
tag flags like <b>/<i>/<u> and the valign= attribute) into cached openpyxl
style objects (Font, PatternFill, Border, Alignment).

Oracle BI Publisher reuses a small, bounded set of CSS classes across an
entire report (tens to a few hundred), but the same *combination* of
class + bold + valign can recur across tens of thousands of cells. We
build every openpyxl style object at most once and hand back the cached
instance, which is what makes writing million-cell reports tractable.
"""
import re
from dataclasses import dataclass, field
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

_COLOR_RE = re.compile(r"^#([0-9a-fA-F]{6}|[0-9a-fA-F]{3})$")

_DEFAULT_FONT_NAME = "Calibri"
_DEFAULT_FONT_SIZE = 11.0


def _norm_color(value):
    """CSS '#rrggbb' -> openpyxl ARGB hex string, or None if not a plain
    hex color (named CSS colors are not used by these reports)."""
    if not value:
        return None
    value = value.strip()
    m = _COLOR_RE.match(value)
    if not m:
        return None
    hexpart = m.group(1)
    if len(hexpart) == 3:
        hexpart = "".join(c * 2 for c in hexpart)
    return "FF" + hexpart.upper()


def _parse_pt(value):
    """'12.5pt' / '12.5px' / '12.5' -> float points (px treated as pt; BI
    Publisher's CSS is pt-based already)."""
    if not value:
        return None
    m = re.match(r"^(-?\d+(?:\.\d+)?)", value.strip())
    if not m:
        return None
    return float(m.group(1))


def _border_style_for_width(width_pt):
    if width_pt is None:
        return "thin"
    if width_pt <= 0:
        return None
    if width_pt < 1.0:
        return "thin"
    if width_pt < 2.25:
        return "medium"
    return "thick"


@dataclass(frozen=True)
class CellStyle:
    font_name: str = _DEFAULT_FONT_NAME
    font_size: float = _DEFAULT_FONT_SIZE
    font_color: str = None
    bold: bool = False
    italic: bool = False
    underline: bool = False
    bg_color: str = None
    border_width: float = None
    border_color: str = None
    border_style_css: str = None
    h_align: str = None
    v_align: str = None
    wrap: bool = False


_SIDE_PROP_NAMES = ("top", "right", "bottom", "left")


def resolve_cell_style(props, bold=False, italic=False, underline=False, valign=None, wrap=False):
    """props: merged CSS declarations dict for one cell (already resolved
    from possibly multiple classes on nested spans/paragraphs/cells).
    """
    font_name = props.get("font-family", _DEFAULT_FONT_NAME)
    size = _parse_pt(props.get("font-size"))
    font_size = size if size else _DEFAULT_FONT_SIZE
    font_color = _norm_color(props.get("color"))
    bg_color = _norm_color(props.get("background-color"))

    border_width = _parse_pt(props.get("border-width"))
    border_color = _norm_color(props.get("border-color")) or "FF000000"
    border_style_css = props.get("border-style")

    h_align = props.get("text-align")
    if h_align not in ("left", "right", "center", "justify"):
        h_align = None

    return CellStyle(
        font_name=font_name,
        font_size=font_size,
        font_color=font_color,
        bold=bold or props.get("font-weight") == "bold",
        italic=italic or props.get("font-style") == "italic",
        underline=underline or props.get("text-decoration") == "underline",
        bg_color=bg_color,
        border_width=border_width,
        border_color=border_color,
        border_style_css=border_style_css,
        h_align=h_align,
        v_align=valign,
        wrap=wrap,
    )


class StyleCache:
    """Caches openpyxl Font/Fill/Border/Alignment objects keyed by their
    resolved CellStyle so identical-looking cells share one style object.
    """

    def __init__(self):
        self._fonts = {}
        self._fills = {}
        self._borders = {}
        self._aligns = {}

    def font(self, style: CellStyle):
        key = (style.font_name, style.font_size, style.font_color, style.bold,
               style.italic, style.underline)
        obj = self._fonts.get(key)
        if obj is None:
            obj = Font(
                name=style.font_name,
                size=style.font_size,
                color=style.font_color,
                bold=style.bold,
                italic=style.italic,
                underline="single" if style.underline else None,
            )
            self._fonts[key] = obj
        return obj

    def fill(self, style: CellStyle):
        if not style.bg_color or style.bg_color == "FFFFFFFF":
            return None
        key = style.bg_color
        obj = self._fills.get(key)
        if obj is None:
            obj = PatternFill(fill_type="solid", fgColor=key)
            self._fills[key] = obj
        return obj

    def border(self, style: CellStyle):
        if style.border_width is None and not style.border_style_css:
            return None
        side_name = _border_style_for_width(style.border_width)
        if side_name is None:
            return None
        key = (side_name, style.border_color)
        obj = self._borders.get(key)
        if obj is None:
            side = Side(style=side_name, color=style.border_color)
            obj = Border(top=side, bottom=side, left=side, right=side)
            self._borders[key] = obj
        return obj

    def alignment(self, style: CellStyle):
        key = (style.h_align, style.v_align, style.wrap)
        obj = self._aligns.get(key)
        if obj is None:
            v = {"top": "top", "middle": "center", "bottom": "bottom"}.get(style.v_align)
            obj = Alignment(horizontal=style.h_align, vertical=v, wrap_text=style.wrap)
            self._aligns[key] = obj
        return obj
