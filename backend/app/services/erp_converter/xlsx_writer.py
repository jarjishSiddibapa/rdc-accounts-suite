"""Writes the ordered Block stream from html_table_parser into a real
.xlsx workbook, preserving fonts, colors, borders, alignment and merged
cells.

The default writer streams worksheet rows as OOXML directly instead of
creating one openpyxl ``WriteOnlyCell`` per value.  openpyxl is retained
only for the small workbook/style package: it remains the proven source of
our fonts, fills, borders, alignments and number formats, while the hot
million-cell path is a few escaped byte writes.  The previous optimized
write-only implementation remains available as ``OpenpyxlXlsxWriter`` for
automatic compatibility fallback and emergency rollback.

Performance note: openpyxl deduplicates style objects (Font/Fill/Border/
Alignment) into shared workbook-level lists via `IndexedList.add()`,
which hashes the *entire* style object on every single `cell.font = ...`
assignment - even when the exact same cached Python object is reused for
millions of cells. Profiling a 430k-cell report showed >70s spent purely
in that repeated hashing. Since ERP reports only ever have a few dozen
distinct style combinations no matter how many rows they have, we bypass
the descriptor machinery: register each distinct Font/Fill/Border/
Alignment/number-format with the workbook exactly once (indices cached
by Python object identity), then write the resulting 9-int StyleArray
onto each cell directly. This is the same internal representation
openpyxl's own descriptors build - just without redoing the expensive
lookup for every cell.
"""
import logging
import os
import shutil
import tempfile
import zipfile
from datetime import date, datetime
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.compat.strings import safe_string
from openpyxl.utils.datetime import to_excel
from openpyxl.utils import get_column_letter
from openpyxl.styles.cell_style import StyleArray
from openpyxl.styles.numbers import BUILTIN_FORMATS_REVERSE, BUILTIN_FORMATS_MAX_SIZE
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange

from .errors import ConversionError
from .style_map import StyleCache, resolve_cell_style
from .value_parser import parse_value
from .html_table_parser import TableBlock, TextBlock
from .io_retry import with_retry

_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 8
_MAX_TRACKED_COLS = 500  # safety cap; reports rarely exceed a few dozen columns

# Excel's hard per-sheet limits (XLSX format spec, unrelated to openpyxl).
# write-only mode has no idea these exist and will happily keep writing
# past them, producing a file that Excel then refuses to open cleanly
# ("unreadable content... repair") - checked as we go so a report that
# would exceed the limit fails fast with a clear, actionable message
# instead of silently shipping a corrupt-on-open output.
_MAX_EXCEL_ROWS = 1_048_576
_MAX_EXCEL_COLS = 16_384

_EMPTY_PROPS = {}  # shared singleton so blank/filler cells share one cache identity

_SHEET_PATH = "xl/worksheets/sheet1.xml"
_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XML_PREFIX = b'<?xml version="1.0" encoding="utf-8" standalone="yes"?>'

logger = logging.getLogger(__name__)


class DirectXlsxWriterError(RuntimeError):
    """The direct OOXML writer failed and the compatibility writer may retry."""


class DirectXlsxWriter:
    """Low-allocation worksheet writer for the HTML-in-XLS ERP reports.

    Row fragments are streamed into a ``SpooledTemporaryFile``.  Only the
    first MiB can live in memory; larger reports transparently roll to disk.
    Finalization asks openpyxl for a tiny, empty one-sheet workbook containing
    our registered styles, then atomically replaces that empty sheet part with
    the streamed XML.  ZIP entry order is irrelevant in OOXML, so styles can
    be discovered while rows are written without buffering the report.
    """

    def __init__(self, progress_cb=None):
        self.wb = Workbook()
        self.wb.active.title = "Report"
        self.style_cache = StyleCache()
        self.col_max_len = {}
        self.current_row = 0
        self.max_col_seen = 0
        self.merges = []
        self._wrote_any = False
        self.progress_cb = progress_cb
        self._rows = tempfile.SpooledTemporaryFile(max_size=1024 * 1024, mode="w+b")
        self._closed = False

        self._font_ids = {}
        self._fill_ids = {}
        self._border_ids = {}
        self._align_ids = {}
        self._numfmt_ids = {}
        self._bundle_cache = {}
        self._style_id_cache = {}
        self._column_letters = {}

    # ------------------------------------------------------------ styles
    def _reg(self, cache, coll, obj):
        if obj is None:
            return 0
        key = id(obj)
        idx = cache.get(key)
        if idx is None:
            idx = coll.add(obj)
            cache[key] = idx
        return idx

    def _numfmt_index(self, fmt):
        if not fmt:
            return 0
        idx = self._numfmt_ids.get(fmt)
        if idx is not None:
            return idx
        if fmt in BUILTIN_FORMATS_REVERSE:
            idx = BUILTIN_FORMATS_REVERSE[fmt]
        else:
            idx = self.wb._number_formats.add(fmt) + BUILTIN_FORMATS_MAX_SIZE
        self._numfmt_ids[fmt] = idx
        return idx

    def _style_bundle(self, props, bold, italic, underline, valign, wrap):
        key = (id(props), bold, italic, underline, valign, wrap)
        bundle = self._bundle_cache.get(key)
        if bundle is not None:
            return bundle
        style = resolve_cell_style(
            props, bold=bold, italic=italic, underline=underline,
            valign=valign, wrap=wrap,
        )
        font = self.style_cache.font(style)
        fill = self.style_cache.fill(style)
        border = self.style_cache.border(style)
        align = self.style_cache.alignment(style)
        if not (align.horizontal or align.vertical or align.wrap_text):
            align = None
        bundle = (
            self._reg(self._font_ids, self.wb._fonts, font),
            self._reg(self._fill_ids, self.wb._fills, fill),
            self._reg(self._border_ids, self.wb._borders, border),
            self._reg(self._align_ids, self.wb._alignments, align),
        )
        self._bundle_cache[key] = bundle
        return bundle

    def _style_id(self, numfmt, props, bold=False, italic=False,
                  underline=False, valign=None, wrap=False):
        font_idx, fill_idx, border_idx, align_idx = self._style_bundle(
            props, bold, italic, underline, valign, wrap,
        )
        numfmt_idx = self._numfmt_index(numfmt)
        key = (font_idx, fill_idx, border_idx, numfmt_idx, align_idx)
        cached = self._style_id_cache.get(key)
        if cached is not None:
            return cached
        style_array = StyleArray(
            [font_idx, fill_idx, border_idx, numfmt_idx, 0, align_idx, 0, 0, 0]
        )
        style_id = self.wb._cell_styles.add(style_array)
        self._style_id_cache[key] = style_id
        return style_id

    # ------------------------------------------------------------ XML helpers
    def _column_letter(self, col_index):
        letter = self._column_letters.get(col_index)
        if letter is None:
            letter = get_column_letter(col_index + 1)
            self._column_letters[col_index] = letter
        return letter

    @staticmethod
    def _checked_text(value):
        value = value[:32767]
        if ILLEGAL_CHARACTERS_RE.search(value):
            raise DirectXlsxWriterError(
                "The report contains a control character that Excel cannot store."
            )
        return value

    def _cell_xml(self, row_number, col_index, value, style_id):
        reference = f"{self._column_letter(col_index)}{row_number}"
        style_attr = f' s="{style_id}"' if style_id else ""
        if value is None:
            if not style_id:
                return b""
            return f'<c r="{reference}"{style_attr}></c>'.encode("utf-8")

        if isinstance(value, str):
            value = self._checked_text(value)
            preserve = ' xml:space="preserve"' if value[:1].isspace() or value[-1:].isspace() else ""
            text = escape(value)
            return (
                f'<c r="{reference}"{style_attr} t="inlineStr"><is>'
                f'<t{preserve}>{text}</t></is></c>'
            ).encode("utf-8")

        if isinstance(value, bool):
            return (
                f'<c r="{reference}"{style_attr} t="b"><v>{1 if value else 0}</v></c>'
            ).encode("ascii")

        if isinstance(value, (date, datetime)):
            value = to_excel(value, self.wb.epoch)
        # Match openpyxl's numeric serialization exactly (16 significant
        # digits and blank NaN/Infinity) so switching writers cannot alter a
        # financial value by one binary-float rounding unit.
        numeric = safe_string(value)
        return f'<c r="{reference}"{style_attr} t="n"><v>{numeric}</v></c>'.encode("ascii")

    def _write_row(self, cells, height=None):
        self._check_row_capacity()
        self.current_row += 1
        attrs = f' r="{self.current_row}"'
        if height is not None:
            attrs += f' ht="{height:.15g}" customHeight="1"'
        self._rows.write(f"<row{attrs}>".encode("ascii"))
        for cell in cells:
            if cell:
                self._rows.write(cell)
        self._rows.write(b"</row>")

    def _track_width(self, col_index, length):
        if col_index >= _MAX_TRACKED_COLS:
            return
        cur = self.col_max_len.get(col_index, 0)
        if length > cur:
            self.col_max_len[col_index] = length

    def _spacer_if_needed(self):
        if self._wrote_any:
            self._write_row([])
        self._wrote_any = True

    def _check_row_capacity(self):
        if self.current_row + 1 > _MAX_EXCEL_ROWS:
            raise ConversionError(
                f"This report converts to more than {_MAX_EXCEL_ROWS:,} rows, "
                "which exceeds Excel's per-sheet limit. Split the source "
                "export into smaller date ranges or batches and convert "
                "each one separately."
            )

    @staticmethod
    def _check_col_capacity(ncols):
        if ncols > _MAX_EXCEL_COLS:
            raise ConversionError(
                f"This report has more than {_MAX_EXCEL_COLS:,} columns, "
                "which exceeds Excel's per-sheet limit."
            )

    # ------------------------------------------------------------- write
    def write_blocks(self, blocks_iter):
        try:
            for blk in blocks_iter:
                if isinstance(blk, TextBlock):
                    self._write_text_block(blk)
                else:
                    self._write_table_block(blk)
                if self.progress_cb:
                    self.progress_cb(self.current_row)
        except ConversionError:
            raise
        except DirectXlsxWriterError:
            raise
        except Exception as exc:
            raise DirectXlsxWriterError("Could not stream the worksheet XML.") from exc

    def _write_text_block(self, blk):
        if blk.text == "":
            return
        self._spacer_if_needed()
        style_id = self._style_id(
            None, blk.props, bold=blk.bold, italic=blk.italic,
            underline=blk.underline, wrap="\n" in blk.text,
        )
        self._write_row([self._cell_xml(self.current_row + 1, 0, blk.text, style_id)])
        self.max_col_seen = max(self.max_col_seen, 1)
        longest = max((len(line) for line in blk.text.split("\n")), default=0)
        self._track_width(0, longest)

    def _write_table_block(self, blk: TableBlock):
        if blk.nrows == 0 or blk.ncols == 0:
            return
        self._check_col_capacity(blk.ncols)
        if self.current_row + blk.nrows + (1 if self._wrote_any else 0) > _MAX_EXCEL_ROWS:
            raise ConversionError(
                f"This report converts to more than {_MAX_EXCEL_ROWS:,} rows, "
                "which exceeds Excel's per-sheet limit. Split the source "
                "export into smaller date ranges or batches and convert "
                "each one separately."
            )
        self._spacer_if_needed()
        base_row = self.current_row
        self.max_col_seen = max(self.max_col_seen, blk.ncols)
        for row_index in range(blk.nrows):
            row_number = self.current_row + 1
            cells = []
            for col_index in range(blk.ncols):
                entry = blk.cells.get((row_index, col_index))
                if entry is None:
                    continue
                if isinstance(entry, tuple):
                    anchor = entry[1]
                    style_id = self._style_id(
                        None, anchor.props, bold=anchor.bold, italic=anchor.italic,
                        underline=anchor.underline, valign=anchor.valign,
                    )
                    xml = self._cell_xml(row_number, col_index, None, style_id)
                    if xml:
                        cells.append(xml)
                    continue

                cell_spec = entry
                value, numfmt = parse_value(cell_spec.text)
                style_id = self._style_id(
                    numfmt, cell_spec.props, bold=cell_spec.bold,
                    italic=cell_spec.italic, underline=cell_spec.underline,
                    valign=cell_spec.valign, wrap=cell_spec.wrap,
                )
                xml = self._cell_xml(row_number, col_index, value, style_id)
                if xml:
                    cells.append(xml)
                text_len = len(str(value)) if value is not None else 0
                self._track_width(col_index, text_len)
                if cell_spec.rowspan > 1 or cell_spec.colspan > 1:
                    r1 = base_row + row_index + 1
                    c1 = col_index + 1
                    self.merges.append((
                        r1, c1,
                        r1 + cell_spec.rowspan - 1,
                        c1 + cell_spec.colspan - 1,
                    ))
            height = blk.row_heights[row_index] if row_index < len(blk.row_heights) else None
            self._write_row(cells, height=height)

    # ------------------------------------------------------------- package
    def _sheet_prefix(self):
        max_ref = f"{self._column_letter(max(self.max_col_seen - 1, 0))}{max(self.current_row, 1)}"
        parts = [
            _XML_PREFIX,
            f'<worksheet xmlns="{_SHEET_NS}">'.encode("ascii"),
            b'<sheetPr><outlinePr summaryBelow="1" summaryRight="1"/>'
            b'<pageSetUpPr/></sheetPr>',
            f'<dimension ref="A1:{max_ref}"/>'.encode("ascii") if max_ref != "A1" else b'<dimension ref="A1"/>',
            b'<sheetViews><sheetView workbookViewId="0"><selection activeCell="A1" sqref="A1"/>'
            b'</sheetView></sheetViews>',
            b'<sheetFormatPr baseColWidth="8" defaultRowHeight="15"/>',
        ]
        # The compatibility writer historically applied calculated widths
        # after write-only rows were emitted. openpyxl cannot serialize column
        # definitions at that point, so production output has always used
        # Excel's default widths. Preserve that visible behavior exactly.
        parts.append(b"<sheetData>")
        return b"".join(parts)

    def _sheet_suffix(self):
        parts = [b"</sheetData>"]
        if self.merges:
            parts.append(f'<mergeCells count="{len(self.merges)}">'.encode("ascii"))
            for r1, c1, r2, c2 in self.merges:
                ref = f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
                parts.append(f'<mergeCell ref="{ref}"/>'.encode("ascii"))
            parts.append(b"</mergeCells>")
        parts.extend([
            b'<pageMargins left="0.75" right="0.75" top="1" bottom="1" header="0.5" footer="0.5"/>',
            b"</worksheet>",
        ])
        return b"".join(parts)

    def _package_once(self, output_path):
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(prefix=".erp-direct-", dir=output_path.parent) as tmp:
            tmp_dir = Path(tmp)
            template_path = tmp_dir / "template.xlsx"
            partial_path = tmp_dir / "output.partial.xlsx"
            self.wb.save(template_path)

            with zipfile.ZipFile(template_path, "r") as source, zipfile.ZipFile(
                # Level 1 preserves the exact workbook contents while avoiding
                # level 6's disproportionate CPU cost on 20-150 MB worksheet
                # XML streams. ERP downloads grow only modestly because the XML
                # remains highly repetitive.
                partial_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=1,
            ) as target:
                for info in source.infolist():
                    if info.filename == _SHEET_PATH:
                        continue
                    target.writestr(info, source.read(info.filename))
                with target.open(_SHEET_PATH, "w", force_zip64=True) as sheet:
                    sheet.write(self._sheet_prefix())
                    self._rows.seek(0)
                    shutil.copyfileobj(self._rows, sheet, length=1024 * 1024)
                    sheet.write(self._sheet_suffix())

            with zipfile.ZipFile(partial_path, "r") as check:
                required = {"[Content_Types].xml", "xl/workbook.xml", "xl/styles.xml", _SHEET_PATH}
                missing = required.difference(check.namelist())
                if missing:
                    raise DirectXlsxWriterError(
                        "The generated workbook package is incomplete: " + ", ".join(sorted(missing))
                    )
            os.replace(partial_path, output_path)

    def finalize(self, output_path):
        try:
            with_retry(lambda: self._package_once(output_path), output_path, "write")
        except ConversionError:
            raise
        except DirectXlsxWriterError:
            raise
        except Exception as exc:
            raise DirectXlsxWriterError("Could not assemble the direct OOXML workbook.") from exc
        finally:
            self.close()

    def close(self):
        if not self._closed:
            self._rows.close()
            self._closed = True



class OpenpyxlXlsxWriter:
    def __init__(self, progress_cb=None):
        self.wb = Workbook(write_only=True)
        self.ws = self.wb.create_sheet(title="Report")
        self.style_cache = StyleCache()
        self.col_max_len = {}
        self.current_row = 0
        self.merges = []
        self._wrote_any = False
        self.progress_cb = progress_cb

        # id(openpyxl style object) -> registered index in the workbook's
        # shared style list. Safe because our StyleCache keeps every such
        # object alive for the lifetime of this writer, so ids never get
        # reused for something else in the meantime.
        self._font_ids = {}
        self._fill_ids = {}
        self._border_ids = {}
        self._align_ids = {}
        self._numfmt_ids = {}

        # A resolved-style *bundle* (font/fill/border/alignment index
        # tuple) keyed by (id(props_dict), bold, italic, underline,
        # valign, wrap) - props dicts are themselves de-duplicated and
        # cached upstream by the parser, so this collapses the huge
        # majority of cells in a report down to a handful of distinct
        # style computations.
        self._bundle_cache = {}

        # (font_idx, fill_idx, border_idx, numfmt_idx, align_idx) -> the one
        # StyleArray instance for that exact combination. ERP reports only
        # ever have a handful of distinct combinations, so reusing the same
        # object for every matching cell (instead of allocating a fresh
        # 9-int array per cell) cuts allocation/GC pressure across the
        # million-cell reports these files can contain.
        self._style_array_cache = {}

    # ------------------------------------------------------------ id-based interning
    def _reg(self, cache, coll, obj):
        if obj is None:
            return 0
        key = id(obj)
        idx = cache.get(key)
        if idx is None:
            idx = coll.add(obj)
            cache[key] = idx
        return idx

    def _numfmt_index(self, fmt):
        if not fmt:
            return 0
        idx = self._numfmt_ids.get(fmt)
        if idx is not None:
            return idx
        if fmt in BUILTIN_FORMATS_REVERSE:
            idx = BUILTIN_FORMATS_REVERSE[fmt]
        else:
            idx = self.wb._number_formats.add(fmt) + BUILTIN_FORMATS_MAX_SIZE
        self._numfmt_ids[fmt] = idx
        return idx

    def _style_bundle(self, props, bold, italic, underline, valign, wrap):
        key = (id(props), bold, italic, underline, valign, wrap)
        bundle = self._bundle_cache.get(key)
        if bundle is not None:
            return bundle
        style = resolve_cell_style(props, bold=bold, italic=italic,
                                    underline=underline, valign=valign, wrap=wrap)
        font = self.style_cache.font(style)
        fill = self.style_cache.fill(style)
        border = self.style_cache.border(style)
        align = self.style_cache.alignment(style)
        if not (align.horizontal or align.vertical or align.wrap_text):
            align = None
        font_idx = self._reg(self._font_ids, self.wb._fonts, font)
        fill_idx = self._reg(self._fill_ids, self.wb._fills, fill)
        border_idx = self._reg(self._border_ids, self.wb._borders, border)
        align_idx = self._reg(self._align_ids, self.wb._alignments, align)
        bundle = (font_idx, fill_idx, border_idx, align_idx)
        self._bundle_cache[key] = bundle
        return bundle

    # ------------------------------------------------------------ helpers
    def _make_cell(self, value, numfmt, props, bold=False, italic=False,
                    underline=False, valign=None, wrap=False):
        cell = WriteOnlyCell(self.ws, value=value)
        font_idx, fill_idx, border_idx, align_idx = self._style_bundle(
            props, bold, italic, underline, valign, wrap)
        numfmt_idx = self._numfmt_index(numfmt)
        if font_idx or fill_idx or border_idx or align_idx or numfmt_idx:
            style_key = (font_idx, fill_idx, border_idx, numfmt_idx, align_idx)
            style_array = self._style_array_cache.get(style_key)
            if style_array is None:
                style_array = StyleArray(
                    [font_idx, fill_idx, border_idx, numfmt_idx, 0, align_idx, 0, 0, 0])
                self._style_array_cache[style_key] = style_array
            cell._style = style_array
        return cell

    def _track_width(self, col_index, length):
        if col_index >= _MAX_TRACKED_COLS:
            return
        cur = self.col_max_len.get(col_index, 0)
        if length > cur:
            self.col_max_len[col_index] = length

    def _spacer_if_needed(self):
        if self._wrote_any:
            self._check_row_capacity()
            self.ws.append([])
            self.current_row += 1
        self._wrote_any = True

    def _check_row_capacity(self):
        if self.current_row + 1 > _MAX_EXCEL_ROWS:
            raise ConversionError(
                f"This report converts to more than {_MAX_EXCEL_ROWS:,} rows, "
                "which exceeds Excel's per-sheet limit. Split the source "
                "export into smaller date ranges or batches and convert "
                "each one separately."
            )

    def _check_col_capacity(self, ncols):
        if ncols > _MAX_EXCEL_COLS:
            raise ConversionError(
                f"This report has more than {_MAX_EXCEL_COLS:,} columns, "
                "which exceeds Excel's per-sheet limit."
            )

    # ------------------------------------------------------------- write
    def write_blocks(self, blocks_iter):
        for blk in blocks_iter:
            if isinstance(blk, TextBlock):
                self._write_text_block(blk)
            else:
                self._write_table_block(blk)
            if self.progress_cb:
                self.progress_cb(self.current_row)

    def _write_text_block(self, blk):
        if blk.text == "":
            return
        self._spacer_if_needed()
        self._check_row_capacity()
        cell = self._make_cell(blk.text, None, blk.props, bold=blk.bold,
                                italic=blk.italic, underline=blk.underline,
                                wrap="\n" in blk.text)
        self.ws.append([cell])
        self.current_row += 1
        longest = max((len(ln) for ln in blk.text.split("\n")), default=0)
        self._track_width(0, longest)

    def _write_table_block(self, blk: TableBlock):
        if blk.nrows == 0 or blk.ncols == 0:
            return
        self._check_col_capacity(blk.ncols)
        if self.current_row + blk.nrows > _MAX_EXCEL_ROWS:
            raise ConversionError(
                f"This report converts to more than {_MAX_EXCEL_ROWS:,} rows, "
                "which exceeds Excel's per-sheet limit. Split the source "
                "export into smaller date ranges or batches and convert "
                "each one separately."
            )
        self._spacer_if_needed()
        base_row = self.current_row  # rows already written before this block (0-based)
        for r in range(blk.nrows):
            row_cells = []
            for c in range(blk.ncols):
                entry = blk.cells.get((r, c))
                if entry is None:
                    row_cells.append(self._make_cell(None, None, _EMPTY_PROPS))
                    continue
                if isinstance(entry, tuple):  # ("span", anchor_cell) - merge continuation
                    anchor = entry[1]
                    row_cells.append(self._make_cell(
                        None, None, anchor.props, bold=anchor.bold,
                        italic=anchor.italic, underline=anchor.underline,
                        valign=anchor.valign))
                    continue
                cell_spec = entry
                value, numfmt = parse_value(cell_spec.text)
                row_cells.append(self._make_cell(
                    value, numfmt, cell_spec.props, bold=cell_spec.bold,
                    italic=cell_spec.italic, underline=cell_spec.underline,
                    valign=cell_spec.valign, wrap=cell_spec.wrap))
                text_len = len(str(value)) if value is not None else 0
                self._track_width(c, text_len)
                if cell_spec.rowspan > 1 or cell_spec.colspan > 1:
                    r1 = base_row + r + 1
                    c1 = c + 1
                    r2 = r1 + cell_spec.rowspan - 1
                    c2 = c1 + cell_spec.colspan - 1
                    self.merges.append((r1, c1, r2, c2))
            self.ws.append(row_cells)
            self.current_row += 1
            rh = blk.row_heights[r] if r < len(blk.row_heights) else None
            if rh:
                self.ws.row_dimensions[base_row + r + 1].height = rh

    # ------------------------------------------------------------- finish
    def finalize(self, output_path):
        # WriteOnlyWorksheet doesn't inherit merge_cells()/self.merged_cells
        # from the regular Worksheet class (it's absent from the small set
        # of methods write-only mode copies over) - calling ws.merge_cells()
        # here has always raised AttributeError, silently swallowed by a
        # bare except, so colspan/rowspan cells were never actually merged
        # in any converted output. WorksheetWriter.write_tail() does read
        # ws.merged_cells if present, so setting it directly (bypassing the
        # missing method) is enough to make merges reach the file.
        if self.merges:
            self.ws.merged_cells = MultiCellRange(
                CellRange(min_col=c1, min_row=r1, max_col=c2, max_row=r2)
                for (r1, c1, r2, c2) in self.merges
            )
        for col_index, length in self.col_max_len.items():
            width = max(_MIN_COL_WIDTH, min(_MAX_COL_WIDTH, length + 2))
            self.ws.column_dimensions[get_column_letter(col_index + 1)].width = width
        with_retry(lambda: self.wb.save(output_path), output_path, "write")

    def close(self):
        self.wb.close()


# Public default used by converter.py. Keeping the alias makes the change
# reversible without touching callers and gives tests an explicit direct class.
XlsxWriter = DirectXlsxWriter
