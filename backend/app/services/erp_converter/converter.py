"""Top-level orchestrator: figure out what kind of file the ERP handed us
and convert it into a real, styled .xlsx.

Oracle report outputs come in three shapes in practice:
  1. "HTML saved as .xls" - what BI Publisher / Oracle Reports actually
     emits for almost every seeded report (all four sample files are
     this). This is the primary, fully-featured path.
  2. Genuine legacy binary .xls (BIFF/OLE2) - occasionally produced by
     older report definitions. Converted via xlrd with best-effort
     formatting; values always come through correctly.
  3. Genuine .xlsx - already a real spreadsheet; copied through as-is.
"""
import os
import shutil

from .errors import ConversionError
from .html_table_parser import HtmlReportParser
from .xlsx_writer import XlsxWriter
from .io_retry import with_retry

XLS_OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
XLSX_ZIP_SIGNATURE = b"PK\x03\x04"

__all__ = ["ConversionError", "convert_file", "detect_kind"]


def _read_head(path):
    with open(path, "rb") as f:
        return f.read(4096)


def detect_kind(path):
    head = with_retry(lambda: _read_head(path), path, "read")
    if head.startswith(XLSX_ZIP_SIGNATURE):
        return "xlsx"
    if head.startswith(XLS_OLE_SIGNATURE):
        return "xls-binary"
    if b"<html" in head[:2048].lower() or b"<table" in head.lower():
        return "html"
    return "html"  # best-effort default: most ERP exports are HTML-in-xls


def convert_file(input_path, output_path, progress_cb=None):
    """progress_cb(fraction 0..1, phase str) is called periodically."""
    kind = detect_kind(input_path)
    if kind == "html":
        _convert_html(input_path, output_path, progress_cb)
    elif kind == "xlsx":
        _passthrough_xlsx(input_path, output_path, progress_cb)
    elif kind == "xls-binary":
        _convert_binary_xls(input_path, output_path, progress_cb)
    else:
        raise ConversionError(f"Unrecognized file format: {input_path}")
    return kind


def _report(progress_cb, frac, phase):
    if progress_cb:
        progress_cb(frac, phase)


def _convert_html(input_path, output_path, progress_cb):
    file_size = os.path.getsize(input_path) or 1
    parser = HtmlReportParser()
    writer = XlsxWriter()

    def bytes_cb(n):
        _report(progress_cb, min(0.95, n / file_size), "Parsing report")

    _report(progress_cb, 0.0, "Parsing report")
    writer.write_blocks(parser.parse(input_path, progress_cb=bytes_cb))
    _report(progress_cb, 0.97, "Writing workbook")
    writer.finalize(output_path)
    _report(progress_cb, 1.0, "Done")


def _passthrough_xlsx(input_path, output_path, progress_cb):
    _report(progress_cb, 0.1, "Copying workbook")
    with_retry(lambda: shutil.copyfile(input_path, output_path), output_path, "write")
    _report(progress_cb, 1.0, "Done")


def _convert_binary_xls(input_path, output_path, progress_cb):
    import xlrd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    _report(progress_cb, 0.05, "Reading legacy workbook")
    book = with_retry(lambda: xlrd.open_workbook(input_path, formatting_info=True),
                       input_path, "read")
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    font_cache = {}
    fill_cache = {}

    def font_for(xf_index):
        xf = book.xf_list[xf_index]
        font = book.font_list[xf.font_index]
        key = (font.bold, font.italic, font.underline_type != 0, font.height)
        obj = font_cache.get(key)
        if obj is None:
            obj = Font(bold=bool(font.bold), italic=bool(font.italic),
                       underline="single" if font.underline_type else None,
                       size=(font.height or 200) / 20.0)
            font_cache[key] = obj
        return obj

    def fill_for(xf_index):
        xf = book.xf_list[xf_index]
        pattern = xf.background
        color_idx = pattern.pattern_colour_index
        if pattern.fill_pattern == 0 or color_idx is None:
            return None
        rgb = book.colour_map.get(color_idx)
        if not rgb:
            return None
        key = rgb
        obj = fill_cache.get(key)
        if obj is None:
            hexval = "FF%02X%02X%02X" % rgb
            obj = PatternFill(fill_type="solid", fgColor=hexval)
            fill_cache[key] = obj
        return obj

    n_sheets = book.nsheets
    for si in range(n_sheets):
        sheet = book.sheet_by_index(si)
        out_ws = out_wb.create_sheet(title=sheet.name[:31] or f"Sheet{si+1}")
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    continue
                out_cell = out_ws.cell(row=r + 1, column=c + 1, value=cell.value)
                try:
                    xf_index = sheet.cell_xf_index(r, c)
                    out_cell.font = font_for(xf_index)
                    fill = fill_for(xf_index)
                    if fill:
                        out_cell.fill = fill
                except Exception:
                    pass
        for c in range(sheet.ncols):
            out_ws.column_dimensions[get_column_letter(c + 1)].width = 14
        _report(progress_cb, 0.1 + 0.8 * (si + 1) / max(n_sheets, 1),
                f"Converting sheet {si+1}/{n_sheets}")

    _report(progress_cb, 0.95, "Writing workbook")
    with_retry(lambda: out_wb.save(output_path), output_path, "write")
    _report(progress_cb, 1.0, "Done")
