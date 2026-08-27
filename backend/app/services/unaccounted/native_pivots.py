"""Native Excel PivotTable support without automating a local Excel process.

OpenPyXL can preserve an existing PivotTable but cannot create one through a
public management API.  Pending MRN therefore starts from a tiny, sanitized
workbook containing two genuine pivot definitions - Vendorwise Pivot and
Locationwise Pivot - and replaces the non-pivot sheets with the freshly
generated report.  Excel refreshes both pivots from that workbook's own data
ranges when the user opens the attachment.

Both pivots must be merged into the output in a single pass: OpenPyXL merges
by keeping the template's pivot sheet(s) as the live objects and cell-copying
every *other* sheet into the same workbook, so a sheet that already contains
a real pivot from an earlier, separate merge would itself get cell-copied
(losing its pivot) if a second merge ran against the same file afterward.
"""

from __future__ import annotations

import logging
from copy import copy
import os
from pathlib import Path
import tempfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.xlsx_formula_cache import cache_formula_values, inject_cached_values

logger = logging.getLogger(__name__)


_PIVOT_SHEET = "Vendorwise Pivot"
_LOCATIONWISE_SHEET = "Locationwise Pivot"
_SUMMARY_SHEET = "Summary"
_MRN_DATA_SHEET = "_PivotSrc"
_MRN_DATA_COLUMNS = ["Dim1", "Dim2", "Period", "Value"]
_TEMPLATE_PATH = Path(__file__).with_name("pivot_templates") / "pending_mrn_vendorwise.xlsx"


def _copy_report_sheet(source, target) -> None:
    """Copy one generated report sheet into the sanitized template workbook."""
    target.sheet_format = copy(source.sheet_format)
    target.sheet_properties = copy(source.sheet_properties)
    target.views = copy(source.views)
    target.freeze_panes = source.freeze_panes
    target.auto_filter.ref = source.auto_filter.ref
    target.sheet_state = source.sheet_state
    target.print_options = copy(source.print_options)
    target.page_margins = copy(source.page_margins)
    target.page_setup = copy(source.page_setup)

    for key, dimension in source.column_dimensions.items():
        copied = target.column_dimensions[key]
        copied.width = dimension.width
        copied.hidden = dimension.hidden
        copied.bestFit = dimension.bestFit
        copied.outlineLevel = dimension.outlineLevel

    for key, dimension in source.row_dimensions.items():
        copied = target.row_dimensions[key]
        copied.height = dimension.height
        copied.hidden = dimension.hidden
        copied.outlineLevel = dimension.outlineLevel

    for row in source.iter_rows():
        for source_cell in row:
            target_cell = target[source_cell.coordinate]
            target_cell.value = source_cell.value
            if source_cell.has_style:
                # Assign components individually so OpenPyXL registers them in
                # the destination workbook's own style tables.
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.protection = copy(source_cell.protection)
                target_cell.number_format = source_cell.number_format
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))


def _validate_mrn_native_pivots(path: Path, expected_summary_ref: str, expected_data_ref: str) -> None:
    workbook = load_workbook(path, data_only=False)
    try:
        vendorwise = workbook[_PIVOT_SHEET]._pivots
        if len(vendorwise) != 1:
            raise ValueError("Pending MRN output does not contain its native Vendorwise PivotTable")
        vsource = vendorwise[0].cache.cacheSource.worksheetSource
        if vsource.sheet != _SUMMARY_SHEET or vsource.ref != expected_summary_ref:
            raise ValueError("Vendorwise PivotTable cache points at the wrong source range")
        if vendorwise[0].cache.refreshOnLoad is not True or vendorwise[0].cache.missingItemsLimit != 0:
            raise ValueError("Vendorwise PivotTable is not configured for a clean open-time refresh")

        locationwise = workbook[_LOCATIONWISE_SHEET]._pivots
        if len(locationwise) != 1:
            raise ValueError("Pending MRN output does not contain its native Locationwise PivotTable")
        lsource = locationwise[0].cache.cacheSource.worksheetSource
        if lsource.sheet != _MRN_DATA_SHEET or lsource.ref != expected_data_ref:
            raise ValueError("Locationwise PivotTable cache points at the wrong source range")
        if locationwise[0].cache.refreshOnLoad is not True or locationwise[0].cache.missingItemsLimit != 0:
            raise ValueError("Locationwise PivotTable is not configured for a clean open-time refresh")
    finally:
        workbook.close()


def attach_pending_mrn_native_pivots(path: str | Path, locationwise_rows: list[tuple]) -> None:
    """Replace both static MRN pivot sheets with refreshable native PivotTables.

    Both pivots live in one template workbook and are merged in a single
    pass: Vendorwise Pivot refreshes from the copied Summary sheet (as
    before); Locationwise Pivot refreshes from a hidden tidy data sheet built
    here from ``locationwise_rows`` (Location, Accounts Incharge, Period,
    Count - already ordered so periods first appear chronologically). This
    function runs entirely through Python/Open XML; the template's only
    other content is one explicitly synthetic row, discarded on first
    refresh via ``missingItemsLimit = 0``.
    """
    if not locationwise_rows:
        raise ValueError("attach_pending_mrn_native_pivots requires at least one Locationwise data row")

    output_path = Path(path)
    if not _TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"Native PivotTable template is missing: {_TEMPLATE_PATH}")

    generated = load_workbook(output_path, data_only=False)
    template = load_workbook(_TEMPLATE_PATH, data_only=False)
    temp_path: Path | None = None
    try:
        if _PIVOT_SHEET not in generated.sheetnames or _SUMMARY_SHEET not in generated.sheetnames:
            raise ValueError("Pending MRN workbook is missing the PivotTable source or destination sheet")

        # IMPORTANT: never call template.remove(...) once the template has
        # more than one live PivotTable - removing any sheet from a
        # multi-pivot workbook (even one unrelated to either pivot's source)
        # leaves the file opening into Excel's "group edit mode", which then
        # blocks PivotCache.Refresh() on every pivot in the workbook with a
        # misleading error. Confirmed by direct bisection: reusing every
        # template sheet in place (delete_rows, never remove/recreate) and
        # only ever *adding* new sheets avoids it entirely. The raw template
        # has exactly four sheets - both pivots, Summary, and _PivotSrc - so
        # there is never anything to remove in the first place.
        kept = {_PIVOT_SHEET, _LOCATIONWISE_SHEET}

        summary = template[_SUMMARY_SHEET]
        if summary.max_row:
            summary.delete_rows(1, summary.max_row)
        _copy_report_sheet(generated[_SUMMARY_SHEET], summary)

        data_ws = template[_MRN_DATA_SHEET]
        if data_ws.max_row:
            data_ws.delete_rows(1, data_ws.max_row)
        data_ws.append(_MRN_DATA_COLUMNS)
        for row in locationwise_rows:
            data_ws.append(list(row))
        data_ws.sheet_state = "hidden"

        for name in generated.sheetnames:
            if name in kept or name == _SUMMARY_SHEET:
                continue
            destination = template.create_sheet(name, len(template.sheetnames))
            _copy_report_sheet(generated[name], destination)

        # Deliberately never reorder sheets here - not even via the public
        # move_sheet() API. Moving *any* sheet in a workbook that already
        # has two live PivotTables leaves the saved file opening into
        # Excel's "group edit mode", which then blocks
        # PivotCache.Refresh() on every pivot with a misleading error, even
        # though the file still looks structurally correct to OpenPyXL
        # itself (confirmed by direct bisection - only skipping every
        # reorder avoids it). The raw template's own sheet order already
        # matches the desired tab order (Locationwise Pivot, Vendorwise
        # Pivot, Summary), _PivotSrc is hidden so its position among them
        # doesn't show up as a tab, and any genuinely new sheet (e.g.
        # Unmapped Sites) lands in the right place simply by being created
        # at the end, after Summary.
        summary_ref = f"A1:{summary.cell(1, summary.max_column).column_letter}{summary.max_row}"
        vendorwise = template[_PIVOT_SHEET]._pivots[0]
        vendorwise.cache.cacheSource.worksheetSource.ref = summary_ref
        vendorwise.cache.cacheSource.worksheetSource.sheet = _SUMMARY_SHEET
        vendorwise.cache.refreshOnLoad = True
        vendorwise.cache.enableRefresh = True
        vendorwise.cache.invalid = False
        vendorwise.cache.missingItemsLimit = 0
        vendorwise.cache.refreshedBy = None
        vendorwise.cache.refreshedDate = None

        data_ref = f"A1:{get_column_letter(len(_MRN_DATA_COLUMNS))}{1 + len(locationwise_rows)}"
        locationwise = template[_LOCATIONWISE_SHEET]._pivots[0]
        locationwise.cache.cacheSource.worksheetSource.ref = data_ref
        locationwise.cache.cacheSource.worksheetSource.sheet = _MRN_DATA_SHEET
        locationwise.cache.refreshOnLoad = True
        locationwise.cache.enableRefresh = True
        locationwise.cache.invalid = False
        locationwise.cache.missingItemsLimit = 0
        locationwise.cache.refreshedBy = None
        locationwise.cache.refreshedDate = None
        for field, caption in zip(locationwise.pivotFields, ["Location", "Accounts Incharge", "Accounting Period"]):
            field.name = caption
        locationwise.dataFields[0].name = "Count"

        template.calculation.fullCalcOnLoad = True
        template.calculation.forceFullCalc = True

        cached_values = cache_formula_values(template)
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{output_path.stem}-pivot-",
            suffix=".xlsx",
            dir=str(output_path.parent),
        )
        os.close(fd)
        temp_path = Path(temp_name)
        template.save(temp_path)
        inject_cached_values(str(temp_path), cached_values)
        _validate_mrn_native_pivots(temp_path, summary_ref, data_ref)
        os.replace(temp_path, output_path)
        temp_path = None
    finally:
        generated.close()
        template.close()
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


def try_attach_native_pivot(attach_fn, *args, log_q=None, **kwargs) -> bool:
    """Run a native-pivot attach function, but never let it fail the report.

    Every attach function only replaces the already-saved static output on
    success (it builds a temp file and ``os.replace``s it in as the last
    step), so on any exception here the static sheet already on disk from
    the plain ``wb.save(path)`` before this call simply stands unchanged.
    Returns True if the native pivot was attached, False if the static
    fallback was kept.
    """
    try:
        attach_fn(*args, **kwargs)
        return True
    except Exception:
        logger.exception("Native pivot attachment failed for %s; keeping static fallback", attach_fn)
        if log_q is not None:
            try:
                log_q.put(("warn", "Could not build a live Excel pivot for this sheet - kept the static table instead"))
            except Exception:
                pass
        return False


_TWO_LEVEL_TEMPLATE_PATH = Path(__file__).with_name("pivot_templates") / "location_incharge_period.xlsx"
_TWO_LEVEL_PIVOT_SHEET = "Pivot"
_TWO_LEVEL_DATA_SHEET = "_PivotSrc"
_TWO_LEVEL_COLUMNS = ["Dim1", "Dim2", "Period", "Value"]


def attach_two_level_pivot(
    path: str | Path,
    *,
    rows: list[tuple[str, str, str, int]],
    target_sheet: str,
    sheet_position: int,
    tab_color: str | None,
    dim1_caption: str,
    dim2_caption: str,
    period_caption: str,
    value_caption: str,
) -> None:
    """Replace a static Location/Incharge x Period count table with a real,
    refreshable Excel PivotTable, built from a sanitized two-row-field x
    one-column-field template (see ``pivot_templates/location_incharge_period.xlsx``).

    ``rows`` must already be in the order periods should first appear in
    (chronological), matching the same manual-sort-field mechanism used by
    ``attach_pending_mrn_native_pivots``. Row order for Dim1/Dim2 does not
    matter - the template's Location/Accounts-Incharge fields are baked in
    with AutoSort descending by the value field, which Excel recomputes on
    every refresh regardless of source row order.
    """
    if not rows:
        raise ValueError("attach_two_level_pivot requires at least one data row")

    output_path = Path(path)
    if not _TWO_LEVEL_TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"Native PivotTable template is missing: {_TWO_LEVEL_TEMPLATE_PATH}")

    generated = load_workbook(output_path, data_only=False)
    template = load_workbook(_TWO_LEVEL_TEMPLATE_PATH, data_only=False)
    temp_path: Path | None = None
    try:
        for name in list(template.sheetnames):
            if name != _TWO_LEVEL_PIVOT_SHEET:
                template.remove(template[name])

        data_ws = template.create_sheet(_TWO_LEVEL_DATA_SHEET)
        data_ws.append(_TWO_LEVEL_COLUMNS)
        for row in rows:
            data_ws.append(list(row))
        data_ws.sheet_state = "hidden"

        for name in generated.sheetnames:
            if name == target_sheet:
                # This is the static sheet the native pivot replaces - skip
                # it, or it would collide by name with the renamed pivot
                # sheet below and wb[target_sheet] would resolve to whichever
                # openpyxl happens to list first.
                continue
            destination = template.create_sheet(name, len(template.sheetnames))
            _copy_report_sheet(generated[name], destination)

        pivot_sheet_obj = template[_TWO_LEVEL_PIVOT_SHEET]
        pivot_sheet_obj.title = target_sheet
        if tab_color:
            pivot_sheet_obj.sheet_properties.tabColor = tab_color

        report_order = [name for name in generated.sheetnames if name != target_sheet]
        insert_at = max(0, min(sheet_position, len(report_order)))
        desired_order = (
            report_order[:insert_at] + [target_sheet] + report_order[insert_at:]
            + [_TWO_LEVEL_DATA_SHEET]
        )
        template._sheets = [template[name] for name in desired_order]

        source_ref = f"A1:{get_column_letter(len(_TWO_LEVEL_COLUMNS))}{1 + len(rows)}"
        pivot = pivot_sheet_obj._pivots[0]
        pivot.cache.cacheSource.worksheetSource.ref = source_ref
        pivot.cache.cacheSource.worksheetSource.sheet = _TWO_LEVEL_DATA_SHEET
        pivot.cache.refreshOnLoad = True
        pivot.cache.enableRefresh = True
        pivot.cache.invalid = False
        pivot.cache.missingItemsLimit = 0
        pivot.cache.refreshedBy = None
        pivot.cache.refreshedDate = None

        captions = [dim1_caption, dim2_caption, period_caption]
        for field, caption in zip(pivot.pivotFields, captions):
            field.name = caption
        pivot.dataFields[0].name = value_caption

        template.calculation.fullCalcOnLoad = True
        template.calculation.forceFullCalc = True

        cached_values = cache_formula_values(template)
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{output_path.stem}-pivot-",
            suffix=".xlsx",
            dir=str(output_path.parent),
        )
        os.close(fd)
        temp_path = Path(temp_name)
        template.save(temp_path)
        inject_cached_values(str(temp_path), cached_values)

        validation = load_workbook(temp_path, data_only=False)
        try:
            pivots = validation[target_sheet]._pivots
            if len(pivots) != 1:
                raise ValueError(f"{target_sheet} output does not contain its native PivotTable")
            vsource = pivots[0].cache.cacheSource.worksheetSource
            if vsource.sheet != _TWO_LEVEL_DATA_SHEET or vsource.ref != source_ref:
                raise ValueError(f"{target_sheet} PivotTable cache points at the wrong source range")
        finally:
            validation.close()

        os.replace(temp_path, output_path)
        temp_path = None
    finally:
        generated.close()
        template.close()
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
