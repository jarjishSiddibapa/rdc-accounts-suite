"""Attaches a genuine, refreshable Excel PivotTable to MRN's Vendorwise
Pivot sheet, matching exactly what the source desktop app builds via live
Excel COM automation (E:\\jarjish-projects\\sneha-raman-unaccounted-
transactions-report\\excel_writers.py, _add_real_pivots) - but authored
directly as OOXML parts here, so the server never runs Excel/COM at request
time. Vendorwise Pivot is the ONLY sheet across Unaccounted/MRN/PO that the
source app ever built as a real pivot; everything else there is - and stays
- a static table.

Design: _add_vendorwise_pivot_sheet already writes the correct, final grid
of plain cell values (verified against real Excel behavior). This module
never touches those cells - it only adds the pivotCacheDefinition,
pivotCacheRecords, and pivotTableDefinition parts alongside them, with the
pivot's own <location> pointing at that same already-correct grid. That
means every reader - openpyxl, the mail-body HTML preview, Excel's
Protected View, or a live Excel session - sees correct data immediately,
with no dependency on a refresh ever happening. If a human later opens it
in full Excel and refreshes, Excel recomputes from the same underlying rows
and should reach the same numbers, confirming (not fixing) what's shown.

Verified by round-tripping through real Excel via COM automation (dev
machine only, never part of the shipped server code) to confirm the file
opens cleanly and refreshes to matching values - see
backend/tests/test_vendorwise_pivot.py.
"""
from __future__ import annotations

import zipfile
from pathlib import Path
from xml.sax.saxutils import quoteattr as _qattr

from openpyxl.utils import get_column_letter

_SUMMARY_SHEET = "Summary"
_VENDORWISE_SHEET = "Vendorwise Pivot"
_ROW_FIELD = "SUPPLIER NAME"
_COL_FIELD = "ACCOUNTING PERIOD"
_PAGE_FIELD = "Location"
_COUNT_SOURCE_FIELD = "SUPPLIER SITE"
_NUMERIC_COLS = {"BASE AMOUNT", "CGST", "SGST", "IGST", "RECEIPT QUANTITY"}

_NS = (
    'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
    'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
)


class VendorwisePivotError(Exception):
    """Raised when the real pivot can't be safely attached - the caller
    must catch this and keep the already-correct static sheet untouched."""


def _cell_value_xml(value) -> str:
    if value is None:
        return "<m/>"
    text = str(value).strip()
    if text == "" or text.lower() == "nan":
        return "<m/>"
    try:
        number = float(value)
        if number == number:  # not NaN
            return f'<n v="{number:g}"/>'
    except (TypeError, ValueError):
        pass
    return f"<s v={_qattr(text)}/>"


def _build_cache_xml(df, cols: list[str], axis_items: dict[str, list[str]]) -> tuple[str, str]:
    """Returns (pivotCacheDefinition XML, pivotCacheRecords XML) for the
    given Summary-sheet rows. axis_items maps each of the 3 axis field
    names to its distinct display values, in the exact order the pivot
    should enumerate them (index == the <x> value records will reference)."""
    n_rows = len(df)
    last_col_letter = get_column_letter(len(cols))
    field_xml_parts = []
    axis_index_lookup: dict[str, dict[str, int]] = {
        field: {value: i for i, value in enumerate(values)} for field, values in axis_items.items()
    }

    for col in cols:
        if col in axis_items:
            items_xml = "".join(f'<s v={_qattr(v)}/>' for v in axis_items[col])
            field_xml_parts.append(
                f'<cacheField name={_qattr(col)} numFmtId="0">'
                f'<sharedItems count="{len(axis_items[col])}">{items_xml}</sharedItems>'
                f"</cacheField>"
            )
        else:
            field_xml_parts.append(f'<cacheField name={_qattr(col)} numFmtId="0"><sharedItems/></cacheField>')

    cache_definition = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<pivotCacheDefinition {_NS} r:id="rId1" refreshOnLoad="1" '
        f'missingItemsLimit="0" createdVersion="3" refreshedVersion="4" '
        f'minRefreshableVersion="3" recordCount="{n_rows}">'
        f'<cacheSource type="worksheet">'
        f'<worksheetSource ref="A1:{last_col_letter}{n_rows + 1}" sheet={_qattr(_SUMMARY_SHEET)}/>'
        f"</cacheSource>"
        f'<cacheFields count="{len(cols)}">{"".join(field_xml_parts)}</cacheFields>'
        f"</pivotCacheDefinition>"
    )

    record_rows = []
    for row in df.itertuples(index=False, name=None):
        cells = []
        for col, value in zip(cols, row):
            if col in axis_index_lookup:
                text = "" if value is None else str(value).strip()
                idx = axis_index_lookup[col].get(text, 0)
                cells.append(f'<x v="{idx}"/>')
            else:
                cells.append(_cell_value_xml(value))
        record_rows.append(f"<r>{''.join(cells)}</r>")

    cache_records = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<pivotCacheRecords {_NS} count="{n_rows}">{"".join(record_rows)}</pivotCacheRecords>'
    )
    return cache_definition, cache_records


def _build_pivot_table_xml(
    cols: list[str],
    suppliers: list[str],
    period_labels: list[str],
    location_ref: str,
    n_locations: int,
) -> str:
    field_index = {name: i for i, name in enumerate(cols)}
    row_idx = field_index[_ROW_FIELD]
    col_idx = field_index[_COL_FIELD]
    page_idx = field_index[_PAGE_FIELD]
    data_idx = field_index[_COUNT_SOURCE_FIELD]

    pivot_fields = []
    for i, col in enumerate(cols):
        if i == data_idx:
            pivot_fields.append('<pivotField dataField="1" showAll="0"/>')
        elif i == row_idx:
            items = "".join(f'<item x="{j}"/>' for j in range(len(suppliers))) + '<item t="default"/>'
            pivot_fields.append(
                f'<pivotField name="Supplier Name" axis="axisRow" showAll="0" sortType="descending">'
                f'<items count="{len(suppliers) + 1}">{items}</items>'
                f'<autoSortScope><pivotArea dataOnly="0" outline="0" fieldPosition="0">'
                f'<references count="1"><reference field="4294967294" count="1" selected="0">'
                f'<x v="0"/></reference></references></pivotArea></autoSortScope>'
                f"</pivotField>"
            )
        elif i == col_idx:
            items = "".join(f'<item x="{j}"/>' for j in range(len(period_labels))) + '<item t="default"/>'
            pivot_fields.append(
                f'<pivotField axis="axisCol" showAll="0"><items count="{len(period_labels) + 1}">{items}</items></pivotField>'
            )
        elif i == page_idx:
            # A page field's item list order is cosmetic (dropdown order) since
            # the filter is left at "(All)") - one <item> per distinct value is
            # still required, matching the cache's own sharedItems count.
            items = "".join(f'<item x="{j}"/>' for j in range(n_locations)) + '<item t="default"/>'
            pivot_fields.append(
                f'<pivotField axis="axisPage" showAll="0"><items count="{n_locations + 1}">{items}</items></pivotField>'
            )
        else:
            pivot_fields.append('<pivotField showAll="0"/>')

    row_items = "".join(
        f'<i><x v="{i}"/></i>' if i else "<i><x/></i>" for i in range(len(suppliers))
    ) + '<i t="grand"><x/></i>'
    col_items = "".join(
        f'<i><x v="{i}"/></i>' if i else "<i><x/></i>" for i in range(len(period_labels))
    ) + '<i t="grand"><x/></i>'

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<pivotTableDefinition {_NS} name="VendorwisePivot" cacheId="1" '
        'applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0" '
        'applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1" '
        'dataCaption="Values" updatedVersion="4" minRefreshableVersion="3" '
        'showCalcMbrs="0" useAutoFormatting="1" itemPrintTitles="1" createdVersion="3" '
        'indent="0" outline="1" outlineData="1" multipleFieldFilters="0">'
        f'<location ref={_qattr(location_ref)} firstHeaderRow="1" firstDataRow="2" '
        'firstDataCol="1" rowPageCount="1" colPageCount="1"/>'
        f'<pivotFields count="{len(cols)}">{"".join(pivot_fields)}</pivotFields>'
        f'<rowFields count="1"><field x="{row_idx}"/></rowFields>'
        f'<rowItems count="{len(suppliers) + 1}">{row_items}</rowItems>'
        f'<colFields count="1"><field x="{col_idx}"/></colFields>'
        f'<colItems count="{len(period_labels) + 1}">{col_items}</colItems>'
        f'<pageFields count="1"><pageField fld="{page_idx}" hier="-1"/></pageFields>'
        '<dataFields count="1"><dataField name="Count of Location" '
        f'fld="{data_idx}" subtotal="count" baseField="0" baseItem="0"/></dataFields>'
        '<pivotTableStyleInfo name="PivotStyleLight16" showRowHeaders="1" '
        'showColHeaders="1" showRowStripes="0" showColStripes="0" showLastColumn="1"/>'
        "</pivotTableDefinition>"
    )


def attach_vendorwise_pivot(
    path: str | Path,
    df,
    cols: list[str],
    pivot_df,
    period_labels: list[str],
) -> None:
    """Add real pivot parts for the already-written, already-correct
    Vendorwise Pivot sheet. ``df`` is the full Summary-sheet data (same rows,
    same column order, as actually written to the Summary sheet). ``pivot_df``
    and ``period_labels`` are exactly what _add_vendorwise_pivot_sheet
    returned - the same suppliers (in the same display order) and period
    labels already rendered onto the sheet, guaranteeing the pivot's
    <location> lines up with the real grid.

    Raises VendorwisePivotError on any problem; never modifies the sheet's
    own cells, so the caller can safely ignore this error and keep the
    static table exactly as already saved.
    """
    output_path = Path(path)
    suppliers = pivot_df["SUPPLIER NAME"].astype(str).tolist()
    if not suppliers or not period_labels:
        raise VendorwisePivotError("No Vendorwise rows/periods to build a pivot from")

    # Normalize the 3 axis columns exactly the way _add_vendorwise_pivot_sheet
    # does (stripped; blank Location -> "Unmapped") before building the cache,
    # so every record resolves to the right shared-item index by construction
    # instead of relying on incidental whitespace matching raw source text.
    normalized_df = df.copy()
    normalized_df[_ROW_FIELD] = normalized_df[_ROW_FIELD].astype(str).str.strip()
    normalized_df[_COL_FIELD] = normalized_df[_COL_FIELD].astype(str).str.strip()
    normalized_df[_PAGE_FIELD] = normalized_df[_PAGE_FIELD].astype(str).str.strip().replace("", "Unmapped")

    locations = sorted(normalized_df[_PAGE_FIELD].unique().tolist())
    # ACCOUNTING PERIOD's *cache* item order must match the Summary sheet's
    # raw values (e.g. "JUL-2026"), not the display captions (e.g. "Jul-26")
    # used for period_labels - the sheet's own header cells already show the
    # formatted captions; the cache only needs to resolve each source row to
    # the right column index.
    periods_raw = list(dict.fromkeys(normalized_df[_COL_FIELD].tolist()))
    periods_in_display_order = sorted(
        periods_raw,
        key=lambda p: period_labels.index(_period_caption(p)) if _period_caption(p) in period_labels else len(period_labels),
    )

    axis_items = {
        _ROW_FIELD: suppliers,
        _COL_FIELD: periods_in_display_order,
        _PAGE_FIELD: locations,
    }
    cache_definition_xml, cache_records_xml = _build_cache_xml(normalized_df, cols, axis_items)

    n_suppliers = len(suppliers)
    n_periods = len(period_labels)
    # B=Supplier Name (col 2), then one column per period, then Grand Total -
    # matches _add_vendorwise_pivot_sheet's own gt_col exactly (that static
    # sheet's gt_col also counts its extra hidden Location column A, which
    # this pivot's <location> deliberately starts after, at B).
    gt_col = 3 + n_periods
    location_ref = f"B3:{get_column_letter(gt_col)}{5 + n_suppliers}"
    pivot_table_xml = _build_pivot_table_xml(cols, suppliers, period_labels, location_ref, len(locations))

    _inject_pivot_parts(output_path, cache_definition_xml, cache_records_xml, pivot_table_xml)


def _period_caption(raw: str) -> str:
    from .excel_writers import _fmt_mrn_period

    return _fmt_mrn_period(raw)


def _existing_rids(rels_xml: str) -> set[str]:
    """Relationship Id/Target attribute order is not guaranteed (openpyxl
    itself writes Target before Id), so this parses properly with
    ElementTree instead of assuming an attribute order via regex."""
    import xml.etree.ElementTree as ET

    root = ET.fromstring(rels_xml)
    return {el.get("Id") for el in root if el.get("Id")}


def _next_rid(rels_xml: str) -> str:
    existing = _existing_rids(rels_xml)
    n = 1
    while f"rId{n}" in existing:
        n += 1
    return f"rId{n}"


def _inject_pivot_parts(
    output_path: Path,
    cache_definition_xml: str,
    cache_records_xml: str,
    pivot_table_xml: str,
) -> None:
    with zipfile.ZipFile(output_path) as zf:
        names = zf.namelist()
        parts = {name: zf.read(name) for name in names}

    sheet_part = _vendorwise_sheet_part(parts)
    if sheet_part is None:
        raise VendorwisePivotError("Could not locate the Vendorwise Pivot worksheet part")

    content_types = parts["[Content_Types].xml"].decode("utf-8")
    if "pivotCacheDefinition1.xml" in content_types:
        raise VendorwisePivotError("A pivot cache part already exists - refusing to overwrite")
    content_types = content_types.replace(
        "</Types>",
        '<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"/>'
        '<Override PartName="/xl/pivotCache/pivotCacheRecords1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"/>'
        '<Override PartName="/xl/pivotTables/pivotTable1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/>'
        "</Types>",
    )

    workbook_xml = parts["xl/workbook.xml"].decode("utf-8")
    workbook_rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
    cache_rid = _next_rid(workbook_rels)
    workbook_rels = workbook_rels.replace(
        "</Relationships>",
        f'<Relationship Id="{cache_rid}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" '
        'Target="pivotCache/pivotCacheDefinition1.xml"/></Relationships>',
    )
    if "<pivotCaches>" in workbook_xml:
        raise VendorwisePivotError("Workbook already declares a pivotCaches section")
    # openpyxl declares xmlns:r locally on each <sheet> element rather than on
    # the <workbook> root, so an element inserted at the workbook's top level
    # (like this one) must declare the namespace itself or "r:id" is an
    # undefined-prefix parse error.
    workbook_xml = workbook_xml.replace(
        "</workbook>",
        '<pivotCaches><pivotCache cacheId="1" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        f'r:id="{cache_rid}"/></pivotCaches></workbook>',
    )

    sheet_xml = parts[sheet_part].decode("utf-8")
    sheet_rels_path = sheet_part.replace("worksheets/", "worksheets/_rels/") + ".rels"
    if sheet_rels_path in parts:
        sheet_rels = parts[sheet_rels_path].decode("utf-8")
        pivot_rid = _next_rid(sheet_rels)
        sheet_rels = sheet_rels.replace(
            "</Relationships>",
            f'<Relationship Id="{pivot_rid}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" '
            'Target="../pivotTables/pivotTable1.xml"/></Relationships>',
        )
    else:
        pivot_rid = "rId1"
        sheet_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'<Relationship Id="{pivot_rid}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" '
            'Target="../pivotTables/pivotTable1.xml"/></Relationships>'
        )

    pivot_table_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" '
        'Target="../pivotCache/pivotCacheDefinition1.xml"/></Relationships>'
    )
    cache_definition_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheRecords" '
        'Target="pivotCacheRecords1.xml"/></Relationships>'
    )

    parts["[Content_Types].xml"] = content_types.encode("utf-8")
    parts["xl/workbook.xml"] = workbook_xml.encode("utf-8")
    parts["xl/_rels/workbook.xml.rels"] = workbook_rels.encode("utf-8")
    parts[sheet_part] = sheet_xml.encode("utf-8")
    parts[sheet_rels_path] = sheet_rels.encode("utf-8")
    parts["xl/pivotTables/pivotTable1.xml"] = pivot_table_xml.encode("utf-8")
    parts["xl/pivotTables/_rels/pivotTable1.xml.rels"] = pivot_table_rels.encode("utf-8")
    parts["xl/pivotCache/pivotCacheDefinition1.xml"] = cache_definition_xml.encode("utf-8")
    parts["xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels"] = cache_definition_rels.encode("utf-8")
    parts["xl/pivotCache/pivotCacheRecords1.xml"] = cache_records_xml.encode("utf-8")

    tmp_path = output_path.with_suffix(".pivot-tmp.xlsx")
    try:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, data in parts.items():
                zf.writestr(name, data)
        tmp_path.replace(output_path)
    finally:
        tmp_path.unlink(missing_ok=True)


def _vendorwise_sheet_part(parts: dict[str, bytes]) -> str | None:
    """Resolve the worksheet part (e.g. "xl/worksheets/sheet2.xml") backing
    the "Vendorwise Pivot" tab, by following workbook.xml's <sheet r:id=...>
    through workbook.xml.rels - via ElementTree, not regex, since attribute
    order (Id vs Target) is not guaranteed (openpyxl itself writes Target
    before Id)."""
    import xml.etree.ElementTree as ET

    r_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"

    workbook_root = ET.fromstring(parts["xl/workbook.xml"])
    rid = None
    for sheet_el in workbook_root.iter(f"{main_ns}sheet"):
        if sheet_el.get("name") == _VENDORWISE_SHEET:
            rid = sheet_el.get(f"{r_ns}id")
            break
    if not rid:
        return None

    rels_root = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    target = None
    for rel_el in rels_root.iter(f"{rel_ns}Relationship"):
        if rel_el.get("Id") == rid:
            target = rel_el.get("Target")
            break
    if not target:
        return None
    return f"xl/{target}" if not target.startswith("/xl/") else target.lstrip("/")
