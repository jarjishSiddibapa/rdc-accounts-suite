"""openpyxl-based Excel writers for both report types."""

import datetime as _dt

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .constants import MRN_SITE_COL, MRN_ANCHOR_COL
from app.regional import format_indian_number, today_ist

# Indian number format: 1,23,456  |  zero displays as  –
# Sections: positive ; negative ; zero
INDIAN_FMT = r'##,##,##0;-##,##,##0;"-"'
INDIAN_DECIMAL_FMT = r'##,##,##0.00;-##,##,##0.00;"-"'


def _thin_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _autofit_header_only(ws, hdr_row: int = 1, min_width: int = 8,
                          max_width: int = 60, padding: int = 4,
                          date_cap: int = 13) -> None:
    """Size each visible column to its header text only (data length ignored),
    then cap any column whose header contains "Date" so only the date portion
    is visible. Pure-Python equivalent of the compact "AutoFit header row
    only" pass this suite used to run through Excel COM automation."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if ws.column_dimensions[col_letter].hidden:
            continue
        header_val = ws.cell(row=hdr_row, column=col[0].column).value
        width = min(max(len(str(header_val)) + padding, min_width), max_width) \
            if header_val is not None else min_width
        if header_val and "Date" in str(header_val):
            width = min(width, date_cap)
        ws.column_dimensions[col_letter].width = width


def _autofit_columns(ws, min_width: int = 8, max_width: int = 60, padding: int = 4) -> None:
    """Set each visible column's width based on header + up to 200 data rows."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if ws.column_dimensions[col_letter].hidden:
            continue
        sample = col[:201]
        max_len = max(
            (len(str(cell.value)) for cell in sample if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + padding, min_width), max_width)


def _add_pivot_sheet(wb, df: "pd.DataFrame") -> None:
    """Add a formatted pivot summary: Location × Accounts Incharge × invoice count.

    Supports **multiple months** — one column is created per distinct month found
    in the GL Date column, sorted chronologically.  A final 'Total' column sums
    across all months.  When only one month is present the layout looks identical
    to the original single-column design (no separate Total column is shown).
    """
    GL_DATE_COL = "GL Date"

    # ── 1. Derive month labels ─────────────────────────────────────────────────
    # Prefer the pre-stamped __month__ column (set from B12 Period Name in
    # processing.py).  Fall back to parsing GL Date if __month__ is absent.
    def _parse_month(val):
        try:
            return pd.to_datetime(val, dayfirst=True, errors="coerce").strftime("%b-%y")
        except Exception:
            return None

    def _month_sort_key(m):
        try:
            return pd.to_datetime(m, format="%b-%y")
        except Exception:
            return pd.Timestamp.max

    if "__month__" in df.columns:
        month_series = df["__month__"].dropna().astype(str).replace("nan", pd.NA).dropna()
    elif GL_DATE_COL in df.columns:
        month_series = df[GL_DATE_COL].dropna().apply(_parse_month).dropna()
    else:
        month_series = pd.Series(dtype=str)

    unique_months = sorted(set(month_series) - {"", "nan"}, key=_month_sort_key)
    if not unique_months:
        unique_months = [today_ist().strftime("%b-%y")]

    multi_month = len(unique_months) > 1

    # ── 2. Add __month__ to working copy ──────────────────────────────────────
    df = df.copy()
    if "__month__" not in df.columns:
        # derive from GL Date as fallback
        if GL_DATE_COL in df.columns:
            df["__month__"] = df[GL_DATE_COL].apply(_parse_month)
        else:
            df["__month__"] = unique_months[0]

    # ── 3. Build pivot: Location × Accounts Incharge × Month → count ──────────
    mapped = df[df["Location"].astype(str).str.strip() != ""].copy()

    pivot_raw = (
        mapped
        .groupby(["Location", "Accounts Incharge", "__month__"], sort=False)
        .size()
        .reset_index(name="Count")
    )

    # Aggregate to Location × Accounts Incharge with one column per month
    pivot = (
        pivot_raw
        .pivot_table(
            index=["Location", "Accounts Incharge"],
            columns="__month__",
            values="Count",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )
    # Ensure month columns are in chronological order
    month_cols = [m for m in unique_months if m in pivot.columns]
    pivot = pivot[["Location", "Accounts Incharge"] + month_cols]
    pivot["__total__"] = pivot[month_cols].sum(axis=1)

    # Sort: locations with most total invoices first; within a location, by total desc
    loc_totals = pivot.groupby("Location")["__total__"].sum()
    pivot["_r"] = pivot["Location"].map(loc_totals)
    pivot = (
        pivot.sort_values(["_r", "__total__"], ascending=[False, False])
             .drop(columns=["_r"])
             .reset_index(drop=True)
    )

    # ── 4. Build the worksheet ─────────────────────────────────────────────────
    ws = wb.create_sheet("Main")
    ws.sheet_properties.tabColor = "1F3864"
    ws.sheet_view.showGridLines  = False

    # Palette
    HDR_BG   = "F4B8C8"
    TOTAL_BG = "F4B8C8"
    bdr      = _thin_border()
    plain    = Font(name="Calibri", size=11)
    bold_f   = Font(name="Calibri", size=11, bold=True)

    # Column layout: Location | Accounts Incharge | month1 [| month2 …] [| Total]
    # If only one month: no separate Total column (matches original design).
    data_cols = month_cols + (["__total__"] if multi_month else [])
    hdr_labels = ["Location", "Accounts Incharge"] + month_cols
    if multi_month:
        hdr_labels.append("Total")

    n_value_cols = len(data_cols)   # number of numeric columns

    # Row 1: headers
    for ci, h in enumerate(hdr_labels, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = bold_f
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr

    # ── 5. Data rows ───────────────────────────────────────────────────────────
    current_row = 2
    grand_totals = {col: 0 for col in data_cols}

    for location, loc_grp in pivot.groupby("Location", sort=False):
        n_incharges = len(loc_grp)
        loc_col_totals = {col: int(loc_grp[col].sum()) for col in data_cols}
        for col in data_cols:
            grand_totals[col] += loc_col_totals[col]

        if n_incharges == 1:
            # Single incharge — one combined row
            incharge = loc_grp.iloc[0]["Accounts Incharge"]
            ws.cell(row=current_row, column=1, value=location).font = plain
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=current_row, column=2, value=incharge).font = plain
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
            for ci_off, col in enumerate(data_cols, 3):
                c = ws.cell(row=current_row, column=ci_off, value=int(loc_grp.iloc[0][col]))
                c.font = plain
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = INDIAN_FMT
            current_row += 1
        else:
            # Location header row
            ws.cell(row=current_row, column=1, value=location).font = plain
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

            detail_start = current_row
            for _, row_data in loc_grp.iterrows():
                ws.row_dimensions[current_row].outline_level = 1
                ws.cell(row=current_row, column=1, value="")
                ws.cell(row=current_row, column=2, value=row_data["Accounts Incharge"]).font = plain
                ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
                for ci_off, col in enumerate(data_cols, 3):
                    c = ws.cell(row=current_row, column=ci_off, value=int(row_data[col]))
                    c.font = plain
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.number_format = INDIAN_FMT
                current_row += 1
            detail_end = current_row - 1

            # Subtotal row
            ws.cell(row=current_row, column=1, value=f"{location} Total").font = plain
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            for ci_off in range(3, 3 + n_value_cols):
                col_letter = get_column_letter(ci_off)
                c = ws.cell(row=current_row, column=ci_off)
                c.value         = f"=SUBTOTAL(9,{col_letter}{detail_start}:{col_letter}{detail_end})"
                c.font          = plain
                c.alignment     = Alignment(horizontal="center", vertical="center")
                c.number_format = INDIAN_FMT
            current_row += 1

    # ── 6. Grand Total row ─────────────────────────────────────────────────────
    grand_total_row = current_row
    c1 = ws.cell(row=grand_total_row, column=1, value="Grand Total")
    c1.font      = bold_f
    c1.fill      = PatternFill("solid", fgColor=TOTAL_BG)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    c2 = ws.cell(row=grand_total_row, column=2, value="")
    c2.fill = PatternFill("solid", fgColor=TOTAL_BG)

    for ci_off in range(3, 3 + n_value_cols):
        col_letter = get_column_letter(ci_off)
        c = ws.cell(row=grand_total_row, column=ci_off)
        c.value         = f"=SUBTOTAL(9,{col_letter}2:{col_letter}{grand_total_row - 1})"
        c.font          = bold_f
        c.fill          = PatternFill("solid", fgColor=TOTAL_BG)
        c.alignment     = Alignment(horizontal="center", vertical="center")
        c.border        = bdr
        c.number_format = INDIAN_FMT

    # ── 7. Apply borders to entire table ──────────────────────────────────────
    total_cols = 2 + n_value_cols
    for r in range(1, grand_total_row + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = bdr

    # ── Freeze, filter, outline ───────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}1"
    ws.sheet_format.outlineLevelRow = 1


def _to_excel_date(val):
    """Parse val as a date and return a Python datetime.date, or None on failure.

    openpyxl converts Python date/datetime objects to Excel serial numbers
    automatically, so Excel's YEAR(), MONTH(), NETWORKDAYS(), DATEDIF() etc.
    all work correctly on cells written with this function.
    """
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("nan", "none", ""):
        return None
    try:
        parsed = pd.to_datetime(val_str, dayfirst=True, errors="coerce")
        if pd.notna(parsed):
            return parsed.date()
    except Exception:
        pass
    return None


def write_formatted_excel(df: "pd.DataFrame", path: str) -> None:
    """Write df to path as a professionally formatted xlsx."""
    wb     = Workbook()
    ws     = wb.active
    ws.title = "Summary"

    # Exclude the internal __month__ column from the data sheet
    cols   = [c for c in list(df.columns) if c != "__month__"]
    n_cols = len(cols)
    n_rows = len(df)
    bdr    = _thin_border()

    HDR_BG   = "C2D8ED"   # light blue header
    HDR_FG   = "000000"   # black text
    ROW_ODD  = "FFFFFF"   # white
    ROW_EVEN = "EAF3FB"   # very light blue
    TITLE_BG = "DAEAF7"   # very light blue title bar

    # Row 1: column headers
    hdr_font  = Font(name="Calibri", size=11, bold=True, color=HDR_FG)
    hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
    for ci, col_name in enumerate(cols, 1):
        c           = ws.cell(row=1, column=ci, value=col_name)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.border    = bdr

    # Data rows
    txt_font      = Font(name="Calibri", size=11, color="000000")
    num_font      = Font(name="Calibri", size=11, color="000000")
    CURRENCY_COLS = {"Amount"}
    DATE_FMT      = "DD-MMM-YYYY"   # renders as 01-Jun-2026 — proper Excel date serial

    for ri, row_vals in enumerate(df[cols].itertuples(index=False), 2):
        for ci, val in enumerate(row_vals, 1):
            col_name = cols[ci - 1]
            c        = ws.cell(row=ri, column=ci)
            c.border = bdr
            if col_name in CURRENCY_COLS:
                c.font          = num_font
                c.alignment     = Alignment(horizontal="right")
                c.number_format = INDIAN_DECIMAL_FMT
                try:
                    c.value = (
                        float(val)
                        if val is not None and str(val).strip() != ""
                        else 0.0
                    )
                except (ValueError, TypeError):
                    c.value = val
            elif "date" in col_name.lower():
                # Write as a real Excel date serial so date functions work
                date_val = _to_excel_date(val)
                if date_val is not None:
                    c.value         = date_val
                    c.number_format = DATE_FMT
                    c.font          = txt_font
                    c.alignment     = Alignment(horizontal="center")
                else:
                    c.font  = txt_font
                    c.value = val if val is not None else ""
            else:
                c.font  = txt_font
                c.value = val if val is not None else ""

    ws.freeze_panes              = "A2"
    ws.auto_filter.ref           = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
    ws.sheet_properties.tabColor = "00B61D"

    # ── Sheet 2: Pivot ────────────────────────────────────────────────────────
    _add_pivot_sheet(wb, df)

    # ── Sheet 3: Unmapped Sites (only created when there are unmapped rows) ─────
    unmatched_df = (
        df[df["Location"].astype(str).str.strip() == ""]
        [["Supplier Name", "Supplier Site", "Invoice Number",
          "Invoice Date", "Amount"]]
        .copy()
        .reset_index(drop=True)
    )

    if len(unmatched_df) > 0:
        ws2          = wb.create_sheet("Unmapped Sites")
        WARN_BG      = "7F0000"
        WARN_HDR_BG  = "C0392B"
        ROW_WARN     = "FFF3F3"
        ROW_WARN_ALT = "FFE5E5"

        ws2.row_dimensions[1].height = 28
        n2_cols = 5
        t2 = ws2.cell(row=1, column=1,
                      value=f"Unmapped Supplier Sites  —  {format_indian_number(len(unmatched_df))} invoice(s) with no Location / Accounts Incharge")
        t2.font      = Font(name="Calibri", size=11, bold=True, color="000000")
        t2.fill      = PatternFill("solid", fgColor=WARN_BG)
        t2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        t2.border    = bdr
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n2_cols)
        for c in range(2, n2_cols + 1):
            ws2.cell(row=1, column=c).fill   = PatternFill("solid", fgColor=WARN_BG)
            ws2.cell(row=1, column=c).border = bdr

        ws2.row_dimensions[2].height = 20
        exp = ws2.cell(
            row=2, column=1,
            value="The Supplier Sites listed below were not found in the Location – Incharges Mapping. "
                  "Please add them via Manage Mappings.")
        exp.font      = Font(name="Calibri", size=11, italic=True, color="000000")
        exp.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        exp.border    = bdr
        ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n2_cols)
        for c in range(2, n2_cols + 1):
            ws2.cell(row=2, column=c).border = bdr

        summary = (
            unmatched_df.groupby("Supplier Site", sort=False)
            .agg(Invoice_Count=("Invoice Number", "count"))
            .reset_index()
            .sort_values("Invoice_Count", ascending=False)
        )

        ws2.row_dimensions[3].height = 8
        ws2.row_dimensions[4].height = 28
        sum_cols = ["Supplier Site", "Invoice Count"]
        sum_hdr_font  = Font(name="Calibri", size=11, bold=True, color="000000")
        sum_hdr_fill  = PatternFill("solid", fgColor=WARN_HDR_BG)
        sum_hdr_align = Alignment(horizontal="center", vertical="center")
        sum_bdr       = _thin_border()
        for ci, h in enumerate(sum_cols, 1):
            c           = ws2.cell(row=4, column=ci, value=h)
            c.font      = sum_hdr_font
            c.fill      = sum_hdr_fill
            c.alignment = sum_hdr_align
            c.border    = sum_bdr

        txt_f  = Font(name="Calibri", size=11)
        num_f  = Font(name="Calibri", size=11)
        for ri, row in enumerate(summary.itertuples(index=False), 5):
            ws2.row_dimensions[ri].height = 18
            bg = ROW_WARN_ALT if ri % 2 == 0 else ROW_WARN
            c1 = ws2.cell(row=ri, column=1, value=row[0])
            c1.font      = txt_f
            c1.fill      = PatternFill("solid", fgColor=bg)
            c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c1.border    = sum_bdr
            c2 = ws2.cell(row=ri, column=2, value=int(row[1]))
            c2.font      = num_f
            c2.fill      = PatternFill("solid", fgColor=bg)
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.border    = sum_bdr

        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 16

        detail_start_row = 5 + len(summary) + 2
        ws2.row_dimensions[detail_start_row - 1].height = 14

        detail_title_row = detail_start_row - 1
        dt = ws2.cell(row=detail_title_row, column=1,
                      value="Full Invoice Detail for Unmapped Sites")
        dt.font      = Font(name="Calibri", size=11, bold=True, color="000000")
        dt.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        dt.border    = sum_bdr
        ws2.merge_cells(start_row=detail_title_row, start_column=1,
                        end_row=detail_title_row, end_column=n2_cols)
        for c in range(2, n2_cols + 1):
            ws2.cell(row=detail_title_row, column=c).border = sum_bdr

        detail_cols = ["Supplier Name", "Supplier Site",
                       "Invoice Number", "Invoice Date", "Amount"]
        ws2.row_dimensions[detail_start_row].height = 28
        for ci, h in enumerate(detail_cols, 1):
            c           = ws2.cell(row=detail_start_row, column=ci, value=h)
            c.font      = sum_hdr_font
            c.fill      = sum_hdr_fill
            c.alignment = sum_hdr_align
            c.border    = sum_bdr

        amt_fmt = Font(name="Calibri", size=11)
        for ri, row in enumerate(unmatched_df.itertuples(index=False),
                                  detail_start_row + 1):
            ws2.row_dimensions[ri].height = 18
            bg = ROW_WARN_ALT if ri % 2 == 0 else ROW_WARN
            for ci, val in enumerate(row, 1):
                c        = ws2.cell(row=ri, column=ci)
                c.border = sum_bdr
                c.fill   = PatternFill("solid", fgColor=bg)
                if detail_cols[ci - 1] == "Amount":
                    try:
                        c.value         = float(val) if val is not None else 0.0
                        c.number_format = INDIAN_DECIMAL_FMT
                        c.font          = amt_fmt
                        c.alignment     = Alignment(horizontal="right", vertical="center")
                    except (ValueError, TypeError):
                        c.value     = val
                        c.font      = txt_f
                        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                else:
                    c.value     = val if val is not None else ""
                    c.font      = txt_f
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        for ci, col_name in enumerate(detail_cols, 1):
            max_len = len(col_name)
            for ri in range(detail_start_row + 1,
                            min(detail_start_row + 1 + len(unmatched_df), detail_start_row + 201)):
                v = ws2.cell(row=ri, column=ci).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws2.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 3, 14), 45)

        ws2.freeze_panes = "A5"
        ws2.sheet_properties.tabColor = "C0392B"

    # Reorder: Main first, then Summary, then Unmapped Sites
    wb.move_sheet("Main", offset=-wb.sheetnames.index("Main"))

    for sheet in wb.worksheets:
        _autofit_columns(sheet)

    wb.save(path)


# ── PO pivot sheet ────────────────────────────────────────────────────────────

def _add_po_pivot_sheet(wb, main_df) -> None:
    """Insert a 'Main' pivot sheet at position 0 (first sheet) of *wb*.

    Mirrors the MRN _add_pivot_sheet structure:
    - Row 1: title bar (no count info)
    - Row 2: column headers (Location | Accounts Incharge | month… | Grand Total)
    - Row 3+: per-location groups with collapsible detail rows + subtotal rows
    - Bottom: Grand Total row
    """
    try:
        import pandas as pd
    except ImportError:
        return

    required = {'PO Number', 'Location', 'Accounts Incharge', 'Month'}
    ws = wb.create_sheet("Main", 0)
    ws.sheet_view.showGridLines  = False
    ws.sheet_properties.tabColor = "1F3864"

    if not required.issubset(main_df.columns) or main_df.empty:
        ws.cell(row=1, column=1, value="No data available for pivot.")
        return

    # ── Deduplicate: each PO counted once per (Location, Incharge, Month) ────
    piv_src = main_df.drop_duplicates(
        subset=['PO Number', 'Location', 'Accounts Incharge', 'Month'])

    # ── Build pivot table ─────────────────────────────────────────────────────
    pivot = (
        piv_src
        .groupby(['Location', 'Accounts Incharge', 'Month'])['PO Number']
        .count()
        .unstack('Month')
        .fillna(0)
        .astype(int)
    )

    def _month_key(m):
        try:
            return pd.to_datetime(m, format='%b-%y')
        except Exception:
            return pd.Timestamp('2099-01-01')

    month_cols = sorted(pivot.columns.tolist(), key=_month_key)
    pivot = pivot.reindex(columns=month_cols, fill_value=0)
    pivot['Grand Total'] = pivot.sum(axis=1)
    piv_reset = pivot.reset_index()
    loc_gt_map = piv_reset.groupby('Location')['Grand Total'].sum()
    piv_reset['_loc_gt'] = piv_reset['Location'].map(loc_gt_map)
    piv_reset = (
        piv_reset
        .sort_values(['_loc_gt', 'Grand Total'], ascending=[False, False])
        .drop(columns=['_loc_gt'])
        .reset_index(drop=True)
    )

    # ── Sheet constants ───────────────────────────────────────────────────────
    bdr    = _thin_border()
    N      = 2 + len(month_cols) + 1   # Location + Incharge + months + Grand Total
    CENTER = Alignment(horizontal="center", vertical="center")

    # MRN palette
    TITLE_BG = "FADADD"
    HDR_BG   = "F4B8C8"
    HDR_FG   = "000000"
    TOTAL_BG = "F4B8C8"
    TOTAL_FG = "000000"

    plain      = Font(name="Calibri", size=11)
    bold_total = Font(name="Calibri", size=11, bold=True, color=TOTAL_FG)

    # ── Row 1: Column headers (no title row) ──────────────────────────────────
    headers = ['Location', 'Accounts Incharge'] + month_cols + ['Grand Total']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = Font(name="Calibri", size=11, bold=True, color=HDR_FG)
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = CENTER

    # ── Data rows ─────────────────────────────────────────────────────────────
    cur_row = 2

    for location, loc_grp in piv_reset.groupby('Location', sort=False):
        n_incharges = len(loc_grp)

        if n_incharges == 1:
            # Single combined row: Location | Incharge | counts… | GT
            rd = loc_grp.iloc[0]
            ws.cell(row=cur_row, column=1, value=location).font = plain
            ws.cell(row=cur_row, column=1).alignment = CENTER
            ws.cell(row=cur_row, column=2,
                    value=str(rd['Accounts Incharge'])).font = plain
            ws.cell(row=cur_row, column=2).alignment = CENTER
            for ci, month in enumerate(month_cols, 3):
                val = int(rd[month]) if month in rd.index else 0
                c = ws.cell(row=cur_row, column=ci, value=val)
                c.font = plain; c.alignment = CENTER; c.number_format = INDIAN_FMT
            c_gt = ws.cell(row=cur_row, column=N, value=int(rd['Grand Total']))
            c_gt.font = plain; c_gt.alignment = CENTER; c_gt.number_format = INDIAN_FMT
            cur_row += 1

        else:
            # Location header row
            ws.cell(row=cur_row, column=1, value=location).font = plain
            ws.cell(row=cur_row, column=1).alignment = CENTER
            for ci in range(2, N + 1):
                ws.cell(row=cur_row, column=ci, value="")
            cur_row += 1

            # Collapsible detail rows
            detail_start = cur_row
            for _, rd in loc_grp.iterrows():
                ws.row_dimensions[cur_row].outline_level = 1
                ws.cell(row=cur_row, column=1, value="")
                ws.cell(row=cur_row, column=2,
                        value=str(rd['Accounts Incharge'])).font = plain
                ws.cell(row=cur_row, column=2).alignment = CENTER
                for ci, month in enumerate(month_cols, 3):
                    val = int(rd[month]) if month in rd.index else 0
                    c = ws.cell(row=cur_row, column=ci, value=val)
                    c.font = plain; c.alignment = CENTER; c.number_format = INDIAN_FMT
                c_gt = ws.cell(row=cur_row, column=N,
                               value=int(rd['Grand Total']))
                c_gt.font = plain; c_gt.alignment = CENTER
                c_gt.number_format = INDIAN_FMT
                cur_row += 1
            detail_end = cur_row - 1

            # Per-location subtotal row
            ws.cell(row=cur_row, column=1,
                    value=f"{location} Total").font = plain
            ws.cell(row=cur_row, column=1).alignment = CENTER
            ws.cell(row=cur_row, column=2, value="")
            for ci, month in enumerate(month_cols, 3):
                col_ltr = get_column_letter(ci)
                c = ws.cell(
                    row=cur_row, column=ci,
                    value=f"=SUBTOTAL(9,{col_ltr}{detail_start}:{col_ltr}{detail_end})")
                c.font = plain; c.alignment = CENTER; c.number_format = INDIAN_FMT
            gt_ltr = get_column_letter(N)
            c_gt = ws.cell(
                row=cur_row, column=N,
                value=f"=SUBTOTAL(9,{gt_ltr}{detail_start}:{gt_ltr}{detail_end})")
            c_gt.font = plain; c_gt.alignment = CENTER; c_gt.number_format = INDIAN_FMT
            cur_row += 1

    # ── Grand Total row ───────────────────────────────────────────────────────
    grand_total_row = cur_row

    c1 = ws.cell(row=grand_total_row, column=1, value="Grand Total")
    c1.font = bold_total
    c1.fill = PatternFill("solid", fgColor=TOTAL_BG)
    c1.alignment = CENTER

    ws.cell(row=grand_total_row, column=2, value="").fill = \
        PatternFill("solid", fgColor=TOTAL_BG)

    for ci, month in enumerate(month_cols, 3):
        col_ltr = get_column_letter(ci)
        c = ws.cell(
            row=grand_total_row, column=ci,
            value=f"=SUBTOTAL(9,{col_ltr}2:{col_ltr}{grand_total_row - 1})")
        c.font = bold_total
        c.fill = PatternFill("solid", fgColor=TOTAL_BG)
        c.alignment = CENTER; c.number_format = INDIAN_FMT

    gt_ltr = get_column_letter(N)
    c_gt = ws.cell(
        row=grand_total_row, column=N,
        value=f"=SUBTOTAL(9,{gt_ltr}2:{gt_ltr}{grand_total_row - 1})")
    c_gt.font = bold_total
    c_gt.fill = PatternFill("solid", fgColor=TOTAL_BG)
    c_gt.alignment = CENTER; c_gt.number_format = INDIAN_FMT

    # ── Borders: one pass over entire table ───────────────────────────────────
    for r in range(1, grand_total_row + 1):
        for c in range(1, N + 1):
            ws.cell(row=r, column=c).border = bdr

    # ── Freeze, filter, outline ───────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(N)}1"
    ws.sheet_format.outlineLevelRow = 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26
    for ci in range(3, N + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 11


# ── Uninvoiced Expense PO Report writer ───────────────────────────────────────

def write_formatted_po_excel(
    main_df: "pd.DataFrame",
    moved_df: "pd.DataFrame",
    unmapped_df: "pd.DataFrame",
    path: str,
) -> None:
    """Write the PO report to *path* with four sheets:
       1. Main             — pivot (Location × Month, unique PO count)
       2. Summary          — full filtered data
       3. Deleted Rows     — keyword / excluded-PO matches
       4. Mapping Not Found — rows whose Supplier Site had no mapping
    """
    wb  = Workbook()
    bdr = _thin_border()

    # ── Palette ───────────────────────────────────────────────────────────────
    TITLE_BG   = "DAEAF7"
    HDR_BG     = "C2D8ED"
    HDR_FG     = "000000"
    ROW_ODD    = "FFFFFF"
    ROW_EVEN   = "EAF3FB"
    DEL_TITLE  = "4A0000"   # dark red for deleted-rows sheet title
    DEL_HDR_BG = "C0392B"
    DEL_ROW1   = "FFF3F3"
    DEL_ROW2   = "FFE5E5"
    UNM_TITLE  = "4A2800"   # dark orange for unmapped sheet title
    UNM_HDR_BG = "D97706"
    UNM_ROW1   = "FFF7ED"
    UNM_ROW2   = "FFEDD5"

    hdr_font  = Font(name="Calibri", size=11, bold=True, color=HDR_FG)
    txt_font  = Font(name="Calibri", size=11, color="000000")
    num_font  = Font(name="Calibri", size=11, color="000000")

    NUM_COLS = {"Unit Price", "Quantity", "Basic Amount", "PO Amount", "GST Amount"}

    def _write_sheet(ws, df, title_text=None, title_bg=None):
        """Write a formatted sheet.  When title_text is None the title row is
        omitted and column headers land on row 1 (freeze at A2).  When a title
        is supplied the structure is title→row 1, headers→row 2, data→row 3+
        (freeze at A3), matching the Deleted Rows / Mapping Not Found layout."""
        cols      = list(df.columns)
        n_cols    = len(cols)
        n_rows    = len(df)
        has_title = title_text is not None

        if has_title:
            # Row 1: merged title bar
            tc = ws.cell(row=1, column=1, value=title_text)
            tc.font      = Font(name="Calibri", size=11, bold=True, color="000000")
            tc.fill      = PatternFill("solid", fgColor=title_bg)
            tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            tc.border    = bdr
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
            for c in range(2, n_cols + 1):
                cell        = ws.cell(row=1, column=c)
                cell.fill   = PatternFill("solid", fgColor=title_bg)
                cell.border = bdr
            hdr_row  = 2
            data_row = 3
        else:
            hdr_row  = 1
            data_row = 2

        # Header row
        hdr_fill = PatternFill("solid", fgColor=HDR_BG)
        for ci, col_name in enumerate(cols, 1):
            c           = ws.cell(row=hdr_row, column=ci, value=col_name)
            c.font      = hdr_font
            c.fill      = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = bdr

        # Data rows — plain white (no alternating background)
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        for ri, row_vals in enumerate(df.itertuples(index=False), data_row):
            for ci, val in enumerate(row_vals, 1):
                col_name = cols[ci - 1]
                c        = ws.cell(row=ri, column=ci)
                c.fill   = white_fill
                c.border = bdr
                if col_name in NUM_COLS:
                    c.font          = num_font
                    c.alignment     = Alignment(horizontal="right", vertical="center")
                    c.number_format = INDIAN_DECIMAL_FMT
                    try:
                        c.value = float(val) if val is not None and str(val).strip() not in ("", "nan") else 0.0
                    except (ValueError, TypeError):
                        c.value = val
                elif col_name == "Month" and val is not None and str(val).strip() not in ("", "nan"):
                    c.font = txt_font
                    try:
                        _mdt = pd.to_datetime(str(val).strip(), format="%b-%y")
                        c.value         = _dt.date(_mdt.year, _mdt.month, 1)
                        c.number_format = "mmm-yy"
                        c.alignment     = Alignment(horizontal="center", vertical="center")
                    except Exception:
                        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                        c.value     = str(val)
                elif "date" in col_name.lower():
                    # Write as a real Excel date serial so date functions work
                    date_val = _to_excel_date(val)
                    c.font = txt_font
                    if date_val is not None:
                        c.value         = date_val
                        c.number_format = "DD-MMM-YYYY"
                        c.alignment     = Alignment(horizontal="center", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                        c.value     = "" if (val is None or str(val).strip() == "nan") else val
                else:
                    c.font      = txt_font
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    c.value     = "" if (val is None or str(val).strip() == "nan") else val

        freeze_row = hdr_row + 1   # freeze below the header row
        ws.freeze_panes              = f"A{freeze_row}"
        if n_rows > 0:
            ws.auto_filter.ref       = f"A{hdr_row}:{get_column_letter(n_cols)}{hdr_row}"
        ws.sheet_properties.tabColor = "00B61D"
        _autofit_columns(ws)

    # ── Sort helper: PO Number (asc) → Line (numeric asc) ───────────────────────
    _PO_COL    = "PO Number"
    _LINE_COLS = ["Line", "PO Line", "Line Number"]   # try in order

    def _sort_by_po_line(df):
        if _PO_COL not in df.columns:
            return df
        _line_col = next((c for c in _LINE_COLS if c in df.columns), None)
        if _line_col:
            df = df.copy()
            df["_line_sort"] = pd.to_numeric(df[_line_col], errors="coerce").fillna(0)
            df = df.sort_values([_PO_COL, "_line_sort"], ascending=[True, True],
                                key=lambda s: s.astype(str).str.strip()
                                             if s.name == _PO_COL else s)
            df = df.drop(columns=["_line_sort"]).reset_index(drop=True)
        else:
            df = df.sort_values(_PO_COL, ascending=True,
                                key=lambda s: s.astype(str).str.strip()
                                ).reset_index(drop=True)
        return df

    # ── Sheet 1 (active): Summary — full filtered data (no title row) ─────────
    ws_main = wb.active
    ws_main.title = "Summary"
    ws_main.sheet_view.showGridLines = False
    main_df = _sort_by_po_line(main_df)
    _write_sheet(ws_main, main_df)   # no title_text → headers at row 1
    _autofit_header_only(ws_main, hdr_row=1)

    # ── Sheet 2: Excluded POs (hidden by default — right-click tab → Unhide) ────
    EXCL_HDR_BG = "F4B8C8"   # same pink as the Main pivot header
    ws_del = wb.create_sheet("Excluded POs")
    ws_del.sheet_view.showGridLines  = False
    ws_del.sheet_properties.tabColor = EXCL_HDR_BG
    ws_del.sheet_state               = "hidden"

    if len(moved_df) == 0:
        ws_del.cell(row=1, column=1,
                    value="No rows were moved to Excluded POs in this run.").font = \
            Font(name="Calibri", size=11, italic=True, color="000000")
    else:
        n_del_cols = len(moved_df.columns)
        moved_df = _sort_by_po_line(moved_df)
        # No title row — headers at row 1, data from row 2
        _write_sheet(ws_del, moved_df)
        _autofit_header_only(ws_del, hdr_row=1)
        # Recolour header row to Main-pivot pink
        excl_hdr_fill = PatternFill("solid", fgColor=EXCL_HDR_BG)
        for ci in range(1, n_del_cols + 1):
            ws_del.cell(row=1, column=ci).fill = excl_hdr_fill
        # Data rows remain white (set by _write_sheet)

    # ── Sheet 3: Mapping Not Found (only created when there are unmapped rows) ─
    if unmapped_df is not None and len(unmapped_df) > 0:
        ws_unm = wb.create_sheet("Mapping Not Found")
        ws_unm.sheet_view.showGridLines   = False
        ws_unm.sheet_properties.tabColor  = "D97706"
        n_unm_cols = len(unmapped_df.columns)
        _write_sheet(
            ws_unm, unmapped_df,
            f"Mapping Not Found  —  {format_indian_number(len(unmapped_df))} row(s) "
            f"(Supplier Site not in mapping table)",
            UNM_TITLE,
        )
        _autofit_header_only(ws_unm, hdr_row=2)
        unm_fill1 = PatternFill("solid", fgColor=UNM_ROW1)
        unm_fill2 = PatternFill("solid", fgColor=UNM_ROW2)
        for ri in range(3, len(unmapped_df) + 3):
            fill = unm_fill2 if ri % 2 == 0 else unm_fill1
            for ci in range(1, n_unm_cols + 1):
                ws_unm.cell(row=ri, column=ci).fill = fill
        unm_hdr_fill = PatternFill("solid", fgColor=UNM_HDR_BG)
        for ci in range(1, n_unm_cols + 1):
            ws_unm.cell(row=2, column=ci).fill = unm_hdr_fill

    # ── Pivot sheet (inserted as position-0, i.e. first tab) ─────────────────
    _add_po_pivot_sheet(wb, main_df)

    wb.save(path)


# ── Shared helpers for MRN pivot sheets ───────────────────────────────────────
def _parse_mrn_period(p):
    """Parse an accounting period in either 'FEB-2026' or 'Feb-26' form.
    Returns a Timestamp, or None if unparseable."""
    s = str(p).strip()
    for fmt in ("%b-%Y", "%b-%y", "%B-%Y", "%B-%y"):
        try:
            return pd.to_datetime(s, format=fmt)
        except Exception:
            continue
    return None


def _mrn_period_sort_key(p):
    ts = _parse_mrn_period(p)
    return ts if ts is not None else pd.Timestamp.max


def _fmt_mrn_period(p):
    """'FEB-2026' or 'Feb-26' → 'Feb-26'"""
    ts = _parse_mrn_period(p)
    return ts.strftime("%b-%y") if ts is not None else str(p)


def _add_mrn_summary_sheet(wb, df: "pd.DataFrame") -> None:
    """Add a Location × Accounts-Incharge × Count pivot named 'Summary'."""
    mapped = df[df["Location"].astype(str).str.strip() != ""].copy()

    pivot = (
        mapped
        .groupby(["Location", "Accounts Incharge"], sort=False)
        .size()
        .reset_index(name="Count")
    )
    loc_totals  = pivot.groupby("Location")["Count"].sum()
    pivot["_r"] = pivot["Location"].map(loc_totals)
    pivot = (
        pivot.sort_values(["_r", "Count"], ascending=[False, False])
             .drop(columns=["_r"])
             .reset_index(drop=True)
    )

    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "1F3864"
    ws.sheet_view.showGridLines  = False

    TITLE_BG = "FADADD"
    HDR_BG   = "F4B8C8"
    HDR_FG   = "000000"
    TOTAL_BG = "F4B8C8"
    TOTAL_FG = "000000"

    bdr = _thin_border()

    # Row 1: title
    tc = ws.cell(row=1, column=1, value="Pending MRN — Summary")
    tc.font      = Font(name="Calibri", size=11, bold=True, color="000000")
    tc.fill      = PatternFill("solid", fgColor=TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    tc.border    = bdr
    ws.merge_cells("A1:C1")
    for c in range(2, 4):
        ws.cell(row=1, column=c).fill   = PatternFill("solid", fgColor=TITLE_BG)
        ws.cell(row=1, column=c).border = bdr

    # Row 2: headers
    for ci, h in enumerate(["Location", "Accounts Incharge", "Count"], 1):
        c           = ws.cell(row=2, column=ci, value=h)
        c.font      = Font(name="Calibri", size=11, bold=True)
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(
            horizontal="right" if ci == 3 else "left",
            vertical="center", indent=1 if ci < 3 else 0)
        c.border = bdr

    plain = Font(name="Calibri", size=11)
    current_row = 3
    grand_total = int(pivot["Count"].sum())

    locs = pivot["Location"].unique()
    for location in locs:
        sub = pivot[pivot["Location"] == location]
        incharges = sub["Accounts Incharge"].unique()

        if len(incharges) == 1:
            incharge = incharges[0]
            count    = int(sub["Count"].sum())
            ws.cell(row=current_row, column=1, value=location).font = plain
            ws.cell(row=current_row, column=2, value=incharge).font  = plain
            c3 = ws.cell(row=current_row, column=3, value=count)
            c3.font = plain
            c3.alignment = Alignment(horizontal="right", vertical="center")
            for ci in range(1, 4):
                ws.cell(row=current_row, column=ci).border = bdr
            current_row += 1
        else:
            detail_start = current_row
            ws.cell(row=current_row, column=1, value=location).font = Font(name="Calibri", size=11, bold=True)
            ws.cell(row=current_row, column=2, value="").font        = plain
            c3 = ws.cell(row=current_row, column=3)
            c3.font      = Font(name="Calibri", size=11, bold=True)
            c3.alignment = Alignment(horizontal="right", vertical="center")
            for ci in range(1, 4):
                ws.cell(row=current_row, column=ci).border = bdr
            current_row += 1

            for _, row in sub.iterrows():
                ws.cell(row=current_row, column=1, value="").font         = plain
                ws.cell(row=current_row, column=2, value=row["Accounts Incharge"]).font = plain
                c3 = ws.cell(row=current_row, column=3, value=int(row["Count"]))
                c3.font      = plain
                c3.alignment = Alignment(horizontal="right", vertical="center")
                for ci in range(1, 4):
                    ws.cell(row=current_row, column=ci).border = bdr
                current_row += 1

            detail_end = current_row - 1
            ws.cell(row=detail_start, column=3).value = (
                f"=SUBTOTAL(9,C{detail_start + 1}:C{detail_end})")

    # Grand Total row
    c1 = ws.cell(row=current_row, column=1, value="Grand Total")
    c1.font      = Font(name="Calibri", size=11, bold=True)
    c1.fill      = PatternFill("solid", fgColor=TOTAL_BG)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c1.border    = bdr
    c2 = ws.cell(row=current_row, column=2)
    c2.fill      = PatternFill("solid", fgColor=TOTAL_BG)
    c2.border    = bdr
    c3 = ws.cell(row=current_row, column=3)
    c3.value     = f"=SUBTOTAL(9,C3:C{current_row - 1})"
    c3.font      = Font(name="Calibri", size=11, bold=True)
    c3.fill      = PatternFill("solid", fgColor=TOTAL_BG)
    c3.alignment = Alignment(horizontal="right", vertical="center")
    c3.border    = bdr

    ws.freeze_panes        = "A3"
    ws.auto_filter.ref     = "A2:C2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10


def _add_vendorwise_pivot_sheet(wb, df: "pd.DataFrame") -> None:
    """Add a Supplier × Accounting-Period count pivot named 'Vendorwise Pivot'.

    Layout matches the reference PivotTable exactly:
      Row 1 : "Location" label  |  "(All)" — Report Filter simulation
      Row 2 : empty
      Row 3 : "Count of Location" label
      Row 4 : Supplier Name ▼ | Feb-26 ▼ | Mar-26 ▼ … | Grand Total ▼  ← AutoFilter
      Row 5… : data rows, sorted descending by Grand Total
      Last   : SUBTOTAL footer row

    Location is stored in a hidden column A so the AutoFilter can filter by it.
    Visible data starts at column B (Supplier Name).
    """
    # ── Build pivot data ──────────────────────────────────────────────────────
    work = df.copy()
    work["Location"] = work["Location"].astype(str).str.strip().replace("", "Unmapped")

    all_periods = [str(p) for p in work[MRN_ANCHOR_COL].dropna().unique()]
    periods     = sorted(all_periods, key=_mrn_period_sort_key)

    grouped = (
        work.groupby(["Location", "SUPPLIER NAME", MRN_ANCHOR_COL], sort=False)
        .size()
        .reset_index(name="Count")
    )

    pivot = grouped.pivot_table(
        index=["Location", "SUPPLIER NAME"],
        columns=MRN_ANCHOR_COL,
        values="Count",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    pivot.columns.name = None

    present = [p for p in periods if p in pivot.columns]
    pivot   = pivot[["Location", "SUPPLIER NAME"] + present]
    pivot["Grand Total"] = pivot[present].sum(axis=1)
    pivot = pivot.sort_values("Grand Total", ascending=False).reset_index(drop=True)

    period_labels = [_fmt_mrn_period(p) for p in present]

    # col A = Location (hidden), col B = Supplier Name, col C+ = periods, last = Grand Total
    n_visible   = 1 + len(present) + 1      # Supplier Name + periods + Grand Total
    n_cols      = 1 + n_visible              # incl. hidden Location col A
    gt_col      = n_cols                     # Grand Total column (1-based)
    # visible header row values (col A excluded from display but in AutoFilter)
    hdr_values  = ["Location", "Supplier Name"] + period_labels + ["Grand Total"]

    # ── Sheet setup ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Vendorwise Pivot")
    ws.sheet_properties.tabColor = "1F6DB1"
    ws.sheet_view.showGridLines  = False

    HDR_BG      = "C2D8ED"
    HDR_FG      = "000000"
    FOOT_BG     = "C2D8ED"
    FILTER_BG   = "EFEFEF"
    bdr         = _thin_border()

    # ── Row 1: Report Filter row — "Location" | "(All)" ───────────────────────
    ws.row_dimensions[1].height = 22
    lbl = ws.cell(row=1, column=1, value="Location")
    lbl.font      = Font(name="Calibri", size=11, bold=True)
    lbl.fill      = PatternFill("solid", fgColor=FILTER_BG)
    lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lbl.border    = bdr

    all_cell = ws.cell(row=1, column=2, value="(All)")
    all_cell.font      = Font(name="Calibri", size=11)
    all_cell.fill      = PatternFill("solid", fgColor="FFFFFF")
    all_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    all_cell.border    = bdr

    # ── Row 2: empty ──────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 8

    # ── Row 3: "Count of Location" label ─────────────────────────────────────
    ws.row_dimensions[3].height = 18
    count_lbl = ws.cell(row=3, column=1, value="Count of Location")
    count_lbl.font      = Font(name="Calibri", size=11, bold=True)
    count_lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    count_lbl.border    = bdr

    # ── Row 4: column headers + AutoFilter ───────────────────────────────────
    hdr_font = Font(name="Calibri", size=11, bold=True, color=HDR_FG)
    hdr_fill = PatternFill("solid", fgColor=HDR_BG)
    for ci, h in enumerate(hdr_values, 1):
        is_num = ci > 2                      # periods + Grand Total are numeric
        c           = ws.cell(row=4, column=ci, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.border    = bdr
        c.alignment = Alignment(
            horizontal="right" if is_num else "left",
            vertical="center", indent=0 if is_num else 1)

    ws.auto_filter.ref = f"A4:{get_column_letter(n_cols)}4"
    ws.freeze_panes    = "B5"               # freeze at Supplier Name; rows 1-4 frozen

    # Hide Location column so it doesn't clutter the view but stays filterable
    ws.column_dimensions["A"].hidden = True

    # ── Data rows ─────────────────────────────────────────────────────────────
    plain   = Font(name="Calibri", size=11)
    txt_aln = Alignment(horizontal="left",  vertical="center", indent=1)
    num_aln = Alignment(horizontal="right", vertical="center")

    data_start = 5
    for ri, row in enumerate(pivot.itertuples(index=False), data_start):
        # col A: Location (hidden, for AutoFilter)
        ws.cell(row=ri, column=1, value=row[0]).font = plain
        # col B: Supplier Name
        c2 = ws.cell(row=ri, column=2, value=row[1])
        c2.font = plain; c2.alignment = txt_aln; c2.border = bdr
        # col C+: period counts
        for ci_off, val in enumerate(row[2 : 2 + len(present)]):
            ci = 3 + ci_off
            c  = ws.cell(row=ri, column=ci, value=int(val) if val else 0)
            c.font = plain; c.alignment = num_aln; c.border = bdr
            c.number_format = INDIAN_FMT
        # last col: Grand Total
        c_gt = ws.cell(row=ri, column=gt_col, value=int(row[-1]))
        c_gt.font          = Font(name="Calibri", size=11, bold=True)
        c_gt.alignment     = num_aln; c_gt.border = bdr
        c_gt.number_format = INDIAN_FMT

    # ── SUBTOTAL footer row ───────────────────────────────────────────────────
    last_data = data_start + len(pivot) - 1
    foot_row  = last_data + 1
    foot_font = Font(name="Calibri", size=11, bold=True)
    foot_fill = PatternFill("solid", fgColor=FOOT_BG)
    foot_aln  = Alignment(horizontal="right", vertical="center")

    # col B footer label
    c2 = ws.cell(row=foot_row, column=2, value="Grand Total")
    c2.font = foot_font; c2.fill = foot_fill; c2.border = bdr
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for ci_off in range(len(present)):
        ci    = 3 + ci_off
        col_l = get_column_letter(ci)
        c     = ws.cell(row=foot_row, column=ci,
                        value=f"=SUBTOTAL(9,{col_l}{data_start}:{col_l}{last_data})")
        c.font = foot_font; c.fill = foot_fill; c.border = bdr; c.alignment = foot_aln
        c.number_format = INDIAN_FMT

    gt_l = get_column_letter(gt_col)
    c_gt = ws.cell(row=foot_row, column=gt_col,
                   value=f"=SUBTOTAL(9,{gt_l}{data_start}:{gt_l}{last_data})")
    c_gt.font = foot_font; c_gt.fill = foot_fill; c_gt.border = bdr; c_gt.alignment = foot_aln
    c_gt.number_format = INDIAN_FMT

    # ── Borders: full grid over header + data + footer (visible cols 2..gt) ───
    for r in range(4, foot_row + 1):
        for c in range(2, gt_col + 1):
            ws.cell(row=r, column=c).border = bdr

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["B"].width = 40   # Supplier Name
    for ci_off in range(len(present)):
        ws.column_dimensions[get_column_letter(3 + ci_off)].width = 11
    ws.column_dimensions[get_column_letter(gt_col)].width = 13


def _add_locationwise_pivot_sheet(wb, df: "pd.DataFrame") -> None:
    """Add a Location × Accounts-Incharge × Period count pivot named 'Locationwise Pivot'.

    Layout mirrors the screenshot:
      Row 1 : title bar
      Row 2 : filter hint
      Row 3 : headers (Location | Accounts Incharge | Feb-26 | … | Grand Total)
              ← AutoFilter applied here
      Row 4… : data rows sorted descending by Grand Total
      Last   : SUBTOTAL footer
    """
    # ── Build pivot data ──────────────────────────────────────────────────────
    work = df[df["Location"].astype(str).str.strip() != ""].copy()

    all_periods = [str(p) for p in work[MRN_ANCHOR_COL].dropna().unique()]
    periods     = sorted(all_periods, key=_mrn_period_sort_key)

    grouped = (
        work.groupby(["Location", "Accounts Incharge", MRN_ANCHOR_COL], sort=False)
        .size()
        .reset_index(name="Count")
    )

    pivot = grouped.pivot_table(
        index=["Location", "Accounts Incharge"],
        columns=MRN_ANCHOR_COL,
        values="Count",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    pivot.columns.name = None

    present = [p for p in periods if p in pivot.columns]
    pivot   = pivot[["Location", "Accounts Incharge"] + present]
    pivot["Grand Total"] = pivot[present].sum(axis=1)
    pivot = pivot.sort_values("Grand Total", ascending=False).reset_index(drop=True)

    period_labels   = [_fmt_mrn_period(p) for p in present]
    all_col_headers = ["Location", "Accounts Incharge"] + period_labels + ["Grand Total"]

    # ── Sheet setup ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Locationwise Pivot")
    ws.sheet_properties.tabColor = "375623"   # dark green — distinct from other tabs
    ws.sheet_view.showGridLines  = False

    TITLE_BG = "FADADD"   # same pink family as Summary
    HDR_BG   = "F4B8C8"
    HDR_FG   = "000000"
    FOOT_BG  = "F4B8C8"
    bdr      = _thin_border()
    n_cols   = len(all_col_headers)
    gt_col   = n_cols

    # Row 1: column headers + AutoFilter
    hdr_font = Font(name="Calibri", size=11, bold=True, color=HDR_FG)
    hdr_fill = PatternFill("solid", fgColor=HDR_BG)
    for ci, h in enumerate(all_col_headers, 1):
        c           = ws.cell(row=1, column=ci, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.border    = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"
    ws.freeze_panes    = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    plain   = Font(name="Calibri", size=11)
    ctr_aln = Alignment(horizontal="center", vertical="center")

    data_start = 2
    for ri, row in enumerate(pivot.itertuples(index=False), data_start):
        # Location (col 1)
        c1 = ws.cell(row=ri, column=1, value=row[0])
        c1.font = plain; c1.alignment = ctr_aln; c1.border = bdr
        # Accounts Incharge (col 2)
        c2 = ws.cell(row=ri, column=2, value=row[1])
        c2.font = plain; c2.alignment = ctr_aln; c2.border = bdr
        # Period counts (cols 3 … n_cols-1)
        for ci_offset, val in enumerate(row[2 : 2 + len(present)]):
            ci  = 3 + ci_offset
            c   = ws.cell(row=ri, column=ci, value=int(val) if val else 0)
            c.font = plain; c.alignment = ctr_aln; c.border = bdr
            c.number_format = INDIAN_FMT
        # Grand Total (last col)
        c_gt = ws.cell(row=ri, column=gt_col, value=int(row[-1]))
        c_gt.font          = Font(name="Calibri", size=11, bold=True)
        c_gt.alignment     = ctr_aln; c_gt.border = bdr
        c_gt.number_format = INDIAN_FMT

    # ── SUBTOTAL footer row ───────────────────────────────────────────────────
    last_data = data_start + len(pivot) - 1
    foot_row  = last_data + 1

    foot_font = Font(name="Calibri", size=11, bold=True)
    foot_fill = PatternFill("solid", fgColor=FOOT_BG)

    c1 = ws.cell(row=foot_row, column=1, value="Grand Total")
    c1.font = foot_font; c1.fill = foot_fill; c1.border = bdr
    c1.alignment = ctr_aln

    ws.cell(row=foot_row, column=2).fill   = foot_fill
    ws.cell(row=foot_row, column=2).border = bdr
    ws.cell(row=foot_row, column=2).alignment = ctr_aln

    for ci_offset in range(len(present)):
        ci    = 3 + ci_offset
        col_l = get_column_letter(ci)
        c     = ws.cell(row=foot_row, column=ci,
                        value=f"=SUBTOTAL(9,{col_l}{data_start}:{col_l}{last_data})")
        c.font = foot_font; c.fill = foot_fill
        c.border = bdr; c.alignment = ctr_aln
        c.number_format = INDIAN_FMT

    gt_l = get_column_letter(gt_col)
    c_gt = ws.cell(row=foot_row, column=gt_col,
                   value=f"=SUBTOTAL(9,{gt_l}{data_start}:{gt_l}{last_data})")
    c_gt.font = foot_font; c_gt.fill = foot_fill
    c_gt.border = bdr; c_gt.alignment = ctr_aln
    c_gt.number_format = INDIAN_FMT

    # ── Borders: full grid over title + header + data + footer ────────────────
    for r in range(1, foot_row + 1):
        for c in range(1, gt_col + 1):
            ws.cell(row=r, column=c).border = bdr

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26   # Location
    ws.column_dimensions["B"].width = 26   # Accounts Incharge
    for ci_offset in range(len(present)):
        ws.column_dimensions[get_column_letter(3 + ci_offset)].width = 11
    ws.column_dimensions[get_column_letter(gt_col)].width = 13


def write_formatted_mrn_excel(df: "pd.DataFrame", path: str) -> None:
    """Write Pending MRN df to formatted xlsx: data sheet + unmapped sites sheet."""
    wb     = Workbook()
    ws     = wb.active
    ws.title = "Summary"

    cols   = list(df.columns)
    n_cols = len(cols)
    n_rows = len(df)
    bdr    = _thin_border()

    HDR_BG   = "C2D8ED"   # light blue header
    HDR_FG   = "000000"   # black text
    TITLE_BG = "DAEAF7"   # very light blue title bar

    # Row 1: column headers
    hdr_font = Font(name="Calibri", size=11, bold=True, color=HDR_FG)
    hdr_fill = PatternFill("solid", fgColor=HDR_BG)
    for ci, col_name in enumerate(cols, 1):
        c        = ws.cell(row=1, column=ci, value=col_name)
        c.font   = hdr_font
        c.fill   = hdr_fill
        c.border = bdr

    # Numeric columns
    NUMERIC_COLS = {"BASE AMOUNT", "CGST", "SGST", "IGST", "RECEIPT QUANTITY"}
    txt_font = Font(name="Calibri", size=11, color="000000")
    num_font = Font(name="Calibri", size=11, color="000000")

    for ri, row_vals in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row_vals, 1):
            col_name = cols[ci - 1]
            c        = ws.cell(row=ri, column=ci)
            c.border = bdr
            if col_name in NUMERIC_COLS:
                c.font          = num_font
                c.alignment     = Alignment(horizontal="right")
                c.number_format = INDIAN_DECIMAL_FMT
                try:
                    c.value = (
                        float(val)
                        if val is not None and str(val).strip() not in ("", "nan")
                        else 0.0
                    )
                except (ValueError, TypeError):
                    c.value = val
            elif "date" in col_name.lower():
                # Write as a real Excel date serial so date functions work
                date_val = _to_excel_date(val)
                c.font = txt_font
                if date_val is not None:
                    c.value         = date_val
                    c.number_format = "DD-MMM-YYYY"
                    c.alignment     = Alignment(horizontal="center")
                else:
                    c.value = val if val is not None else ""
            else:
                c.font  = txt_font
                c.value = val if val is not None else ""

    ws.freeze_panes              = "A2"
    ws.auto_filter.ref           = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
    ws.sheet_properties.tabColor = "1F6DB1"

    # ── Sheet 2: Unmapped Sites ───────────────────────────────────────────────
    show_cols    = [c for c in ["SUPPLIER NAME", MRN_SITE_COL, "ORGANIZATION NAME",
                                 "RECEIPT NUMBER", "BASE AMOUNT"] if c in df.columns]
    unmatched_df = (
        df[df["Location"].astype(str).str.strip() == ""][show_cols]
        .copy()
        .reset_index(drop=True)
    )

    if len(unmatched_df) > 0:
        ws2         = wb.create_sheet("Unmapped Sites")
        WARN_BG     = "7F0000"
        WARN_HDR_BG = "C0392B"
        ROW_WARN    = "FFF3F3"
        ROW_WARN_ALT= "FFE5E5"
        n2_cols     = max(len(show_cols), 1)

        ws2.row_dimensions[1].height = 28
        t2 = ws2.cell(row=1, column=1,
                      value=f"Unmapped Supplier Sites  —  {format_indian_number(len(unmatched_df))} row(s) with no Location / Accounts Incharge")
        t2.font      = Font(name="Calibri", size=11, bold=True, color="000000")
        t2.fill      = PatternFill("solid", fgColor=WARN_BG)
        t2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        t2.border    = bdr
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n2_cols)
        for c in range(2, n2_cols + 1):
            ws2.cell(row=1, column=c).fill   = PatternFill("solid", fgColor=WARN_BG)
            ws2.cell(row=1, column=c).border = bdr

        ws2.row_dimensions[2].height = 20
        exp = ws2.cell(
            row=2, column=1,
            value="The Supplier Sites listed below were not found in the mapping. "
                  "Please add them via Manage Mappings.")
        exp.font      = Font(name="Calibri", size=11, italic=True, color="000000")
        exp.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        exp.border    = bdr
        ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n2_cols)
        for c in range(2, n2_cols + 1):
            ws2.cell(row=2, column=c).border = bdr

        summary = (
            unmatched_df.groupby(MRN_SITE_COL, sort=False)
            .size()
            .reset_index(name="Row Count")
            .sort_values("Row Count", ascending=False)
        )
        ws2.row_dimensions[3].height = 8
        ws2.row_dimensions[4].height = 28
        sum_cols     = [MRN_SITE_COL, "Row Count"]
        sum_hdr_font = Font(name="Calibri", size=11, bold=True, color="000000")
        sum_hdr_fill = PatternFill("solid", fgColor=WARN_HDR_BG)
        sum_hdr_aln  = Alignment(horizontal="center", vertical="center")
        sum_bdr      = _thin_border()
        for ci, h in enumerate(sum_cols, 1):
            c           = ws2.cell(row=4, column=ci, value=h)
            c.font      = sum_hdr_font
            c.fill      = sum_hdr_fill
            c.alignment = sum_hdr_aln
            c.border    = sum_bdr

        txt_f = Font(name="Calibri", size=11)
        for ri, row in enumerate(summary.itertuples(index=False), 5):
            ws2.row_dimensions[ri].height = 18
            bg = ROW_WARN_ALT if ri % 2 == 0 else ROW_WARN
            c1 = ws2.cell(row=ri, column=1, value=row[0])
            c1.font      = txt_f
            c1.fill      = PatternFill("solid", fgColor=bg)
            c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c1.border    = sum_bdr
            c2 = ws2.cell(row=ri, column=2, value=int(row[1]))
            c2.font      = txt_f
            c2.fill      = PatternFill("solid", fgColor=bg)
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.border    = sum_bdr

        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 16

        detail_start_row = 5 + len(summary) + 2
        ws2.row_dimensions[detail_start_row - 1].height = 14

        dt = ws2.cell(row=detail_start_row - 1, column=1,
                      value="Full Row Detail for Unmapped Sites")
        dt.font      = Font(name="Calibri", size=11, bold=True, color="000000")
        dt.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        dt.border    = sum_bdr
        ws2.merge_cells(start_row=detail_start_row - 1, start_column=1,
                        end_row=detail_start_row - 1, end_column=n2_cols)
        for c in range(2, n2_cols + 1):
            ws2.cell(row=detail_start_row - 1, column=c).border = sum_bdr

        ws2.row_dimensions[detail_start_row].height = 28
        for ci, h in enumerate(show_cols, 1):
            c           = ws2.cell(row=detail_start_row, column=ci, value=h)
            c.font      = sum_hdr_font
            c.fill      = sum_hdr_fill
            c.alignment = sum_hdr_aln
            c.border    = sum_bdr

        amt_fmt = Font(name="Calibri", size=11)
        for ri, row in enumerate(unmatched_df.itertuples(index=False),
                                  detail_start_row + 1):
            ws2.row_dimensions[ri].height = 18
            bg = ROW_WARN_ALT if ri % 2 == 0 else ROW_WARN
            for ci, val in enumerate(row, 1):
                c        = ws2.cell(row=ri, column=ci)
                c.border = sum_bdr
                c.fill   = PatternFill("solid", fgColor=bg)
                if show_cols[ci - 1] == "BASE AMOUNT":
                    try:
                        c.value         = float(val) if val is not None else 0.0
                        c.number_format = INDIAN_DECIMAL_FMT
                        c.font          = amt_fmt
                        c.alignment     = Alignment(horizontal="right", vertical="center")
                    except (ValueError, TypeError):
                        c.value     = val
                        c.font      = txt_f
                        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                else:
                    c.value     = val if val is not None else ""
                    c.font      = txt_f
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        for ci, col_name in enumerate(show_cols, 1):
            max_len = len(col_name)
            for ri in range(detail_start_row + 1,
                            min(detail_start_row + 201, detail_start_row + 1 + len(unmatched_df))):
                v = ws2.cell(row=ri, column=ci).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws2.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 3, 14), 45)

        ws2.freeze_panes = "A5"
        ws2.sheet_properties.tabColor = "C0392B"

    # ── Sheet 3: Vendorwise Pivot (static fallback) ───────────────────────────
    _add_vendorwise_pivot_sheet(wb, df)

    # ── Sheet 4: Locationwise Pivot (static fallback) ─────────────────────────
    _add_locationwise_pivot_sheet(wb, df)

    # ── Reorder sheets: Locationwise, Vendorwise, Summary, Unmapped Sites ──────
    _desired = ["Locationwise Pivot", "Vendorwise Pivot", "Summary", "Unmapped Sites"]
    _ordered = [wb[n] for n in _desired if n in wb.sheetnames]
    _ordered += [ws for ws in wb._sheets if ws not in _ordered]   # keep any extras
    wb._sheets = _ordered

    for sheet in wb.worksheets:
        _autofit_columns(sheet)

    wb.save(path)


