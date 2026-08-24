"""GST Invoice Number Adder — ported from the desktop app "RDCTRAK GST
Invoice Number Enricher v12" (single-file customtkinter script).

Framework-agnostic port: every GUI class (GSTEnricherApp, THEMES, the whole
customtkinter widget tree) is dropped entirely — none of it is business
logic. What's ported is exactly the v9-proven query/matching logic plus the
v11/v12 performance layer on top of it, verbatim:

  - GST_QUERY: byte-for-byte identical to the original. Per the original's
    own docstring, this is "the only approach that works correctly" —
    filtering by BOTH invoice number AND transaction date, with
    _to_oracle_date() generating the date string used as BOTH the SQL bind
    variable and the gst_map dict key, so they're identical by construction
    and immune to date-format mismatches. Do not "simplify" this query.
  - Vectorized (inv_no, trx_date) key extraction, unique-date-parsed-once
    lookup, and the pooled/parallel Oracle fetch (ThreadPoolExecutor +
    oracledb connection pool) — all unchanged from the original.
  - openpyxl-based GST column insertion with style/merge/width preservation
    — unchanged.

Adaptations for the web suite (matching this suite's established
conventions, not desktop-app choices):
  - Oracle connectivity is an explicit OracleConfig value object built by
    the router from app.config's ORACLE_* env vars, instead of hardcoded
    DB_USER/DB_PASS/DB_HOST/DB_PORT/DB_SERVICE globals and a
    sys.frozen/_MEIPASS-based instant-client path resolver — same pattern
    already established in services/unapplied_receipts/processor.py (this
    suite's other Oracle-backed tool, same actual Oracle server).
  - No tkinter progress_cb/status_cb/cancel_ev split. Every function that
    used to report progress now takes the suite's standard `log_q` (an
    object exposing .put((level, message)); see e.g.
    app.routers.unaccounted_txn's _LogQueue adapter, which the router wires
    up to translate these messages into the background job's progress_cb).
  - .xlsb uploads no longer go through Excel COM automation at all. The
    original app (and an earlier version of this port) converted .xlsb to
    .xlsx via win32com so openpyxl could write into it - but Excel COM
    automation turned out to be too fragile running unattended on a server
    (needs an interactive desktop session; breaks under Task Scheduler
    with nobody logged in, or when a leftover EXCEL.EXE process is stuck).
    .xlsb genuinely doesn't expose cell styling to any Python library
    (verified against both pyxlsb and python-calamine - neither returns
    anything beyond raw cell values), so build_gst_report_pure_python()
    builds a brand-new .xlsx directly from the data pyxlsb already reads,
    styled with this suite's own report look instead of trying to
    preserve the source .xlsb's specific formatting - a deliberate,
    explicitly-agreed tradeoff in exchange for zero external dependency.
    .xlsx uploads are unaffected: insert_gst_column() still edits the
    real uploaded file in place via openpyxl, preserving its exact
    original styling, exactly as before.
"""

from __future__ import annotations

import copy
import re as _re
from dataclasses import dataclass
from datetime import date as _date_type, datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

import oracledb
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

try:
    import pyxlsb as _pyxlsb
    _PYXLSB_OK = True
except ImportError:
    _PYXLSB_OK = False


# ── Oracle connection config (same shape as unapplied_receipts.processor) ────

@dataclass(frozen=True)
class OracleConfig:
    host: str
    port: str
    service_name: str
    user: str
    password: str
    instant_client_dir: str

    @property
    def dsn(self) -> str:
        return f"{self.host}:{self.port}/{self.service_name}"


def init_oracle_client(instant_client_dir: str) -> None:
    """init_oracle_client() may only be called once per process and raises
    on subsequent calls - safe to ignore, same as the original's bare
    try/except Exception: pass."""
    try:
        oracledb.init_oracle_client(lib_dir=instant_client_dir)
    except Exception:
        pass


# ── Constants (verbatim from the original) ───────────────────────────────────

SKIP_ROWS = 13
COL_INV_NO = "Invoice No/ Receipt No"
COL_INV_DATE = "Invoice/ Receipt Date"
COL_GST_NEW = "GST Invoice Number"

THREAD_WORKERS = 8
POOL_MIN = 2
POOL_MAX = THREAD_WORKERS

# ── GST query (byte-for-byte identical to the original v9/v12 query) ─────────

GST_QUERY = """
SELECT
    rtrim(
        regexp_replace(
            LISTAGG(jtl.tax_invoice_num, ', ')
                WITHIN GROUP (ORDER BY jtl.tax_invoice_num),
            '([^-]*)(-\\1)+($|-)',
            '\\1\\3'
        ),
        '-'
    ) AS gst_inv_num
FROM
    apps.jai_tax_det_factors      jtl,
    apps.ra_customer_trx_lines_all rctl,
    apps.ra_customer_trx_all       rcta
WHERE
        jtl.trx_line_id      = rctl.customer_trx_line_id
    AND jtl.trx_id           = rctl.customer_trx_id
    AND rctl.customer_trx_id = rcta.customer_trx_id
    AND rcta.trx_number      = :inv_no
    AND rctl.line_type       = 'LINE'
    AND jtl.entity_code      = 'TRANSACTIONS'
    AND jtl.trx_date         = :trx_date
"""


# ── Date helpers (verbatim logic from the original) ──────────────────────────

def _xlsb_serial_to_datetime(serial):
    try:
        n = float(serial)
        if n <= 0:
            return None
        return _pyxlsb.convert_date(n)
    except Exception:
        return None


def _to_oracle_date(val) -> str:
    """Convert any date representation to Oracle DD-Mon-YY string."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d-%b-%y")
    if isinstance(val, _date_type):
        return val.strftime("%d-%b-%y")
    if hasattr(val, "strftime"):
        try:
            return pd.Timestamp(val).strftime("%d-%b-%y")
        except Exception:
            pass
    s = str(val).strip()
    try:
        f = float(s)
        if f > 1000 and _PYXLSB_OK:
            dt = _xlsb_serial_to_datetime(f)
            if dt:
                return dt.strftime("%d-%b-%y")
    except ValueError:
        pass
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d",
                "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y",
                "%d %b %Y", "%d %b %y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d-%b-%y")
        except ValueError:
            pass
    return s


def _clean_inv_no(val) -> str:
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


# ── Vectorized key extraction (verbatim from the original) ───────────────────

_INVALID_INV = frozenset({"nan", "none", "nat", ""})


def _extract_pairs_vectorized(df, inv_no_col, inv_date_col) -> list:
    """Extract unique (inv_no, oracle_date) pairs using vectorised pandas
    ops. Each unique date value is parsed only once (lookup table)."""
    inv = (df[inv_no_col]
           .astype(str)
           .str.strip()
           .str.replace(r'\.0$', '', regex=True))

    valid_mask = ~inv.str.lower().isin(_INVALID_INV)
    inv = inv[valid_mask]
    raw_dates = df.loc[valid_mask, inv_date_col]

    unique_raw = raw_dates.unique()
    date_lookup = {v: _to_oracle_date(v) for v in unique_raw}
    oracle_dates = raw_dates.map(date_lookup)

    pairs = set(zip(inv, oracle_dates))
    pairs.discard(("nan", ""))
    pairs.discard(("None", ""))
    return list(pairs)


# ── Read input file for key extraction (verbatim logic) ──────────────────────

def read_data_df(input_path: str):
    ext = input_path.rsplit(".", 1)[-1].lower()
    engine = "pyxlsb" if ext == "xlsb" else "openpyxl"
    df = pd.read_excel(input_path, skiprows=SKIP_ROWS,
                        engine=engine, dtype=object, header=0)
    df = df.dropna(how="all").reset_index(drop=True)
    if "Customer Name" in df.columns:
        df = df[~df["Customer Name"].astype(str)
                 .str.strip().str.lower().str.startswith("grand total")]
        df = df.reset_index(drop=True)

    col_map = {str(c).strip().lower(): c for c in df.columns}
    inv_no_col = col_map.get(COL_INV_NO.strip().lower())
    inv_date_col = col_map.get(COL_INV_DATE.strip().lower())

    if inv_no_col is None:
        raise ValueError(f"Column '{COL_INV_NO}' not found.\nFound: {list(df.columns)}")
    if inv_date_col is None:
        raise ValueError(f"Column '{COL_INV_DATE}' not found.\nFound: {list(df.columns)}")

    return df, inv_no_col, inv_date_col


# ── Pure-Python .xlsb output builder (no Excel/COM, no external app) ─────────
# .xlsb genuinely doesn't expose cell formatting (font/fill/border/number
# format) to any Python library - verified directly against both pyxlsb
# (Cell = namedtuple('Cell', ['r', 'c', 'v']) - value only) and
# python-calamine (same: values + merged-cell ranges, no style info) against
# a real sample file. Only a full spreadsheet engine (Excel, LibreOffice)
# actually parses .xlsb's style records, and Excel COM automation turned out
# to be too fragile server-side (requires an interactive desktop session,
# breaks under Task Scheduler / no logged-in console, leftover EXCEL.EXE
# processes get stuck). So for .xlsb uploads this builds a brand-new .xlsx
# from the data pyxlsb already gives us, styled with this suite's own
# established report look, instead of trying to preserve the source file's
# specific original formatting - a deliberate accepted tradeoff (data is
# 100% unchanged; only the visual styling differs from the source .xlsb).
# .xlsx uploads still go through insert_gst_column() below, which preserves
# the original file's exact styling - openpyxl can read/write .xlsx natively,
# no conversion needed there at all.

_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DAT_FONT = Font(name="Calibri", size=10)
_GST_FONT = Font(name="Calibri", size=10, bold=True, color="1F3864")
_THIN_BORDER = Border(*(Side(style="thin", color="D8DEE8") for _ in range(4)))


def _cell_value_for_output(raw_val, is_date_col: bool):
    """Returns (value, needs_date_format). pyxlsb doesn't expose cell number
    formats, so a date column often comes back as a raw Excel serial float
    (e.g. 46032.0) rather than a real date - convert it to an actual date so
    the output shows "24-Aug-2026", not a bare number."""
    if pd.isna(raw_val):
        return None, False
    if is_date_col:
        if isinstance(raw_val, (int, float)) and not isinstance(raw_val, bool):
            dt = _xlsb_serial_to_datetime(raw_val)
            if dt is not None:
                return dt, True
        if hasattr(raw_val, "strftime"):
            return raw_val, True
    return raw_val, False


def build_gst_report_pure_python(df, inv_no_col: str, inv_date_col: str,
                                  gst_map: dict, output_path: str, log_q) -> tuple[int, int, int]:
    """Build the enriched output workbook directly from the already-read
    .xlsb data, with no Excel/COM/LibreOffice involved at all. Returns
    (total, found, blank), same contract as insert_gst_column()."""
    log_q.put(("info", "Building output workbook (pure Python, no Excel required)..."))

    orig_cols = list(df.columns)
    inv_idx = orig_cols.index(inv_no_col)
    # GST column goes immediately before Invoice No, matching where
    # insert_gst_column() places it for .xlsx uploads.
    new_cols = orig_cols[:inv_idx] + [COL_GST_NEW] + orig_cols[inv_idx:]

    wb = Workbook()
    ws = wb.active
    ws.title = "GST Enriched"

    for ci, name in enumerate(new_cols, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = _HDR_ALIGN
        c.border = _THIN_BORDER
    ws.row_dimensions[1].height = 26

    total = found = blank = 0
    row_idx = 2
    for row in df.itertuples(index=False):
        row_dict = dict(zip(orig_cols, row))
        raw_inv = row_dict.get(inv_no_col)
        raw_date = row_dict.get(inv_date_col)
        if pd.isna(raw_inv) and pd.isna(raw_date):
            continue

        inv_no = _clean_inv_no(raw_inv) if not pd.isna(raw_inv) else ""
        trx_date = _to_oracle_date(raw_date) if not pd.isna(raw_date) else ""
        if not inv_no or inv_no in ("nan", "None", ""):
            continue

        total += 1
        gst = gst_map.get((inv_no, trx_date), "")
        if gst:
            found += 1
        else:
            blank += 1

        for ci, name in enumerate(new_cols, 1):
            cell = ws.cell(row=row_idx, column=ci)
            cell.border = _THIN_BORDER
            if name == COL_GST_NEW:
                cell.value = gst or None
                cell.font = _GST_FONT
                continue
            val, is_date = _cell_value_for_output(row_dict.get(name), name == inv_date_col)
            cell.value = val
            cell.font = _DAT_FONT
            if is_date:
                cell.number_format = "DD-MMM-YYYY"
        row_idx += 1

    n_data_rows = row_idx - 1
    for ci, name in enumerate(new_cols, 1):
        col_letter = get_column_letter(ci)
        max_len = len(str(name))
        for r in range(2, min(n_data_rows, 200) + 1):
            v = ws.cell(row=r, column=ci).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"
    if n_data_rows >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(new_cols))}{n_data_rows}"

    log_q.put(("info", "Saving enriched output..."))
    wb.save(output_path)
    log_q.put(("ok", f"Saved - total={total} found={found} blank={blank}"))
    return total, found, blank


# ── Style copy helper (verbatim) ──────────────────────────────────────────────

def _copy_cell_style(src, dst):
    if not src.has_style:
        return
    try:
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)
    except Exception:
        pass


# ── Insert GST column (verbatim logic) ────────────────────────────────────────

def insert_gst_column(wb_path: str, output_path: str, gst_map: dict,
                       header_row: int, log_q) -> tuple[int, int, int]:
    log_q.put(("info", "Loading workbook for column insertion..."))
    wb = load_workbook(wb_path)
    ws = wb.active

    merged_before = [str(m) for m in ws.merged_cells.ranges]

    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None:
            header_map[str(v).strip()] = c

    def _require(name):
        if name not in header_map:
            raise ValueError(
                f"Header '{name}' not found in row {header_row}.\n"
                f"Found: {list(header_map.keys())}")
        return header_map[name]

    insert_col = _require(COL_INV_NO)
    inv_no_col = insert_col
    inv_date_col = _require(COL_INV_DATE)

    col_widths = {column_index_from_string(k): v.width
                  for k, v in ws.column_dimensions.items()}
    row_heights = {r: ws.row_dimensions[r].height
                   for r in ws.row_dimensions if ws.row_dimensions[r].height}

    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    ws.insert_cols(insert_col)

    if inv_no_col >= insert_col:
        inv_no_col += 1
    if inv_date_col >= insert_col:
        inv_date_col += 1

    for ci_before, w in sorted(col_widths.items(), reverse=True):
        ci_after = ci_before + 1 if ci_before >= insert_col else ci_before
        ws.column_dimensions[get_column_letter(ci_after)].width = w

    gst_w = col_widths.get(inv_no_col - 1, 28)
    ws.column_dimensions[get_column_letter(insert_col)].width = max(gst_w, 24)

    for r, h in row_heights.items():
        ws.row_dimensions[r].height = h

    cell_ref_pat = _re.compile(r'([A-Z]+)(\d+)')

    def _shift_col_str(col_str: str) -> str:
        ci = column_index_from_string(col_str)
        return get_column_letter(ci + 1 if ci >= insert_col else ci)

    def _shift_ref(ref: str) -> str:
        return cell_ref_pat.sub(lambda m: _shift_col_str(m.group(1)) + m.group(2), ref)

    for mr_str in merged_before:
        try:
            ws.merge_cells(_shift_ref(mr_str))
        except Exception:
            pass

    inv_hdr = ws.cell(row=header_row, column=inv_no_col)
    gst_hdr = ws.cell(row=header_row, column=insert_col)
    gst_hdr.value = COL_GST_NEW
    _copy_cell_style(inv_hdr, gst_hdr)
    gst_hdr.font = Font(
        name=inv_hdr.font.name or "Calibri",
        size=inv_hdr.font.size or 9,
        bold=inv_hdr.font.bold,
        italic=inv_hdr.font.italic,
        underline=inv_hdr.font.underline,
        color=inv_hdr.font.color.rgb
        if inv_hdr.font.color and inv_hdr.font.color.type == "rgb"
        else "FFFFFF",
    )

    log_q.put(("info", "Writing GST values..."))
    found = blank = total = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        inv_no_cell = ws.cell(row=row_idx, column=inv_no_col)
        inv_date_cell = ws.cell(row=row_idx, column=inv_date_col)
        raw_inv = inv_no_cell.value
        raw_date = inv_date_cell.value

        if raw_inv is None and raw_date is None:
            continue

        inv_no = _clean_inv_no(raw_inv) if raw_inv is not None else ""
        trx_date = _to_oracle_date(raw_date) if raw_date is not None else ""

        if not inv_no or inv_no in ("nan", "None", ""):
            continue

        total += 1
        gst = gst_map.get((inv_no, trx_date), "")

        new_cell = ws.cell(row=row_idx, column=insert_col)
        new_cell.value = gst
        _copy_cell_style(inv_no_cell, new_cell)

        if gst:
            found += 1
        else:
            blank += 1

    log_q.put(("info", "Saving enriched output..."))
    wb.save(output_path)
    log_q.put(("ok", f"Saved - total={total} found={found} blank={blank}"))
    return total, found, blank


# ── Oracle fetch: pooled + parallel (verbatim logic) ──────────────────────────

_ORACLE_CONNECT_ERROR = (
    "Could not connect to the Oracle ERP database - this isn't an application "
    "bug. Check that the ERP server is reachable from this network and that "
    "ORACLE_HOST/ORACLE_SERVICE_NAME/ORACLE_USER/ORACLE_PASSWORD in backend/.env "
    "are correct, then try again."
)


def _make_pool(oracle_cfg: OracleConfig, log_q):
    log_q.put(("info", f"Creating connection pool (min={POOL_MIN}, max={POOL_MAX})..."))
    try:
        pool = oracledb.create_pool(
            user=oracle_cfg.user, password=oracle_cfg.password, dsn=oracle_cfg.dsn,
            min=POOL_MIN, max=POOL_MAX, increment=1,
            timeout=30,
        )
    except Exception as exc:
        # Catches bare Exception, not just oracledb.Error: a DNS/host
        # resolution failure surfaces as a plain socket.gaierror in thin
        # mode, not an oracledb exception (verified against a real failed
        # connection). A cryptic driver string here reads as "the
        # application is broken" - say plainly that Oracle is unreachable.
        raise RuntimeError(_ORACLE_CONNECT_ERROR) from exc
    log_q.put(("ok", "Pool ready - connections established"))
    return pool


def _fetch_batch_pooled(pairs: list, pool, log_q) -> dict:
    """Fetch GST numbers for a batch of (inv_no, trx_date) pairs. gst_map
    key is (inv_no, trx_date) - the SAME strings used as SQL bind
    variables, so the lookup in insert_gst_column always hits."""
    results = {}
    try:
        with pool.acquire() as conn:
            cursor = conn.cursor()
            cursor.prefetchrows = 2
            for inv_no, trx_date in pairs:
                key = (inv_no, trx_date)
                try:
                    cursor.execute(GST_QUERY, inv_no=inv_no, trx_date=trx_date)
                    row = cursor.fetchone()
                    gst = str(row[0]).strip() if (row and row[0] is not None) else ""
                    results[key] = gst
                except Exception as e:
                    results[key] = ""
                    log_q.put(("err", f"inv={inv_no!r} date={trx_date!r} -> {e}"))
            cursor.close()
    except Exception as conn_err:
        # pool.acquire() itself failed (ERP dropped mid-run, network blip,
        # DNS failure) - distinct from a single bad row above. Let this
        # propagate instead of blank-filling the batch: silently returning
        # "" for every pair here looks identical to "no GST number found",
        # which hides a real connectivity failure behind what looks like a
        # normal empty result.
        log_q.put(("err", f"Pool error: {conn_err}"))
        raise
    return results


def _fetch_all_parallel(pairs_unique: list, pool, log_q, progress_cb) -> dict:
    n = len(pairs_unique)
    batch_sz = max(20, min(50, max(1, n // (THREAD_WORKERS * 4))))
    batches = [pairs_unique[i:i + batch_sz] for i in range(0, n, batch_sz)]

    log_q.put(("info", f"Fetching {n} unique pairs in {len(batches)} batch(es), "
                        f"{THREAD_WORKERS} workers"))

    merged: dict = {}
    done = 0
    lock = Lock()
    conn_failures = 0

    with ThreadPoolExecutor(max_workers=THREAD_WORKERS) as executor:
        futures = {executor.submit(_fetch_batch_pooled, b, pool, log_q): b for b in batches}
        for fut in as_completed(futures):
            try:
                res = fut.result()
                with lock:
                    merged.update(res)
                    done += len(res)
                if progress_cb:
                    pct = 0.10 + 0.78 * (done / max(n, 1))
                    progress_cb(min(pct, 0.88), f"Fetched {done:,} / {n:,} GST numbers")
            except Exception as e:
                # _fetch_batch_pooled only ever propagates here on a
                # pool.acquire()-level failure - every per-row query error is
                # already caught and blank-filled inside it.
                conn_failures += 1
                log_q.put(("err", f"Future error: {e}"))

    # Every batch failed to even acquire a connection - Oracle went
    # unreachable mid-run. Raise instead of returning a "successful" job
    # whose report has a GST Number column that's blank on every row, which
    # reads as the application failing to do its job rather than as Oracle
    # being down.
    if batches and conn_failures == len(batches):
        raise RuntimeError(_ORACLE_CONNECT_ERROR)

    return merged


# ── Orchestrator ───────────────────────────────────────────────────────────────

def process_report(input_path: str, output_path: str, oracle_cfg: OracleConfig,
                    log_q, progress_cb=None) -> dict:
    """Enrich the uploaded RDC Receivable Aging Report with a GST Invoice
    Number column, querying Oracle for each unique (invoice no, invoice
    date) pair. Returns {"total", "found", "blank"}."""
    if progress_cb:
        progress_cb(0.02, "Reading input file...")
    df, inv_no_col_pd, inv_date_col_pd = read_data_df(input_path)
    log_q.put(("info", f"{len(df)} data rows found"))
    if len(df) == 0:
        raise ValueError("No data rows found after skipping header rows.")

    if progress_cb:
        progress_cb(0.04, "Extracting unique keys...")
    pairs_unique = _extract_pairs_vectorized(df, inv_no_col_pd, inv_date_col_pd)
    log_q.put(("info", f"{len(pairs_unique)} unique (invoice no, date) pairs"))

    if progress_cb:
        progress_cb(0.08, "Connecting to Oracle...")
    init_oracle_client(oracle_cfg.instant_client_dir)
    pool = _make_pool(oracle_cfg, log_q)

    try:
        gst_map = _fetch_all_parallel(pairs_unique, pool, log_q, progress_cb)
    finally:
        try:
            pool.close()
        except Exception:
            pass

    if progress_cb:
        progress_cb(0.90, "Preparing output workbook...")
    ok_cnt = sum(1 for v in gst_map.values() if v)
    log_q.put(("info", f"{ok_cnt}/{len(gst_map)} keys returned a GST number"))

    ext = input_path.rsplit(".", 1)[-1].lower()
    if ext == "xlsb":
        if progress_cb:
            progress_cb(0.92, "Building enriched output (pure Python)...")
        total, found, blank = build_gst_report_pure_python(
            df, inv_no_col_pd, inv_date_col_pd, gst_map, output_path, log_q,
        )
    else:
        if progress_cb:
            progress_cb(0.94, "Inserting GST Invoice Number column...")
        header_row_1based = SKIP_ROWS + 1
        total, found, blank = insert_gst_column(
            wb_path=input_path, output_path=output_path,
            gst_map=gst_map, header_row=header_row_1based, log_q=log_q,
        )

    if progress_cb:
        progress_cb(1.0, "Done")
    return {"total": total, "found": found, "blank": blank}
