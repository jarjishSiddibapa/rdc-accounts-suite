"""Closing Period Report Generator - combine pipeline.

Ports the desktop app's parsing/aggregation logic (see the sibling source
repo `kishore-sir-closing-period-report-generator/script.py`) onto the
shared FastAPI backend. Each uploaded file is an Oracle BI Publisher report
saved as HTML wrapped in a `.xls` extension:

  Row 0  : Location Name in cell [0]
  Row 2  : Header - Organization | Sub Inventory | Item | Description | UOM |
            <DD-MON-YY> Quantity | Current Quantity | Change Quantity |
            <DD-MON-YY> Value   | Current value    | Change Value

The date label (e.g. "31-MAY-26") is auto-detected from the header columns.
Only rows whose Sub Inventory is in ALLOWED_SUBINV are kept.
"""

from __future__ import annotations

import re

from lxml import etree, html
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.jobs import JobUserError
from app.services.xlsx_formula_cache import inject_cached_values

ALLOWED_SUBINV = {"MOD-RM", "NMOD-RM", "STORES"}
SKIP_FIRST_COL = {"Sub Inventory Total", "Report Total", "Sorted by", "Organization"}
DATE_RE = re.compile(r"\d{2}-[A-Z]{3}-\d{2,4}")  # e.g. 31-MAY-26

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM2 = "#,##0.00"


def _detect_date_label_and_cols(header_cells: list[str]) -> tuple[str, int, int]:
    qty_col = val_col = None
    date_label = None
    for i, h in enumerate(header_cells):
        m = DATE_RE.search(h)
        if not m:
            continue
        if "Quantity" in h and qty_col is None:
            qty_col = i
            date_label = m.group()
        elif "Value" in h and val_col is None:
            val_col = i
            if date_label is None:
                date_label = m.group()
    if qty_col is None:
        qty_col = 5
    if val_col is None:
        val_col = 8
    if date_label is None:
        date_label = "UNKNOWN"
    return date_label, qty_col, val_col


def _to_float(s: str) -> float:
    try:
        return float(s.replace(",", "")) if s else 0.0
    except ValueError:
        return 0.0


def _cell_text(cell) -> str:
    """Match BeautifulSoup's ``get_text(strip=True)`` without adding a new
    production dependency: trim each text node and concatenate the pieces."""
    return "".join(part.strip() for part in cell.itertext())


def _parse_file(original_filename: str, saved_path: str, log_q) -> tuple[list[dict], str]:
    """Parse one uploaded file. ``original_filename`` (not the uuid-prefixed
    on-disk name) is used for the "File Name" column and log messages."""
    with open(saved_path, "r", encoding="utf-8", errors="ignore") as fh:
        payload = fh.read()
    try:
        document = html.fromstring(payload)
    except (etree.ParserError, ValueError):
        document = None

    tables = document.xpath("//table") if document is not None else []
    if not tables:
        log_q.put(("warn", f"No table found in {original_filename} - skipped."))
        return [], ""

    rows = tables[0].xpath(".//tr")
    if len(rows) < 3:
        log_q.put(("warn", f"Too few rows in {original_filename} - skipped."))
        return [], ""

    first_row_cells = [_cell_text(td) for td in rows[0].xpath("./td|./th")]
    location_name = next((c for c in first_row_cells if c), "")

    header_cells = [_cell_text(td) for td in rows[2].xpath("./td|./th")]
    date_label, qty_col, val_col = _detect_date_label_and_cols(header_cells)

    records = []
    for row in rows[3:]:
        cells = [_cell_text(td) for td in row.xpath("./td|./th")]
        if not cells:
            continue
        if any(cells[0].startswith(kw) for kw in SKIP_FIRST_COL):
            continue

        subinv = cells[1] if len(cells) > 1 else ""
        if subinv not in ALLOWED_SUBINV:
            continue

        item = cells[2] if len(cells) > 2 else ""
        qty = _to_float(cells[qty_col]) if len(cells) > qty_col else 0.0
        val = _to_float(cells[val_col]) if len(cells) > val_col else 0.0
        rate = (val / qty) if qty != 0 else 0.0

        records.append({
            "File Name": original_filename,
            "Location Name": location_name,
            "Sub Inventory": subinv,
            "Item": item,
            f"{date_label} Quantity": qty,
            f"{date_label} Value": val,
            f"{date_label} Rate": rate,
        })

    return records, date_label


def _style_main_sheet(ws, headers, data_rows, date_label) -> None:
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tc = ws.cell(row=1, column=1, value=f"Closing Period Inventory Report - {date_label}")
    tc.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="172B4D")
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 28

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[2].height = 32

    num_cols = {
        headers.index(f"{date_label} Quantity") + 1,
        headers.index(f"{date_label} Value") + 1,
        headers.index(f"{date_label} Rate") + 1,
    }

    for ri, row_data in enumerate(data_rows, 3):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = DATA_FONT
            c.fill = fill
            c.border = BORDER
            if ci in num_cols:
                c.alignment = RIGHT
                c.number_format = NUM2
            else:
                c.alignment = LEFT
        ws.row_dimensions[ri].height = 16

    widths = {"File Name": 20, "Location Name": 24, "Sub Inventory": 14, "Item": 20}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 18)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(data_rows) + 2}"


def _style_summary_sheet(ws, date_label, agg, main_sheet_name, main_data_start, main_data_end) -> dict[str, float]:
    """Location x Sub Inventory summary. Quantity/Value are live
    =SUMIFS(...) formulas over the Main sheet's own data range, rather than
    a static pre-computed value, so the totals recalculate in Excel like a
    professionally hand-built report - a plain SUMIFS is used instead of a
    native PivotTable here since this shape (two independent value fields,
    no column dimension) still deserves live numbers without the
    OOXML pivot-merge risk that a from-scratch PivotTable would carry for a
    brand-new report with no prior verification history.

    ``agg`` maps (location, sub_inventory) -> (quantity_sum, value_sum),
    already computed in Python. Returns {cell_coordinate: value} for these
    cells (and the Grand Total row) - openpyxl never caches a formula
    result itself, and SUMIFS isn't one of the few forms
    xlsx_formula_cache's generic resolver understands, so this sheet's
    formulas would otherwise show blank until Excel recalculates them
    (which downloaded files can be as slow to do as email attachments,
    since both can open in Protected View)."""
    headers = ["Location Name", "Sub Inventory",
               f"{date_label} Quantity", f"{date_label} Value"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[1].height = 28

    qty_range = f"'{main_sheet_name}'!$E${main_data_start}:$E${main_data_end}"
    val_range = f"'{main_sheet_name}'!$F${main_data_start}:$F${main_data_end}"
    loc_range = f"'{main_sheet_name}'!$B${main_data_start}:$B${main_data_end}"
    sub_range = f"'{main_sheet_name}'!$C${main_data_start}:$C${main_data_end}"

    cached_values: dict[str, float] = {}
    qty_total = val_total = 0.0

    for ri, (loc, sub) in enumerate(sorted(agg.keys()), 2):
        qty_sum, val_sum = agg[(loc, sub)]
        qty_total += qty_sum
        val_total += val_sum
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        loc_esc = loc.replace('"', '""')
        sub_esc = sub.replace('"', '""')

        c1 = ws.cell(row=ri, column=1, value=loc)
        c1.font = DATA_FONT; c1.fill = fill; c1.border = BORDER; c1.alignment = LEFT
        c2 = ws.cell(row=ri, column=2, value=sub)
        c2.font = DATA_FONT; c2.fill = fill; c2.border = BORDER; c2.alignment = LEFT

        c3 = ws.cell(row=ri, column=3,
                     value=f'=SUMIFS({qty_range},{loc_range},"{loc_esc}",{sub_range},"{sub_esc}")')
        c3.font = DATA_FONT; c3.fill = fill; c3.border = BORDER
        c3.alignment = RIGHT; c3.number_format = NUM2
        cached_values[c3.coordinate] = qty_sum

        c4 = ws.cell(row=ri, column=4,
                     value=f'=SUMIFS({val_range},{loc_range},"{loc_esc}",{sub_range},"{sub_esc}")')
        c4.font = DATA_FONT; c4.fill = fill; c4.border = BORDER
        c4.alignment = RIGHT; c4.number_format = NUM2
        cached_values[c4.coordinate] = val_sum

    last_row = 1 + len(agg)
    if last_row >= 2:
        gt = ws.cell(row=last_row + 1, column=1, value="Grand Total")
        gt.font = Font(name="Arial", bold=True, size=10)
        gt.fill = HEADER_FILL
        gt.alignment = LEFT
        gt.border = BORDER
        ws.cell(row=last_row + 1, column=2).fill = HEADER_FILL
        ws.cell(row=last_row + 1, column=2).border = BORDER
        for ci, total in ((3, qty_total), (4, val_total)):
            col_letter = get_column_letter(ci)
            c = ws.cell(row=last_row + 1, column=ci,
                        value=f"=SUBTOTAL(9,{col_letter}2:{col_letter}{last_row})")
            c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            c.fill = HEADER_FILL
            c.alignment = RIGHT
            c.border = BORDER
            c.number_format = NUM2
            cached_values[c.coordinate] = total

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{last_row}"
    return cached_values


def run_combine(files: list[tuple[str, str]], output_path: str, log_q) -> dict:
    """``files`` is (original_filename, saved_disk_path) pairs - see
    _parse_file's docstring for why the original name has to survive to
    here. Writes the combined workbook to ``output_path`` and returns a
    JSON-serializable summary dict."""
    if not files:
        raise JobUserError("No files were uploaded")

    log_q.put(("info", f"Processing {len(files)} file(s)"))

    all_records: list[dict] = []
    date_labels: set[str] = set()
    ok = skip = 0

    for original_filename, saved_path in files:
        records, date_label = _parse_file(original_filename, saved_path, log_q)
        if records:
            all_records.extend(records)
            date_labels.add(date_label)
            log_q.put(("success", f"{original_filename}  ->  {len(records)} rows  "
                                   f"| period: {date_label}  | location: {records[0]['Location Name']}"))
            ok += 1
        else:
            log_q.put(("warn", f"SKIP  {original_filename}"))
            skip += 1

    if not all_records:
        raise JobUserError("No data extracted. Check sub-inventory names and file format.")

    if len(date_labels) > 1:
        log_q.put(("warn", f"Multiple period labels found: {sorted(date_labels)} - using most common"))

    date_label = max(date_labels, key=lambda d: sum(
        1 for r in all_records if f"{d} Quantity" in r))

    norm_records = []
    for r in all_records:
        qty = r.get(f"{date_label} Quantity") or next(
            (v for k, v in r.items() if "Quantity" in k), 0.0)
        val = r.get(f"{date_label} Value") or next(
            (v for k, v in r.items() if k.endswith("Value")), 0.0)
        rate = (val / qty) if qty else 0.0
        norm_records.append({
            "File Name": r["File Name"],
            "Location Name": r["Location Name"],
            "Sub Inventory": r["Sub Inventory"],
            "Item": r["Item"],
            f"{date_label} Quantity": qty,
            f"{date_label} Value": val,
            f"{date_label} Rate": rate,
        })

    headers = [
        "File Name", "Location Name", "Sub Inventory", "Item",
        f"{date_label} Quantity", f"{date_label} Value", f"{date_label} Rate",
    ]

    wb = Workbook()
    ws = wb.active
    main_sheet_name = f"Report {date_label}"
    ws.title = main_sheet_name

    for ci, header in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=header)
    data_rows = [[r[h] for h in headers] for r in norm_records]
    for ri, row_data in enumerate(data_rows, 2):
        for ci, v in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci, value=v)

    log_q.put(("info", "Writing output workbook..."))
    _style_main_sheet(ws, headers, data_rows, date_label)
    # _style_main_sheet inserted a title row above the data, so the actual
    # data now sits on rows 3..(3 + len(data_rows) - 1).
    main_data_start = 3
    main_data_end = 3 + len(data_rows) - 1

    agg: dict[tuple, list[float]] = {}
    for r in norm_records:
        key = (r["Location Name"], r["Sub Inventory"])
        entry = agg.setdefault(key, [0.0, 0.0])
        entry[0] += r[f"{date_label} Quantity"]
        entry[1] += r[f"{date_label} Value"]
    agg = {key: tuple(values) for key, values in agg.items()}

    ws_sum = wb.create_sheet("Summary")
    summary_cached_values = _style_summary_sheet(
        ws_sum, date_label, agg, main_sheet_name, main_data_start, main_data_end)

    log_q.put(("success", f"Period detected : {date_label}"))
    log_q.put(("success", f"Files processed : {ok} ok / {skip} skipped"))
    log_q.put(("success", f"Total data rows : {len(norm_records)}"))

    wb.save(output_path)
    inject_cached_values(str(output_path), {"Summary": summary_cached_values})
    log_q.put(("success", f"Saved  ->  {output_path}"))

    return {
        "files": ok,
        "skipped": skip,
        "total_rows": len(norm_records),
        "date_label": date_label,
        "output_path": str(output_path),
    }
