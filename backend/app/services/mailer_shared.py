"""Email sending, credential management, and pivot-to-HTML helpers.

Ported from the desktop app's mailer.py. Behavior (STARTTLS send, HTML body
composition, sheet_to_html rendering) is preserved verbatim; the only real
change is where credentials come from: instead of reading/writing
%APPDATA%\\UnaccountedTransactions\\email_settings.json, ``get_email_settings(user_id)``
reads that user's OWN row in the suite's EmailSettings table (one row per
user, keyed by user_id — each logged-in user configures their own Gmail
sender identity, never a shared one) via SessionLocal, decrypting the app
password with app.security.decrypt.

Default recipients are stored per application and managed by admins. The
legacy singleton is read once as a non-destructive migration source.
"""

import json
import logging
import os
import re
import time

import openpyxl
from openpyxl.utils.cell import range_boundaries

from app.database import SessionLocal
from app.models import ApplicationEmailRecipient, EmailSettings, ReportRecipientDefaults
from app.regional import format_indian_number
from app.security import decrypt
from app.validation import normalize_email

logger = logging.getLogger(__name__)


# ── Defaults (seeding only — not read at send-time once configured) ──────────
DEFAULT_TO = ["accountsincharges@rdc.in", "accountsgroup@rdc.in"]
DEFAULT_CC = ["manish.modani@rdc.in", "umesh.gawade@rdc.in"]


def _normalized_recipients(to_list: list[str], cc_list: list[str]) -> tuple[list[str], list[str]]:
    normalized_to = [normalize_email(value) for value in to_list]
    normalized_cc = [normalize_email(value) for value in cc_list]
    all_addresses = normalized_to + normalized_cc
    if len(set(all_addresses)) != len(all_addresses):
        raise ValueError("Each default recipient email must be unique within an application")
    return normalized_to, normalized_cc


def _legacy_defaults(db) -> tuple[list[str], list[str]]:
    row = db.query(ReportRecipientDefaults).filter(
        ReportRecipientDefaults.id == 1,
        ReportRecipientDefaults.is_deleted == False,  # noqa: E712
    ).first()
    if row is None:
        return list(DEFAULT_TO), list(DEFAULT_CC)
    try:
        to_list = json.loads(row.default_to) if row.default_to else []
    except (ValueError, TypeError):
        to_list = []
    try:
        cc_list = json.loads(row.default_cc) if row.default_cc else []
    except (ValueError, TypeError):
        cc_list = []
    return to_list or list(DEFAULT_TO), cc_list or list(DEFAULT_CC)


def set_report_recipient_defaults(
    db,
    app_key: str,
    to_list: list[str],
    cc_list: list[str],
) -> dict:
    normalized_to, normalized_cc = _normalized_recipients(to_list, cc_list)
    desired = {email: "to" for email in normalized_to}
    desired.update({email: "cc" for email in normalized_cc})
    existing = {
        row.email.lower(): row
        for row in db.query(ApplicationEmailRecipient)
        .filter(ApplicationEmailRecipient.app_key == app_key)
        .all()
    }

    for email, recipient_type in desired.items():
        row = existing.get(email)
        if row is None:
            row = ApplicationEmailRecipient(app_key=app_key, email=email)
            db.add(row)
        row.recipient_type = recipient_type
        row.is_deleted = False

    for email, row in existing.items():
        if email not in desired:
            row.is_deleted = True

    db.flush()
    return {"to": normalized_to, "cc": normalized_cc}


def get_report_recipient_defaults(db=None, app_key: str = "unaccounted") -> dict:
    """Return active admin-managed default To/Cc recipients for one app."""
    owns_session = db is None
    if owns_session:
        db = SessionLocal()
    try:
        rows = (
            db.query(ApplicationEmailRecipient)
            .filter(
                ApplicationEmailRecipient.app_key == app_key,
                ApplicationEmailRecipient.is_deleted == False,  # noqa: E712
            )
            .order_by(ApplicationEmailRecipient.id)
            .all()
        )
        if not rows and app_key == "unaccounted":
            to_list, cc_list = _legacy_defaults(db)
            set_report_recipient_defaults(db, app_key, to_list, cc_list)
            db.commit()
            rows = (
                db.query(ApplicationEmailRecipient)
                .filter(
                    ApplicationEmailRecipient.app_key == app_key,
                    ApplicationEmailRecipient.is_deleted == False,  # noqa: E712
                )
                .order_by(ApplicationEmailRecipient.id)
                .all()
            )
        return {
            "to": [row.email for row in rows if row.recipient_type == "to"],
            "cc": [row.email for row in rows if row.recipient_type == "cc"],
        }
    finally:
        if owns_session:
            db.close()


# ── Credential helpers ────────────────────────────────────────────────────────
def get_email_settings(user_id: int) -> dict:
    """Return the logged-in user's OWN EmailSettings row (one row per user,
    keyed by user_id — each user configures their own Gmail sender identity,
    nothing here is shared across accounts), shaped like the old
    email_settings.json file did so the rest of the ported mailer logic needs
    minimal changes:

        {
            "email": str, "app_password": str, "signature": str,
            "to": list[str], "cc": list[str], "configured": bool,
        }

    "to"/"cc" come from the shared, admin-controlled ReportRecipientDefaults
    (see get_report_recipient_defaults) — recipients are an application-wide
    business rule, not a per-user preference.

    "configured" is False (and "email"/"app_password" are "") when this user
    hasn't set up their sender identity yet — callers must check this and
    surface a clear error ("You haven't set up your email sender yet — go to
    Settings") rather than attempting to send with blank credentials.

    Never logs/prints the decrypted app_password anywhere. Opens and closes
    its own SessionLocal() session.
    """
    db = SessionLocal()
    try:
        recipients = get_report_recipient_defaults(db, "unaccounted")
        row = db.query(EmailSettings).filter(
            EmailSettings.user_id == user_id,
            EmailSettings.is_deleted == False,  # noqa: E712
        ).first()
        if row is None or not row.sender_email or not row.app_password_encrypted:
            return {
                "email": "",
                "app_password": "",
                "signature": "",
                "to": recipients["to"],
                "cc": recipients["cc"],
                "configured": False,
            }

        return {
            "email": row.sender_email or "",
            "app_password": decrypt(row.app_password_encrypted or ""),
            "signature": row.signature or "",
            "to": recipients["to"],
            "cc": recipients["cc"],
            "configured": True,
        }
    finally:
        db.close()


def test_smtp_connection(email: str, app_password: str, **_) -> tuple:
    """Return (ok: bool, message: str)."""
    import smtplib
    try:
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=60) as s:
            s.ehlo(); s.starttls()
            s.login(email.strip(), app_password.strip())
        return True, "Connection and authentication successful."
    except smtplib.SMTPAuthenticationError:
        return (False,
                "Authentication failed.\n"
                "Check your Gmail address and app password.\n"
                "For Gmail: Google Account → Security → 2-Step Verification → App Passwords.")
    except Exception:  # noqa: BLE001 - normalize network/provider errors at this boundary
        logger.exception("Report-sender SMTP connection test failed")
        return False, "Connection failed. Check the network and email settings, then try again."


# ── Email subject / intro helpers ────────────────────────────────────────────

def _join_with_and(items: list) -> str:
    """Natural-language list join: 'A' / 'A and B' / 'A, B and C'."""
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} and {items[1]}"
    return ", ".join(items[:-1]) + f" and {items[-1]}"


def build_default_subject(
    month_subject: str,
    include_ua: bool = True,
    include_mrn: bool = True,
    include_po: bool = True,
) -> str:
    parts = []
    if include_ua:
        parts.append("Unaccounted Transactions")
    if include_mrn:
        parts.append("Pending MRN")
    if include_po:
        parts.append("Uninvoiced Expense")
    label = _join_with_and(parts) if parts else "Report"
    return f"{label} till {month_subject}"


def build_default_intro_text(
    include_ua: bool = True,
    include_mrn: bool = True,
    include_po: bool = True,
    month_ua: str = "",
    month_mrn: str = "",
    month_po: str = "",
) -> str:
    raw_items = []
    if include_ua:
        raw_items.append(f"Unaccounted transactions till {month_ua or 'Month'}")
    if include_mrn:
        raw_items.append(f"Pending MRN till {month_mrn or 'Month'}")
    if include_po:
        raw_items.append(f"Uninvoiced Expenses till {month_po or 'Month'}")

    # Numbering only makes sense for multiple reports
    if len(raw_items) == 1:
        items = raw_items
    else:
        items = [f"{i + 1}. {item}" for i, item in enumerate(raw_items)]

    text = "Dear Team,\n\nPlease find below:\n\n" + "\n".join(items)
    text += "\n\nKindly clear the same on priority basis and confirm."
    text += (
        "\n\n**Also note that Rent and Land Lease have been excluded from the "
        "Uninvoiced Expense Report. Accordingly, the remaining expenses are "
        "required to be booked.**"
    )
    return text


def _plain_to_html(text: str) -> str:
    import html as _html_mod
    import re as _re
    escaped = _html_mod.escape(text.strip())
    escaped = escaped.replace('\r\n', '\n').replace('\r', '\n')
    # **text** → <strong>text</strong>
    escaped = _re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', escaped, flags=_re.DOTALL)
    parts   = []
    for para in escaped.split('\n\n'):
        para = para.strip()
        if not para:
            continue
        parts.append(
            '<p style="margin:0 0 12px 0;">'
            + '<br>'.join(para.split('\n'))
            + '</p>'
        )
    return '\n'.join(parts)


# ── Excel-sheet → HTML table (reads actual cell formatting from file) ─────────
def sheet_to_html(path: str, sheet_name: str) -> str:
    """Open the saved Excel file and convert the pivot sheet to an HTML table
    that exactly mirrors the formatting in the report: fill colors, bold,
    font size, font name, text alignment — all read directly from the cells.

    The suite deliberately writes live ``SUBTOTAL`` formulas into its report
    footers.  openpyxl does not calculate or cache those formulas, so loading
    the file with ``data_only=True`` makes otherwise valid totals appear blank.
    Read the formula workbook instead and evaluate the small, controlled set of
    aggregate formulas used by these email tables for display only.  The saved
    attachment is never modified and retains every Excel formula.
    """
    from openpyxl.cell.cell import MergedCell as _MergedCell

    try:
        # Do NOT use read_only=True — that skips loading cell styles.  Formula
        # text is required so totals can be resolved without Excel/COM.
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        return f"<p style='color:red'>(Cannot open {os.path.basename(path)}: {e})</p>"

    if sheet_name not in wb.sheetnames:
        wb.close()
        return f"<p style='color:red'>(Sheet &quot;{sheet_name}&quot; not found)</p>"

    ws = wb[sheet_name]

    _cell_or_range = r"\$?[A-Z]{1,3}\$?\d+(?:\s*:\s*\$?[A-Z]{1,3}\$?\d+)?"
    _subtotal_re = re.compile(
        r"^\s*=\s*SUBTOTAL\s*\(\s*(9|109)\s*,\s*"
        r"(\$?[A-Z]{1,3}\$?\d+\s*:\s*\$?[A-Z]{1,3}\$?\d+)\s*\)\s*$",
        re.IGNORECASE,
    )
    # One or more comma-separated ranges/cells, e.g. SUM(D5:E5,G5:H5) for a
    # row total that skips non-adjacent columns - mirrors the same extension
    # in app.services.xlsx_formula_cache._SUM_RE.
    _sum_re = re.compile(
        r"^\s*=\s*SUM\s*\(\s*"
        r"(" + _cell_or_range + r"(?:\s*,\s*" + _cell_or_range + r")*)"
        r"\s*\)\s*$",
        re.IGNORECASE,
    )

    def _iter_sum_parts(parts_text):
        for part in re.split(r"\s*,\s*", parts_text.strip()):
            part = part.replace(" ", "")
            ref = part if ":" in part else f"{part}:{part}"
            min_col, min_row, max_col, max_row = range_boundaries(ref)
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                yield from row

    _resolved_values = {}
    _resolving = set()

    def _display_value(cell):
        """Return a server-computed display value for supported formulas."""
        coordinate = cell.coordinate
        if coordinate in _resolved_values:
            return _resolved_values[coordinate]

        raw = cell.value
        if not (isinstance(raw, str) and raw.startswith("=")):
            _resolved_values[coordinate] = raw
            return raw
        if coordinate in _resolving:
            return None

        subtotal_match = _subtotal_re.match(raw)
        sum_match = _sum_re.match(raw)
        match = subtotal_match or sum_match
        if not match:
            # Keep unsupported formulas out of the HTML instead of leaking the
            # formula text.  Current email pivot sheets use SUBTOTAL only.
            _resolved_values[coordinate] = None
            return None

        function_num = int(subtotal_match.group(1)) if subtotal_match else None

        _resolving.add(coordinate)
        try:
            total = 0
            if subtotal_match:
                min_col, min_row, max_col, max_row = range_boundaries(
                    re.sub(r"\s+", "", subtotal_match.group(2))
                )
                source_cells = (
                    source_cell
                    for row in ws.iter_rows(
                        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col,
                    )
                    for source_cell in row
                )
            else:
                source_cells = _iter_sum_parts(sum_match.group(1))
            for source_cell in source_cells:
                source_raw = source_cell.value
                # Excel SUBTOTAL ignores nested SUBTOTAL formulas.  This is
                # essential for Grand Total ranges that also contain each
                # location's subtotal row.
                if (
                    subtotal_match
                    and isinstance(source_raw, str)
                    and _subtotal_re.match(source_raw)
                ):
                    continue
                if function_num == 109 and ws.row_dimensions[source_cell.row].hidden:
                    continue
                value = _display_value(source_cell)
                if (
                    isinstance(value, (int, float))
                    and not isinstance(value, bool)
                ):
                    total += value
        finally:
            _resolving.remove(coordinate)

        _resolved_values[coordinate] = total
        return total

    # ── helpers ───────────────────────────────────────────────────────────────
    def _hex6(color_obj):
        """Return 6-char RRGGBB hex from an openpyxl Color, or None."""
        if not color_obj:
            return None
        try:
            if color_obj.type == "rgb":
                v = str(color_obj.rgb)          # may be 8-char AARRGGBB or 6-char RRGGBB
                # openpyxl stores a plain hex string passed to PatternFill(fgColor=...)
                # as "00RRGGBB" - that leading "00" is just its placeholder for "no
                # alpha specified", not a real transparent alpha channel. A solid
                # PatternFill is never actually invisible, so always take the color
                # itself rather than treating a "00" prefix as "no fill".
                if len(v) == 8:
                    return v[2:]
                if len(v) == 6:
                    return v
        except Exception:
            pass
        return None

    def _cell_css(cell):
        # Background ────────────────────────────────────────────────────────
        bg = "FFFFFF"
        try:
            f = cell.fill
            if f and getattr(f, "patternType", None) == "solid":
                c = _hex6(f.fgColor)
                if c and c.upper() not in ("FFFFFF", "000000"):
                    bg = c
        except Exception:
            pass

        # Font ──────────────────────────────────────────────────────────────
        bold       = False
        font_size  = 11
        font_name  = "Calibri"
        font_color = "000000"
        try:
            fn = cell.font
            if fn:
                bold      = bool(fn.bold)
                font_size = int(fn.size) if fn.size else 11
                font_name = fn.name  if fn.name  else "Calibri"
                c = _hex6(fn.color)
                if c:
                    font_color = c
        except Exception:
            pass

        # Alignment ─────────────────────────────────────────────────────────
        align = "center"
        try:
            a = cell.alignment
            if a and a.horizontal in ("left", "right", "center"):
                align = a.horizontal
        except Exception:
            pass

        parts = [
            f"background-color:#{bg}",
            f"color:#{font_color}",
            f"font-family:{font_name},Arial,sans-serif",
            f"font-size:{font_size}pt",
            f"text-align:{align}",
            "border:1px solid #000000",
            "padding:5px 10px",
            "white-space:nowrap",
        ]
        if bold:
            parts.append("font-weight:bold")
        return ";".join(parts)

    def _fmt_val(v):
        if v is None:
            return ""
        if isinstance(v, (int, float)):
            if v == 0:
                return "–"
            n = int(v) if v == int(v) else v
            return format_indian_number(n, 0 if isinstance(n, int) else 2)
        return str(v)

    # ── build table ───────────────────────────────────────────────────────────
    lines = [
        '<table style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;'
        'font-size:11pt;margin:4px 0 18px 0;">'
    ]

    for row in ws.iter_rows():
        display_row = [
            None if isinstance(cell, _MergedCell) else _display_value(cell)
            for cell in row
        ]
        # Skip rows where every rendered cell is blank (value=None).
        if all(value is None for value in display_row):
            continue

        lines.append("<tr>")
        for cell, display_value in zip(row, display_row):
            if isinstance(cell, _MergedCell):
                # continuation cell of a merge — render empty, no extra border
                lines.append('<td style="border:none;padding:0;"></td>')
                continue
            css = _cell_css(cell)
            val = _fmt_val(display_value)
            lines.append(f'  <td style="{css}">{val}</td>')
        lines.append("</tr>")

    lines.append("</table>")
    wb.close()
    return "\n".join(lines)


# ── Email composition ─────────────────────────────────────────────────────────
def build_email_content(
    unaccounted_path: str = "",
    mrn_path: str = "",
    po_path: str = "",
    month_subject: str = "",
    month_unaccounted: str = "",
    month_mrn: str = "",
    month_po: str = "",
    signature: str = "",
    include_ua: bool = True,
    include_mrn: bool = True,
    include_po: bool = True,
    custom_subject: str = "",
    custom_intro: str = "",
) -> tuple:
    """Return (subject: str, html_body: str)."""
    import html as _html_mod

    subject = custom_subject or build_default_subject(
        month_subject, include_ua, include_mrn, include_po)

    # Intro section
    if custom_intro and custom_intro.strip():
        intro_html = _plain_to_html(custom_intro)
    else:
        intro_html = _plain_to_html(build_default_intro_text(
            include_ua, include_mrn, include_po,
            month_unaccounted, month_mrn, month_po))

    # Tables for included reports only
    _report_count = sum([
        bool(include_ua and unaccounted_path),
        bool(include_mrn and mrn_path),
        bool(include_po and po_path),
    ])
    _single = (_report_count == 1)

    table_sections = []
    n = 1
    if include_ua and unaccounted_path:
        header = "" if _single else f'<p><b>{n}. Unaccounted Transactions:</b></p>\n'
        table_sections.append(header + sheet_to_html(unaccounted_path, "Main"))
        n += 1
    if include_mrn and mrn_path:
        header = "" if _single else f'<p><b>{n}. Pending MRN:</b></p>\n'
        table_sections.append(header + sheet_to_html(mrn_path, "Locationwise Pivot"))
        n += 1
    if include_po and po_path:
        header = "" if _single else f'<p><b>{n}. Uninvoiced Expenses:</b></p>\n'
        table_sections.append(header + sheet_to_html(po_path, "Main"))

    # Signature
    sig_html = ""
    if signature and signature.strip():
        escaped = _html_mod.escape(signature.strip())
        lines   = escaped.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        sig_html = (
            '<p style="margin-top:24px;font-family:Calibri,Arial,sans-serif;'
            'font-size:12pt;color:#1a1a1a;line-height:1.65;">'
            + "<br>".join(lines)
            + "</p>"
        )

    html_body = (
        '<html>\n'
        '<body style="font-family:Calibri,Arial,sans-serif;font-size:12pt;'
        'color:#1a1a1a;line-height:1.65;">\n\n'
        + intro_html + "\n\n"
        + "\n\n".join(table_sections) + "\n\n"
        + sig_html + "\n"
        + "</body>\n</html>"
    )
    return subject, html_body


# ── Send ──────────────────────────────────────────────────────────────────────
class MailAttachmentError(Exception):
    """Raised when a requested attachment can't be read as a complete, non-
    empty file at send time - e.g. the scratch-cleanup sweep reclaimed a
    generated report before a slow "review the preview, then confirm send"
    workflow got around to actually sending it, or (seen in production even
    on an immediate send) a brief window right after the file is written
    where a fresh cross-process read sees 0 bytes - most likely antivirus
    on-access scanning holding the file, a known Windows phenomenon. Callers
    must surface this to the user and ask them to regenerate rather than
    silently sending a mail that claims an attachment it doesn't actually
    have."""


def read_attachment_bytes(path: str, *, max_attempts: int = 6, retry_delay_seconds: float = 0.5) -> bytes:
    """Read a report file's full bytes for an email attachment, tolerating
    the brief post-write window described in MailAttachmentError above by
    retrying a few times before giving up. A genuinely missing, deleted, or
    corrupt file still fails - this only rides out a transient race, it
    never invents content."""
    last_error: Exception | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            expected_size = os.path.getsize(path)
        except OSError as exc:
            last_error = exc
            expected_size = None
        if expected_size == 0:
            last_error = OSError(f"{path} is 0 bytes")
        elif expected_size is not None:
            with open(path, "rb") as fh:
                data = fh.read()
            if data and len(data) == expected_size:
                return data
            last_error = OSError(f"read {len(data)} of {expected_size} expected bytes from {path}")
        if attempt < max_attempts:
            time.sleep(retry_delay_seconds)
    raise MailAttachmentError(
        f"Could not read a complete copy of {os.path.basename(path)} after {max_attempts} "
        f"attempts ({last_error}). Regenerate the report and send again."
    )


def send_mail(
    from_email: str,
    app_password: str,
    to_addresses: list,
    cc_addresses: list,
    subject: str,
    html_body: str,
    attachments: list,
    **_,   # absorb any extra kwargs (smtp_server / smtp_port) — always use Gmail
) -> None:
    """Send HTML email with Excel attachments via Gmail SMTP STARTTLS.

    Every path in ``attachments`` is read via read_attachment_bytes, which
    raises MailAttachmentError (before contacting SMTP at all, since every
    attachment is read while still only building the message) if any file
    can't be read as complete and non-empty even after retrying - never
    silently sends an email with a named-but-empty attachment."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText
    from email.mime.base      import MIMEBase
    from email                import encoders

    msg            = MIMEMultipart("mixed")
    msg["From"]    = from_email
    msg["To"]      = ", ".join(str(a).strip() for a in to_addresses if a)
    if cc_addresses:
        msg["Cc"]  = ", ".join(str(a).strip() for a in cc_addresses if a)
    msg["Subject"] = subject

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alt)

    for path in attachments:
        if not path:
            continue
        data = read_attachment_bytes(path)
        part = MIMEBase("application", "octet-stream")
        part.set_payload(data)
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{os.path.basename(path)}"',
        )
        msg.attach(part)

    all_rcpt = to_addresses + cc_addresses

    with smtplib.SMTP("smtp.gmail.com", 587, timeout=120) as smtp:
        smtp.ehlo(); smtp.starttls()
        smtp.login(from_email.strip(), app_password.strip())
        smtp.sendmail(from_email.strip(), all_rcpt, msg.as_string())
