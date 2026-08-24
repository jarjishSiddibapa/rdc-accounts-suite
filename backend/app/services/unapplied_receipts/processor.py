"""RDC Unapplied Receipts Report — processing pipeline.

Ports the desktop app's (``unapplied_processor.py``) core, framework-agnostic
pipeline:

  1. ``process_report``               – read the uploaded Unapplied Receipts
                                         Register export, drop metadata rows
                                         and unwanted columns, split out
                                         "***** Unidentified" rows, fetch
                                         Location (+ Comments for the
                                         Unidentified rows) live from Oracle
                                         ERP, and compute Due Days / Ageing
                                         Bucket.
  2. ``classify_advance_customers``   – read the Ageing export and split the
                                         main report into "main" vs "advance
                                         of customers" rows using the Z→A
                                         Type-sort / first-occurrence rule.
  3. ``write_formatted_excel``        – write the 4-sheet formatted workbook
                                         (Summary pivot, Unapplied Receipts,
                                         Advance of Customers, Unidentified
                                         Customers).
  4. ``_validate_before_save``        – same missing-mapping check the
                                         desktop app ran just before saving
                                         (mirrors ``ValidationErrorDialog``).

Deliberate deviations from the original ``unapplied_processor.py``:

  - No tkinter/customtkinter, no module-level mutable state, no AppData JSON
    files. Every function takes its inputs (file paths, Oracle connection
    settings, mapping dicts, an "as on" date, a log sink) as parameters.
    ``LOCATION_MAP``/``ACCOUNT_INCHARGE_MAP`` are no longer hardcoded module
    constants — callers load them from
    ``app.services.unapplied_receipts.mapping_store.load_all()`` (backed by
    MySQL) and pass them in as ``supplier_site_map`` / ``incharge_map``.
  - ``log_q`` may be any object exposing ``.put((level, message))`` — the
    original's ``queue.Queue`` works unchanged, and so does this suite's
    ``_LogQueue`` adapter (see e.g. ``app.routers.unaccounted_txn``).
  - ``_build_excel_pivot`` (the win32com/pywin32 real-Excel-PivotTable
    generator) and the unused ``_BUCKET_FILL`` constant are intentionally
    NOT ported: grepping the original source shows ``_build_excel_pivot`` is
    never called from anywhere (the GUI worker only ever calls
    ``write_formatted_excel``, which builds its own Summary sheet via the
    openpyxl-based ``_write_pivot_sheet``) and ``_BUCKET_FILL`` is defined
    but never referenced. Both were already dead code in the desktop app,
    and ``_build_excel_pivot`` additionally requires a real local Excel
    installation via COM automation, which is not available on a server —
    porting it would just be unreachable, unrunnable code.
  - Oracle connectivity (host/port/service/user/password/instant client
    dir) is expressed as an explicit ``OracleConfig`` value object instead
    of hardcoded ``DB_USER``/``DB_PASS``/``DB_HOST``/``DB_PORT``/``DB_SERVICE``
    globals and a ``_instant_client()`` path resolver tied to
    ``sys.frozen``/``sys._MEIPASS``. The router builds one from
    ``app.config`` (which reads the ``ORACLE_*`` env vars).
"""

from __future__ import annotations

import datetime as _dt
import os
from dataclasses import dataclass
from html.parser import HTMLParser as _HTMLParser

import oracledb
import pandas as pd
import pyxlsb  # noqa: F401 - registers the xlsb engine with pandas
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Oracle connection config ────────────────────────────────────────────────

@dataclass(frozen=True)
class OracleConfig:
    """Explicit stand-in for the original's hardcoded DB_USER/DB_PASS/
    DB_HOST/DB_PORT/DB_SERVICE globals + _instant_client() path resolver."""

    host: str
    port: str
    service_name: str
    user: str
    password: str
    instant_client_dir: str

    @property
    def dsn(self) -> str:
        return f"{self.host}:{self.port}/{self.service_name}"


_ORACLE_CONNECT_ERROR = (
    "Could not connect to the Oracle ERP database - this isn't an application "
    "bug. Check that the ERP server is reachable from this network and that "
    "ORACLE_HOST/ORACLE_SERVICE_NAME/ORACLE_USER/ORACLE_PASSWORD in backend/.env "
    "are correct, then try again."
)


def _connect_oracle(oracle_cfg: "OracleConfig"):
    """oracledb.connect(), but a connection-level failure (host unreachable,
    bad credentials, listener down) raises a message that says plainly
    "Oracle is unreachable" instead of a cryptic driver string - otherwise
    this shows up in the job's error field looking like the application
    itself is broken. Catches bare Exception, not just oracledb.Error: a
    DNS/host-resolution failure surfaces as a plain socket.gaierror in
    thin mode, not an oracledb exception - verified by triggering a real
    failed connection against a nonexistent host."""
    try:
        return oracledb.connect(
            user=oracle_cfg.user, password=oracle_cfg.password, dsn=oracle_cfg.dsn
        )
    except Exception as exc:
        raise RuntimeError(_ORACLE_CONNECT_ERROR) from exc


def _init_oracle_client(instant_client_dir: str) -> None:
    """Same best-effort init as the original: init_oracle_client() may only
    be called once per process and raises on subsequent calls (or if thick
    mode isn't needed) — both are safe to ignore, exactly like the original
    ``try/except Exception: pass``."""
    try:
        oracledb.init_oracle_client(lib_dir=instant_client_dir)
    except Exception:
        pass


# ── ERP queries (verbatim from the original) ────────────────────────────────

_ERP_SQL = """
SELECT
    acr.receipt_number,
    acr.global_attribute10 AS location
FROM
    apps.ar_cash_receipts_all acr
WHERE
    acr.receipt_number IN ({placeholders})
"""

_COMMENTS_SQL = """
SELECT
    acr.receipt_number,
    acr.comments AS description
FROM
    apps.ar_cash_receipts_all acr
WHERE
    acr.receipt_number IN ({placeholders})
"""

_BATCH_SIZE = 999  # Oracle IN clause hard limit is 1000


# ── Core processing ──────────────────────────────────────────────────────────

class _OracleHTMLParser(_HTMLParser):
    """
    Event-driven streaming HTML parser for Oracle ERP HTML-disguised XLS exports.

    WHY: Oracle HTML reports can exceed 500 MB. DOM-based parsers (lxml,
    html5lib, pd.read_html) try to build the full tree in RAM first — they
    crash or OOM on large files. This parser processes 1 MB at a time and
    only keeps the current row in memory.

    Uses Python stdlib only — no lxml, html5lib, or bs4 required.
    """
    def __init__(self):
        super().__init__(convert_charrefs=True)   # auto-handles &nbsp; &#160; etc.
        self.rows    = []
        self._row    = None
        self._cell   = None
        self._tdepth = 0

    def handle_starttag(self, tag, attrs):
        t = tag.lower()
        if t == "table":
            self._tdepth += 1
        elif t == "tr" and self._tdepth:
            self._row = []
        elif t in ("td", "th") and self._row is not None:
            self._cell = []
        elif t == "br" and self._cell is not None:
            self._cell.append(" ")   # <br> → space (preserves word boundary)

    def handle_endtag(self, tag):
        t = tag.lower()
        if t == "table":
            self._tdepth = max(0, self._tdepth - 1)
        elif t in ("td", "th") and self._cell is not None:
            # Collapse any whitespace (including \n from raw HTML) to single space
            val = " ".join("".join(self._cell).split()).strip()
            self._row.append(val)
            self._cell = None
        elif t == "tr" and self._row is not None:
            if any(v.strip() for v in self._row):   # skip fully-blank rows
                self.rows.append(self._row)
            self._row = None

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)


def _read_html_streaming(path: str, log_q=None) -> pd.DataFrame:
    """
    Parse an Oracle HTML-disguised XLS file with Python's stdlib HTMLParser.
    Reads in 1 MB chunks — works on files of any size without blowing RAM.
    """
    CHUNK = 1_048_576   # 1 MB
    for enc in ("utf-8", "cp1252", "latin-1"):
        parser = _OracleHTMLParser()
        try:
            with open(path, "r", encoding=enc, errors="strict") as fh:
                while True:
                    chunk = fh.read(CHUNK)
                    if not chunk:
                        break
                    parser.feed(chunk)
            if parser.rows:
                break               # decoded successfully and found rows
        except (UnicodeDecodeError, UnicodeError):
            continue                # try next encoding

    if not parser.rows:
        raise ValueError(
            f"No table rows extracted from HTML file: {os.path.basename(path)}")

    if log_q:
        log_q.put(("info", f"Streaming parse complete: {len(parser.rows):,} rows extracted"))

    max_cols = max(len(r) for r in parser.rows)
    padded   = [r + [""] * (max_cols - len(r)) for r in parser.rows]
    return pd.DataFrame(padded, dtype=object)


def _is_html(raw_sig: bytes) -> bool:
    """Return True if file bytes start with an HTML signature."""
    return raw_sig.lstrip()[:1] in (b"<", b"\xef")


def _read_file(path: str) -> pd.DataFrame:
    """Read unapplied receipts file — auto-detects xls/xlsx/xlsb/HTML."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return pd.read_excel(path, header=None, engine="openpyxl")
    if ext == ".xlsb":
        return pd.read_excel(path, header=None, engine="pyxlsb")
    # .xls — check signature to distinguish HTML vs BIFF
    with open(path, "rb") as f:
        sig = f.read(8)
    if _is_html(sig):
        return _read_html_streaming(path)
    return pd.read_excel(path, header=None, engine="xlrd")


def _find_header_row(df: pd.DataFrame, marker: str,
                     default: int = 13, search_range: int = 25) -> int:
    """
    Scan the first `search_range` rows for the one that contains `marker`
    as a cell value.  Returns `default` if not found.
    """
    for i in range(min(search_range, len(df))):
        if any(marker in str(v) for v in df.iloc[i].tolist()):
            return i
    return default


def _read_ageing(path: str) -> pd.DataFrame:
    """
    Read the ageing file — auto-detects format, auto-locates the header row
    (normally row 13 for Oracle ERP exports, but verified dynamically).
    """
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xlsb":
        return pd.read_excel(path, engine="pyxlsb", header=13)
    if ext == ".xlsx":
        return pd.read_excel(path, engine="openpyxl", header=13)

    # .xls (and any other extension) — read first 8 bytes to detect format
    with open(path, "rb") as f:
        sig = f.read(8)

    if _is_html(sig):
        # Stream-parse — never loads the full DOM (handles 500 MB+ files)
        df_raw = _read_html_streaming(path)
        # Locate the actual header row (normally row 13 in Oracle ERP exports)
        hdr = _find_header_row(df_raw, "Customer Account", default=13)
        if len(df_raw) <= hdr:
            raise ValueError(
                f"Ageing file has only {len(df_raw)} rows; "
                f"could not find header row (tried row {hdr}).")
        # Promote that row to column names and strip whitespace
        df_raw.columns = [str(c).strip() for c in df_raw.iloc[hdr].tolist()]
        df_raw = df_raw.iloc[hdr + 1:].reset_index(drop=True)
        return df_raw

    # True BIFF .xls
    return pd.read_excel(path, engine="xlrd", header=13)


def _fetch_locations(receipts: list, log_q, oracle_cfg: OracleConfig,
                      location_map: dict) -> dict:
    _init_oracle_client(oracle_cfg.instant_client_dir)

    log_q.put(("info", f"Sample receipts sent to ERP: {receipts[:3]}"))

    conn    = _connect_oracle(oracle_cfg)
    loc_map = {}   # receipt_number → mapped location string
    try:
        cur       = conn.cursor()
        col_names = None
        batches   = [receipts[i:i + _BATCH_SIZE]
                     for i in range(0, len(receipts), _BATCH_SIZE)]
        log_q.put(("info", f"Querying ERP in {len(batches)} batch(es) of ≤{_BATCH_SIZE}"))

        for batch in batches:
            escaped = [r.replace("'", "''") for r in batch]
            sql     = _ERP_SQL.format(placeholders=",".join(f"'{r}'" for r in escaped))
            cur.execute(sql)
            if col_names is None:
                col_names = [d[0] for d in cur.description]
                log_q.put(("info", f"ERP columns: {col_names}"))
            for row in cur.fetchall():
                rd      = dict(zip(col_names, row))
                key     = str(rd.get("RECEIPT_NUMBER") or "").strip()
                raw_loc = str(rd.get("LOCATION") or "").strip().upper()
                mapped  = location_map.get(raw_loc, raw_loc)   # fallback = raw value
                if key and (key not in loc_map or mapped):
                    loc_map[key] = mapped

        cur.close()
        if loc_map:
            sample_key = next(iter(loc_map))
            log_q.put(("info", f"Sample ERP key: '{sample_key}' → '{loc_map[sample_key]}'"))
    finally:
        conn.close()

    return loc_map


def _fetch_comments(receipts: list, log_q, oracle_cfg: OracleConfig) -> dict:
    """Fetch the Comments field from ar_cash_receipts_all for a list of receipt numbers.
    Returns {receipt_number: comments_string}."""
    _init_oracle_client(oracle_cfg.instant_client_dir)

    conn         = _connect_oracle(oracle_cfg)
    comments_map = {}
    try:
        cur     = conn.cursor()
        batches = [receipts[i:i + _BATCH_SIZE]
                   for i in range(0, len(receipts), _BATCH_SIZE)]
        log_q.put(("info", f"Fetching ERP Comments in {len(batches)} batch(es)"))

        col_names = None
        for batch in batches:
            escaped = [r.replace("'", "''") for r in batch]
            sql     = _COMMENTS_SQL.format(
                placeholders=",".join(f"'{r}'" for r in escaped))
            cur.execute(sql)
            if col_names is None:
                col_names = [d[0] for d in cur.description]
            for row in cur.fetchall():
                rd  = dict(zip(col_names, row))
                key = str(rd.get("RECEIPT_NUMBER") or "").strip()
                val = str(rd.get("DESCRIPTION") or "").strip()
                if key:
                    comments_map[key] = val

        cur.close()
        log_q.put(("ok", f"ERP Comments fetched for {len(comments_map)} receipts"))
    finally:
        conn.close()

    return comments_map


def process_report(input_path: str, log_q, as_on_date: _dt.date = None, *,
                    oracle_cfg: OracleConfig, supplier_site_map: dict) -> tuple:
    """
    Returns (df, total_input_rows, unidentified_removed_count, df_unidentified, erp_ok).

    ``supplier_site_map`` replaces the original's hardcoded ``LOCATION_MAP``
    (raw Oracle ``global_attribute10`` value → display Location) — load it
    via ``mapping_store.load_all()``.

    ``erp_ok`` is False when the Oracle ERP lookup itself failed (as
    opposed to succeeding but returning no match for a given receipt) - in
    that case every row's Location comes back blank, which also collapses
    the Summary sheet's Location x Ageing Bucket pivot to a single blank
    row (see write_formatted_excel). The caller must surface erp_ok to the
    user as a clear "Oracle was unreachable" signal, not just leave it in
    the log - a report with every Location blank looks like the
    application is broken, not like an ERP connectivity issue.
    """
    if as_on_date is None:
        as_on_date = _dt.date.today()

    log_q.put(("info", f"Reading file: {os.path.basename(input_path)} …"))
    df = _read_file(input_path)
    log_q.put(("info", f"Raw read: {len(df)} rows × {len(df.columns)} columns"))

    # ── Locate and promote the real header row (Oracle puts 8 metadata rows) ──
    HDR_ROW = 8
    if len(df) <= HDR_ROW:
        raise ValueError(
            f"File has only {len(df)} rows — expected at least {HDR_ROW + 2}. "
            "Is this the correct unapplied receipts file?")
    # Strip whitespace from column names to avoid hidden-space KeyErrors
    df.columns = [str(c).strip() if not pd.isna(c) else ""
                  for c in df.iloc[HDR_ROW].tolist()]
    df = df.iloc[HDR_ROW + 1:].reset_index(drop=True)

    # ── Drop fully-blank trailing rows (Oracle sometimes appends empty rows) ──
    df = df.dropna(how="all").reset_index(drop=True)

    # ── Drop unwanted columns ─────────────────────────────────────────────────
    for col in ["Company:", "GL Date", "Batch Source", "Batch Name", "Claim Amount"]:
        if col in df.columns:
            df = df.drop(columns=col)

    # ── Separate "***** Unidentified" rows → own sheet, not deleted ──────────
    cust_col = "Customer Name"
    if cust_col not in df.columns:
        raise ValueError(
            f"Column '{cust_col}' not found after reading.\n"
            f"Columns found: {list(df.columns)[:10]}")
    before         = len(df)
    cleaned        = df[cust_col].astype(str).str.strip()
    is_unidentified = cleaned.isin({"***** Unidentified", "*****  Unidentified"})
    df_unidentified = df[is_unidentified].reset_index(drop=True)
    df              = df[~is_unidentified].reset_index(drop=True)
    removed         = len(df_unidentified)
    log_q.put(("ok", f"Separated {removed} Unidentified rows → Unidentified Customers sheet  "
                     f"| {len(df)} rows remaining for processing"))

    # ── Fetch Description (Comments) from ERP for Unidentified rows ───────────
    pay_num_col = "Payment Number"
    if not df_unidentified.empty and pay_num_col in df_unidentified.columns:
        unid_receipts = [
            str(r).strip() for r in df_unidentified[pay_num_col].dropna().unique()
            if str(r).strip() and str(r).strip().lower() not in ("nan", "none", "")
        ]
        if unid_receipts:
            try:
                comments_map = _fetch_comments(unid_receipts, log_q, oracle_cfg)
                df_unidentified["Bank Description"] = df_unidentified[pay_num_col].apply(
                    lambda x: comments_map.get(str(x).strip(), ""))
            except Exception as exc:
                log_q.put(("warn", f"Could not fetch ERP Comments for Unidentified rows: {exc}"))
                df_unidentified["Bank Description"] = ""
        else:
            df_unidentified["Bank Description"] = ""

    # ── Fetch Location from Oracle ERP ────────────────────────────────────────
    if pay_num_col not in df.columns:
        raise ValueError(
            f"Column '{pay_num_col}' not found.\n"
            f"Columns found: {list(df.columns)[:10]}")
    receipts = [str(r).strip() for r in df[pay_num_col].dropna().unique()
                if str(r).strip() and str(r).strip().lower() not in ("nan", "none", "")]
    log_q.put(("info", f"Fetching Location for {len(receipts)} unique receipts from ERP …"))

    erp_ok  = True
    loc_map = {}
    try:
        loc_map = _fetch_locations(receipts, log_q, oracle_cfg, supplier_site_map)
        log_q.put(("ok", f"ERP returned data for {len(loc_map)} receipts"))
    except Exception as exc:
        erp_ok = False
        log_q.put(("warn", f"ERP connection failed — Location will be blank  ({exc})"))

    # ── Insert Location column immediately after Payment Date ─────────────────
    pay_date_col = "Payment Date"
    df["Location"] = df[pay_num_col].apply(lambda x: loc_map.get(str(x).strip(), ""))
    cols = list(df.columns)
    cols.remove("Location")
    if pay_date_col in cols:
        cols.insert(cols.index(pay_date_col) + 1, "Location")
    else:
        cols.append("Location")   # fallback: append if Payment Date not found
    df = df[cols]

    # ── Due Days & Ageing Bucket ──────────────────────────────────────────────
    def _parse_date(val):
        s = str(val).strip()
        if s.lower() in ("nan", "none", "nat", ""):
            return pd.NaT
        for fmt in ("%d-%b-%y", "%d-%B-%y", "%d-%b-%Y", "%d-%B-%Y",
                    "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return pd.to_datetime(s, format=fmt)
            except Exception:
                pass
        try:
            return pd.to_datetime(s, dayfirst=True)
        except Exception:
            return pd.NaT

    if pay_date_col not in df.columns:
        log_q.put(("warn", f"'{pay_date_col}' column not found — Due Days will be blank"))
        df["Due Days"]      = ""
        df["Ageing Bucket"] = ""
    else:
        parsed_dates   = df[pay_date_col].apply(_parse_date)
        as_on_ts       = pd.Timestamp(as_on_date)
        df["Due Days"] = (as_on_ts - parsed_dates).dt.days

        def _bucket(days):
            if pd.isna(days):
                return ""
            d = int(days)
            if d <= 30:   return "0 - 30"
            if d <= 60:   return "31 - 60"
            if d <= 90:   return "61 - 90"
            if d <= 120:  return "91 - 120"
            if d <= 150:  return "121 - 150"
            if d <= 180:  return "151 - 180"
            return ">180 days"

        df["Ageing Bucket"] = df["Due Days"].apply(_bucket)
        log_q.put(("ok", f"Ageing buckets computed (as on {as_on_date.strftime('%d-%b-%Y')})"))

    log_q.put(("ok" if erp_ok else "warn", "Report processing complete"))
    return df, before, removed, df_unidentified, erp_ok


def classify_advance_customers(df: pd.DataFrame, ageing_path: str, log_q) -> tuple:
    """
    Split df into (df_main, df_advance).

    Manual-equivalent steps:
      1. Read ageing file (header at row 13).
      2. Drop blank / subtotal rows (no valid Customer Account).
      3. Sort Z→A by Type  →  'Transactions' sorts before 'Receipts'.
      4. For each customer, take the FIRST occurrence in that sorted list.
      5. If the first occurrence is 'Receipts'  →  advance customer
         (meaning the customer has NO transaction entries at all).
      6. Rows in File 1 whose Customer Number matches → moved to Advance sheet.
    """
    log_q.put(("info", f"Reading ageing file ({os.path.splitext(ageing_path)[1].lower()}) …"))
    ageing = _read_ageing(ageing_path)

    # ── find the key columns by name ─────────────────────────────────────────
    type_col  = next((c for c in ageing.columns if "Type" in str(c)), None)
    cust_col  = next((c for c in ageing.columns if "Customer Account" in str(c)), None)
    sales_col = next((c for c in ageing.columns
                      if "sales" in str(c).lower() and "person" in str(c).lower()), None)
    if sales_col is None:
        # Broader fallback: any column with "sales" in the name
        sales_col = next((c for c in ageing.columns if "sales" in str(c).lower()), None)

    log_q.put(("info", f"Ageing — Sales Person column: '{sales_col}'"))

    if not type_col or not cust_col:
        log_q.put(("warn", f"Could not find Type / Customer Account columns. "
                            f"Columns: {list(ageing.columns)}"))
        return df, pd.DataFrame(columns=df.columns), {}

    log_q.put(("info", f"Ageing columns — Type: '{type_col}'  |  Customer: '{cust_col}'"))
    log_q.put(("info", f"Ageing raw rows before cleaning: {len(ageing):,}"))

    # ── helper: normalise a customer code to plain integer string ─────────────
    def _norm(val):
        try:
            return str(int(float(str(val).strip())))
        except Exception:
            return ""

    # ── drop rows that have no valid Customer Account (subtotals / blanks) ────
    ageing["_cust"] = ageing[cust_col].apply(_norm)
    ageing = ageing[ageing["_cust"] != ""].copy()
    log_q.put(("info", f"Ageing rows after dropping blanks: {len(ageing):,}"))

    # ── sample to verify Type values ──────────────────────────────────────────
    sample_types = ageing[type_col].dropna().unique()[:5].tolist()
    log_q.put(("info", f"Sample Type values in ageing: {sample_types}"))

    # ── Step 3: sort Z→A by Type (Transactions > Receipts) ───────────────────
    ageing_sorted = ageing.sort_values(type_col, ascending=False).reset_index(drop=True)

    # ── Step 4: first occurrence per customer in the sorted list ──────────────
    keep_cols = ["_cust", type_col]
    if sales_col:
        keep_cols.append(sales_col)

    first_occ = (
        ageing_sorted
        .drop_duplicates(subset=["_cust"], keep="first")
        [keep_cols]
    )
    # dict: normalised_customer_code → first type after Z→A sort
    first_type_map = dict(zip(first_occ["_cust"], first_occ[type_col].astype(str).str.strip()))

    # dict: normalised_customer_code → salesperson name (from same Z→A-sorted first row)
    if sales_col:
        salesperson_map = {
            k: str(v).strip()
            for k, v in zip(first_occ["_cust"], first_occ[sales_col])
            if str(v).strip() not in ("", "nan", "None", "NaT")
        }
        log_q.put(("info", f"Salesperson map built: {len(salesperson_map):,} entries"))
    else:
        salesperson_map = {}
        log_q.put(("warn", "No Sales Person column found in ageing — Sales Person will be blank"))

    log_q.put(("info", f"Unique customers in ageing: {len(first_type_map):,}"))

    # ── Step 5: which customers have Receipt as their first occurrence? ────────
    advance_customers = {
        cust for cust, typ in first_type_map.items()
        if typ.lower() in ("receipts", "receipt")
    }
    log_q.put(("info", f"Advance customers found: {len(advance_customers):,}"))

    # ── debug: sample advance customers ───────────────────────────────────────
    sample_adv = list(advance_customers)[:5]
    log_q.put(("info", f"Sample advance customer codes: {sample_adv}"))

    # ── Step 6: split File 1 rows ─────────────────────────────────────────────
    df = df.copy()
    df["_cust"] = df["Customer Number"].apply(_norm)

    ageing_customer_set = set(first_type_map.keys())

    # How many are found vs not found in ageing
    matched   = df["_cust"].isin(ageing_customer_set).sum()
    unmatched = (df["_cust"] != "") & ~df["_cust"].isin(ageing_customer_set)
    log_q.put(("info", f"File 1 rows matched in ageing: {matched:,} / {len(df):,}"))
    log_q.put(("info", f"Not found in ageing (VLOOKUP #N/A equivalent): {unmatched.sum():,} rows"
                       f" → treated as Advance"))

    # ── Add Sales Person column (from Z→A-sorted ageing, customer-ID → salesperson) ─
    df["Sales Person"] = df["_cust"].apply(lambda c: salesperson_map.get(c, ""))
    # Insert Sales Person immediately after Location column
    _cols = list(df.columns)
    _cols.remove("Sales Person")
    _loc_idx = _cols.index("Location") if "Location" in _cols else len(_cols) - 1
    _cols.insert(_loc_idx + 1, "Sales Person")
    df = df[_cols]

    # Advance = Receipt-first customers OR not found in ageing at all (= VLOOKUP #N/A)
    mask_receipt_first = df["_cust"].isin(advance_customers)
    mask_not_in_ageing = (df["_cust"] != "") & ~df["_cust"].isin(ageing_customer_set)
    mask_advance       = mask_receipt_first | mask_not_in_ageing

    df_advance = df[mask_advance].drop(columns=["_cust"]).reset_index(drop=True)
    df_main    = df[~mask_advance].drop(columns=["_cust"]).reset_index(drop=True)

    log_q.put(("ok", f"Advance of Customers: {len(df_advance):,} rows  "
                     f"(Receipt-first: {mask_receipt_first.sum():,}  |  Not in ageing: {mask_not_in_ageing.sum():,})"))
    log_q.put(("ok", f"Main (non-advance): {len(df_main):,} rows"))
    return df_main, df_advance, salesperson_map


def _validate_before_save(df_main: pd.DataFrame,
                           incharge_map: dict,
                           supplier_site_map: dict) -> list:
    """
    Validate that all locations and supplier sites in df_main are mapped.
    Returns a list of (category_title, [items]) tuples — empty list = no errors.
    Mirrors the desktop app's ``ValidationErrorDialog`` check, run just
    before the workbook is written.
    """
    errors = []

    # ── Check 1: Accounts Incharge mapping ────────────────────────────────────
    if "Location" in df_main.columns:
        uniq_locs = (
            df_main["Location"]
            .astype(str).str.strip()
            .replace({"": pd.NA, "nan": pd.NA})
            .dropna()
            .unique()
        )
        missing = sorted(
            loc for loc in uniq_locs
            if loc and loc not in incharge_map
        )
        if missing:
            errors.append(("Accounts Incharge Not Mapped", missing))

    # ── Check 2: Supplier Site mapping (only if such a column exists) ─────────
    supp_col = next(
        (c for c in df_main.columns
         if "supplier" in str(c).lower() and "site" in str(c).lower()), None)
    if supp_col and supplier_site_map is not None:
        uniq_sites = (
            df_main[supp_col]
            .astype(str).str.strip()
            .replace({"": pd.NA, "nan": pd.NA})
            .dropna()
            .unique()
        )
        missing_sites = sorted(
            site for site in uniq_sites
            if site and site not in supplier_site_map
        )
        if missing_sites:
            errors.append(("Supplier Site Not Mapped", missing_sites))

    return errors


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL FORMATTING
# ══════════════════════════════════════════════════════════════════════════════

# Columns that hold currency / numeric values.
# Includes every variant that Oracle HTML may produce depending on export format
# (slash present/absent, single/double space, <br> vs raw newline in header).
_CURRENCY_COLS = {
    "On Account/ Prepayment Amount",   # slash on same line as "On Account/"
    "On Account Prepayment Amount",    # slash absent / on separate HTML line
    "On Account/Prepayment Amount",    # no space after slash
    "Unapplied  Amount",               # double-space (older Oracle exports)
    "Unapplied Amount",                # single-space (normalised)
    "Claim Amount",
}
_INT_COLS = {"Customer Number", "Due Days"}


def _thin_border():
    s = Side(style="thin", color="D0D8E4")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_totals_row(ws, df: pd.DataFrame, cols: list, n_rows: int,
                      bdr, hdr_bg: str = "1F3864"):
    """
    Append a bold Grand Total row immediately after the last data row.
    Sums all _CURRENCY_COLS; other cells are left blank (except column 1 = 'Total').
    Row layout: title(1) + header(2) + data(3 … n_rows+2) + totals(n_rows+3)
    """
    tr          = n_rows + 3
    tot_font_w  = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
    tot_num_f   = Font(name="Consolas",  bold=True, size=10, color="FFFFFF")
    tot_fill    = PatternFill("solid", fgColor=hdr_bg)
    r_align     = Alignment(horizontal="right", vertical="center")
    l_align     = Alignment(horizontal="left",  vertical="center", indent=1)

    ws.row_dimensions[tr].height = 22

    for ci, col_name in enumerate(cols, 1):
        c = ws.cell(row=tr, column=ci)
        c.border = bdr
        c.fill   = tot_fill
        if ci == 1:
            c.value     = "Grand Total"
            c.font      = tot_font_w
            c.alignment = l_align
        elif col_name in _CURRENCY_COLS:
            col_sum = pd.to_numeric(df[col_name], errors="coerce").fillna(0).sum()
            c.value         = float(col_sum) if col_sum else None
            c.number_format = "#,##0.00"
            c.font          = tot_num_f
            c.alignment     = r_align
        else:
            c.font      = tot_font_w
            c.alignment = r_align


def write_formatted_excel(df_main: pd.DataFrame, df_advance: pd.DataFrame,
                          path: str, as_on_date: _dt.date,
                          df_unidentified: pd.DataFrame = None,
                          incharge_map: dict = None,
                          log_q=None):
    """Write df to path as a professionally formatted xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Unapplied Receipts"

    df       = df_main
    cols     = list(df.columns)
    n_cols   = len(cols)
    n_rows   = len(df)
    bdr      = _thin_border()

    # ── palette ───────────────────────────────────────────────────────────────
    HDR_BG   = "1F3864"   # deep navy
    HDR_FG   = "FFFFFF"
    ROW_ODD  = "FFFFFF"
    ROW_EVEN = "EDF2FB"   # light periwinkle
    TITLE_BG = "152748"

    # ── Row 1: report title ───────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    title_cell = ws.cell(row=1, column=1,
                          value=f"RDC Unapplied Receipts  —  As On {as_on_date.strftime('%d-%b-%Y')}")
    title_cell.font      = Font(name="Segoe UI", bold=True, size=13, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor=TITLE_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                      indent=1)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=n_cols)
    for c in range(2, n_cols + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=TITLE_BG)

    # ── Row 2: column headers ─────────────────────────────────────────────────
    ws.row_dimensions[2].height = 32
    hdr_font  = Font(name="Segoe UI", bold=True, size=10, color=HDR_FG)
    hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center",
                           wrap_text=True)
    for ci, col_name in enumerate(cols, 1):
        c            = ws.cell(row=2, column=ci, value=col_name)
        c.font       = hdr_font
        c.fill       = hdr_fill
        c.alignment  = hdr_align
        c.border     = bdr

    # ── Data rows ─────────────────────────────────────────────────────────────
    txt_font = Font(name="Segoe UI", size=10)
    num_font = Font(name="Consolas", size=10)

    for ri, row_vals in enumerate(df.itertuples(index=False), 3):
        ws.row_dimensions[ri].height = 18
        is_even = (ri % 2 == 0)
        row_bg  = ROW_EVEN if is_even else ROW_ODD

        for ci, val in enumerate(row_vals, 1):
            col_name = cols[ci - 1]
            c        = ws.cell(row=ri, column=ci)
            c.border = bdr

            # ── Ageing Bucket: plain centred text ─────────────────────────
            if col_name == "Ageing Bucket":
                c.fill      = PatternFill("solid", fgColor=row_bg)
                c.font      = Font(name="Segoe UI", size=10)
                c.value     = str(val) if val is not None else ""
                c.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # ── Due Days: plain number ─────────────────────────────────────
            if col_name == "Due Days":
                c.fill          = PatternFill("solid", fgColor=row_bg)
                c.font          = num_font
                c.alignment     = Alignment(horizontal="right", vertical="center")
                c.number_format = "#,##0"
                try:
                    c.value = int(val) if val is not None and str(val).strip() != "" else ""
                except (TypeError, ValueError):
                    c.value = val
                continue

            # ── Currency columns ───────────────────────────────────────────
            if col_name in _CURRENCY_COLS:
                c.fill         = PatternFill("solid", fgColor=row_bg)
                c.font         = num_font
                c.alignment    = Alignment(horizontal="right", vertical="center")
                c.number_format = '#,##0.00'
                try:
                    c.value = float(val) if val is not None and str(val).strip() != "" else 0.0
                except (ValueError, TypeError):
                    c.value = val
                continue

            # ── Customer Number ────────────────────────────────────────────
            if col_name in _INT_COLS:
                c.fill      = PatternFill("solid", fgColor=row_bg)
                c.font      = num_font
                c.alignment = Alignment(horizontal="right", vertical="center")
                try:
                    c.value = int(float(val)) if val is not None and str(val).strip() != "" else ""
                except (ValueError, TypeError):
                    c.value = val
                continue

            # ── Default text cell ──────────────────────────────────────────
            c.fill      = PatternFill("solid", fgColor=row_bg)
            c.font      = txt_font
            c.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
            c.value     = val if val is not None else ""

    # ── Auto column widths ────────────────────────────────────────────────────
    for ci, col_name in enumerate(cols, 1):
        col_letter = get_column_letter(ci)
        # measure header + sample values
        max_len = len(col_name)
        for ri in range(3, min(3 + n_rows, 3 + 200)):   # sample up to 200 rows
            v = ws.cell(row=ri, column=ci).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        width = min(max(max_len + 3, 10), 45)
        ws.column_dimensions[col_letter].width = width

    # ── Grand Total row (Sheet 1) ─────────────────────────────────────────────
    _write_totals_row(ws, df_main, cols, n_rows, bdr, HDR_BG)

    # ── Freeze panes & auto filter ────────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = (
        f"A2:{get_column_letter(n_cols)}{n_rows + 2}"
    )

    # ── Tab colour ────────────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = "00B61D"

    # ── Advance of Customers sheet ────────────────────────────────────────────
    _write_advance_sheet(wb, df_advance, as_on_date)

    # ── Unidentified Customers sheet ──────────────────────────────────────────
    _write_unidentified_sheet(
        wb,
        df_unidentified if df_unidentified is not None else pd.DataFrame(),
        as_on_date)

    # Save with openpyxl first (no Summary sheet yet)
    wb.save(path)

    # ── Summary sheet: openpyxl formatted table (Location × Ageing Bucket) ──────
    _eff_incharge = incharge_map or {}
    wb2 = load_workbook(path)
    _write_pivot_sheet(wb2, df_main, as_on_date, incharge_map=_eff_incharge)
    summary_idx = wb2.sheetnames.index("Summary")
    wb2.move_sheet("Summary", offset=-summary_idx)
    wb2.save(path)
    if log_q:
        log_q.put(("ok", "Summary sheet written ✓"))


def _write_pivot_sheet(wb, df: pd.DataFrame, as_on_date: _dt.date,
                       incharge_map: dict = None):
    """Add a Summary sheet: Location × Ageing Bucket, ALL values in Lakhs.
    Grand Total (in Lakhs) is the FIRST column after Location Name."""
    incharge_map = incharge_map or {}

    # ── Compute Total Amount = On Account/Prepayment + Unapplied ─────────────
    df = df.copy()
    unapplied_col = next((c for c in df.columns if "Unapplied" in c and "Amount" in c), None)
    prepay_col    = next((c for c in df.columns if "On Account" in c or "Prepayment" in c), None)

    def _to_num(col):
        return pd.to_numeric(df[col], errors="coerce").fillna(0) if col else 0

    df["_total"] = _to_num(unapplied_col) + _to_num(prepay_col)

    # ── Build pivot ───────────────────────────────────────────────────────────
    BUCKET_ORDER = ["0 - 30", "31 - 60", "61 - 90",
                    "91 - 120", "121 - 150", "151 - 180", ">180 days"]

    pivot = (
        df.groupby(["Location", "Ageing Bucket"])["_total"]
        .sum()
        .unstack(fill_value=0)
    )

    ordered = [b for b in BUCKET_ORDER if b in pivot.columns]
    extras  = [c for c in pivot.columns if c not in BUCKET_ORDER]
    pivot   = pivot[ordered + extras].sort_index()

    # Convert ALL bucket values to Lakhs (÷ 1,00,000)
    LAKH  = 100_000
    pivot = pivot / LAKH                          # every bucket now in Lakhs

    # Grand Total = sum of already-Lakh'd columns → insert first
    GT_COL = "Grand Total"
    pivot.insert(0, GT_COL, pivot.sum(axis=1))
    pivot = pivot.sort_values(GT_COL, ascending=False)   # sort rows by Grand Total desc
    grand_row = pivot.sum(axis=0)

    all_cols  = list(pivot.columns)  # [Grand Total, 0-30, 31-60, …]
    n_cols    = len(all_cols)
    # Total sheet columns: Location(1) + Accounts Incharge(2) + data columns
    total_sheet_cols = n_cols + 2

    # ── Styles ────────────────────────────────────────────────────────────────
    HDR_BG   = "1F3864"
    TITLE_BG = "152748"
    ROW_ODD  = "FFFFFF"
    ROW_EVEN = "EDF2FB"
    GTOT_BG  = "D6E4F0"   # light blue for Grand Total cells
    bdr      = _thin_border()

    hdr_font   = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
    hdr_fill   = PatternFill("solid", fgColor=HDR_BG)
    num_font   = Font(name="Consolas", size=10)
    tot_font   = Font(name="Segoe UI", bold=True, size=10)
    grand_font = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
    txt_font   = Font(name="Segoe UI", size=10)

    c_right  = Alignment(horizontal="right",  vertical="center")
    c_left   = Alignment(horizontal="left",   vertical="center", indent=1)
    c_center = Alignment(horizontal="center", vertical="center")

    # ── Sheet ─────────────────────────────────────────────────────────────────
    ws = wb.create_sheet(title="Summary")
    ws.sheet_properties.tabColor = "1F3864"
    ws.sheet_view.showGridLines   = False

    # Row 1 — title
    ws.row_dimensions[1].height = 26
    tc = ws.cell(row=1, column=1,
                 value=f"Unapplied Receipts  |  Location × Ageing  |  As On {as_on_date.strftime('%d-%b-%Y')}")
    tc.font      = Font(name="Segoe UI", bold=True, size=12, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor=TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_sheet_cols)
    for c in range(2, total_sheet_cols + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=TITLE_BG)

    # Row 2 — sub-label
    ws.row_dimensions[2].height = 14
    sl = ws.cell(row=2, column=1,
                 value="All amounts in Lakhs (÷ 1,00,000)")
    sl.font      = Font(name="Segoe UI", size=9, italic=True, color="888888")
    sl.alignment = c_left

    # Row 3 — column headers
    # Col 1: Location Name | Col 2: Accounts Incharge | Col 3+: data columns
    ws.row_dimensions[3].height = 30
    corner = ws.cell(row=3, column=1, value="Location Name")
    corner.font = hdr_font; corner.fill = hdr_fill
    corner.alignment = c_left; corner.border = bdr

    ic_hdr = ws.cell(row=3, column=2, value="Accounts Incharge")
    ic_hdr.font = hdr_font; ic_hdr.fill = hdr_fill
    ic_hdr.alignment = c_center; ic_hdr.border = bdr

    for ci, col_name in enumerate(all_cols, 3):   # data cols start at column 3
        c = ws.cell(row=3, column=ci, value=col_name)
        c.border    = bdr
        c.alignment = c_center
        # All column headers use the same navy style (Grand Total included)
        c.font = hdr_font
        c.fill = hdr_fill

    # Data rows
    for ri, (loc, row_data) in enumerate(pivot.iterrows(), 4):
        ws.row_dimensions[ri].height = 17
        bg = ROW_EVEN if ri % 2 == 0 else ROW_ODD

        # Col 1: location name
        lbl = ws.cell(row=ri, column=1, value=str(loc))
        lbl.font = txt_font
        lbl.fill = PatternFill("solid", fgColor=bg)
        lbl.alignment = c_left; lbl.border = bdr

        # Col 2: Accounts Incharge (look up from mapping dict)
        incharge = incharge_map.get(str(loc).strip(), "")
        ic = ws.cell(row=ri, column=2, value=incharge)
        ic.font = txt_font
        ic.fill = PatternFill("solid", fgColor=bg)
        ic.alignment = c_left; ic.border = bdr

        # Col 3+: data columns
        for ci, (col_name, val) in enumerate(zip(all_cols, row_data), 3):
            is_grand = (col_name == GT_COL)
            c = ws.cell(row=ri, column=ci)
            c.value         = float(val) if val and val != 0 else None
            c.number_format = "#,##0.00"
            c.border        = bdr
            c.alignment     = c_right
            if is_grand:
                c.font = tot_font
                c.fill = PatternFill("solid", fgColor=GTOT_BG)
            else:
                c.font = num_font
                c.fill = PatternFill("solid", fgColor=bg)

    # Grand Total row (bottom)
    gr = 4 + len(pivot)
    ws.row_dimensions[gr].height = 22

    gt = ws.cell(row=gr, column=1, value="Grand Total")
    gt.font = grand_font; gt.fill = hdr_fill
    gt.alignment = c_left; gt.border = bdr

    # Col 2 (Incharge) — blank in totals row
    gt2 = ws.cell(row=gr, column=2, value="")
    gt2.font = grand_font; gt2.fill = hdr_fill
    gt2.alignment = c_left; gt2.border = bdr

    # Col 3+: column sums
    for ci, (col_name, val) in enumerate(zip(all_cols, grand_row), 3):
        c = ws.cell(row=gr, column=ci)
        c.value         = float(val) if val and val != 0 else None
        c.number_format = "#,##0.00"
        c.font          = grand_font
        c.fill          = PatternFill("solid", fgColor=HDR_BG)
        c.alignment     = c_right
        c.border        = bdr

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28   # Location Name
    ws.column_dimensions["B"].width = 18   # Accounts Incharge
    ws.column_dimensions["C"].width = 22   # Grand Total (in Lakhs)
    for ci in range(4, total_sheet_cols + 1):   # bucket columns
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # No freeze panes on Summary sheet (user preference)


def _write_advance_sheet(wb, df_advance: pd.DataFrame, as_on_date: _dt.date):
    """Add Advance of Customers sheet — same formatting as main sheet."""
    ws = wb.create_sheet(title="Advance of Customers")
    ws.sheet_properties.tabColor = "FF7043"
    ws.sheet_view.showGridLines   = False

    if df_advance.empty:
        ws.cell(row=1, column=1, value="No advance customer entries found.")
        return

    cols   = list(df_advance.columns)
    n_cols = len(cols)
    n_rows = len(df_advance)

    HDR_BG   = "7B2D00"   # burnt orange — distinct from main sheet
    TITLE_BG = "5C1F00"
    ROW_ODD  = "FFFFFF"
    ROW_EVEN = "FFF3EE"
    bdr      = _thin_border()

    hdr_font  = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    txt_font  = Font(name="Segoe UI",  size=10)
    num_font  = Font(name="Consolas",  size=10)

    # Row 1: title
    ws.row_dimensions[1].height = 28
    tc = ws.cell(row=1, column=1,
                 value=f"Advance of Customers  —  As On {as_on_date.strftime('%d-%b-%Y')}")
    tc.font      = Font(name="Segoe UI", bold=True, size=13, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor=TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    for c in range(2, n_cols + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=TITLE_BG)

    # Row 2: headers
    ws.row_dimensions[2].height = 32
    for ci, col_name in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=col_name)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = bdr

    # Data rows
    for ri, row_vals in enumerate(df_advance.itertuples(index=False), 3):
        ws.row_dimensions[ri].height = 18
        bg = ROW_EVEN if ri % 2 == 0 else ROW_ODD

        for ci, val in enumerate(row_vals, 1):
            col_name = cols[ci - 1]
            c = ws.cell(row=ri, column=ci)
            c.border = bdr

            if col_name in _CURRENCY_COLS:
                c.fill = PatternFill("solid", fgColor=bg); c.font = num_font
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '#,##0.00'
                try:
                    c.value = float(val) if val is not None and str(val).strip() != "" else 0.0
                except Exception:
                    c.value = val
            elif col_name in _INT_COLS:
                c.fill = PatternFill("solid", fgColor=bg); c.font = num_font
                c.alignment = Alignment(horizontal="right", vertical="center")
                try:
                    c.value = int(float(val)) if val is not None and str(val).strip() != "" else ""
                except Exception:
                    c.value = val
            else:
                c.fill = PatternFill("solid", fgColor=bg); c.font = txt_font
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                c.value = val if val is not None else ""

    # Auto column widths
    for ci, col_name in enumerate(cols, 1):
        max_len = len(col_name)
        for ri in range(3, min(3 + n_rows, 203)):
            v = ws.cell(row=ri, column=ci).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 3, 10), 45)

    # Grand Total row — Advance sheet uses burnt-orange header colour
    _write_totals_row(ws, df_advance, cols, n_rows, bdr, "7B2D00")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{n_rows + 2}"


def _write_unidentified_sheet(wb, df_unid: pd.DataFrame, as_on_date: _dt.date):
    """Add Unidentified Customers sheet — rows where Customer Name was ***** Unidentified."""
    ws = wb.create_sheet(title="Unidentified Customers")
    ws.sheet_properties.tabColor = "607D8B"   # slate grey — visually distinct
    ws.sheet_view.showGridLines   = False

    if df_unid.empty:
        ws.cell(row=1, column=1, value="No unidentified customer entries found.")
        return

    cols   = list(df_unid.columns)
    n_cols = len(cols)
    n_rows = len(df_unid)

    HDR_BG   = "37474F"   # dark slate
    TITLE_BG = "263238"
    ROW_ODD  = "FFFFFF"
    ROW_EVEN = "ECEFF1"   # very light grey
    bdr      = _thin_border()

    hdr_font  = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    txt_font  = Font(name="Segoe UI",  size=10)
    num_font  = Font(name="Consolas",  size=10)

    # Row 1: title
    ws.row_dimensions[1].height = 28
    tc = ws.cell(row=1, column=1,
                 value=f"Unidentified Customers  —  As On {as_on_date.strftime('%d-%b-%Y')}")
    tc.font      = Font(name="Segoe UI", bold=True, size=13, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor=TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    for c in range(2, n_cols + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=TITLE_BG)

    # Row 2: headers
    ws.row_dimensions[2].height = 32
    for ci, col_name in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=col_name)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = bdr

    # Data rows
    for ri, row_vals in enumerate(df_unid.itertuples(index=False), 3):
        ws.row_dimensions[ri].height = 18
        bg = ROW_EVEN if ri % 2 == 0 else ROW_ODD

        for ci, val in enumerate(row_vals, 1):
            col_name = cols[ci - 1]
            c = ws.cell(row=ri, column=ci)
            c.border = bdr
            if col_name in _CURRENCY_COLS:
                c.fill = PatternFill("solid", fgColor=bg); c.font = num_font
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '#,##0.00'
                try:
                    c.value = float(val) if val is not None and str(val).strip() != "" else 0.0
                except Exception:
                    c.value = val
            elif col_name in _INT_COLS:
                c.fill = PatternFill("solid", fgColor=bg); c.font = num_font
                c.alignment = Alignment(horizontal="right", vertical="center")
                try:
                    c.value = int(float(val)) if val is not None and str(val).strip() != "" else ""
                except Exception:
                    c.value = val
            else:
                c.fill = PatternFill("solid", fgColor=bg); c.font = txt_font
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                c.value = val if val is not None else ""

    # Auto column widths
    for ci, col_name in enumerate(cols, 1):
        max_len = len(col_name)
        for ri in range(3, min(3 + n_rows, 203)):
            v = ws.cell(row=ri, column=ci).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 3, 10), 45)

    # Grand Total row — Unidentified sheet uses slate header colour
    _write_totals_row(ws, df_unid, cols, n_rows, bdr, "37474F")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{n_rows + 2}"
