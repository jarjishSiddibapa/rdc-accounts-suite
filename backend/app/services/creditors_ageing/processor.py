"""Pure-Python port of the desktop Creditors Ageing report engine.

The authoritative reference is
``E:\\jarjish-projects\\rakesh-sir-creditors-ageing\\main.py``.  This module
preserves its sheet discovery, Tally Cr/Dr handling, as-on-date inference,
vendor classification, workbook layout, formulas, ageing buckets, and
unclassified-vendor output.  Desktop GUI/AppData concerns are intentionally
replaced by the suite's durable job runtime and centralized MySQL mapping.
"""

from __future__ import annotations

import copy
import csv
import datetime as dt
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.config import SEED_DIR
from app.services.xlsx_formula_cache import inject_cached_values

BUCKETS = (
    "0-30 Days",
    "31-60 Days",
    "61-90 Days",
    "91-120 Days",
    "121-150 Days",
    "151-180 Days",
    "181-365 Days",
    "Above 365 Days",
)
TEMPLATE_SHEETS = ("TB", "Bill Wise", "Only Creditors", "Advances", "Intercompany")
TEMPLATE_PATH = SEED_DIR / "creditors-ageing-report-template.xlsx"


class AgeingReportError(Exception):
    """Expected input/template problem safe to display to a user."""


def _die(message: str):
    raise AgeingReportError(message)


def _norm(value):
    return value.strip().lower() if isinstance(value, str) else None


def _stamp(destination, template):
    destination._style = copy.copy(template._style)


def _clear_rows(sheet, first: int, last: int):
    for row_number in range(first, last + 1):
        for column in range(1, sheet.max_column + 1):
            sheet.cell(row_number, column).value = None


def _read_sheet_rows(path: str, sheet_name: str, include_number_formats: bool = False):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return None
        rows = []
        number_formats = [] if include_number_formats else None
        for row in workbook[sheet_name].iter_rows():
            rows.append([cell.value for cell in row])
            if include_number_formats:
                number_formats.append([cell.number_format for cell in row])
        return {"rows": rows, "number_formats": number_formats}
    finally:
        workbook.close()


def _find_header_row(rows, must_have: tuple[str, ...], max_scan: int = 20):
    for index, row in enumerate(rows[:max_scan]):
        values = [_norm(value) for value in row if value is not None]
        values = [value for value in values if value]
        if all(any(token in value for value in values) for token in must_have):
            return index
    return None


def _discover_input_sheets(path: str):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        tb_sheet = None
        bill_wise_sheet = None
        for sheet in workbook.worksheets:
            rows = []
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                rows.append(list(row))
                if index >= 19:
                    break
            if tb_sheet is None and _find_header_row(rows, ("debit", "credit")) is not None:
                tb_sheet = sheet.title
            if bill_wise_sheet is None and _find_header_row(rows, ("party",)) is not None:
                bill_wise_sheet = sheet.title
        return tb_sheet, bill_wise_sheet, list(workbook.sheetnames)
    finally:
        workbook.close()


def _column_index(row, tokens: tuple[str, ...]):
    for index, value in enumerate(row):
        normalized = _norm(value)
        if normalized and any(token in normalized for token in tokens):
            return index
    return None


def _extract_tb_rows(rows):
    header_index = _find_header_row(rows, ("debit", "credit"))
    if header_index is None:
        _die("Could not locate the Debit/Credit header row in the TB sheet.")
    header = rows[header_index]
    debit_column = _column_index(header, ("debit",))
    credit_column = _column_index(header, ("credit",))
    if debit_column is None or credit_column is None:
        _die("Could not identify the Debit and Credit columns in the TB sheet.")

    result = []
    for row in rows[header_index + 1 :]:
        name = row[0] if row else None
        if not isinstance(name, str) or not name.strip() or name.strip().lower() == "grand total":
            continue
        debit = (row[debit_column] if debit_column < len(row) else 0) or 0
        credit = (row[credit_column] if credit_column < len(row) else 0) or 0
        if not isinstance(debit, (int, float)) or isinstance(debit, bool):
            debit = 0
        if not isinstance(credit, (int, float)) or isinstance(credit, bool):
            credit = 0
        result.append((name.strip(), debit, credit))
    return result


def _signed_amount(value, number_format):
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return 0
    display_format = str(number_format or "").lower()
    if re.search(r"\bdr\b", display_format):
        return -abs(value)
    if re.search(r"\bcr\b", display_format):
        return abs(value)
    return value


def _as_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if not isinstance(value, str):
        return None
    text = re.sub(r"(?i)(\d)(st|nd|rd|th)\b", r"\1", value.strip())
    text = re.sub(r"\s+", " ", text)
    for date_format in (
        "%d-%B-%Y", "%d-%b-%Y", "%d %B %Y", "%d %b %Y",
        "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
        "%d-%B-%y", "%d-%b-%y", "%d %B %y", "%d %b %y",
        "%d/%m/%y", "%d-%m-%y",
    ):
        try:
            return dt.datetime.strptime(text, date_format).date()
        except ValueError:
            pass
    return None


def _ageing_bucket(overdue_days: int) -> str:
    for maximum, bucket in zip((30, 60, 90, 120, 150, 180, 365), BUCKETS):
        if overdue_days <= maximum:
            return bucket
    return BUCKETS[-1]


def _extract_bill_wise_rows(rows, number_formats, ageing_date: dt.date):
    header_index = _find_header_row(rows, ("party",))
    if header_index is None:
        _die("Could not locate the Party's Name header row in the Bill Wise sheet.")
    header = rows[header_index]
    columns = {
        "date": _column_index(header, ("date",)),
        "ref": _column_index(header, ("ref",)),
        "party": _column_index(header, ("party",)),
        "cost_center": _column_index(header, ("cost center", "location")),
        "opening": _column_index(header, ("opening",)),
        "pending": _column_index(header, ("pending amount", "pending")),
        "due_on": _column_index(header, ("due on",)),
    }
    missing = [name for name in ("date", "party", "opening", "pending") if columns[name] is None]
    if missing:
        _die(f"Bill Wise sheet is missing expected column(s): {', '.join(missing)}")

    def get(row, index):
        return row[index] if index is not None and index < len(row) else None

    result = []
    for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 1):
        party = get(row, columns["party"])
        if not isinstance(party, str) or party.strip() in ("", "-"):
            continue
        invoice_date = _as_date(get(row, columns["date"]))
        if invoice_date is None:
            continue
        overdue_days = (ageing_date - invoice_date).days
        pending = get(row, columns["pending"]) or 0
        pending_format = ""
        pending_column = columns["pending"]
        if (
            number_formats
            and row_index < len(number_formats)
            and pending_column is not None
            and pending_column < len(number_formats[row_index])
        ):
            pending_format = number_formats[row_index][pending_column]
        result.append({
            "date": get(row, columns["date"]),
            "ref": get(row, columns["ref"]),
            "party": party.strip(),
            "cc": get(row, columns["cost_center"]) or "",
            "opening": get(row, columns["opening"]) or 0,
            "pending": pending,
            "overdue_days": overdue_days,
            "bucket": _ageing_bucket(overdue_days),
            "new_amt": _signed_amount(pending, pending_format),
            "due_on": get(row, columns["due_on"]),
        })
    return result


def _infer_as_on_from_rows(rows, fallback=None):
    header_index = _find_header_row(rows, ("party",))
    header_rows = rows[: (header_index + 1 if header_index is not None else 10)]
    for row in header_rows:
        for value in row:
            if not isinstance(value, str):
                continue
            match = re.search(
                r"(?i)\bto\s+(\d{1,2}(?:st|nd|rd|th)?[- /][A-Za-z]{3,9}[- /]\d{2,4})\b",
                value,
            )
            if match:
                parsed = _as_date(match.group(1))
                if parsed:
                    return parsed
    for row in header_rows:
        for value in row:
            if isinstance(value, (dt.datetime, dt.date)):
                return _as_date(value)
    return fallback


def _infer_as_on_from_filename(path: str):
    stem = Path(path).stem
    tokens = re.findall(
        r"\d{1,2}(?:st|nd|rd|th)?[- /][A-Za-z]{3,9}[- /]\d{2,4}|\d{1,2}[- /]\d{1,2}[- /]\d{2,4}",
        stem,
        re.I,
    )
    for token in tokens:
        parsed = _as_date(token)
        if parsed:
            return parsed
    return None


def read_tb_export(path: str, as_on_date=None, log=None):
    """Return desktop-compatible ``(tb_rows, bill_rows, as_on_date)``."""
    tb_sheet, bill_sheet, sheet_names = _discover_input_sheets(path)
    if tb_sheet is None or bill_sheet is None:
        _die(
            "Could not identify the TB and Bill Wise data sheets. Expected one sheet "
            "with Debit/Credit headers and one with a Party's Name header "
            f"(found: {', '.join(sheet_names)})."
        )
    tb_raw = _read_sheet_rows(path, tb_sheet, False)
    bill_raw = _read_sheet_rows(path, bill_sheet, True)
    if tb_raw is None or bill_raw is None:
        _die("Could not read the detected TB and Bill Wise sheets.")
    if log and (tb_sheet != "TB" or bill_sheet != "Bill Wise"):
        log("info", f"Detected raw Tally sheets: {tb_sheet} (TB) and {bill_sheet} (Bill Wise)")

    inferred = _infer_as_on_from_rows(bill_raw["rows"])
    inferred = inferred or _infer_as_on_from_rows(tb_raw["rows"])
    inferred = inferred or _infer_as_on_from_filename(path)
    if as_on_date:
        report_date = _as_date(as_on_date)
        if report_date is None:
            _die("Could not understand the report as-on date. Choose a valid calendar date.")
    else:
        report_date = inferred or dt.date.today()
    ageing_date = report_date - dt.timedelta(days=1)
    tb_rows = _extract_tb_rows(tb_raw["rows"])
    bill_rows = _extract_bill_wise_rows(bill_raw["rows"], bill_raw["number_formats"], ageing_date)
    if not tb_rows:
        _die("The TB sheet did not contain any ledger rows.")
    if not bill_rows:
        _die("The Bill Wise sheet did not contain any valid pending-bill rows.")
    return tb_rows, bill_rows, report_date


def _build_workbook(
    input_path: str,
    mapping: dict[str, dict],
    *,
    as_on_date=None,
    template_path=TEMPLATE_PATH,
    progress_cb=None,
    log_cb=None,
):
    def log(level: str, message: str):
        if log_cb:
            log_cb(level, message)

    def progress(fraction: float, message: str):
        if progress_cb:
            progress_cb(fraction, message)
        log("info", message)

    template_path = Path(template_path)
    if not template_path.is_file():
        _die("The packaged Creditors Ageing report template is missing.")
    if not Path(input_path).is_file():
        _die("The uploaded Tally export is no longer available. Please upload it again.")

    progress(0.04, "Validating the Tally export...")
    progress(0.10, "Reading TB and Bill Wise sheets...")
    tb_rows, bill_rows, report_date = read_tb_export(input_path, as_on_date=as_on_date, log=log)
    report_date_text = report_date.strftime("%d-%B-%Y")
    ageing_date = report_date - dt.timedelta(days=1)
    log("info", f"Report as on {report_date_text}; ageing calculated through {ageing_date:%d-%B-%Y}")
    log("success", f"Read {len(tb_rows):,} TB ledgers and {len(bill_rows):,} Bill Wise rows")

    progress(0.22, "Loading the report format...")
    workbook = openpyxl.load_workbook(template_path)
    for sheet_name in TEMPLATE_SHEETS:
        if sheet_name not in workbook.sheetnames:
            _die(f"The report template is missing the expected sheet '{sheet_name}'.")

    progress(0.30, "Classifying vendors with centralized mappings...")
    tb_net = {name: credit - debit for name, debit, credit in tb_rows}
    tb_formula_totals = defaultdict(float)
    for name, debit, credit in tb_rows:
        tb_formula_totals[name] += credit - debit

    cost_center_guesses: dict[str, dict[str, int]] = {}
    for row in bill_rows:
        guesses = cost_center_guesses.setdefault(row["party"], {})
        guesses[row["cc"]] = guesses.get(row["cc"], 0) + 1

    new_vendors: list[str] = []
    new_vendor_info: dict[str, dict] = {}

    def lookup(name: str):
        key = name.upper()
        if key in mapping:
            entry = mapping[key]
            return {
                "loc": entry.get("loc", ""),
                "vt": entry.get("vt", ""),
                "vst": entry.get("vst", ""),
                "intercompany": entry.get("intercompany", False),
            }
        if key in new_vendor_info:
            return new_vendor_info[key]
        guesses = cost_center_guesses.get(name, {})
        guessed_location = max(guesses, key=guesses.get) if guesses else ""
        new_vendors.append(name)
        new_vendor_info[key] = {
            "loc": guessed_location,
            "vt": "",
            "vst": "",
            "intercompany": False,
        }
        return new_vendor_info[key]

    order_index = {key: index for index, key in enumerate(mapping)}
    only_creditors = []
    advances = []
    intercompany = []
    for name, _debit, _credit in tb_rows:
        classification = lookup(name)
        if classification.get("intercompany"):
            intercompany.append((name, classification))
        elif tb_net[name] >= 0:
            only_creditors.append((name, classification))
        else:
            advances.append((name, classification))
    only_creditors.sort(key=lambda item: item[0].lower())
    advances.sort(key=lambda item: item[0].lower())
    intercompany.sort(key=lambda item: order_index.get(item[0].upper(), 999_999))
    log(
        "success",
        "Classified vendors: "
        f"Only Creditors {len(only_creditors):,}, Advances {len(advances):,}, "
        f"Intercompany {len(intercompany):,}",
    )
    if new_vendors:
        log("warning", f"{len(new_vendors):,} vendor(s) need classification")

    cached_values: dict[str, dict[str, float]] = defaultdict(dict)

    progress(0.44, "Rebuilding the TB sheet...")
    sheet = workbook["TB"]
    old_max = sheet.max_row
    header_row = None
    for row_number in range(1, min(20, sheet.max_row) + 1):
        values = [_norm(sheet.cell(row_number, column).value) for column in range(1, sheet.max_column + 1)]
        values = [value for value in values if value]
        if values and all(any(token in value for value in values) for token in ("debit", "credit")):
            header_row = row_number
            break
    if header_row is None:
        _die("The report template's TB sheet has no Debit/Credit header.")
    template_cells = [sheet.cell(header_row + 1, column) for column in range(1, sheet.max_column + 1)]
    data_start = header_row + 1
    _clear_rows(sheet, data_start, old_max)
    row_number = data_start
    for name, debit, credit in tb_rows:
        for column, value in enumerate((name, debit, credit), start=1):
            cell = sheet.cell(row_number, column)
            _stamp(cell, template_cells[column - 1])
            cell.value = value
        row_number += 1
    last_tb_row = row_number - 1
    tb_grand_total_row = last_tb_row + 2
    sheet.cell(tb_grand_total_row, 1).value = "Grand Total"
    sheet.cell(tb_grand_total_row, 1).font = Font(bold=True)
    for column in (2, 3):
        cell = sheet.cell(tb_grand_total_row, column)
        _stamp(cell, template_cells[column - 1])
        letter = get_column_letter(column)
        cell.value = f"=SUM({letter}{data_start}:{letter}{last_tb_row})"
        cell.font = Font(bold=True)
        cached_values["TB"][cell.coordinate] = sum(
            value[column - 1] for value in tb_rows if isinstance(value[column - 1], (int, float))
        )
    sheet.auto_filter.ref = f"A{header_row}:C{last_tb_row}"
    for row in range(1, header_row):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            if isinstance(cell.value, str) and cell.value.startswith("=SUBTOTAL"):
                letter = get_column_letter(column)
                cell.value = f"=SUBTOTAL(9,{letter}{data_start}:{letter}{last_tb_row})"
                cached_values["TB"][cell.coordinate] = sum(
                    values[column - 1]
                    for values in tb_rows
                    if column - 1 < len(values) and isinstance(values[column - 1], (int, float))
                )

    progress(0.56, "Rebuilding the Bill Wise sheet...")
    sheet = workbook["Bill Wise"]
    old_max = sheet.max_row
    bill_header_row = None
    for row in range(1, min(20, sheet.max_row) + 1):
        values = [_norm(sheet.cell(row, column).value) for column in range(1, sheet.max_column + 1)]
        if any(value and "party" in value for value in values):
            bill_header_row = row
            break
    if bill_header_row is None:
        _die("The report template's Bill Wise sheet has no Party's Name header.")
    template_row = [sheet.cell(bill_header_row + 1, column) for column in range(1, sheet.max_column + 1)]
    column_count = sheet.max_column
    columns = {
        _norm(sheet.cell(bill_header_row, column).value): column
        for column in range(1, column_count + 1)
    }
    date_column = columns.get("date")
    reference_column = columns.get("ref. no.") or columns.get("ref no") or columns.get("ref no.")
    party_column = columns.get("party's name") or columns.get("partys name")
    cost_center_column = columns.get("cost center/location")
    opening_column = next((value for key, value in columns.items() if "opening" in (key or "")), None)
    pending_column = next((value for key, value in columns.items() if "pending" in (key or "")), None)
    overdue_column = next((value for key, value in columns.items() if "overdue" in (key or "") and "day" in (key or "")), None)
    bucket_column = next((value for key, value in columns.items() if "bucket" in (key or "")), None)
    new_amount_column = next((value for key, value in columns.items() if "new amt" in (key or "")), None)
    if party_column is None or new_amount_column is None:
        _die("The report template's Bill Wise sheet is missing Party's Name or New Amt.")
    _clear_rows(sheet, bill_header_row + 1, old_max)
    row_number = bill_header_row + 1
    for item in bill_rows:
        for column in range(1, column_count + 1):
            _stamp(sheet.cell(row_number, column), template_row[column - 1])
        assignments = (
            (date_column, item["date"]),
            (reference_column, item["ref"]),
            (party_column, item["party"]),
            (cost_center_column, item["cc"]),
            (opening_column, item["opening"]),
            (pending_column, item["pending"]),
            (overdue_column, item["overdue_days"]),
            (bucket_column, item["bucket"]),
            (new_amount_column, item["new_amt"]),
        )
        for column, value in assignments:
            if column:
                sheet.cell(row_number, column).value = value
        row_number += 1
    last_bill_row = row_number - 1
    for row in range(1, bill_header_row):
        for column in range(1, column_count + 1):
            cell = sheet.cell(row, column)
            if isinstance(cell.value, str) and cell.value.startswith("=SUBTOTAL"):
                letter = get_column_letter(new_amount_column)
                cell.value = f"=SUBTOTAL(9,{letter}{bill_header_row + 1}:{letter}{last_bill_row})"
                cached_values["Bill Wise"][cell.coordinate] = sum(float(item["new_amt"]) for item in bill_rows)
    sheet.auto_filter.ref = f"A{bill_header_row}:{get_column_letter(column_count)}{last_bill_row}"

    bill_totals = defaultdict(float)
    for item in bill_rows:
        bill_totals[(item["party"], item["bucket"])] += float(item["new_amt"])

    def build_ageing_sheet(sheet_name: str, data):
        sheet = workbook[sheet_name]
        old_max = sheet.max_row
        header = None
        for row in range(1, min(20, sheet.max_row) + 1):
            values = [_norm(sheet.cell(row, column).value) for column in range(1, sheet.max_column + 1)]
            if any(value and "vendor name" in value for value in values):
                header = row
                break
        if header is None:
            _die(f"The report template's {sheet_name} sheet has no Vendor Name header.")
        columns = {
            _norm(sheet.cell(header, column).value): column
            for column in range(1, sheet.max_column + 1)
        }
        name_column = columns.get("vendor name")
        location_column = columns.get("location")
        type_column = columns.get("vendor type")
        sub_type_column = columns.get("vendor sub type")
        tb_balance_column = columns.get("tb balance (cr)")
        if not name_column or not tb_balance_column:
            _die(f"The report template's {sheet_name} sheet is missing Vendor Name or TB Balance (Cr).")

        for row in (1, 2):
            value = sheet.cell(row, 1).value
            if isinstance(value, str) and "|" in value and ("as at" in value.lower() or "as on" in value.lower()):
                prefix = value.split("|")[0].strip().replace("�", "—")
                sheet.cell(row, 1).value = f"{prefix} | As on {report_date_text}"

        data_start = header + 1
        template_cells = {column: sheet.cell(data_start, column) for column in range(1, sheet.max_column + 1)}
        _clear_rows(sheet, data_start, old_max)
        row_number = data_start
        numeric_values_by_column: dict[int, list[float]] = defaultdict(list)
        last_formula_column = None
        for serial, (name, classification) in enumerate(data, start=1):
            for column in range(1, sheet.max_column + 1):
                _stamp(sheet.cell(row_number, column), template_cells[column])
            sheet.cell(row_number, 1).value = serial
            sheet.cell(row_number, name_column).value = name
            if location_column:
                sheet.cell(row_number, location_column).value = classification.get("loc", "")
            if type_column:
                sheet.cell(row_number, type_column).value = classification.get("vt", "")
            if sub_type_column:
                sheet.cell(row_number, sub_type_column).value = classification.get("vst", "")

            letter = get_column_letter(name_column)
            tb_cell = sheet.cell(row_number, tb_balance_column)
            tb_cell.value = (
                f"=IFERROR(SUMIFS(TB!C:C,TB!A:A,{letter}{row_number})-"
                f"SUMIFS(TB!B:B,TB!A:A,{letter}{row_number}),0)"
            )
            tb_value = float(tb_formula_totals.get(name, 0))
            cached_values[sheet_name][tb_cell.coordinate] = tb_value
            numeric_values_by_column[tb_balance_column].append(tb_value)

            bucket_values = []
            for bucket_index, bucket in enumerate(BUCKETS):
                column = tb_balance_column + 1 + bucket_index
                cell = sheet.cell(row_number, column)
                cell.value = (
                    f"=SUMIFS('Bill Wise'!I:I,'Bill Wise'!C:C,{letter}{row_number},"
                    f"'Bill Wise'!H:H,\"{bucket}\")"
                )
                value = float(bill_totals.get((name, bucket), 0))
                cached_values[sheet_name][cell.coordinate] = value
                numeric_values_by_column[column].append(value)
                bucket_values.append(value)

            bill_total_column = tb_balance_column + 1 + len(BUCKETS)
            difference_column = bill_total_column + 1
            last_formula_column = difference_column
            bill_total_cell = sheet.cell(row_number, bill_total_column)
            bill_total_cell.value = (
                f"=SUM({get_column_letter(tb_balance_column + 1)}{row_number}:"
                f"{get_column_letter(bill_total_column - 1)}{row_number})"
            )
            bill_total = sum(bucket_values)
            cached_values[sheet_name][bill_total_cell.coordinate] = bill_total
            numeric_values_by_column[bill_total_column].append(bill_total)

            difference_cell = sheet.cell(row_number, difference_column)
            difference_cell.value = (
                f"={get_column_letter(tb_balance_column)}{row_number}-"
                f"{get_column_letter(bill_total_column)}{row_number}"
            )
            difference = tb_value - bill_total
            cached_values[sheet_name][difference_cell.coordinate] = difference
            numeric_values_by_column[difference_column].append(difference)
            row_number += 1

        last_data_row = row_number - 1
        subtotal_columns = list(range(tb_balance_column, last_formula_column + 1)) if data else []
        for column in subtotal_columns:
            cell = sheet.cell(3, column)
            letter = get_column_letter(column)
            cell.value = f"=SUBTOTAL(9,{letter}{data_start}:{letter}{last_data_row})"
            cached_values[sheet_name][cell.coordinate] = sum(numeric_values_by_column[column])

        grand_total_row = last_data_row + 2
        sheet.cell(grand_total_row, name_column).value = "Grand Total"
        sheet.cell(grand_total_row, name_column).font = Font(bold=True)
        for column in subtotal_columns:
            cell = sheet.cell(grand_total_row, column)
            _stamp(cell, template_cells[column])
            letter = get_column_letter(column)
            cell.value = f"=SUM({letter}{data_start}:{letter}{last_data_row})"
            cell.font = Font(bold=True)
            cached_values[sheet_name][cell.coordinate] = sum(numeric_values_by_column[column])
        sheet.auto_filter.ref = f"A{header}:{get_column_letter(sheet.max_column)}{last_data_row}"
        if sheet.max_row > grand_total_row:
            sheet.delete_rows(grand_total_row + 1, sheet.max_row - grand_total_row)

    progress(0.68, "Rebuilding the Only Creditors sheet...")
    build_ageing_sheet("Only Creditors", only_creditors)
    progress(0.76, "Rebuilding the Advances sheet...")
    build_ageing_sheet("Advances", advances)
    progress(0.84, "Rebuilding the Intercompany sheet...")
    build_ageing_sheet("Intercompany", intercompany)

    progress(0.92, "Finalizing formulas and workbook formatting...")
    for sheet_name, last_row in (("TB", tb_grand_total_row), ("Bill Wise", last_bill_row)):
        sheet = workbook[sheet_name]
        if sheet.max_row > last_row:
            sheet.delete_rows(last_row + 1, sheet.max_row - last_row)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    new_vendor_guesses = {
        name: new_vendor_info[name.upper()]["loc"]
        for name in new_vendors
    }
    return workbook, {
        "as_on_date": report_date.isoformat(),
        "as_on_label": report_date_text,
        "ageing_through_date": ageing_date.isoformat(),
        "new_vendors": new_vendors,
        "new_vendor_guesses": new_vendor_guesses,
        "counts": {
            "only_creditors": len(only_creditors),
            "advances": len(advances),
            "intercompany": len(intercompany),
        },
        "tb_ledgers": len(tb_rows),
        "bill_wise_rows": len(bill_rows),
        "default_name": f"Ultrafine Creditors Ageing as on {report_date_text}.xlsx",
        "tb_net": tb_net,
    }, dict(cached_values)


def write_new_vendors_csv(path: str | Path, meta: dict, mapping: dict[str, dict]) -> int:
    unresolved = [name for name in meta["new_vendors"] if name.upper() not in mapping]
    with Path(path).open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.writer(output)
        writer.writerow(("Vendor Name", "Guessed Location (from Bill Wise)", "Net Balance", "Suggested Sheet"))
        for name in unresolved:
            balance = meta["tb_net"].get(name, 0)
            writer.writerow((
                name,
                meta["new_vendor_guesses"].get(name, ""),
                round(balance, 2),
                "Only Creditors" if balance >= 0 else "Advances",
            ))
    return len(unresolved)


def generate_report(
    input_path: str,
    output_path: str,
    mapping: dict[str, dict],
    *,
    as_on_date=None,
    new_vendors_csv_path: str | None = None,
    progress_cb=None,
    log_cb=None,
) -> dict:
    workbook, meta, cached_values = _build_workbook(
        input_path,
        mapping,
        as_on_date=as_on_date,
        progress_cb=progress_cb,
        log_cb=log_cb,
    )
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    inject_cached_values(output_path, cached_values)

    unresolved_count = 0
    if meta["new_vendors"] and new_vendors_csv_path:
        unresolved_count = write_new_vendors_csv(new_vendors_csv_path, meta, mapping)

    result = {key: value for key, value in meta.items() if key != "tb_net"}
    result["output_path"] = str(output_path)
    result["new_vendors_csv_path"] = str(new_vendors_csv_path) if unresolved_count else None
    return result
