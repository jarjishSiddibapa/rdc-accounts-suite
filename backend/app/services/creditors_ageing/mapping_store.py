"""MySQL-backed vendor classifications for Creditors Ageing.

The reference desktop app stored these rows in
``%APPDATA%\\CreditorsAgeingApp\\vendor-mapping.xlsx`` and seeded that file
from its bundled report template.  The suite keeps the same normalized key
and insertion order, but MySQL is now the only runtime source of truth.
The packaged template is a one-way, additive deployment seed: startup never
overwrites administrator edits and never revives archived rows.
"""

from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from app.config import SEED_DIR
from app.soft_delete import delete_keyed_row, restore_keyed_row, seed_missing_keyed_rows

from .models import VendorMapping

SEED_TEMPLATE_PATH = SEED_DIR / "creditors-ageing-report-template.xlsx"
REPORT_SHEETS = ("Only Creditors", "Advances", "Intercompany")


class ArchivedMappingError(ValueError):
    """A row with this normalized key exists in the archive."""


class DuplicateMappingError(ValueError):
    """A different active row already owns the requested normalized key."""


def _s(value) -> str:
    return str(value).strip() if value is not None else ""


def vendor_key(value: str) -> str:
    return _s(value).upper()


def _find_header_row(ws, must_have: tuple[str, ...], max_scan: int = 20):
    for row_number in range(1, min(max_scan, ws.max_row) + 1):
        values = [
            _s(ws.cell(row_number, column).value).lower()
            for column in range(1, ws.max_column + 1)
            if _s(ws.cell(row_number, column).value)
        ]
        if all(any(token in value for value in values) for token in must_have):
            return row_number
    return None


def parse_seed_template(path=SEED_TEMPLATE_PATH) -> dict[str, dict]:
    """Read the desktop app's existing mappings in their original order."""
    source = Path(path)
    if not source.is_file():
        return {}

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    mappings: dict[str, dict] = {}
    try:
        for sheet_name in REPORT_SHEETS:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            header_row = _find_header_row(sheet, ("vendor name",))
            if header_row is None:
                continue
            columns = {
                _s(sheet.cell(header_row, column).value).lower(): column
                for column in range(1, sheet.max_column + 1)
            }
            name_column = columns.get("vendor name")
            if not name_column:
                continue
            location_column = columns.get("location")
            type_column = columns.get("vendor type")
            sub_type_column = columns.get("vendor sub type")
            for row_number in range(header_row + 1, sheet.max_row + 1):
                name = sheet.cell(row_number, name_column).value
                if not isinstance(name, str) or not name.strip():
                    continue
                name = name.strip()
                if name.lower() == "grand total":
                    continue
                key = vendor_key(name)
                mappings[key] = {
                    "vendor_name": name,
                    "location": _s(sheet.cell(row_number, location_column).value) if location_column else "",
                    "vendor_type": _s(sheet.cell(row_number, type_column).value) if type_column else "",
                    "vendor_sub_type": _s(sheet.cell(row_number, sub_type_column).value) if sub_type_column else "",
                    "intercompany": sheet_name == "Intercompany",
                }
    finally:
        workbook.close()
    return mappings


def seed_missing_from_template(db: Session, path=None) -> int:
    seed = parse_seed_template(path or SEED_TEMPLATE_PATH)
    inserted = seed_missing_keyed_rows(
        db,
        VendorMapping,
        ("vendor_key",),
        seed,
    )
    db.commit()
    return inserted


def load_all(db: Session) -> dict[str, dict]:
    """Return the processor's desktop-compatible ordered mapping shape."""
    rows = (
        db.query(VendorMapping)
        .filter(VendorMapping.is_deleted.is_(False))
        .order_by(VendorMapping.id)
        .all()
    )
    return {
        row.vendor_key: {
            "name": row.vendor_name,
            "loc": row.location,
            "vt": row.vendor_type,
            "vst": row.vendor_sub_type,
            "intercompany": bool(row.intercompany),
        }
        for row in rows
    }


def list_rows(db: Session, *, archived: bool = False) -> list[dict]:
    rows = (
        db.query(VendorMapping)
        .filter(VendorMapping.is_deleted.is_(archived))
        .order_by(VendorMapping.vendor_name)
        .all()
    )
    return [
        {
            "vendor_name": row.vendor_name,
            "location": row.location,
            "vendor_type": row.vendor_type,
            "vendor_sub_type": row.vendor_sub_type,
            "intercompany": "Yes" if row.intercompany else "No",
        }
        for row in rows
    ]


def upsert_mapping(
    db: Session,
    *,
    original_name: str | None,
    vendor_name: str,
    location: str,
    vendor_type: str,
    vendor_sub_type: str,
    intercompany: bool,
) -> None:
    """Create or edit one mapping without racing unrelated rows."""
    name = _s(vendor_name)
    if not name:
        raise ValueError("Vendor name is required")

    new_key = vendor_key(name)
    old_key = vendor_key(original_name) if original_name is not None else None
    target = db.query(VendorMapping).filter(VendorMapping.vendor_key == new_key).first()

    if old_key is None:
        if target is not None and target.is_deleted:
            raise ArchivedMappingError("This vendor exists in the archive. Restore it instead of adding a duplicate.")
    elif new_key != old_key:
        original = db.query(VendorMapping).filter(
            VendorMapping.vendor_key == old_key,
            VendorMapping.is_deleted.is_(False),
        ).first()
        if original is None:
            raise LookupError("Mapping not found")
        if target is not None:
            if target.is_deleted:
                raise ArchivedMappingError("The new vendor name exists in the archive. Restore that row first.")
            raise DuplicateMappingError("A mapping already exists for the new vendor name.")
        original.is_deleted = True
        target = None

    if target is None:
        target = VendorMapping(vendor_key=new_key)
        db.add(target)
    target.vendor_name = name
    target.location = _s(location)
    target.vendor_type = _s(vendor_type)
    target.vendor_sub_type = _s(vendor_sub_type)
    target.intercompany = bool(intercompany)
    target.is_deleted = False
    db.commit()


def archive_mapping(db: Session, name: str) -> bool:
    changed = delete_keyed_row(db, VendorMapping, ("vendor_key",), vendor_key(name))
    db.commit()
    return changed


def restore_mapping(db: Session, name: str) -> bool:
    changed = restore_keyed_row(db, VendorMapping, ("vendor_key",), vendor_key(name))
    db.commit()
    return changed

