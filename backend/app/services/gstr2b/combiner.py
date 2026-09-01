"""Core GSTR-2B combine pipeline.

Ports the desktop app's processing logic (sumeet-sir-gstr-2b-file-
combinator/app.py, lines 50-259: TARGET_TABS, TAB_CFG, MONTHS, _clean,
_build_col_names, _read_tab, _write_sheet, _parse_filename, run_combine)
verbatim wherever the original logic doesn't need to change, with two
adaptations required for the web port:

  1. run_combine takes an explicit list of uploaded file paths instead of
     globbing a local input_dir - the browser uploads files individually,
     there is no server-side "folder" to scan.
  2. State names are resolved from a `state_codes: dict[int, str]` argument
     (loaded by the caller via state_codes_store.load_all(db)) instead of
     the hardcoded module-level STATE_CODES constant, so admins can edit the
     mapping table without a code change. The exact fallback string for an
     unmapped code, f"Unknown state ({sc})", is preserved unchanged -
     including for a code a user has since removed via the CRUD UI.

Logging/progress follows this suite's established convention (see
app/services/unaccounted/processing.py): callers pass a `log_q` object with
a `.put((tag, message))` method instead of the original app's two separate
on_progress/on_log callbacks; the router's job wrapper adapts that to the
background job's progress_cb(frac, phase).
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.jobs import JobUserError

TARGET_TABS = ["B2B", "B2BA", "B2B-CDNR", "B2B-CDNRA", "IMPG"]

TAB_CFG: dict[str, dict] = {
    "B2B":       {"header_rows": [4, 5],    "data_start": 6},
    "B2BA":      {"header_rows": [4, 5, 6], "data_start": 7},
    "B2B-CDNR":  {"header_rows": [4, 5],    "data_start": 6},
    "B2B-CDNRA": {"header_rows": [4, 5, 6], "data_start": 7},
    "IMPG":      {"header_rows": [4, 5],    "data_start": 6},
}

MONTHS: dict[str, str] = {
    "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
    "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
    "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
}

# Excel output styling - identical to the desktop app's _HDR_FILL/_HDR_FONT/
# _HDR_ALIGN/_DAT_FONT/_META_FILL/_META_FONT constants.
_HDR_FILL  = PatternFill("solid", start_color="1F497D", end_color="1F497D")
_HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DAT_FONT  = Font(name="Arial", size=10)
_META_FILL = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
_META_FONT = Font(bold=True, color="1F497D", name="Arial", size=10)


def _clean(v) -> str | None:
    """Return stripped string value or None for blank / NaN cells."""
    if not pd.notna(v):
        return None
    s = str(v).strip()
    return None if s in ("", "nan") else s


def _build_col_names(raw: pd.DataFrame, header_rows: list[int]) -> list[str]:
    """
    Flatten multi-level header rows into unique, readable column names.
    Non-leaf header rows are forward-filled (they span child columns via
    merged cells, which arrive as NaN after the first cell).
    The leaf level is kept literal; nulls mean 'no sub-header - use parent'.
    """
    n = len(header_rows)
    levels: list[list[str | None]] = []

    for idx, row_i in enumerate(header_rows):
        raw_vals = raw.iloc[row_i].tolist()
        if idx < n - 1:                         # non-leaf: forward-fill
            buf: list[str | None] = []
            last: str | None = None
            for v in raw_vals:
                sv = _clean(v)
                if sv:
                    last = sv
                buf.append(last)
        else:                                   # leaf: keep nulls
            buf = [_clean(v) for v in raw_vals]
        levels.append(buf)

    names: list[str] = []
    seen: dict[str, int] = {}
    for i in range(len(levels[0])):
        parts: list[str] = []
        for lvl in levels:
            val = lvl[i]
            if val and (not parts or val != parts[-1]):
                parts.append(val)
        base  = " - ".join(parts) if parts else f"Column_{i + 1}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        names.append(base if count == 1 else f"{base}.{count - 1}")
    return names


_CURRENCY_MARKER = "₹"  # the GST portal marks every monetary column with a "(₹)" suffix


def _coerce_currency_columns(df: pd.DataFrame) -> pd.DataFrame:
    """The raw sheet is read with dtype=str so the multi-row header parsing
    in _build_col_names() stays uniform - but that also turns every DATA
    cell into a string, including the real amount columns (Invoice Value,
    Taxable Value, Integrated/Central/State/Cess Tax). Those then land in
    the output as text - can't be summed, shows left-aligned in Excel - a
    bug confirmed present in the original desktop app's output too (not a
    web-porting regression), reported by a user and fixed here for both
    paths since this port now diverges from the original on purpose.

    The GST portal itself marks every monetary column with a "(₹)" suffix
    in its own header text (preserved verbatim through _build_col_names) -
    a precise, low-risk signal for which columns are genuinely numeric,
    unlike guessing from position or a fuzzy name match that could
    misfire on an identifier column (e.g. Invoice Number can be purely
    numeric text and must NOT be converted - it has no ₹ in its header,
    so this leaves it untouched).

    Column-wide, not per-cell: if even one non-blank value in a ₹ column
    fails to parse as numeric, the whole column is left as text rather
    than risk silently blanking a malformed cell.
    """
    for col in df.columns:
        if _CURRENCY_MARKER not in str(col):
            continue
        series = df[col]
        non_blank = series.notna() & (series.astype(str).str.strip() != "")
        if not non_blank.any():
            continue
        numeric = pd.to_numeric(series[non_blank], errors="coerce")
        if numeric.isna().any():
            continue  # something in this "money" column didn't parse - leave the whole column as text
        # dtype=str above produces pandas' strict StringDtype column, which
        # raises on assigning a non-string value via .loc - cast to plain
        # object dtype first so the numeric values can actually be written.
        df[col] = df[col].astype(object)
        df.loc[non_blank, col] = numeric
    return df


def _read_tab(fp: Path, tab: str) -> pd.DataFrame:
    """Read one tab from a GSTR-2B file, returning clean data rows only."""
    cfg  = TAB_CFG[tab]
    raw  = pd.read_excel(fp, sheet_name=tab, header=None, dtype=str)
    cols = _build_col_names(raw, cfg["header_rows"])
    ds   = cfg["data_start"]
    if ds >= len(raw):
        return pd.DataFrame(columns=cols)
    df = raw.iloc[ds:].copy()
    df.columns = cols
    df = df.dropna(how="all").reset_index(drop=True)
    return _coerce_currency_columns(df)


def _write_sheet(ws, df: pd.DataFrame, log_q=None) -> None:
    """Write a combined DataFrame to an openpyxl worksheet with styling."""
    # Header row
    for c, h in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill      = _META_FILL if c <= 3 else _HDR_FILL
        cell.font      = _META_FONT if c <= 3 else _HDR_FONT
        cell.alignment = _HDR_ALIGN
    ws.row_dimensions[1].height = 32

    # Every log_q.put() nudges the job's progress fraction along (see
    # _LogQueue in the router) - without this, a tab with many rows would
    # go completely silent between "Writing output workbook..." and
    # "Saved", the same frozen-progress symptom the ERP converter had
    # before its own fix. Throttled to ~50 updates per tab regardless of
    # row count so it doesn't spam the log panel.
    n_rows = len(df)
    log_step = max(1, n_rows // 50)

    # Data rows - itertuples is ~10x faster than iterrows
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, v in enumerate(row, 1):
            safe: object = None
            if not (isinstance(v, float) and np.isnan(v)):
                safe = v
            if isinstance(safe, str) and safe.lower() in ("nan", "none"):
                safe = None
            ws.cell(row=r, column=c, value=safe).font = _DAT_FONT
        if log_q is not None:
            pos = r - 1
            if pos % log_step == 0 or pos == n_rows:
                log_q.put(("dim", f"  [{ws.title}]  writing row {pos:,}/{n_rows:,}"))

    # Column widths - sample first 100 data rows + header to keep it fast
    sample_rows = min(ws.max_row, 101)
    for c, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=sample_rows), 1):
        w = max((len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells)
        ws.column_dimensions[get_column_letter(c)].width = min(w + 2, 45)

    ws.freeze_panes = "A2"


def _parse_filename(fp: Path) -> tuple[str, int]:
    """
    Extract (month_year, state_code) from a GSTR-2B filename.
    Pattern: MMYYYY_SSGSTIN_GSTR2B_DDMMYYYY.xlsx
    """
    parts  = fp.stem.split("_")
    mmyyyy = parts[0]
    mm, yyyy = mmyyyy[:2], mmyyyy[2:]
    return f"{MONTHS[mm]}-{yyyy}", int(parts[1][:2])


def run_combine(
    files: list[tuple[str, str]],
    output_path: Path | str,
    state_codes: dict[int, str],
    log_q,          # object with .put((tag: str, message: str)) -> None
) -> dict:
    """
    Combine the given GSTR-2B .xlsx files into a single workbook.

    `files` is a list of (original_filename, saved_path) pairs - the
    caller (router) must pass the filename the user actually uploaded, not
    the uuid-prefixed name save_upload() writes to disk (mirrors
    ultrafine_balance_confirmation/processor.py's build_pdf_lookup).
    _parse_filename() reads MMYYYY_SSGSTIN_... out of the *original* name;
    passing the on-disk name here made every file fail to parse (the uuid
    prefix isn't a valid "01".."12" month) and silently produced an empty
    workbook with no error, since a zero-files-parsed run was never treated
    as a failure.

    `state_codes` resolves the 2-digit state code embedded in each
    filename's GSTIN segment to a state name; any code missing from it
    falls back to the original desktop app's exact
    f"Unknown state ({sc})" string (including for a code the user has since
    removed via the state-code CRUD UI).
    """
    pairs = sorted(files, key=lambda pair: pair[0])
    if not pairs:
        raise JobUserError("No .xlsx files were uploaded")

    frames: dict[str, list[pd.DataFrame]] = {t: [] for t in TARGET_TABS}
    counts: dict[str, int]                = {t: 0  for t in TARGET_TABS}
    n_files = 0
    unresolved_codes: set[int] = set()

    for original_name, saved_path in pairs:
        fp = Path(saved_path)
        try:
            month_year, sc = _parse_filename(Path(original_name))
        except Exception as exc:
            log_q.put(("warn", f"Skipped {original_name}  (cannot parse: {exc})"))
            continue

        state = state_codes.get(sc, f"Unknown state ({sc})")
        if sc not in state_codes:
            unresolved_codes.add(sc)
            log_q.put(("warn", f"  State code {sc} has no mapping - using \"{state}\" as a placeholder"))
        log_q.put(("info", f"Processing  {original_name}"))
        log_q.put(("dim", f"  State: {sc} - {state}   Period: {month_year}"))
        n_files += 1

        for tab in TARGET_TABS:
            try:
                df = _read_tab(fp, tab)
                if not df.empty:
                    df.insert(0, "Month & Year", month_year)
                    df.insert(0, "State Name",   state)
                    df.insert(0, "State Code",   sc)
                    frames[tab].append(df)
                    counts[tab] += len(df)
                    log_q.put(("success", f"  [{tab}]  {len(df)} rows"))
                else:
                    log_q.put(("dim", f"  [{tab}]  (empty)"))
            except Exception as exc:
                log_q.put(("error", f"  [{tab}]  ERROR - {exc}"))

    if n_files == 0:
        raise JobUserError(
            "None of the uploaded files could be parsed as GSTR-2B exports "
            "(expected filename pattern MMYYYY_SSGSTIN_GSTR2B_DDMMYYYY.xlsx) "
            "- see the log above for why each file was skipped."
        )

    log_q.put(("info", "Writing output workbook..."))
    wb = Workbook()
    wb.remove(wb.active)

    for tab in TARGET_TABS:
        ws = wb.create_sheet(title=tab)
        out_df = (pd.concat(frames[tab], ignore_index=True) if frames[tab]
                  else pd.DataFrame(columns=["State Code", "State Name", "Month & Year"]))
        _write_sheet(ws, out_df, log_q=log_q)

    output_path = Path(output_path)
    wb.save(output_path)
    log_q.put(("success", f"Saved  ->  {output_path}"))

    return {
        "files": n_files,
        "counts": counts,
        "output_path": str(output_path),
        "unresolved_state_codes": sorted(unresolved_codes),
    }
