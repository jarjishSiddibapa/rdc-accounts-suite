"""
Persistence layer for all mapping tables — now backed by the suite's shared
MySQL database (see models.py for the 6 ORM models) instead of a single
vendor-site-code-mapping.xlsx workbook read/written with a file lock.

Original xlsx sheet layout (kept only for the internal startup seed parser)
────────────────────────────────────────────────────────────────────────────
1. "Vendor Site Codes"          : Vendor Site Code | Region | Accounts Incharge
2. "Location Code Map"          : Location Code    | Location (Region) | Accounts Incharge
3. "Row Exclusions"             : Vendor Number    | Invoice Number
4. "Invoice Overrides"          : Vendor Number    | Invoice Number | Region | Accounts Incharge
5. "Region Incharge Map"        : Region           | Accounts Incharge
6. "Transaction Type Overrides" : Vendor Number    | Invoice Number | Transaction Type

Internal in-memory formats (unchanged by the DB migration — processor.py's
business logic consumes these exact shapes and needed zero changes)
────────────────────────────────────────────────────────────────────────────
vendor_mapping      : { UPPER_KEY: {"Region": str, "Accounts Incharge": str, "_key": original_case} }
loc_code_map        : { "code-str": {"Region": str, "Accounts Incharge": str} }
exclusions          : set of (vendor_number_upper, invoice_number_upper) tuples
invoice_overrides   : { (vendor_number_upper, invoice_number_upper): {"Region": str, "Accounts Incharge": str} }
region_incharge_map : { region: accounts_incharge }
txn_type_overrides  : { (vendor_number_upper, invoice_number_upper): "Transaction Type" }
"""

from pathlib import Path

import openpyxl
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.soft_delete import delete_keyed_row, seed_missing_keyed_rows, sync_keyed_rows, upsert_keyed_row

from . import paths
from .models import (
    InvoiceOverride,
    LocationCodeMap,
    RegionInchargeMap,
    RowExclusion,
    TransactionTypeOverride,
    VendorSiteMapping,
)

SHEET_VENDOR       = "Vendor Site Codes"
SHEET_LOC          = "Location Code Map"
SHEET_EXCL         = "Row Exclusions"
SHEET_INVOICE      = "Invoice Overrides"
SHEET_REGION       = "Region Incharge Map"
SHEET_TXN_OVERRIDE = "Transaction Type Overrides"


# ── helpers ────────────────────────────────────────────────────────────────

def _s(v): return str(v).strip() if v is not None else ""


# ── DB -> in-memory ──────────────────────────────────────────────────────────

def _maybe_seed(db: Session) -> None:
    """First-run only: if the vendor site mapping table is completely
    empty (a brand-new MySQL database), load the existing real mapping
    data from the seed workbook (backend/seed_data/vendor-site-code-
    mapping.xlsx) so the suite launches with zero data loss from what the
    desktop app already had on disk."""
    if db.query(VendorSiteMapping.id).first() is not None:
        return
    seed_path = paths.SEED_MAPPING_PATH
    if seed_path.exists():
        seed_missing_from_excel(db, seed_path)


def load_all(db: Session) -> tuple:
    """Return (vendor_mapping, loc_code_map, exclusions_set,
    invoice_overrides, region_incharge_map, txn_type_overrides) read from
    the shared MySQL tables. Seeds from the seed workbook on first run if
    the database is empty."""
    _maybe_seed(db)

    vendor_mapping = {}
    for row in db.query(VendorSiteMapping).filter(VendorSiteMapping.is_deleted == False).all():  # noqa: E712
        key = _s(row.vendor_site_code)
        vendor_mapping[key.upper()] = {
            "Region":            _s(row.region),
            "Accounts Incharge": _s(row.accounts_incharge),
            "_key":              key,
        }

    loc_code_map = {}
    for row in db.query(LocationCodeMap).filter(LocationCodeMap.is_deleted == False).all():  # noqa: E712
        loc_code_map[_s(row.location_code)] = {
            "Region":            _s(row.region),
            "Accounts Incharge": _s(row.accounts_incharge),
        }

    exclusions = set()
    for row in db.query(RowExclusion).filter(RowExclusion.is_deleted == False).all():  # noqa: E712
        exclusions.add((_s(row.vendor_number).upper(), _s(row.invoice_number).upper()))

    invoice_overrides = {}
    for row in db.query(InvoiceOverride).filter(InvoiceOverride.is_deleted == False).all():  # noqa: E712
        key = (_s(row.vendor_number).upper(), _s(row.invoice_number).upper())
        invoice_overrides[key] = {
            "Region":            _s(row.region),
            "Accounts Incharge": _s(row.accounts_incharge),
        }

    region_incharge_map = {}
    for row in db.query(RegionInchargeMap).filter(RegionInchargeMap.is_deleted == False).all():  # noqa: E712
        region_incharge_map[_s(row.region)] = _s(row.accounts_incharge)

    txn_type_overrides = {}
    for row in db.query(TransactionTypeOverride).filter(TransactionTypeOverride.is_deleted == False).all():  # noqa: E712
        key = (_s(row.vendor_number).upper(), _s(row.invoice_number).upper())
        txn_type_overrides[key] = _s(row.transaction_type)

    return (vendor_mapping, loc_code_map, exclusions, invoice_overrides,
            region_incharge_map, txn_type_overrides)


def load(db: Session) -> dict:
    """Backward-compat: load only vendor_mapping."""
    m, _, _, _, _, _ = load_all(db)
    return m


# ── in-memory -> DB (always rewrites all 6 tables, mirroring the original
#    xlsx save_all's always-rewrite-everything semantics) ───────────────────

def save_all(db: Session, vendor_mapping: dict, loc_code_map: dict, exclusions: set,
             invoice_overrides: dict = None, region_incharge_map: dict = None,
             txn_type_overrides: dict = None) -> None:
    """Sync all 6 mapping tables in MySQL to match these in-memory
    structures - soft-delete-aware (see app/soft_delete.py): a row missing
    from the incoming dict/set gets is_deleted=True, never actually removed;
    a row whose key reappears (even if previously soft-deleted) is revived
    and updated in place."""
    invoice_overrides   = invoice_overrides or {}
    region_incharge_map = region_incharge_map or {}
    txn_type_overrides  = txn_type_overrides or {}

    sync_keyed_rows(db, VendorSiteMapping, ("vendor_site_code",), {
        (e.get("_key", upper_key) or upper_key): {
            "region": e.get("Region", ""),
            "accounts_incharge": e.get("Accounts Incharge", ""),
        }
        for upper_key, e in vendor_mapping.items()
    })

    sync_keyed_rows(db, LocationCodeMap, ("location_code",), {
        code: {"region": e.get("Region", ""), "accounts_incharge": e.get("Accounts Incharge", "")}
        for code, e in loc_code_map.items()
    })

    sync_keyed_rows(db, RowExclusion, ("vendor_number", "invoice_number"), {
        (vnum, invno): {} for vnum, invno in exclusions
    })

    sync_keyed_rows(db, InvoiceOverride, ("vendor_number", "invoice_number"), {
        (vnum, invno): {"region": e.get("Region", ""), "accounts_incharge": e.get("Accounts Incharge", "")}
        for (vnum, invno), e in invoice_overrides.items()
    })

    sync_keyed_rows(db, RegionInchargeMap, ("region",), {
        region: {"accounts_incharge": incharge}
        for region, incharge in region_incharge_map.items()
    })

    sync_keyed_rows(db, TransactionTypeOverride, ("vendor_number", "invoice_number"), {
        (vnum, invno): {"transaction_type": txn_type}
        for (vnum, invno), txn_type in txn_type_overrides.items()
    })

    db.commit()


def save(db: Session, vendor_mapping: dict, loc_code_map: dict = None, exclusions: set = None,
         invoice_overrides: dict = None, region_incharge_map: dict = None,
         txn_type_overrides: dict = None) -> None:
    """Save vendor_mapping. Preserves the other 5 tables' current DB
    contents for any not supplied."""
    if (loc_code_map is None or exclusions is None or invoice_overrides is None
            or region_incharge_map is None or txn_type_overrides is None):
        (_, existing_loc, existing_excl, existing_ov,
         existing_rim, existing_txn) = load_all(db)
        if loc_code_map is None:
            loc_code_map = existing_loc
        if exclusions is None:
            exclusions = existing_excl
        if invoice_overrides is None:
            invoice_overrides = existing_ov
        if region_incharge_map is None:
            region_incharge_map = existing_rim
        if txn_type_overrides is None:
            txn_type_overrides = existing_txn
    save_all(db, vendor_mapping, loc_code_map, exclusions, invoice_overrides,
             region_incharge_map, txn_type_overrides)


# ── Single-row CRUD (safe under concurrent edits — see app/soft_delete.py's
#    upsert_keyed_row/delete_keyed_row) ──────────────────────────────────────
#
# save_all() above always rewrites all 6 tables from one combined in-memory
# snapshot - fixing/adding/editing/deleting a single row anywhere used to go
# through load_all() -> mutate one entry -> save_all(), which races: two
# such requests close together (completely normal usage, e.g. clearing two
# "Mapping Not Found" rows back-to-back) can each read a snapshot that
# doesn't yet include the other's write, then each write their own stale
# snapshot back over all 6 tables - whichever commits last silently
# soft-deletes every row the other one had just added. The functions below
# touch only the ONE row being changed and are what every single-row
# endpoint should call instead.

def get_region_incharge(db: Session, region: str) -> str:
    """Single-row lookup — avoids loading all 6 tables just to derive the
    Accounts Incharge for one Region."""
    row = (
        db.query(RegionInchargeMap)
        .filter(RegionInchargeMap.region == region.strip(), RegionInchargeMap.is_deleted == False)  # noqa: E712
        .first()
    )
    return row.accounts_incharge if row else ""


def upsert_vendor_site_code(db: Session, site: str, region: str, accounts_incharge: str) -> None:
    """Insert/update exactly ONE Vendor Site Code. Identity is
    case-insensitive (matching the desktop app) while storage is
    case-preserving, so this looks the row up case-insensitively rather
    than via the generic upsert_keyed_row helper."""
    site = site.strip()
    existing = (
        db.query(VendorSiteMapping)
        .filter(func.upper(VendorSiteMapping.vendor_site_code) == site.upper())
        .first()
    )
    if existing is not None:
        existing.vendor_site_code = site
        existing.region = region
        existing.accounts_incharge = accounts_incharge
        existing.is_deleted = False
    else:
        db.add(VendorSiteMapping(
            vendor_site_code=site, region=region,
            accounts_incharge=accounts_incharge, is_deleted=False,
        ))
    db.commit()


def delete_vendor_site_code(db: Session, site: str) -> bool:
    existing = (
        db.query(VendorSiteMapping)
        .filter(func.upper(VendorSiteMapping.vendor_site_code) == site.strip().upper())
        .first()
    )
    if existing is None or existing.is_deleted:
        return False
    existing.is_deleted = True
    db.commit()
    return True


def upsert_location_code(db: Session, code: str, region: str, accounts_incharge: str) -> None:
    upsert_keyed_row(db, LocationCodeMap, ("location_code",), code.strip(), {
        "region": region, "accounts_incharge": accounts_incharge,
    })
    db.commit()


def delete_location_code(db: Session, code: str) -> bool:
    result = delete_keyed_row(db, LocationCodeMap, ("location_code",), code.strip())
    db.commit()
    return result


def upsert_row_exclusion(db: Session, vendor_number: str, invoice_number: str) -> None:
    upsert_keyed_row(db, RowExclusion, ("vendor_number", "invoice_number"),
                      (vendor_number, invoice_number), {})
    db.commit()


def delete_row_exclusion(db: Session, vendor_number: str, invoice_number: str) -> bool:
    result = delete_keyed_row(db, RowExclusion, ("vendor_number", "invoice_number"),
                               (vendor_number, invoice_number))
    db.commit()
    return result


def upsert_invoice_override(db: Session, vendor_number: str, invoice_number: str,
                             region: str, accounts_incharge: str) -> None:
    upsert_keyed_row(db, InvoiceOverride, ("vendor_number", "invoice_number"),
                      (vendor_number, invoice_number),
                      {"region": region, "accounts_incharge": accounts_incharge})
    db.commit()


def delete_invoice_override(db: Session, vendor_number: str, invoice_number: str) -> bool:
    result = delete_keyed_row(db, InvoiceOverride, ("vendor_number", "invoice_number"),
                               (vendor_number, invoice_number))
    db.commit()
    return result


def upsert_region_incharge(db: Session, region: str, accounts_incharge: str) -> None:
    upsert_keyed_row(db, RegionInchargeMap, ("region",), region.strip(), {
        "accounts_incharge": accounts_incharge,
    })
    db.commit()


def delete_region_incharge(db: Session, region: str) -> bool:
    result = delete_keyed_row(db, RegionInchargeMap, ("region",), region.strip())
    db.commit()
    return result


def upsert_txn_type_override(db: Session, vendor_number: str, invoice_number: str,
                              transaction_type: str) -> None:
    upsert_keyed_row(db, TransactionTypeOverride, ("vendor_number", "invoice_number"),
                      (vendor_number, invoice_number), {"transaction_type": transaction_type})
    db.commit()


def delete_txn_type_override(db: Session, vendor_number: str, invoice_number: str) -> bool:
    result = delete_keyed_row(db, TransactionTypeOverride, ("vendor_number", "invoice_number"),
                               (vendor_number, invoice_number))
    db.commit()
    return result


# ── Internal legacy workbook parser ─────────────────────────────────────────
# The workbook is a bundled, one-way deployment seed. The centralized database
# is the sole runtime source of truth; there is no user-facing import/export.

def _parse_workbook(path) -> tuple:
    """Parse a 6-sheet vendor-site-code-mapping.xlsx workbook into the same
    6 in-memory structures load_all() returns. Tolerates a missing file or
    missing sheets gracefully (same tolerance the original file-based
    mapping_store had)."""
    path = Path(path)
    if not path.exists():
        return {}, {}, set(), {}, {}, {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    vendor_mapping = {}
    ws_name = SHEET_VENDOR if SHEET_VENDOR in wb.sheetnames else (wb.sheetnames[0] if wb.sheetnames else None)
    if ws_name:
        for row in wb[ws_name].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            key = _s(row[0])
            vendor_mapping[key.upper()] = {
                "Region":            _s(row[1]) if len(row) > 1 else "",
                "Accounts Incharge": _s(row[2]) if len(row) > 2 else "",
                "_key":              key,
            }

    loc_code_map = {}
    if SHEET_LOC in wb.sheetnames:
        for row in wb[SHEET_LOC].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            key = _s(row[0])
            loc_code_map[key] = {
                "Region":            _s(row[1]) if len(row) > 1 else "",
                "Accounts Incharge": _s(row[2]) if len(row) > 2 else "",
            }

    exclusions = set()
    if SHEET_EXCL in wb.sheetnames:
        for row in wb[SHEET_EXCL].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            vnum  = _s(row[0]).upper()
            invno = _s(row[1]).upper() if len(row) > 1 and row[1] else ""
            exclusions.add((vnum, invno))

    invoice_overrides = {}
    if SHEET_INVOICE in wb.sheetnames:
        for row in wb[SHEET_INVOICE].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            vnum  = _s(row[0]).upper()
            invno = _s(row[1]).upper() if len(row) > 1 and row[1] else ""
            invoice_overrides[(vnum, invno)] = {
                "Region":            _s(row[2]) if len(row) > 2 else "",
                "Accounts Incharge": _s(row[3]) if len(row) > 3 else "",
            }

    region_incharge_map = {}
    if SHEET_REGION in wb.sheetnames:
        for row in wb[SHEET_REGION].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            region_incharge_map[_s(row[0])] = _s(row[1]) if len(row) > 1 else ""

    txn_type_overrides = {}
    if SHEET_TXN_OVERRIDE in wb.sheetnames:
        for row in wb[SHEET_TXN_OVERRIDE].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            vnum  = _s(row[0]).upper()
            invno = _s(row[1]).upper() if len(row) > 1 and row[1] else ""
            txn_type_overrides[(vnum, invno)] = _s(row[2]) if len(row) > 2 else ""

    wb.close()
    return (vendor_mapping, loc_code_map, exclusions, invoice_overrides,
            region_incharge_map, txn_type_overrides)


def seed_missing_from_excel(db: Session, path=None) -> dict[str, int]:
    """Add legacy workbook rows that have never existed in the database.

    It never changes an existing row and never revives an archived row, so it
    is safe to run on every application startup and remains idempotent.
    """
    source = Path(path or paths.SEED_MAPPING_PATH)
    (vendor_mapping, loc_code_map, exclusions, invoice_overrides,
     region_incharge_map, txn_type_overrides) = _parse_workbook(source)

    inserted = {
        "vendor_site_codes": seed_missing_keyed_rows(
            db, VendorSiteMapping, ("vendor_site_code",), {
                (entry.get("_key", upper_key) or upper_key): {
                    "region": entry.get("Region", ""),
                    "accounts_incharge": entry.get("Accounts Incharge", ""),
                }
                for upper_key, entry in vendor_mapping.items()
            },
        ),
        "location_codes": seed_missing_keyed_rows(
            db, LocationCodeMap, ("location_code",), {
                code: {
                    "region": entry.get("Region", ""),
                    "accounts_incharge": entry.get("Accounts Incharge", ""),
                }
                for code, entry in loc_code_map.items()
            },
        ),
        "row_exclusions": seed_missing_keyed_rows(
            db, RowExclusion, ("vendor_number", "invoice_number"), {
                (vendor_number, invoice_number): {}
                for vendor_number, invoice_number in exclusions
            },
        ),
        "invoice_overrides": seed_missing_keyed_rows(
            db, InvoiceOverride, ("vendor_number", "invoice_number"), {
                (vendor_number, invoice_number): {
                    "region": entry.get("Region", ""),
                    "accounts_incharge": entry.get("Accounts Incharge", ""),
                }
                for (vendor_number, invoice_number), entry in invoice_overrides.items()
            },
        ),
        "region_incharge": seed_missing_keyed_rows(
            db, RegionInchargeMap, ("region",), {
                region: {"accounts_incharge": accounts_incharge}
                for region, accounts_incharge in region_incharge_map.items()
            },
        ),
        "transaction_type_overrides": seed_missing_keyed_rows(
            db, TransactionTypeOverride, ("vendor_number", "invoice_number"), {
                (vendor_number, invoice_number): {"transaction_type": transaction_type}
                for (vendor_number, invoice_number), transaction_type
                in txn_type_overrides.items()
            },
        ),
    }
    db.commit()
    return inserted
