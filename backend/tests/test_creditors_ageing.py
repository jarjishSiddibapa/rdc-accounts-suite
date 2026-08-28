import datetime as dt
import tempfile
import unittest
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.services.creditors_ageing import mapping_store, processor
from app.services.creditors_ageing.models import VendorMapping


class CreditorsAgeingTests(unittest.TestCase):
    def _write_tally_export(self, path: Path) -> None:
        workbook = openpyxl.Workbook()
        tb = workbook.active
        tb.title = "Sheet1"
        tb.append(["Trial Balance 1-Apr-26 to 28-Aug-26"])
        tb.append(["Particulars", "Debit", "Credit"])
        tb.append(["Vendor A", 0, 1000])
        tb.append(["Vendor B", 200, 0])

        bills = workbook.create_sheet("Sheet2")
        bills.append(["Pending bills 1-Apr-26 to 28-Aug-26"])
        bills.append([
            "Date", "Ref. No.", "Party's Name", "Cost Center/Location",
            "Opening Amount", "Pending Amount", "Due On",
        ])
        bills.append([dt.date(2026, 8, 1), "A-1", "Vendor A", "Goa", 600, 600, None])
        bills.append([dt.date(2026, 6, 1), "A-2", "Vendor A", "Goa", 400, 400, None])
        bills.append([dt.date(2026, 8, 1), "B-1", "Vendor B", "Wada", 200, 200, None])
        bills["F3"].number_format = '0.00 "Cr"'
        bills["F4"].number_format = '0.00 "Cr"'
        bills["F5"].number_format = '0.00 "Dr"'
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _find_vendor_row(sheet, vendor: str) -> int:
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row, 2).value == vendor:
                return row
        raise AssertionError(f"Vendor {vendor!r} not found")

    def test_report_preserves_formulas_and_injects_visible_cached_values(self):
        mapping = {
            "VENDOR A": {"name": "Vendor A", "loc": "Goa", "vt": "Expenses", "vst": "Testing", "intercompany": False},
            "VENDOR B": {"name": "Vendor B", "loc": "Wada", "vt": "Advance", "vst": "Regular", "intercompany": False},
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "raw.xlsx"
            output_path = Path(temp_dir) / "report.xlsx"
            csv_path = Path(temp_dir) / "new-vendors.csv"
            self._write_tally_export(input_path)

            result = processor.generate_report(
                str(input_path),
                str(output_path),
                mapping,
                as_on_date="2026-08-28",
                new_vendors_csv_path=str(csv_path),
            )

            self.assertEqual(result["as_on_date"], "2026-08-28")
            self.assertEqual(result["ageing_through_date"], "2026-08-27")
            self.assertEqual(result["counts"], {"only_creditors": 1, "advances": 1, "intercompany": 0})
            self.assertEqual(result["new_vendors"], [])
            self.assertIsNone(result["new_vendors_csv_path"])

            formulas = openpyxl.load_workbook(output_path, data_only=False)
            values = openpyxl.load_workbook(output_path, data_only=True)
            try:
                row = self._find_vendor_row(formulas["Only Creditors"], "Vendor A")
                self.assertTrue(str(formulas["Only Creditors"].cell(row, 6).value).startswith("=IFERROR(SUMIFS"))
                self.assertEqual(values["Only Creditors"].cell(row, 6).value, 1000)
                self.assertEqual(values["Only Creditors"].cell(row, 7).value, 600)
                self.assertEqual(values["Only Creditors"].cell(row, 9).value, 400)
                self.assertEqual(values["Only Creditors"].cell(row, 15).value, 1000)
                self.assertEqual(values["Only Creditors"].cell(row, 16).value, 0)
                self.assertIn("As on 28-August-2026", formulas["Only Creditors"]["A2"].value)

                advance_row = self._find_vendor_row(formulas["Advances"], "Vendor B")
                self.assertEqual(values["Advances"].cell(advance_row, 6).value, -200)
                self.assertEqual(values["Advances"].cell(advance_row, 7).value, -200)
            finally:
                formulas.close()
                values.close()

    def test_invalid_workbook_has_clear_user_error(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invalid.xlsx"
            openpyxl.Workbook().save(path)
            with self.assertRaisesRegex(processor.AgeingReportError, "Could not identify"):
                processor.read_tb_export(str(path))


class VendorMappingStoreTests(unittest.TestCase):
    """Direct coverage of mapping_store.upsert_mapping's rename/duplicate/
    archived-conflict handling — previously exercised only indirectly
    through the router, with no unit test of its own anywhere in the suite."""

    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(bind=self.engine, tables=[VendorMapping.__table__])
        self.Session = sessionmaker(bind=self.engine, expire_on_commit=False)

    def tearDown(self):
        self.engine.dispose()

    def test_add_new_mapping(self):
        with self.Session() as db:
            mapping_store.upsert_mapping(
                db, original_name=None, vendor_name="Acme Traders",
                location="Goa", vendor_type="Expenses", vendor_sub_type="Freight",
                intercompany=False,
            )
            row = db.query(VendorMapping).filter_by(vendor_key="ACME TRADERS").one()
            self.assertEqual(row.location, "Goa")
            self.assertFalse(row.is_deleted)

    def test_edit_without_rename_updates_the_same_row(self):
        with self.Session() as db:
            mapping_store.upsert_mapping(
                db, original_name=None, vendor_name="Acme Traders",
                location="Goa", vendor_type="", vendor_sub_type="", intercompany=False,
            )
            mapping_store.upsert_mapping(
                db, original_name="Acme Traders", vendor_name="Acme Traders",
                location="Wada", vendor_type="Expenses", vendor_sub_type="", intercompany=False,
            )
            rows = db.query(VendorMapping).filter_by(vendor_key="ACME TRADERS").all()
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0].location, "Wada")

    def test_rename_archives_the_old_key_and_activates_the_new_one(self):
        with self.Session() as db:
            mapping_store.upsert_mapping(
                db, original_name=None, vendor_name="Acme Traders",
                location="Goa", vendor_type="", vendor_sub_type="", intercompany=False,
            )
            mapping_store.upsert_mapping(
                db, original_name="Acme Traders", vendor_name="Acme Traders Pvt Ltd",
                location="Goa", vendor_type="", vendor_sub_type="", intercompany=False,
            )
            old = db.query(VendorMapping).filter_by(vendor_key="ACME TRADERS").one()
            new = db.query(VendorMapping).filter_by(vendor_key="ACME TRADERS PVT LTD").one()
            self.assertTrue(old.is_deleted)
            self.assertFalse(new.is_deleted)

    def test_rename_onto_an_existing_active_vendor_is_rejected(self):
        with self.Session() as db:
            mapping_store.upsert_mapping(
                db, original_name=None, vendor_name="Acme Traders",
                location="Goa", vendor_type="", vendor_sub_type="", intercompany=False,
            )
            mapping_store.upsert_mapping(
                db, original_name=None, vendor_name="Beta Traders",
                location="Wada", vendor_type="", vendor_sub_type="", intercompany=False,
            )
            with self.assertRaises(mapping_store.DuplicateMappingError):
                mapping_store.upsert_mapping(
                    db, original_name="Acme Traders", vendor_name="Beta Traders",
                    location="Goa", vendor_type="", vendor_sub_type="", intercompany=False,
                )
            # Neither row should have moved.
            self.assertFalse(db.query(VendorMapping).filter_by(vendor_key="ACME TRADERS").one().is_deleted)
            self.assertEqual(db.query(VendorMapping).filter_by(vendor_key="BETA TRADERS").one().location, "Wada")

    def test_rename_onto_an_archived_vendor_is_rejected(self):
        with self.Session() as db:
            mapping_store.upsert_mapping(
                db, original_name=None, vendor_name="Acme Traders",
                location="Goa", vendor_type="", vendor_sub_type="", intercompany=False,
            )
            mapping_store.upsert_mapping(
                db, original_name=None, vendor_name="Old Vendor",
                location="Goa", vendor_type="", vendor_sub_type="", intercompany=False,
            )
            mapping_store.archive_mapping(db, "Old Vendor")
            with self.assertRaises(mapping_store.ArchivedMappingError):
                mapping_store.upsert_mapping(
                    db, original_name="Acme Traders", vendor_name="Old Vendor",
                    location="Goa", vendor_type="", vendor_sub_type="", intercompany=False,
                )

    def test_adding_a_name_that_matches_an_archived_row_is_rejected(self):
        with self.Session() as db:
            mapping_store.upsert_mapping(
                db, original_name=None, vendor_name="Old Vendor",
                location="Goa", vendor_type="", vendor_sub_type="", intercompany=False,
            )
            mapping_store.archive_mapping(db, "Old Vendor")
            with self.assertRaises(mapping_store.ArchivedMappingError):
                mapping_store.upsert_mapping(
                    db, original_name=None, vendor_name="Old Vendor",
                    location="Wada", vendor_type="", vendor_sub_type="", intercompany=False,
                )

    def test_archive_then_restore_round_trip(self):
        with self.Session() as db:
            mapping_store.upsert_mapping(
                db, original_name=None, vendor_name="Acme Traders",
                location="Goa", vendor_type="", vendor_sub_type="", intercompany=True,
            )
            self.assertTrue(mapping_store.archive_mapping(db, "Acme Traders"))
            self.assertEqual(mapping_store.list_rows(db), [])
            self.assertEqual(len(mapping_store.list_rows(db, archived=True)), 1)

            self.assertTrue(mapping_store.restore_mapping(db, "Acme Traders"))
            active = mapping_store.list_rows(db)
            self.assertEqual(len(active), 1)
            self.assertEqual(active[0]["intercompany"], "Yes")

    def test_archiving_an_unknown_vendor_is_a_no_op(self):
        with self.Session() as db:
            self.assertFalse(mapping_store.archive_mapping(db, "Nobody"))
            self.assertFalse(mapping_store.restore_mapping(db, "Nobody"))


if __name__ == "__main__":
    unittest.main()

