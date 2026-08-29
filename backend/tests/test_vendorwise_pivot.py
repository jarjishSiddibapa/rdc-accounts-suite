"""Regression coverage for the real Vendorwise Pivot restoration.

Context: the source desktop app (E:\\jarjish-projects\\sneha-raman-
unaccounted-transactions-report\\excel_writers.py) builds MRN's Vendorwise
Pivot as a genuine, refreshable Excel PivotTable via live win32com
automation - the ONLY sheet across Unaccounted/MRN/PO it ever does this for.
A prior, now-reverted attempt at porting this ("native_pivots.py") only set
the pivot's cache pointer and refresh flags, never the actual cache records
or displayed grid - so any reader that doesn't trigger a live Excel refresh
(the suite's own mail-body HTML preview, or Excel's Protected View on a
downloaded/emailed file) showed the pivot template's placeholder seed data
forever. That shipped a real email with fake "SYNTHETIC"/"Jan-00" data.

vendorwise_pivot.py fixes this by never touching the sheet's own cells (kept
exactly as the already-correct static builder writes them) and instead
authoring the pivotCacheDefinition/pivotCacheRecords/pivotTableDefinition
parts as a pure additive OOXML overlay, so every reader - openpyxl, the
mail-body preview, Protected View, or live Excel - sees correct data
immediately, with no dependency on Excel ever running.

TestPivotStructureIsCorrect runs everywhere (no Excel required) and checks
the file openpyxl gets back is well-formed and points at the right data.
TestRealExcelRoundTrip is skipped unless pywin32 + a real Excel install are
available (this dev machine only, never CI/production) - it opens the file
in actual Excel, confirms no repair prompt, and confirms a live refresh
reaches the exact same numbers already shown, proving the cache and the
pre-written grid genuinely agree.
"""
import os
import re
import tempfile
import unittest
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd

from app.services import mailer_shared
from app.services.unaccounted import excel_writers

try:
    import pythoncom
    import win32com.client as win32

    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False


def _sample_mrn_df() -> pd.DataFrame:
    return pd.DataFrame({
        "SUPPLIER SITE": ["SITE-A1", "SITE-B1", "SITE-A1", "SITE-D1"],
        "SUPPLIER NAME": ["Acme Traders", "Bharat Steel", "Acme Traders", "Delta Fabricators"],
        "ACCOUNTING PERIOD": ["JUL-2026", "JUL-2026", "AUG-2026", "AUG-2026"],
        "Location": ["Mumbai Plant", "Chennai Plant", "Mumbai Plant", "Pune Plant"],
        "Accounts Incharge": ["Rakesh D", "Priya S", "Rakesh D", "Sneha R"],
        "MRN Number": ["MRN-1", "MRN-2", "MRN-3", "MRN-4"],
    })


class VendorwisePivotStructureTests(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.out = str(Path(self._tmp.name) / "mrn.xlsx")
        excel_writers.write_formatted_mrn_excel(_sample_mrn_df(), self.out)

    def test_a_genuine_pivottable_is_attached(self):
        wb = openpyxl.load_workbook(self.out, data_only=False)
        pivots = wb["Vendorwise Pivot"]._pivots
        self.assertEqual(len(pivots), 1)
        pivot = pivots[0]
        self.assertEqual(pivot.name, "VendorwisePivot")
        self.assertEqual(pivot.cache.cacheSource.worksheetSource.sheet, "Summary")
        self.assertTrue(pivot.cache.refreshOnLoad)

    def test_pivot_location_matches_the_written_grid_exactly(self):
        wb = openpyxl.load_workbook(self.out, data_only=False)
        ws = wb["Vendorwise Pivot"]
        pivot = ws._pivots[0]
        # 3 suppliers -> rows 5,6,7 + grand total row 8; 2 periods -> cols B,C + Grand Total D
        self.assertEqual(pivot.location.ref, "A3:D8")

    def test_mail_body_preview_shows_real_data_not_placeholder(self):
        html = mailer_shared.sheet_to_html(self.out, "Vendorwise Pivot")
        self.assertNotIn("SYNTHETIC", html)
        self.assertNotIn("Synthetic placeholder", html)
        self.assertIn("Acme Traders", html)
        self.assertIn(">2<", html)  # Acme Traders: Jul-26 + Aug-26 = 2
        self.assertIn("Grand Total", html)
        self.assertIn(">4<", html)

    def test_falls_back_silently_to_the_static_sheet_if_pivot_attach_fails(self):
        from app.services.unaccounted import vendorwise_pivot

        original = vendorwise_pivot.attach_vendorwise_pivot

        def _boom(*args, **kwargs):
            raise vendorwise_pivot.VendorwisePivotError("simulated failure")

        out = str(Path(self._tmp.name) / "mrn_fallback.xlsx")
        vendorwise_pivot.attach_vendorwise_pivot = _boom
        try:
            excel_writers.write_formatted_mrn_excel(_sample_mrn_df(), out)  # must not raise
        finally:
            vendorwise_pivot.attach_vendorwise_pivot = original

        wb = openpyxl.load_workbook(out, data_only=False)
        self.assertEqual(wb["Vendorwise Pivot"]._pivots, [])  # no pivot, but...
        html = mailer_shared.sheet_to_html(out, "Vendorwise Pivot")
        self.assertIn("Acme Traders", html)
        self.assertIn(">2<", html)  # ...the static data is still correct

    def test_handles_ampersands_and_quotes_in_supplier_names_without_corrupting_the_xml(self):
        df = _sample_mrn_df()
        df.loc[0, "SUPPLIER NAME"] = 'Acme & "Sons" Traders'
        df.loc[2, "SUPPLIER NAME"] = 'Acme & "Sons" Traders'
        out = str(Path(self._tmp.name) / "mrn_special_chars.xlsx")
        excel_writers.write_formatted_mrn_excel(df, out)

        wb = openpyxl.load_workbook(out, data_only=False)
        self.assertEqual(len(wb["Vendorwise Pivot"]._pivots), 1)
        html = mailer_shared.sheet_to_html(out, "Vendorwise Pivot")
        self.assertIn('Acme & "Sons" Traders', html)

    def test_supplier_used_from_multiple_locations_collapses_to_one_row(self):
        """Regression test for a real production bug: a supplier ordering
        from several different locations (e.g. a marketplace vendor like
        Amazon) used to get one row per (Location, Supplier) pair instead of
        one aggregated row - which also duplicated its name in the real
        PivotTable's row-field cache, and real Excel then repaired the file
        by stripping the PivotTable entirely, leaving exactly the flawed flat
        table behind (reproduced against the actual generated report - see
        vendorwise_pivot.py's module docstring)."""
        df = pd.DataFrame({
            "SUPPLIER SITE": [f"SITE-{i}" for i in range(1, 6)],
            "SUPPLIER NAME": ["Amazon Wholesale (India) Private Limited"] * 4 + ["Bharat Steel"],
            "ACCOUNTING PERIOD": ["JUL-2026", "JUL-2026", "AUG-2026", "AUG-2026", "JUL-2026"],
            "Location": ["Mumbai Plant", "Chennai Plant", "Pune Plant", "Delhi Plant", "Chennai Plant"],
            "Accounts Incharge": ["Rakesh D", "Priya S", "Sneha R", "Divya M", "Priya S"],
            "MRN Number": [f"MRN-{i}" for i in range(1, 6)],
        })
        out = str(Path(self._tmp.name) / "mrn_multi_location_supplier.xlsx")
        pivot_df, _ = excel_writers._add_vendorwise_pivot_sheet(openpyxl.Workbook(), df)
        rows = pivot_df[pivot_df["SUPPLIER NAME"] == "Amazon Wholesale (India) Private Limited"]
        self.assertEqual(len(rows), 1, "one supplier across 4 locations must collapse to one row")
        self.assertEqual(rows.iloc[0]["Grand Total"], 4)

        excel_writers.write_formatted_mrn_excel(df, out)
        wb = openpyxl.load_workbook(out, data_only=False)
        self.assertEqual(len(wb["Vendorwise Pivot"]._pivots), 1)

        with zipfile.ZipFile(out) as zf:
            cache_xml = zf.read("xl/pivotCache/pivotCacheDefinition1.xml").decode("utf-8")
        supplier_names = re.findall(
            r'<cacheField name="SUPPLIER NAME"[^>]*><sharedItems count="\d+">(.*?)</sharedItems>',
            cache_xml, re.DOTALL,
        )[0]
        items = re.findall(r'<s v="(.*?)"/>', supplier_names)
        self.assertEqual(
            len(items), len(set(items)),
            "the pivot cache's SUPPLIER NAME axis must contain only distinct values",
        )


@unittest.skipUnless(_EXCEL_AVAILABLE, "pywin32/Excel not available - this check is dev-machine-only")
class VendorwisePivotRealExcelRoundTripTests(unittest.TestCase):
    """Opens the generated file in real Excel to confirm it is not
    corrupted and that a live refresh reaches the same numbers already
    shown. This never runs in production - it exists purely to catch OOXML
    mistakes that only real Excel would notice."""

    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)

    def _open_refresh_and_read(self, path: str):
        pythoncom.CoInitialize()
        xl = win32.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        try:
            wb = xl.Workbooks.Open(os.path.abspath(path), CorruptLoad=0)
            ws = wb.Sheets("Vendorwise Pivot")
            pivots = ws.PivotTables()
            self.assertEqual(pivots.Count, 1)
            pt = ws.PivotTables(1)

            used = ws.UsedRange
            last_row = used.Row + used.Rows.Count - 1
            last_col = used.Column + used.Columns.Count - 1
            full_ref = f"B4:{openpyxl.utils.get_column_letter(last_col)}{last_row}"
            before = ws.Range(full_ref).Value

            pt.RefreshTable()
            xl.CalculateFull()
            after = ws.Range(full_ref).Value

            wb.Save()  # raises if Excel considers the result invalid
            return before, after
        finally:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                xl.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()

    def test_opens_without_repair_and_refresh_matches_the_pre_written_grid(self):
        out = str(Path(self._tmp.name) / "mrn.xlsx")
        excel_writers.write_formatted_mrn_excel(_sample_mrn_df(), out)

        before, after = self._open_refresh_and_read(out)
        self.assertEqual(before, after)
        # Grand Total row's Grand Total cell = total row count (4)
        self.assertEqual(after[-1][-1], 4.0)

    def test_larger_dataset_with_unmapped_locations_and_special_characters(self):
        df = pd.DataFrame({
            "SUPPLIER SITE": [f"SITE-{i}" for i in range(1, 9)],
            "SUPPLIER NAME": [
                "Acme & Sons Traders", 'Bharat "Steel" Co.', "Coastal Logistics", "Delta Fabricators",
                "Acme & Sons Traders", "Everest Minerals", 'Bharat "Steel" Co.', "Ganges Textiles",
            ],
            "ACCOUNTING PERIOD": ["JUL-2026", "JUL-2026", "JUL-2026", "AUG-2026",
                                   "AUG-2026", "AUG-2026", "SEP-2026", "SEP-2026"],
            "Location": ["Mumbai Plant", "Chennai Plant", "", "Pune Plant",
                         "Mumbai Plant", "Delhi Plant", "Chennai Plant", "Kolkata Plant"],
            "Accounts Incharge": ["Rakesh D", "Priya S", "Anil K", "Sneha R",
                                   "Rakesh D", "Divya M", "Priya S", "Vikram T"],
            "MRN Number": [f"MRN-{i}" for i in range(1, 9)],
        })
        out = str(Path(self._tmp.name) / "mrn_stress.xlsx")
        excel_writers.write_formatted_mrn_excel(df, out)

        before, after = self._open_refresh_and_read(out)
        self.assertEqual(before, after)
        self.assertEqual(after[-1][-1], 8.0)  # Grand Total of all 8 rows

    def test_supplier_spanning_many_locations_opens_without_repair(self):
        """The real-Excel counterpart to
        test_supplier_used_from_multiple_locations_collapses_to_one_row: a
        supplier repeated across many distinct locations used to produce
        duplicate <s v="..."/> entries in the pivot cache's SUPPLIER NAME
        sharedItems, which real Excel considers invalid and silently repairs
        by stripping the PivotTable - _open_refresh_and_read's own
        `pivots.Count == 1` assertion is what catches that regression, since
        a repaired-away PivotTable would make that count 0."""
        df = pd.DataFrame({
            "SUPPLIER SITE": [f"SITE-{i}" for i in range(1, 18)],
            "SUPPLIER NAME": ["Amazon Wholesale (India) Private Limited"] * 17,
            "ACCOUNTING PERIOD": (["JUL-2026"] * 6 + ["AUG-2026"] * 6 + ["SEP-2026"] * 5),
            "Location": [f"Location-{i}" for i in range(1, 18)],
            "Accounts Incharge": ["Rakesh D"] * 17,
            "MRN Number": [f"MRN-{i}" for i in range(1, 18)],
        })
        out = str(Path(self._tmp.name) / "mrn_many_locations.xlsx")
        excel_writers.write_formatted_mrn_excel(df, out)

        before, after = self._open_refresh_and_read(out)
        self.assertEqual(before, after)
        self.assertEqual(after[-1][-1], 17.0)  # Grand Total: all 17 rows, one supplier


if __name__ == "__main__":
    unittest.main()
