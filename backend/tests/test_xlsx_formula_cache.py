import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.xlsx_formula_cache import cache_formula_values, inject_cached_values


class XlsxFormulaCacheTests(unittest.TestCase):
    def test_subtotal_and_sum_get_a_real_cached_value_formula_stays_live(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Main"
            ws.append(["Location", "Amount"])
            ws.append(["West", 10])
            ws.append(["East", 5])
            ws.append(["Grand Total", "=SUBTOTAL(9,B2:B3)"])
            ws.append(["Sum Row", "=SUM(B2:B3)"])

            cached = cache_formula_values(wb)
            wb.save(path)
            inject_cached_values(str(path), cached)

            wb_f = load_workbook(path, data_only=False)
            self.assertEqual(wb_f["Main"]["B4"].value, "=SUBTOTAL(9,B2:B3)")
            self.assertEqual(wb_f["Main"]["B5"].value, "=SUM(B2:B3)")
            wb_f.close()

            wb_v = load_workbook(path, data_only=True)
            self.assertEqual(wb_v["Main"]["B4"].value, 15)
            self.assertEqual(wb_v["Main"]["B5"].value, 15)
            wb_v.close()

    def test_sum_supports_comma_separated_non_adjacent_ranges_and_cells(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Main"
            # A2=1, B2=2, C2=99 (skipped), D2=4, E2=5 - a row total that
            # skips a non-adjacent column, matching the shape of e.g. RDC
            # Payables' per-row "Total LE30" column, which sums only every
            # other transaction-type block.
            ws.append([1, 2, 99, 4, 5])
            ws["G1"] = "=SUM(A1,B1:B1,D1:E1)"

            cached = cache_formula_values(wb)
            wb.save(path)
            inject_cached_values(str(path), cached)

            wb_f = load_workbook(path, data_only=False)
            self.assertEqual(wb_f["Main"]["G1"].value, "=SUM(A1,B1:B1,D1:E1)")
            wb_f.close()

            wb_v = load_workbook(path, data_only=True)
            self.assertEqual(wb_v["Main"]["G1"].value, 1 + 2 + 4 + 5)
            wb_v.close()

    def test_grand_total_excludes_nested_subtotal_like_excel_does(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Main"
            ws.append(["Location", "Amount"])
            ws.append(["West-1", 10])
            ws.append(["West-2", 20])
            ws.append(["West Total", "=SUBTOTAL(9,B2:B3)"])
            ws.append(["East", 5])
            ws.append(["Grand Total", "=SUBTOTAL(9,B2:B5)"])

            cached = cache_formula_values(wb)
            wb.save(path)
            inject_cached_values(str(path), cached)

            wb_v = load_workbook(path, data_only=True)
            self.assertEqual(wb_v["Main"]["B4"].value, 30)   # West Total: 10 + 20
            self.assertEqual(wb_v["Main"]["B6"].value, 35)   # Grand Total: 10+20+5, NOT 30+5+30=65
            wb_v.close()

    def test_multiple_sheets_each_get_their_own_cells_cached(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "report.xlsx"
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Main"
            ws1.append(["A"])
            ws1.append([1])
            ws1.append(["=SUM(A1:A2)"])
            ws2 = wb.create_sheet("Summary")
            ws2.append(["X", "Y"])
            ws2.append([3, 4])
            ws2.append(["=SUM(A2:B2)"])

            cached = cache_formula_values(wb)
            wb.save(path)
            inject_cached_values(str(path), cached)

            wb_v = load_workbook(path, data_only=True)
            self.assertEqual(wb_v["Main"]["A3"].value, 1)  # "A" (text) + 1 -> non-numeric term skipped
            self.assertEqual(wb_v["Summary"]["A3"].value, 7)
            wb_v.close()

    def test_file_remains_structurally_valid_after_injection(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Main"
            for i in range(1, 21):
                ws.append([f"Row {i}", i])
            ws.append(["Grand Total", "=SUBTOTAL(9,B1:B20)"])

            cached = cache_formula_values(wb)
            wb.save(path)
            inject_cached_values(str(path), cached)

            # A clean re-open + re-save is a strong structural-validity signal.
            wb2 = load_workbook(path)
            resave_path = Path(temp_dir) / "resaved.xlsx"
            wb2.save(resave_path)
            wb2.close()
            self.assertTrue(resave_path.is_file())


if __name__ == "__main__":
    unittest.main()
