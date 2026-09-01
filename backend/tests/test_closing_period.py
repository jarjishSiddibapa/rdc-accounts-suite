import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.jobs import JobUserError
from app.routers.closing_period import _run_combine_job
from app.services.closing_period.combiner import run_combine


class _LogQueue:
    def __init__(self):
        self.messages = []

    def put(self, item):
        self.messages.append(item)


def _report_html(location: str, rows: list[tuple[str, str, float, float]]) -> str:
    body = []
    for sub_inventory, item, quantity, value in rows:
        body.append(
            "<tr>"
            f"<td>ORG</td><td>{sub_inventory}</td><td><span>{item}</span></td>"
            "<td>Description</td><td>EA</td>"
            f"<td>{quantity}</td><td>0</td><td>0</td>"
            f"<td>{value}</td><td>0</td><td>0</td>"
            "</tr>"
        )
    return (
        "<html><body><table>"
        f"<tr><td>{location}</td></tr>"
        "<tr><td>metadata</td></tr>"
        "<tr><th>Organization</th><th>Sub Inventory</th><th>Item</th>"
        "<th>Description</th><th>UOM</th><th>31-MAY-26 Quantity</th>"
        "<th>Current Quantity</th><th>Change Quantity</th>"
        "<th>31-MAY-26 Value</th><th>Current Value</th><th>Change Value</th></tr>"
        + "".join(body)
        + "</table></body></html>"
    )


class ClosingPeriodTests(unittest.TestCase):
    def test_combine_preserves_first_data_row_and_caches_summary_formulas(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first = root / "A1.xls"
            second = root / "B1.xls"
            output = root / "combined.xlsx"
            first.write_text(
                _report_html(
                    "Location A",
                    [("MOD-RM", "FIRST-ITEM", 2, 20), ("IGNORED", "DROP-ME", 9, 90)],
                ),
                encoding="utf-8",
            )
            second.write_text(
                _report_html("Location B", [("STORES", "SECOND-ITEM", 3, 45)]),
                encoding="utf-8",
            )

            result = run_combine(
                [(first.name, str(first)), (second.name, str(second))],
                str(output),
                _LogQueue(),
            )

            self.assertEqual(result["files"], 2)
            self.assertEqual(result["total_rows"], 2)
            formulas = load_workbook(output, data_only=False)
            try:
                main = formulas["Report 31-MAY-26"]
                self.assertEqual(main["A2"].value, "File Name")
                self.assertEqual(main["D3"].value, "FIRST-ITEM")
                self.assertEqual(main["D4"].value, "SECOND-ITEM")
                self.assertTrue(str(formulas["Summary"]["C2"].value).startswith("=SUMIFS("))
                self.assertTrue(str(formulas["Summary"]["C4"].value).startswith("=SUBTOTAL("))
            finally:
                formulas.close()

            values = load_workbook(output, data_only=True)
            try:
                summary = values["Summary"]
                self.assertEqual(summary["C2"].value, 2)
                self.assertEqual(summary["D2"].value, 20)
                self.assertEqual(summary["C3"].value, 3)
                self.assertEqual(summary["D3"].value, 45)
                self.assertEqual(summary["C4"].value, 5)
                self.assertEqual(summary["D4"].value, 65)
            finally:
                values.close()

    def test_a_large_combine_logs_periodic_write_progress(self):
        """Same gap class as the ERP converter before its fix: the main
        data-row write loop used to run silently between the "Processing"
        and "Writing output workbook..." log lines, with nothing in
        between on a report with many rows."""
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            path = root / "A1.xls"
            output = root / "combined.xlsx"
            subinvs = ["MOD-RM", "NMOD-RM", "STORES"]
            rows = [(subinvs[i % 3], f"ITEM-{i}", 1, 10) for i in range(300)]
            path.write_text(_report_html("Location A", rows), encoding="utf-8")

            log_q = _LogQueue()
            run_combine([(path.name, str(path))], str(output), log_q)

            progress_messages = [m for _tag, m in log_q.messages if "Writing row" in m]
            self.assertGreater(len(progress_messages), 1)
            self.assertIn("300/300", progress_messages[-1])

    def test_background_job_cleans_saved_upload_when_input_is_invalid(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            upload = root / "bad.xls"
            output = root / "combined.xlsx"
            upload.write_text("this is not an HTML table", encoding="utf-8")

            with self.assertRaises(JobUserError):
                _run_combine_job([(upload.name, str(upload))], str(output))

            self.assertFalse(upload.exists())
            self.assertFalse(output.exists())


if __name__ == "__main__":
    unittest.main()
