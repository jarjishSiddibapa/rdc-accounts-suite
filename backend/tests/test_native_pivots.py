import tempfile
import unittest
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from app.services.unaccounted.excel_writers import write_formatted_mrn_excel
from app.services.unaccounted.native_pivots import (
    _TEMPLATE_PATH,
    _TWO_LEVEL_TEMPLATE_PATH,
    attach_two_level_pivot,
)


MRN_COLUMNS = [
    "ACCOUNTING PERIOD",
    "SUPPLIER NUMBER",
    "SUPPLIER NAME",
    "ORGANIZATION NAME",
    "SUPPLIER SITE",
    "Location",
    "Accounts Incharge",
    "PO NUMBER",
    "ITEM CODE",
    "LINE DESCRIPTION",
    "GL ACCOUNT",
    "RECEIPT NUMBER",
    "Expected Invoice Number (From MRN)",
    "RECEIPT DATE",
    "UOM",
    "RECEIPT QUANTITY",
    "BASE AMOUNT",
    "ITEM GST ENABLED FLAG",
    "CGST",
    "SGST",
    "IGST",
    "STATUS",
]


def _row(period: str, supplier: str, site: str, location: str, owner: str, number: int):
    row = {column: "" for column in MRN_COLUMNS}
    row.update(
        {
            "ACCOUNTING PERIOD": period,
            "SUPPLIER NUMBER": f"SUP-{number}",
            "SUPPLIER NAME": supplier,
            "ORGANIZATION NAME": "Test Organization",
            "SUPPLIER SITE": site,
            "Location": location,
            "Accounts Incharge": owner,
            "PO NUMBER": f"PO-{number}",
            "ITEM CODE": f"ITEM-{number}",
            "LINE DESCRIPTION": "Test row",
            "GL ACCOUNT": "1000",
            "RECEIPT NUMBER": f"MRN-{number}",
            "RECEIPT DATE": datetime(2026, 7, number),
            "UOM": "EA",
            "RECEIPT QUANTITY": number,
            "BASE AMOUNT": 100 * number,
            "ITEM GST ENABLED FLAG": "Y",
            "CGST": 0,
            "SGST": 0,
            "IGST": 18 * number,
            "STATUS": "",
        }
    )
    return row


class NativePivotTests(unittest.TestCase):
    def test_template_is_sanitized_and_contains_both_native_pivots(self):
        self.assertTrue(_TEMPLATE_PATH.is_file())
        with zipfile.ZipFile(_TEMPLATE_PATH) as archive:
            pivot_parts = [name for name in archive.namelist() if "pivot" in name.lower()]
            xml_payload = b"\n".join(
                archive.read(name)
                for name in archive.namelist()
                if name.endswith((".xml", ".rels"))
            )
        self.assertIn("xl/pivotTables/pivotTable1.xml", pivot_parts)
        self.assertIn("xl/pivotTables/pivotTable2.xml", pivot_parts)
        self.assertIn("xl/pivotCache/pivotCacheDefinition1.xml", pivot_parts)
        self.assertIn("xl/pivotCache/pivotCacheDefinition2.xml", pivot_parts)
        self.assertNotIn(b"Jarjish", xml_payload)
        self.assertNotIn(b"Sidhabappa", xml_payload)

        workbook = load_workbook(_TEMPLATE_PATH)
        try:
            for sheet_name in ("Vendorwise Pivot", "Locationwise Pivot"):
                pivots = workbook[sheet_name]._pivots
                self.assertEqual(len(pivots), 1)
                self.assertTrue(pivots[0].cache.refreshOnLoad)
                self.assertEqual(pivots[0].cache.missingItemsLimit, 0)
        finally:
            workbook.close()

    def test_generic_two_level_template_is_sanitized_and_attachable(self):
        self.assertTrue(_TWO_LEVEL_TEMPLATE_PATH.is_file())
        with zipfile.ZipFile(_TWO_LEVEL_TEMPLATE_PATH) as archive:
            names = set(archive.namelist())
            xml_payload = b"\n".join(
                archive.read(name)
                for name in names
                if name.endswith((".xml", ".rels"))
            )
        self.assertIn("xl/pivotTables/pivotTable1.xml", names)
        self.assertIn("xl/pivotCache/pivotCacheDefinition1.xml", names)
        self.assertNotIn(b"Jarjish", xml_payload)
        self.assertNotIn(b"Sidhabappa", xml_payload)

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "report.xlsx"
            workbook = Workbook()
            workbook.active.title = "Main"
            workbook["Main"]["A1"] = "Static fallback"
            workbook.create_sheet("Details")["A1"] = "Preserved report sheet"
            workbook.save(path)

            rows = [
                ("CHENNAI", "Analyst One", "Jun-26", 1),
                ("CHENNAI", "Analyst One", "Jul-26", 2),
                ("PUNE", "Analyst Two", "Jul-26", 3),
            ]
            attach_two_level_pivot(
                path,
                rows=rows,
                target_sheet="Main",
                sheet_position=0,
                tab_color="1F3864",
                dim1_caption="Location",
                dim2_caption="Accounts Incharge",
                period_caption="Month",
                value_caption="Count",
            )

            output = load_workbook(path, data_only=False)
            try:
                self.assertEqual(
                    [name for name in output.sheetnames if output[name].sheet_state == "visible"],
                    ["Main", "Details"],
                )
                self.assertEqual(output["_PivotSrc"].sheet_state, "hidden")
                pivot = output["Main"]._pivots[0]
                source = pivot.cache.cacheSource.worksheetSource
                self.assertEqual(source.sheet, "_PivotSrc")
                self.assertEqual(source.ref, "A1:D4")
                self.assertTrue(pivot.cache.refreshOnLoad)
                self.assertEqual(
                    [field.name for field in pivot.pivotFields[:3]],
                    ["Location", "Accounts Incharge", "Month"],
                )
                self.assertEqual(pivot.dataFields[0].name, "Count")
                self.assertEqual(
                    set(tuple(cell.value for cell in row) for row in output["_PivotSrc"].iter_rows(min_row=2)),
                    set(rows),
                )
            finally:
                output.close()

    def test_pending_mrn_output_contains_both_refreshable_native_pivots(self):
        rows = [
            _row("Jun-26", "Supplier Gamma", "SITE-G", "CHENNAI", "Analyst Three", 1),
            _row("Jul-26", "Supplier Gamma", "SITE-G", "CHENNAI", "Analyst Three", 2),
            _row("Jul-26", "Supplier Delta", "SITE-D", "PUNE", "Analyst Four", 3),
        ]
        dataframe = pd.DataFrame(rows, columns=MRN_COLUMNS)

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "pending-mrn.xlsx"
            write_formatted_mrn_excel(dataframe, str(path))

            with zipfile.ZipFile(path) as archive:
                names = set(archive.namelist())
            self.assertIn("xl/pivotTables/pivotTable1.xml", names)
            self.assertIn("xl/pivotTables/pivotTable2.xml", names)
            self.assertIn("xl/pivotCache/pivotCacheDefinition1.xml", names)
            self.assertIn("xl/pivotCache/pivotCacheDefinition2.xml", names)

            workbook = load_workbook(path, data_only=False)
            try:
                visible = [name for name in workbook.sheetnames if workbook[name].sheet_state == "visible"]
                self.assertEqual(visible, ["Locationwise Pivot", "Vendorwise Pivot", "Summary"])
                self.assertEqual(workbook["_PivotSrc"].sheet_state, "hidden")

                vendorwise = workbook["Vendorwise Pivot"]._pivots
                self.assertEqual(len(vendorwise), 1)
                vsource = vendorwise[0].cache.cacheSource.worksheetSource
                self.assertEqual(vsource.sheet, "Summary")
                self.assertEqual(vsource.ref, "A1:V4")
                self.assertTrue(vendorwise[0].cache.refreshOnLoad)
                self.assertTrue(vendorwise[0].cache.enableRefresh)
                self.assertEqual(vendorwise[0].cache.missingItemsLimit, 0)
                self.assertEqual(workbook["Summary"]["C2"].value, "Supplier Gamma")

                locationwise = workbook["Locationwise Pivot"]._pivots
                self.assertEqual(len(locationwise), 1)
                lsource = locationwise[0].cache.cacheSource.worksheetSource
                self.assertEqual(lsource.sheet, "_PivotSrc")
                self.assertEqual(lsource.ref, "A1:D4")
                self.assertTrue(locationwise[0].cache.refreshOnLoad)
                self.assertTrue(locationwise[0].cache.enableRefresh)
                self.assertEqual(locationwise[0].cache.missingItemsLimit, 0)
                self.assertEqual(len(locationwise[0].rowFields), 2)

                data_rows = [
                    tuple(cell.value for cell in row)
                    for row in workbook["_PivotSrc"].iter_rows(min_row=2)
                ]
                self.assertEqual(
                    set(data_rows),
                    {
                        ("CHENNAI", "Analyst Three", "Jun-26", 1),
                        ("CHENNAI", "Analyst Three", "Jul-26", 1),
                        ("PUNE", "Analyst Four", "Jul-26", 1),
                    },
                )
            finally:
                workbook.close()

            # OpenPyXL must also be able to preserve the native parts on a
            # normal re-open/re-save cycle without corrupting the package.
            resaved = Path(temp_dir) / "pending-mrn-resaved.xlsx"
            workbook = load_workbook(path)
            workbook.save(resaved)
            workbook.close()
            preserved = load_workbook(resaved)
            try:
                self.assertEqual(len(preserved["Vendorwise Pivot"]._pivots), 1)
            finally:
                preserved.close()

    def test_summary_rows_are_ordered_chronologically_by_period(self):
        # The native pivot has no live Excel session to re-sort its period
        # columns after opening (unlike the desktop app's COM-built pivot),
        # so a PivotField without an explicit AutoSort shows new items in
        # the order Excel meets them scanning the Summary sheet top-to-
        # bottom. Feed periods out of both chronological and alphabetical
        # order to prove the Summary sheet - and therefore the pivot's
        # column order on refresh - comes out chronological regardless of
        # upload order.
        rows = [
            _row("Jan-26", "Supplier A", "SITE-A", "CHENNAI", "Analyst One", 1),
            _row("Nov-25", "Supplier B", "SITE-B", "PUNE", "Analyst Two", 2),
            _row("Feb-26", "Supplier C", "SITE-C", "CHENNAI", "Analyst One", 3),
            _row("Dec-25", "Supplier D", "SITE-D", "PUNE", "Analyst Two", 4),
        ]
        dataframe = pd.DataFrame(rows, columns=MRN_COLUMNS)

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "pending-mrn-period-order.xlsx"
            write_formatted_mrn_excel(dataframe, str(path))

            workbook = load_workbook(path)
            try:
                summary = workbook["Summary"]
                periods = [summary.cell(r, 1).value for r in range(2, summary.max_row + 1)]
                self.assertEqual(periods, ["Nov-25", "Dec-25", "Jan-26", "Feb-26"])
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
