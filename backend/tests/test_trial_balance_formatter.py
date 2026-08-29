import json
import tempfile
import unittest
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.services.trial_balance_formatter import mapping_store, processor
from app.services.trial_balance_formatter.models import LedgerNature


BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_DIR = BACKEND_DIR.parent
RAW_REFERENCE = REPO_DIR / "Raw Trial Balance June 2026.xlsx"
EXPECTED_REFERENCE = REPO_DIR / "Ultrafine Trial Balance as on 30th June 2026.xlsx"
SEED_PATH = BACKEND_DIR / "seed_data" / "trial-balance-formatter-ledger-natures.json"


def _color(color):
    if color is None:
        return None
    return (
        color.type,
        color.rgb if color.type == "rgb" else None,
        color.indexed if color.type == "indexed" else None,
        color.theme if color.type == "theme" else None,
        color.tint,
    )


def _side(side):
    if side is None or side.style is None:
        return None
    return side.style, _color(side.color)


def _visual_style(cell):
    font = cell.font
    alignment = cell.alignment
    fill = cell.fill
    border = cell.border
    return {
        "font": (
            font.name,
            font.sz,
            font.b,
            font.i,
            font.u,
            font.strike,
            _color(font.color),
        ),
        "fill": (fill.fill_type, _color(fill.fgColor), _color(fill.bgColor)),
        "border": tuple(_side(side) for side in (border.left, border.right, border.top, border.bottom)),
        "alignment": (
            alignment.horizontal,
            alignment.vertical,
            alignment.text_rotation,
            alignment.wrap_text,
            alignment.shrink_to_fit,
            alignment.indent,
        ),
        "number_format": cell.number_format,
        "protection": (cell.protection.locked, cell.protection.hidden),
    }


class TrialBalanceFormatterMappingTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(bind=self.engine, tables=[LedgerNature.__table__])
        self.Session = sessionmaker(bind=self.engine, expire_on_commit=False)

    def tearDown(self):
        self.engine.dispose()

    def test_reference_seed_is_complete_additive_and_archive_safe(self):
        seed = json.loads(SEED_PATH.read_text(encoding="utf-8"))
        self.assertEqual(len(seed), 202)
        self.assertEqual(sum(bool(row["is_subgroup"]) for row in seed.values()), 6)
        self.assertEqual(seed["Advance From Customers"]["nature"], "Dr")
        self.assertEqual(seed["Sundry Creditors for Expenses"]["nature"], "Dr")

        with self.Session() as db:
            db.add(LedgerNature(
                ledger_key="CAPITAL ACCOUNT",
                ledger_name="Capital Account",
                nature="Dr",
                is_subgroup=False,
            ))
            db.add(LedgerNature(
                ledger_key="SECURED LOANS",
                ledger_name="Secured Loans",
                nature="Cr",
                is_subgroup=True,
                is_deleted=True,
            ))
            db.commit()

            self.assertEqual(mapping_store.seed_missing_from_json(db, SEED_PATH), 200)
            self.assertEqual(mapping_store.seed_missing_from_json(db, SEED_PATH), 0)
            self.assertEqual(db.query(LedgerNature).filter_by(ledger_key="CAPITAL ACCOUNT").one().nature, "Dr")
            self.assertTrue(db.query(LedgerNature).filter_by(ledger_key="SECURED LOANS").one().is_deleted)

    def test_mapping_crud_uses_soft_delete_and_explicit_restore(self):
        with self.Session() as db:
            mapping_store.set_nature(db, "Example Ledger", "Cr", True)
            row = db.query(LedgerNature).one()
            self.assertEqual(row.nature, "Cr")
            self.assertTrue(row.is_subgroup)

            self.assertTrue(mapping_store.archive_nature(db, "Example Ledger"))
            self.assertTrue(db.query(LedgerNature).one().is_deleted)
            self.assertTrue(mapping_store.restore_nature(db, "Example Ledger"))
            self.assertFalse(db.query(LedgerNature).one().is_deleted)

    def test_mapping_rename_archives_original_without_hard_deleting_it(self):
        with self.Session() as db:
            mapping_store.set_nature(db, "Old Ledger", "Dr")
            mapping_store.set_nature(
                db,
                "Renamed Ledger",
                "Cr",
                True,
                original_name="Old Ledger",
            )
            rows = db.query(LedgerNature).order_by(LedgerNature.id).all()
            self.assertEqual(len(rows), 2)
            self.assertTrue(rows[0].is_deleted)
            self.assertEqual(rows[1].ledger_name, "Renamed Ledger")
            self.assertEqual(rows[1].nature, "Cr")
            self.assertTrue(rows[1].is_subgroup)


@unittest.skipUnless(
    RAW_REFERENCE.is_file() and EXPECTED_REFERENCE.is_file(),
    "Supplied confidential parity workbooks are not committed to source control",
)
class SuppliedTrialBalanceParityTests(unittest.TestCase):
    def test_generated_visible_report_matches_supplied_reference(self):
        seed = json.loads(SEED_PATH.read_text(encoding="utf-8"))
        nature_map = {key.strip().upper(): value for key, value in seed.items()}
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "output.xlsx"
            result = processor.generate_report(
                str(RAW_REFERENCE),
                str(output),
                nature_map,
            )
            self.assertEqual(result["row_count"], 202)
            self.assertEqual(result["needs_review"], [])
            self.assertTrue(result["reference_adjustments_applied"])
            self.assertAlmostEqual(result["tb_balance"], 6_924_135.65, places=2)

            expected_formula = openpyxl.load_workbook(EXPECTED_REFERENCE, data_only=False)["June26"]
            actual_formula = openpyxl.load_workbook(output, data_only=False)["June26"]
            expected_values = openpyxl.load_workbook(EXPECTED_REFERENCE, data_only=True)["June26"]
            actual_values = openpyxl.load_workbook(output, data_only=True)["June26"]

            self.assertEqual(expected_formula.calculate_dimension(), actual_formula.calculate_dimension())
            self.assertEqual(
                {str(value) for value in expected_formula.merged_cells.ranges},
                {str(value) for value in actual_formula.merged_cells.ranges},
            )
            self.assertEqual(expected_formula.page_setup.orientation, actual_formula.page_setup.orientation)
            self.assertEqual(expected_formula.sheet_view.selection[0].sqref, actual_formula.sheet_view.selection[0].sqref)

            for column in expected_formula.column_dimensions:
                self.assertEqual(
                    expected_formula.column_dimensions[column].width,
                    actual_formula.column_dimensions[column].width,
                    column,
                )

            for row in range(1, expected_formula.max_row + 1):
                for column in range(1, expected_formula.max_column + 1):
                    coordinate = expected_formula.cell(row, column).coordinate
                    self.assertEqual(
                        expected_formula[coordinate].value,
                        actual_formula[coordinate].value,
                        coordinate,
                    )
                    self.assertEqual(
                        _visual_style(expected_formula[coordinate]),
                        _visual_style(actual_formula[coordinate]),
                        coordinate,
                    )
                    if expected_formula[coordinate].data_type == "f":
                        self.assertAlmostEqual(
                            float(expected_values[coordinate].value),
                            float(actual_values[coordinate].value),
                            places=2,
                            msg=coordinate,
                        )


if __name__ == "__main__":
    unittest.main()
