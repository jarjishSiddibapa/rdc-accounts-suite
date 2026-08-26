import tempfile
import unittest
from html.parser import HTMLParser
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.routers import unaccounted_txn
from app.services import mailer_shared


class _EmailTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []
        self._row = None
        self._cell = None

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            self._cell = []

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag):
        if tag in {"td", "th"} and self._cell is not None:
            self._row.append(" ".join("".join(self._cell).split()))
            self._cell = None
        elif tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None


class MailTemplateDefaultsTests(unittest.TestCase):
    def test_compose_defaults_select_all_reports_and_return_editable_recipients(self):
        recipients = {
            "to": ["accountsincharges@rdc.in", "accountsgroup@rdc.in"],
            "cc": ["manish.modani@rdc.in", "umesh.gawade@rdc.in"],
        }
        with patch.object(
            mailer_shared,
            "get_report_recipient_defaults",
            return_value=recipients,
        ):
            defaults = unaccounted_txn.get_mail_defaults(db=object())

        self.assertEqual(defaults["to"], recipients["to"])
        self.assertEqual(defaults["cc"], recipients["cc"])
        self.assertTrue(defaults["include_ua"])
        self.assertTrue(defaults["include_mrn"])
        self.assertTrue(defaults["include_po"])
        self.assertIn("Unaccounted Transactions", defaults["subject"])
        self.assertIn("1. Unaccounted transactions", defaults["intro"])
        self.assertIn("2. Pending MRN", defaults["intro"])
        self.assertIn("3. Uninvoiced Expenses", defaults["intro"])

    def test_subject_tracks_each_report_selection(self):
        cases = [
            ((True, False, False), "Unaccounted Transactions till Aug-26"),
            ((False, True, False), "Pending MRN till Aug-26"),
            ((False, False, True), "Uninvoiced Expense till Aug-26"),
            (
                (True, True, False),
                "Unaccounted Transactions and Pending MRN till Aug-26",
            ),
            (
                (True, False, True),
                "Unaccounted Transactions and Uninvoiced Expense till Aug-26",
            ),
            (
                (True, True, True),
                "Unaccounted Transactions, Pending MRN and Uninvoiced Expense till Aug-26",
            ),
        ]

        for selections, expected in cases:
            with self.subTest(selections=selections):
                self.assertEqual(
                    mailer_shared.build_default_subject("Aug-26", *selections),
                    expected,
                )

    def test_single_report_intro_is_not_numbered(self):
        intro = mailer_shared.build_default_intro_text(
            include_ua=False,
            include_mrn=True,
            include_po=False,
            month_mrn="Aug-26",
        )

        self.assertIn("Pending MRN till Aug-26", intro)
        self.assertNotIn("1. Pending MRN", intro)
        self.assertNotIn("Unaccounted transactions", intro)

    def test_multiple_report_intro_is_numbered_in_selection_order(self):
        intro = mailer_shared.build_default_intro_text(
            include_ua=True,
            include_mrn=False,
            include_po=True,
            month_ua="Jul-26",
            month_po="Aug-26",
        )

        self.assertIn("1. Unaccounted transactions till Jul-26", intro)
        self.assertIn("2. Uninvoiced Expenses till Aug-26", intro)
        self.assertNotIn("Pending MRN till", intro)
        self.assertIn("Rent and Land Lease have been excluded", intro)

    def test_email_tables_calculate_subtotals_without_replacing_excel_formulas(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "report.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Location", "Accounts Incharge", "Count"])
            sheet.append(["West", "A", 10])
            sheet.append(["", "B", 20])
            sheet.append(["West Total", "", "=SUBTOTAL(9,C2:C3)"])
            sheet.append(["East", "C", 5])
            sheet.append(["Grand Total", "", "=SUBTOTAL(9,C2:C5)"])
            workbook.save(workbook_path)
            workbook.close()

            _, html_body = mailer_shared.build_email_content(
                unaccounted_path=str(workbook_path),
                include_ua=True,
                include_mrn=False,
                include_po=False,
            )

            parser = _EmailTableParser()
            parser.feed(html_body)
            self.assertIn(["West Total", "", "30"], parser.rows)
            self.assertEqual(parser.rows[-1], ["Grand Total", "", "35"])

            saved = load_workbook(workbook_path, data_only=False)
            self.assertEqual(saved["Main"]["C4"].value, "=SUBTOTAL(9,C2:C3)")
            self.assertEqual(saved["Main"]["C6"].value, "=SUBTOTAL(9,C2:C5)")
            saved.close()


if __name__ == "__main__":
    unittest.main()
