"""Progress-reporting coverage for GSTR2B's _write_sheet: same gap class
as the ERP converter before its fix - the per-tab row-write loop had no
progress feedback at all, so a tab with many rows would go completely
silent between the "Writing output workbook..." and "Saved" log lines
(the only two messages logged around the entire multi-tab write). The
router's _LogQueue nudges the job's progress fraction on every log_q.put
call, so periodic writes from inside the row loop are enough to keep it
moving without any new plumbing.
"""
import unittest

import pandas as pd
from openpyxl import Workbook

from app.services.gstr2b.combiner import _write_sheet


class _LogQueue:
    def __init__(self):
        self.messages = []

    def put(self, item):
        self.messages.append(item)


class WriteSheetProgressTests(unittest.TestCase):
    def test_logs_periodic_progress_during_a_large_write(self):
        df = pd.DataFrame({"A": range(500), "B": [f"row-{i}" for i in range(500)]})
        wb = Workbook()
        ws = wb.active
        ws.title = "B2B"
        log_q = _LogQueue()

        _write_sheet(ws, df, log_q=log_q)

        progress_messages = [m for tag, m in log_q.messages if "writing row" in m]
        self.assertGreater(len(progress_messages), 1, "expected more than one progress update")
        self.assertIn("B2B", progress_messages[0])
        self.assertIn("500/500", progress_messages[-1])

    def test_without_log_q_is_unaffected(self):
        df = pd.DataFrame({"A": range(5)})
        wb = Workbook()
        ws = wb.active
        _write_sheet(ws, df)
        self.assertEqual(ws.cell(row=2, column=1).value, 0)

    def test_empty_dataframe_writes_only_the_header(self):
        df = pd.DataFrame(columns=["A", "B"])
        wb = Workbook()
        ws = wb.active
        log_q = _LogQueue()
        _write_sheet(ws, df, log_q=log_q)
        self.assertEqual(ws.cell(row=1, column=1).value, "A")
        self.assertEqual(log_q.messages, [])


if __name__ == "__main__":
    unittest.main()
