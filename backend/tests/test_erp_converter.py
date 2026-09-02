"""End-to-end coverage for the ERP HTML/XLS -> Excel converter: datatype and
formatting fidelity on the real conversion path, plus the CPU-pool progress
relay added so large conversions report real, moving progress instead of
appearing frozen (see app.jobs.run_cpu_phase / app.routers.erp_converter).
"""
import tempfile
import unittest
import zipfile
from contextlib import nullcontext
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from fastapi import HTTPException
from openpyxl import load_workbook

from app import config, jobs as jobs_module
from app.jobs import JobUserError
from app.routers.erp_converter import DownloadAllBody, _job_convert, download_all
from app.services.erp_converter import converter, xlsx_writer
from app.services.erp_converter.errors import ConversionError

_SAMPLE_HTML = """<html><head><style type="text/css">
.c0 {font-family: 'Calibri';font-size: 14.0pt;font-weight: bold;}
.c2 {font-family: 'Calibri';font-size: 9.0pt;font-weight: bold;background-color: #D9E1F2;}
.c3 {font-family: 'Calibri';font-size: 9.0pt;}
</style></head><body>
<p class="c0">Vendor Ledger Report</p>
<table class="c3" cellspacing="0">
<tr><td class="c2" colspan="2">Ledger Code</td><td class="c2">Amount</td><td class="c2">Invoice Date</td><td class="c2">TDS Note</td></tr>
<tr><td class="c3">00123</td><td class="c3"></td><td class="c3">1,234.50</td><td class="c3">31-Aug-2026</td><td class="c3">Not a date</td></tr>
<tr><td class="c3">00456</td><td class="c3"></td><td class="c3">(500.00)</td><td class="c3">01-Jan-2026</td><td class="c3">Plain text</td></tr>
<tr><td class="c3">00789</td><td class="c3"></td><td class="c3">223883.8983050808</td><td class="c3">02-Jan-2026</td><td class="c3">Precision value</td></tr>
</table>
</body></html>"""


class ErpConverterFidelityTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.in_path = Path(self.tmp.name) / "report.xls"
        self.out_path = Path(self.tmp.name) / "report.xlsx"
        self.in_path.write_text(_SAMPLE_HTML, encoding="utf-8")

    def test_html_export_is_detected_and_converted(self):
        kind = converter.convert_file(str(self.in_path), str(self.out_path))
        self.assertEqual(kind, "html")
        self.assertTrue(self.out_path.exists())

    # Row 1: title. Row 2: a blank spacer row before the table. Row 3:
    # header. Rows 4-5: data.

    def test_leading_zero_code_stays_text(self):
        converter.convert_file(str(self.in_path), str(self.out_path))
        wb = load_workbook(self.out_path)
        ws = wb["Report"]
        self.assertEqual(ws["A4"].value, "00123")
        self.assertIsInstance(ws["A4"].value, str)

    def test_amount_becomes_a_real_number_with_indian_grouping_format(self):
        converter.convert_file(str(self.in_path), str(self.out_path))
        wb = load_workbook(self.out_path)
        ws = wb["Report"]
        self.assertEqual(ws["C4"].value, 1234.5)
        self.assertIn("#", ws["C4"].number_format)

    def test_parenthesized_amount_becomes_negative_number(self):
        converter.convert_file(str(self.in_path), str(self.out_path))
        wb = load_workbook(self.out_path)
        ws = wb["Report"]
        self.assertEqual(ws["C5"].value, -500)

    def test_date_becomes_a_real_date_value(self):
        converter.convert_file(str(self.in_path), str(self.out_path))
        wb = load_workbook(self.out_path)
        ws = wb["Report"]
        self.assertEqual(ws["D4"].value.strftime("%Y-%m-%d"), "2026-08-31")

    def test_non_date_text_is_left_as_plain_text(self):
        converter.convert_file(str(self.in_path), str(self.out_path))
        wb = load_workbook(self.out_path)
        ws = wb["Report"]
        self.assertEqual(ws["E4"].value, "Not a date")

    def test_header_row_keeps_bold_font_and_fill(self):
        converter.convert_file(str(self.in_path), str(self.out_path))
        wb = load_workbook(self.out_path)
        ws = wb["Report"]
        header_cell = ws["A3"]
        self.assertTrue(header_cell.font.bold)
        self.assertIsNotNone(header_cell.fill.fgColor.rgb)

    def test_colspan_becomes_a_merged_range(self):
        converter.convert_file(str(self.in_path), str(self.out_path))
        wb = load_workbook(self.out_path)
        ws = wb["Report"]
        merged_ranges = {str(r) for r in ws.merged_cells.ranges}
        self.assertIn("A3:B3", merged_ranges)


class DirectOoxmlWriterTests(unittest.TestCase):
    """The low-allocation writer must remain interchangeable with the
    previous openpyxl writer, including the details users can see in Excel.
    """

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.in_path = Path(self.tmp.name) / "report.xls"
        self.direct_path = Path(self.tmp.name) / "direct.xlsx"
        self.compat_path = Path(self.tmp.name) / "compat.xlsx"
        self.in_path.write_text(_SAMPLE_HTML, encoding="utf-8")

    def _convert(self, output_path, mode=None):
        env = patch.dict("os.environ", {"ERP_CONVERTER_WRITER": mode}) if mode else nullcontext()
        with env:
            converter.convert_file(str(self.in_path), str(output_path))

    @staticmethod
    def _cell_signature(cell):
        return (
            cell.value,
            cell.data_type,
            cell.number_format,
            cell.font.name,
            cell.font.sz,
            cell.font.bold,
            cell.font.italic,
            cell.font.underline,
            cell.fill.fill_type,
            cell.fill.fgColor.type,
            cell.fill.fgColor.rgb,
            cell.alignment.horizontal,
            cell.alignment.vertical,
            cell.alignment.wrap_text,
            cell.border.left.style,
            cell.border.right.style,
            cell.border.top.style,
            cell.border.bottom.style,
        )

    def test_direct_output_matches_openpyxl_compatibility_output(self):
        self._convert(self.direct_path)
        self._convert(self.compat_path, "openpyxl")

        direct = load_workbook(self.direct_path)["Report"]
        compat = load_workbook(self.compat_path)["Report"]
        self.assertEqual(direct.max_row, compat.max_row)
        self.assertEqual(direct.max_column, compat.max_column)
        self.assertEqual(
            {str(value) for value in direct.merged_cells.ranges},
            {str(value) for value in compat.merged_cells.ranges},
        )
        for row in range(1, direct.max_row + 1):
            self.assertEqual(direct.row_dimensions[row].height, compat.row_dimensions[row].height)
            for column in range(1, direct.max_column + 1):
                self.assertEqual(
                    self._cell_signature(direct.cell(row, column)),
                    self._cell_signature(compat.cell(row, column)),
                    f"cell mismatch at row {row}, column {column}",
                )
        for column in range(1, direct.max_column + 1):
            letter = direct.cell(1, column).column_letter
            self.assertEqual(
                direct.column_dimensions[letter].width,
                compat.column_dimensions[letter].width,
            )

    def test_direct_package_has_valid_crc_and_required_ooxml_parts(self):
        self._convert(self.direct_path)
        with zipfile.ZipFile(self.direct_path) as archive:
            self.assertIsNone(archive.testzip())
            self.assertTrue(
                {"[Content_Types].xml", "xl/workbook.xml", "xl/styles.xml",
                 "xl/worksheets/sheet1.xml"}.issubset(archive.namelist())
            )

    def test_direct_writer_preserves_xml_characters_and_edge_whitespace(self):
        writer = xlsx_writer.DirectXlsxWriter()
        writer.write_blocks([
            xlsx_writer.TextBlock("  A & B <C>  ", {}, False, False, False)
        ])
        writer.finalize(self.direct_path)
        value = load_workbook(self.direct_path)["Report"]["A1"].value
        self.assertEqual(value, "  A & B <C>  ")

    def test_direct_writer_failure_retries_with_compatibility_writer(self):
        with (
            patch.object(
                xlsx_writer.DirectXlsxWriter,
                "finalize",
                side_effect=xlsx_writer.DirectXlsxWriterError("synthetic packaging failure"),
            ),
            self.assertLogs(converter.logger, level="ERROR") as captured,
        ):
            converter.convert_file(str(self.in_path), str(self.direct_path))
        self.assertIn("retrying with openpyxl", "\n".join(captured.output))
        workbook = load_workbook(self.direct_path)
        self.assertEqual(workbook["Report"]["A4"].value, "00123")

    def test_openpyxl_environment_override_skips_direct_writer(self):
        with (
            patch.dict("os.environ", {"ERP_CONVERTER_WRITER": "openpyxl"}),
            patch.object(
                xlsx_writer.DirectXlsxWriter,
                "__init__",
                side_effect=AssertionError("direct writer should not be instantiated"),
            ),
        ):
            converter.convert_file(str(self.in_path), str(self.compat_path))
        self.assertTrue(self.compat_path.is_file())


class ExcelSheetLimitTests(unittest.TestCase):
    """write-only mode has no idea Excel caps a sheet at 1,048,576 rows /
    16,384 columns and will happily keep writing past that, producing a
    file Excel then refuses to open cleanly. XlsxWriter checks as it goes
    so this fails fast with an actionable message instead."""

    def test_row_limit_raises_a_clear_conversion_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            in_path = Path(tmp) / "report.xls"
            in_path.write_text(_SAMPLE_HTML, encoding="utf-8")
            with patch.object(xlsx_writer, "_MAX_EXCEL_ROWS", 3):
                with self.assertRaises(ConversionError) as ctx:
                    converter.convert_file(str(in_path), str(Path(tmp) / "out.xlsx"))
            self.assertIn("exceeds Excel's per-sheet limit", str(ctx.exception))

    def test_job_convert_translates_conversion_error_into_job_user_error(self):
        with patch(
            "app.routers.erp_converter.jobs.run_cpu_phase",
            side_effect=ConversionError("This report converts to too many rows."),
        ):
            with self.assertRaises(JobUserError) as ctx:
                _job_convert("in.xls", "out.xlsx", "report.xls")
        self.assertIn("too many rows", str(ctx.exception))


class DownloadAllStreamsFromDiskTests(unittest.TestCase):
    """/download-all used to buffer the whole combined zip of up to 100
    converted workbooks in an in-memory BytesIO, then copy it again via
    .getvalue() - the largest realistic memory exposure found in this
    converter. It now writes straight to a scratch file and streams that
    via FileResponse, with a background task cleaning it up afterward."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.user = SimpleNamespace(id=1)

    def _fake_job(self, output_path, original_filename):
        return {
            "status": "done",
            "result": {"output_path": str(output_path), "original_filename": original_filename},
        }

    def test_bundles_two_outputs_into_a_zip_file_on_disk(self):
        out1 = Path(self.tmp.name) / "a.xlsx"
        out2 = Path(self.tmp.name) / "b.xlsx"
        out1.write_bytes(b"fake-xlsx-1")
        out2.write_bytes(b"fake-xlsx-2")
        jobs_by_id = {
            "job1": self._fake_job(out1, "Report A.xls"),
            "job2": self._fake_job(out2, "Report B.xls"),
        }
        with patch(
            "app.routers.erp_converter.jobs.get_job",
            side_effect=lambda job_id, owner_id: jobs_by_id[job_id],
        ):
            response = download_all(DownloadAllBody(job_ids=["job1", "job2"]), user=self.user)

        zip_path = Path(response.path)
        self.addCleanup(lambda: zip_path.unlink(missing_ok=True))
        self.assertTrue(zip_path.is_file(), "zip should be written to a real scratch file, not BytesIO")
        with zipfile.ZipFile(zip_path) as archive:
            names = set(archive.namelist())
            self.assertEqual(names, {"Report A.xlsx", "Report B.xlsx"})
            self.assertEqual(archive.read("Report A.xlsx"), b"fake-xlsx-1")
        self.assertIsNotNone(response.background, "should clean up the scratch zip after sending it")

    def test_error_mid_bundle_cleans_up_the_partial_zip(self):
        out1 = Path(self.tmp.name) / "a.xlsx"
        out1.write_bytes(b"fake-xlsx-1")
        jobs_by_id = {
            "job1": self._fake_job(out1, "Report A.xls"),
            "job2": None,
        }
        written_paths = []
        real_get_job = jobs_by_id.get

        def fake_get_job(job_id, owner_id):
            return real_get_job(job_id)

        with patch("app.routers.erp_converter.jobs.get_job", side_effect=fake_get_job):
            with self.assertRaises(HTTPException):
                download_all(DownloadAllBody(job_ids=["job1", "job2"]), user=self.user)

        leftover = list(config.SCRATCH_DIR.glob("*_ERP_Excel_Conversions.zip"))
        self.assertEqual(leftover, [], "a failed bundle should not leave a partial zip behind")


def _toy_cpu_task(steps: int, progress_cb=None) -> str:
    """Module-level (picklable) stand-in for a real CPU-phase job body -
    reports several distinct progress updates like convert_file does, each
    spaced out enough for run_cpu_phase's 0.25s poll loop to observe more
    than just the final one."""
    import time

    for i in range(steps):
        if progress_cb:
            progress_cb((i + 1) / steps, f"step {i + 1}/{steps}")
        time.sleep(0.12)
    return "done"


class RunCpuPhaseProgressRelayTests(unittest.TestCase):
    """app.jobs.run_cpu_phase relays a CPU-pool worker's own progress_cb
    calls back to the caller via a cross-process queue - this is what makes
    large ERP conversions show real, moving progress instead of sitting at
    a single fixed percentage for the whole call (the originally reported
    "stuck at 2%" symptom)."""

    def test_progress_crosses_the_process_boundary(self):
        events = []
        result = jobs_module.run_cpu_phase(
            _toy_cpu_task, 8, progress_cb=lambda frac, phase: events.append((frac, phase))
        )
        self.assertEqual(result, "done")
        self.assertGreaterEqual(len(events), 2, "expected more than one distinct progress update")
        self.assertEqual(events[-1][0], 1.0)
        fractions = [e[0] for e in events]
        self.assertEqual(fractions, sorted(fractions))

    def test_omitting_progress_cb_is_unaffected(self):
        result = jobs_module.run_cpu_phase(_toy_cpu_task, 3)
        self.assertEqual(result, "done")


if __name__ == "__main__":
    unittest.main()
